import io
import json
from decimal import Decimal
from datetime import date, timedelta, time
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth.models import User

import openpyxl

from django.core.cache import cache

from scheduling.models import Agent, Five9Profile, AgentSeparation, Shift
from adherence.models import AdherenceRecord, DailyUpload, DailyAgentHours, Coding, PayrollAdjustment
from finance.models import BillingSettings
from finance.views import _split_agent_display_name


_WEEK_START = date(2025, 1, 6)
_WEEK = [_WEEK_START + timedelta(days=i) for i in range(7)]


def _make_agent(username, role='agent', role_type='agent'):
    user = User.objects.create_user(username, password='x')
    return Agent.objects.create(
        user=user, role=role, role_type=role_type,
        agent_name=username, status='active', track_attendance=True,
    )


def _settings(**overrides):
    obj, _ = BillingSettings.objects.get_or_create(pk=1)
    for k, v in overrides.items():
        setattr(obj, k, v)
    obj.save()
    return obj


def _add_hours(agent, login_seconds, nr_seconds, week_day=0):
    upload, _ = DailyUpload.objects.get_or_create(date=_WEEK[week_day])
    DailyAgentHours.objects.create(
        upload=upload, agent=agent,
        five9_username=agent.agent_name,
        login_seconds=login_seconds,
        not_ready_seconds=nr_seconds,
    )


def _status(agent, week_day, status):
    return AdherenceRecord.objects.create(agent=agent, date=_WEEK[week_day], status=status)


class NRCapCheck1Tests(TestCase):
    """check1: deduct NR hours above the absolute weekly cap."""

    def setUp(self):
        self.agent = _make_agent('check1_agent')
        # nr_ratio_max_hours=200 ensures check2 is disabled (pre_total always < 200)
        # Actually set it low so check2 doesn't trigger when login >> 48
        self.s = _settings(
            nr_cap_regular_hours=Decimal('6.00'),
            nr_ratio=Decimal('0.1250'),
            nr_ratio_max_hours=Decimal('200.00'),
        )

    def _call(self):
        from finance.views import _get_billable_weekly_data
        return _get_billable_weekly_data([self.agent], _WEEK, self.s)

    def test_nr_above_cap_triggers_check1(self):
        # 100 h login, 8 h NR → check1=2, check2=8-12.5=0 → deduction=2
        _add_hours(self.agent, 100 * 3600, 8 * 3600)
        result = self._call()[self.agent.pk]
        self.assertAlmostEqual(float(result['final_hrs']), 100.0 - 2.0, places=2)

    def test_nr_at_cap_no_check1(self):
        # 100 h login, 6 h NR → check1=0, check2=6-12.5=0 → no deduction
        _add_hours(self.agent, 100 * 3600, 6 * 3600)
        result = self._call()[self.agent.pk]
        self.assertAlmostEqual(float(result['final_hrs']), 100.0, places=2)


class NRCapCheck2Tests(TestCase):
    """check2: deduct NR hours above 12.5% of login time."""

    def setUp(self):
        self.agent = _make_agent('check2_agent')
        # cap=99 so check1 never triggers; ratio=0.125; max_hours=48
        self.s = _settings(
            nr_cap_regular_hours=Decimal('99.00'),
            nr_ratio=Decimal('0.1250'),
            nr_ratio_max_hours=Decimal('48.00'),
        )

    def _call(self):
        from finance.views import _get_billable_weekly_data
        return _get_billable_weekly_data([self.agent], _WEEK, self.s)

    def test_nr_above_ratio_triggers_check2(self):
        # 20 h login, 6 h NR → allowance=20*0.125=2.5 → excess=3.5
        _add_hours(self.agent, 20 * 3600, 6 * 3600)
        result = self._call()[self.agent.pk]
        # final = 20 - 3.5 = 16.5
        self.assertAlmostEqual(float(result['final_hrs']), 16.5, places=2)

    def test_nr_within_ratio_no_deduction(self):
        # 40 h login, 4 h NR → allowance=40*0.125=5 → no excess
        _add_hours(self.agent, 40 * 3600, 4 * 3600)
        result = self._call()[self.agent.pk]
        self.assertAlmostEqual(float(result['final_hrs']), 40.0, places=2)

    def test_check2_disabled_above_max_hours(self):
        # login=50 h (above nr_ratio_max_hours=48) → check2 disabled
        # cap=99 so check1 also won't fire → no deduction at all
        _add_hours(self.agent, 50 * 3600, 10 * 3600)
        result = self._call()[self.agent.pk]
        self.assertAlmostEqual(float(result['final_hrs']), 50.0, places=2)


class NRCapMaxOfTwoTests(TestCase):
    """The larger of check1 and check2 is applied, never both."""

    def setUp(self):
        self.agent = _make_agent('max_test')
        self.s = _settings(
            nr_cap_regular_hours=Decimal('6.00'),
            nr_ratio=Decimal('0.1250'),
            nr_ratio_max_hours=Decimal('48.00'),
        )

    def _call(self):
        from finance.views import _get_billable_weekly_data
        return _get_billable_weekly_data([self.agent], _WEEK, self.s)

    def test_check2_wins_when_login_low(self):
        # 20 h login, 8 h NR
        # check1 = max(0, 8-6) = 2
        # check2 = max(0, 8 - 20*0.125) = max(0, 8-2.5) = 5.5
        # max = 5.5 → final = 20 - 5.5 = 14.5
        _add_hours(self.agent, 20 * 3600, 8 * 3600)
        result = self._call()[self.agent.pk]
        self.assertAlmostEqual(float(result['final_hrs']), 14.5, places=2)

    def test_check1_wins_when_login_high(self):
        # 100 h login (above max_hours=48 → check2 disabled), 8 h NR
        # check1 = 2, check2 = 0 (disabled) → final = 100 - 2 = 98
        _add_hours(self.agent, 100 * 3600, 8 * 3600)
        result = self._call()[self.agent.pk]
        self.assertAlmostEqual(float(result['final_hrs']), 98.0, places=2)

    def test_final_hours_never_negative(self):
        # Pathological: 5 h login, 100 h NR
        _add_hours(self.agent, 5 * 3600, 100 * 3600)
        result = self._call()[self.agent.pk]
        self.assertGreaterEqual(float(result['final_hrs']), 0.0)


class VTOWeeklyAllowanceTests(TestCase):
    """Any VTO-type status (VTO/P+VTO/T+VTO) in the week raises the weekly
    NR allowance to the flat cap and skips the 12.5% ratio check entirely."""

    def setUp(self):
        self.s = _settings(
            nr_cap_regular_hours=Decimal('6.00'),
            nr_cap_kill_team_hours=Decimal('7.00'),
            nr_ratio=Decimal('0.1250'),
            nr_ratio_max_hours=Decimal('48.00'),
        )

    def _call(self, agent):
        from finance.views import _get_billable_weekly_data
        return _get_billable_weekly_data([agent], _WEEK, self.s)

    def test_vto_exactly_at_cap_no_deduction(self):
        # 20 h login, 6 h NR, VTO present → deduction 0 (flat cap, not ratio)
        agent = _make_agent('vto_at_cap')
        _add_hours(agent, 20 * 3600, 6 * 3600)
        _status(agent, 0, 'VTO')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_deduction']), 0.0, places=2)
        self.assertAlmostEqual(float(result['final_hrs']), 20.0, places=2)

    def test_vto_above_cap_deducts_excess_only(self):
        # 20 h login, 6 h 30 m NR, VTO present → deduction 30 m
        agent = _make_agent('vto_above_cap')
        _add_hours(agent, 20 * 3600, int(6.5 * 3600))
        _status(agent, 0, 'VTO')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_deduction']), 0.5, places=2)
        self.assertAlmostEqual(float(result['final_hrs']), 19.5, places=2)

    def test_kill_team_vto_uses_7h_cap(self):
        agent = _make_agent('vto_kill_team', role_type='kill_team')
        _add_hours(agent, 20 * 3600, int(7.5 * 3600))
        _status(agent, 0, 'VTO')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_cap_hrs']), 7.0, places=2)
        self.assertAlmostEqual(float(result['nr_deduction']), 0.5, places=2)

    def test_p_vto_triggers_flat_cap(self):
        agent = _make_agent('p_vto_test')
        _add_hours(agent, 20 * 3600, int(6.5 * 3600))
        _status(agent, 0, 'P+VTO')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_deduction']), 0.5, places=2)

    def test_t_vto_triggers_flat_cap(self):
        agent = _make_agent('t_vto_test')
        _add_hours(agent, 20 * 3600, int(6.5 * 3600))
        _status(agent, 0, 'T+VTO')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_deduction']), 0.5, places=2)

    def test_no_vto_ratio_path_unchanged(self):
        # Regression: same shape as NRCapCheck2Tests.test_nr_above_ratio_triggers_check2 —
        # 20 h login, 6 h NR, no VTO → ratio check still binds (unaffected by this change)
        agent = _make_agent('no_vto_test')
        _add_hours(agent, 20 * 3600, 6 * 3600)
        _status(agent, 0, 'P')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_deduction']), 3.5, places=2)
        self.assertAlmostEqual(float(result['final_hrs']), 16.5, places=2)

    def test_nr_allowed_hrs_is_flat_cap_when_vto(self):
        agent = _make_agent('allowed_vto_test')
        _add_hours(agent, 20 * 3600, int(6.5 * 3600))
        _status(agent, 0, 'VTO')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_allowed_hrs']), 6.0, places=2)

    def test_nr_allowed_hrs_is_ratio_when_no_vto(self):
        # min(cap=6, login*ratio=20*0.125=2.5) = 2.5 — what Billing Report v2 displays
        agent = _make_agent('allowed_no_vto_test')
        _add_hours(agent, 20 * 3600, 6 * 3600)
        _status(agent, 0, 'P')
        result = self._call(agent)[agent.pk]
        self.assertAlmostEqual(float(result['nr_allowed_hrs']), 2.5, places=2)


XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

HEADERS = [
    'Username', 'ID', 'LEGAL NAMES', 'Supervisor',
    'Mon Total', 'Tues Total', 'Wed Total', 'Thur Total', 'Fri Total', 'Sat Total', 'Sun Total',
    None,
    'Total', 'Total Decimal',
]


class CodingsExportTests(TestCase):
    """Export Codings — one row per agent, Mon-Sun day totals + a weekly total."""

    def setUp(self):
        # role='admin'/role_type='supervisor' keeps AgentAccessMiddleware from
        # treating this login as a portal agent restricted to /agent/ paths.
        self.boss = _make_agent('codeboss', role='admin', role_type='supervisor')
        self.boss.is_super_admin = True
        self.boss.save()
        self.client.login(username='codeboss', password='x')

    def _export_ws(self, week_start=_WEEK_START):
        resp = self.client.get(reverse('codings_export') + f'?week={week_start.isoformat()}')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        return openpyxl.load_workbook(io.BytesIO(resp.content)).active

    def _rows(self, ws):
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def _row_index_for(self, ws, agent):
        legal_name = agent.user.get_full_name() or agent.agent_name
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=3).value == legal_name:
                return r
        raise AssertionError(f"No row found for {legal_name!r}")

    def _row_for(self, ws, agent):
        return self._rows(ws)[self._row_index_for(ws, agent) - 1]

    def test_header_row(self):
        ws = self._export_ws()
        self.assertEqual(self._rows(ws)[0], HEADERS)

    def test_filename_and_content_type(self):
        resp = self.client.get(reverse('codings_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        self.assertIn(f'codings_{_WEEK_START.isoformat()}.xlsx', resp['Content-Disposition'])

    def test_daily_and_weekly_totals(self):
        agent = _make_agent('dow1')
        # Monday: regular (2h) + admin (1h) on the SAME day -> summed into one 3h total
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(11, 0),
                               is_admin_coding=False)
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(13, 0), end_time=time(14, 0),
                               is_admin_coding=True)
        # Wednesday: 0.5h regular
        Coding.objects.create(agent=agent, date=_WEEK_START + timedelta(days=2), start_time=time(9, 0),
                               end_time=time(9, 30), is_admin_coding=False)
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(row[4], time(3, 0, 0))    # Mon Total
        self.assertEqual(row[5], time(0, 0, 0))    # Tues Total
        self.assertEqual(row[6], time(0, 30, 0))   # Wed Total
        self.assertEqual(row[7], time(0, 0, 0))    # Thur Total
        self.assertEqual(row[8], time(0, 0, 0))    # Fri Total
        self.assertEqual(row[9], time(0, 0, 0))    # Sat Total
        self.assertEqual(row[10], time(0, 0, 0))   # Sun Total
        self.assertEqual(row[12], timedelta(hours=3, minutes=30))  # Total
        self.assertEqual(row[13], 3.5)              # Total Decimal

    def test_agent_with_no_codings_is_zero_filled(self):
        agent = _make_agent('zerofill1')
        row = self._row_for(self._export_ws(), agent)
        for i in range(4, 11):
            self.assertEqual(row[i], time(0, 0, 0))
        self.assertEqual(row[12], timedelta(0))
        self.assertEqual(row[13], 0.0)

    def test_weekly_total_over_24_hours(self):
        agent = _make_agent('over24')
        for i, day_date in enumerate(_WEEK):
            Coding.objects.create(agent=agent, date=day_date, start_time=time(8, 0), end_time=time(12, 0),
                                   is_admin_coding=(i % 2 == 0))  # 4h/day * 7 days = 28h
        ws = self._export_ws()
        r = self._row_index_for(ws, agent)
        self.assertEqual(ws.cell(row=r, column=13).value, timedelta(hours=28))
        self.assertEqual(ws.cell(row=r, column=13).number_format, '[h]:mm:ss')
        self.assertEqual(ws.cell(row=r, column=14).value, 28.0)

    def test_time_cells_are_real_duration_values_not_text(self):
        agent = _make_agent('durfmt1')
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 30),
                               is_admin_coding=False)
        ws = self._export_ws()
        r = self._row_index_for(ws, agent)
        for col in range(5, 12):
            self.assertEqual(ws.cell(row=r, column=col).number_format, 'h:mm:ss')
        self.assertIsInstance(ws.cell(row=r, column=5).value, time)
        self.assertEqual(ws.cell(row=r, column=13).number_format, '[h]:mm:ss')
        self.assertIsInstance(ws.cell(row=r, column=13).value, timedelta)
        self.assertNotEqual(ws.cell(row=r, column=14).number_format, '@')
        self.assertIsInstance(ws.cell(row=r, column=14).value, float)

    def test_employee_id_and_username_populate_and_blank(self):
        agent = _make_agent('idtest1')
        agent.employee_id = 'EMP-500'
        agent.save()
        Five9Profile.objects.create(agent=agent, five9_username='idtest1.f9', is_primary=True)
        agent2 = _make_agent('idtest2')  # no employee_id, no Five9 profile
        ws = self._export_ws()
        row1 = self._row_for(ws, agent)
        row2 = self._row_for(ws, agent2)
        self.assertEqual(row1[1], 'EMP-500')
        self.assertEqual(row1[0], 'idtest1.f9')
        # openpyxl round-trips a written '' as a blank cell (None) — both mean "blank"
        self.assertFalse(row2[1])
        self.assertFalse(row2[0])

    def test_coding_outside_week_excluded_from_totals(self):
        agent = _make_agent('outside1')
        Coding.objects.create(agent=agent, date=_WEEK_START - timedelta(days=1), start_time=time(9, 0),
                               end_time=time(10, 0), is_admin_coding=False)
        Coding.objects.create(agent=agent, date=_WEEK_START + timedelta(days=7), start_time=time(9, 0),
                               end_time=time(10, 0), is_admin_coding=False)
        row = self._row_for(self._export_ws(), agent)  # row exists — zero-fill roster
        for i in range(4, 11):
            self.assertEqual(row[i], time(0, 0, 0))
        self.assertEqual(row[12], timedelta(0))

    def test_non_super_admin_denied(self):
        self.client.logout()
        _make_agent('plainadmin', role='admin', role_type='supervisor')
        self.client.login(username='plainadmin', password='x')
        resp = self.client.get(reverse('codings_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 302)

    def test_unauthenticated_denied(self):
        self.client.logout()
        resp = self.client.get(reverse('codings_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 302)

    def test_pay_window_inclusion(self):
        agent = _make_agent('payin1')
        agent.status = 'inactive'
        agent.save()
        AgentSeparation.objects.create(
            agent=agent, status='finalized', separation_type='quit',
            last_day_worked=_WEEK_START - timedelta(days=3),
            remove_from_adherence_date=_WEEK_START + timedelta(days=1),
        )
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
                               is_admin_coding=False)
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(row[4], time(1, 0, 0))

    def test_pay_window_exclusion(self):
        agent = _make_agent('payout1')
        agent.status = 'inactive'
        agent.save()
        AgentSeparation.objects.create(
            agent=agent, status='finalized', separation_type='quit',
            last_day_worked=_WEEK_START - timedelta(days=10),
            remove_from_adherence_date=_WEEK_START,  # not > week_start -> already passed
        )
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
                               is_admin_coding=False)
        rows = self._rows(self._export_ws())
        names = [r[2] for r in rows[1:]]
        self.assertNotIn('payout1', names)

    def test_sort_by_supervisor_then_legal_name(self):
        sup_a = _make_agent('supa', role='admin', role_type='supervisor')
        sup_a.agent_name = 'Sup Alpha'
        sup_a.save()
        sup_b = _make_agent('supb', role='admin', role_type='supervisor')
        sup_b.agent_name = 'Sup Bravo'
        sup_b.save()

        under_a_z = _make_agent('undera_z')
        under_a_z.agent_name = 'Zeta Under A'
        under_a_z.supervisor = sup_a
        under_a_z.save()
        under_a_a = _make_agent('undera_a')
        under_a_a.agent_name = 'Alpha Under A'
        under_a_a.supervisor = sup_a
        under_a_a.save()
        under_b = _make_agent('underb')
        under_b.agent_name = 'Someone Under B'
        under_b.supervisor = sup_b
        under_b.save()

        rows = self._rows(self._export_ws())[1:]
        names_under_a = [r[2] for r in rows if r[3] == 'Sup Alpha']
        self.assertEqual(names_under_a, ['Alpha Under A', 'Zeta Under A'])
        idx_a = next(i for i, r in enumerate(rows) if r[3] == 'Sup Alpha')
        idx_b = next(i for i, r in enumerate(rows) if r[3] == 'Sup Bravo')
        self.assertLess(idx_a, idx_b)

    def test_lcc_employer_agent_excluded(self):
        lcc_agent = _make_agent('lccworker1')
        lcc_agent.employer = 'LCC'
        lcc_agent.save()
        infinity_agent = _make_agent('infworker1')

        ws = self._export_ws()
        with self.assertRaises(AssertionError):
            self._row_index_for(ws, lcc_agent)
        self._row_index_for(ws, infinity_agent)  # still present


V2_HEADERS = [
    'AGENT/ADMIN user name', 'AGENT FIRST NAME', 'AGENT LAST NAME',
    'LOGIN TIME', 'NOT READY TIME', 'Coded time', 'Total connected time',
    'Allowed Not Ready', 'Time that should be deducted for going over NR Allowed',
    'Total work time after deduction of Not Ready in Decimal',
    'Total work time after deduction of Not Ready',
]


def _as_timedelta(value):
    """Normalize an openpyxl round-tripped duration cell (time, timedelta, or a bare
    numeric zero — openpyxl round-trips a zero-valued duration cell as plain 0) to a timedelta."""
    if isinstance(value, timedelta):
        return value
    if isinstance(value, time):
        return timedelta(hours=value.hour, minutes=value.minute, seconds=value.second)
    if isinstance(value, (int, float)):
        return timedelta(days=float(value))
    raise AssertionError(f"expected a duration-like value, got {value!r}")


class BillingV2ExportTests(TestCase):
    """Billing Report v2 — one row per agent, weekly payroll matrix reusing the exact
    hours-engine NR/final-hours math. Purely additive; must not touch billing_report/export."""

    def setUp(self):
        self.boss = _make_agent('v2boss', role='admin', role_type='supervisor')
        self.boss.is_super_admin = True
        self.boss.save()
        self.client.login(username='v2boss', password='x')
        _settings(
            nr_cap_regular_hours=Decimal('6.00'),
            nr_cap_kill_team_hours=Decimal('7.00'),
            nr_ratio=Decimal('0.1250'),
            nr_ratio_max_hours=Decimal('48.00'),
        )

    def _export_ws(self, week_start=_WEEK_START):
        resp = self.client.get(reverse('billing_export_v2') + f'?week={week_start.isoformat()}')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        return openpyxl.load_workbook(io.BytesIO(resp.content)).active

    def _rows(self, ws):
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def _row_index_for(self, ws, agent):
        first, last = _split_agent_display_name(agent.agent_name)
        for r in range(2, ws.max_row + 1):
            # openpyxl round-trips a written '' as a blank cell (None) — normalize both.
            if (ws.cell(row=r, column=2).value or '') == first and (ws.cell(row=r, column=3).value or '') == last:
                return r
        raise AssertionError(f"No row found for {agent.agent_name!r}")

    def _row_for(self, ws, agent):
        return self._rows(ws)[self._row_index_for(ws, agent) - 1]

    def test_header_row(self):
        ws = self._export_ws()
        self.assertEqual(self._rows(ws)[0], V2_HEADERS)
        # A not bold, B-K bold
        self.assertFalse(ws.cell(row=1, column=1).font.bold)
        for col in range(2, 12):
            self.assertTrue(ws.cell(row=1, column=col).font.bold)

    def test_filename_and_content_type(self):
        resp = self.client.get(reverse('billing_export_v2') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        self.assertIn(f'billing_v2_{_WEEK_START.isoformat()}.xlsx', resp['Content-Disposition'])

    def test_g_k_j_identities_for_real_hours(self):
        # 44h login, 3h coded, 7h NR, regular agent -> matches the worked example.
        agent = _make_agent('identity1')
        _add_hours(agent, 44 * 3600, 7 * 3600)
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(12, 0))
        row = self._row_for(self._export_ws(), agent)
        d = _as_timedelta(row[3])
        f = _as_timedelta(row[5])
        g = _as_timedelta(row[6])
        i = _as_timedelta(row[8])
        k = _as_timedelta(row[10])
        j = row[9]
        self.assertEqual(g, d + f)
        self.assertEqual(k, g - i)
        self.assertAlmostEqual(j, k.total_seconds() / 3600, places=2)
        self.assertAlmostEqual(k.total_seconds() / 3600, 45.5, places=2)

    def test_allowed_nr_regular_vs_kill_team_cap(self):
        # pre_total=50h (>48h ratio ceiling) -> H collapses to the absolute cap exactly.
        regular = _make_agent('capregular', role_type='regular_agent')
        kill = _make_agent('capkill', role_type='kill_team')
        _add_hours(regular, 50 * 3600, 10 * 3600)
        _add_hours(kill, 50 * 3600, 10 * 3600)
        ws = self._export_ws()
        row_r = self._row_for(ws, regular)
        row_k = self._row_for(ws, kill)
        self.assertEqual(_as_timedelta(row_r[7]), timedelta(hours=6))   # Allowed NR
        self.assertEqual(_as_timedelta(row_k[7]), timedelta(hours=7))   # kill team +1h
        self.assertEqual(_as_timedelta(row_r[8]), timedelta(hours=4))   # deduction 10-6
        self.assertEqual(_as_timedelta(row_k[8]), timedelta(hours=3))   # deduction 10-7
        self.assertEqual(_as_timedelta(row_r[10]), timedelta(hours=46))
        self.assertEqual(_as_timedelta(row_k[10]), timedelta(hours=47))

    def test_nr_under_allowance_zero_deduction(self):
        agent = _make_agent('underallow1')
        _add_hours(agent, 40 * 3600, 2 * 3600)  # allowance = min(6, 40*.125=5) = 5h > 2h NR
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(_as_timedelta(row[8]), timedelta(0))
        self.assertEqual(_as_timedelta(row[10]), _as_timedelta(row[6]))  # K == G

    def test_nr_over_allowance_deduction_reduces_final(self):
        agent = _make_agent('overallow1')
        _add_hours(agent, 20 * 3600, 8 * 3600)  # allowance = min(6, 2.5) = 2.5h; NR=8h
        row = self._row_for(self._export_ws(), agent)
        allowed = _as_timedelta(row[7])
        ded = _as_timedelta(row[8])
        nr = _as_timedelta(row[4])
        self.assertEqual(allowed, timedelta(hours=2, minutes=30))
        self.assertEqual(ded, nr - allowed)
        self.assertEqual(_as_timedelta(row[10]), timedelta(hours=14, minutes=30))

    def test_zero_fill_for_agent_with_no_hours(self):
        agent = _make_agent('zerofillv2')
        row = self._row_for(self._export_ws(), agent)
        for col in (3, 4, 5, 6, 8, 9):
            self.assertEqual(_as_timedelta(row[col]), timedelta(0))
        self.assertEqual(row[9], 0.0)

    def test_admin_with_no_login_shows_coded_as_connected_time(self):
        agent = _make_agent('adminnologin', role='admin', role_type='supervisor')
        Coding.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(14, 0))
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(_as_timedelta(row[3]), timedelta(0))            # login 0
        self.assertEqual(_as_timedelta(row[5]), timedelta(hours=5))     # coded 5h
        self.assertEqual(_as_timedelta(row[6]), timedelta(hours=5))     # connected == coded

    def test_durations_are_real_values_not_text_and_j_is_number(self):
        agent = _make_agent('durfmtv2')
        _add_hours(agent, 10 * 3600, 1 * 3600)
        ws = self._export_ws()
        r = self._row_index_for(ws, agent)
        self.assertEqual(ws.cell(row=r, column=4).number_format, '[h]:mm:ss')
        self.assertEqual(ws.cell(row=r, column=5).number_format, 'h:mm:ss')
        self.assertEqual(ws.cell(row=r, column=6).number_format, '[h]:mm:ss;@')
        self.assertEqual(ws.cell(row=r, column=7).number_format, '[h]:mm:ss;@')
        self.assertEqual(ws.cell(row=r, column=8).number_format, '[h]:mm:ss')
        self.assertEqual(ws.cell(row=r, column=9).number_format, '[h]:mm:ss')
        self.assertEqual(ws.cell(row=r, column=10).number_format, '0.00')
        self.assertEqual(ws.cell(row=r, column=11).number_format, '[h]:mm:ss')
        for col in (4, 5, 6, 7, 8, 9, 11):
            self.assertNotIsInstance(ws.cell(row=r, column=col).value, str)
        j_value = ws.cell(row=r, column=10).value
        self.assertIsInstance(j_value, (int, float))
        self.assertNotIsInstance(j_value, bool)

    def test_name_split_parenthetical_middle_name(self):
        agent = _make_agent('nametest1')
        agent.agent_name = 'David (IB) Green'
        agent.save()
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(row[1], 'David (IB)')
        self.assertEqual(row[2], 'Green')

    def test_name_split_single_word(self):
        agent = _make_agent('nametest2')
        agent.agent_name = 'Cher'
        agent.save()
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(row[1], 'Cher')
        self.assertEqual(row[2] or '', '')  # openpyxl round-trips a written '' as blank (None)

    def test_pay_window_inclusion(self):
        agent = _make_agent('v2payin1')
        agent.status = 'inactive'
        agent.save()
        AgentSeparation.objects.create(
            agent=agent, status='finalized', separation_type='quit',
            last_day_worked=_WEEK_START - timedelta(days=3),
            remove_from_adherence_date=_WEEK_START + timedelta(days=1),
        )
        _add_hours(agent, 5 * 3600, 0)
        row = self._row_for(self._export_ws(), agent)
        self.assertEqual(_as_timedelta(row[3]), timedelta(hours=5))

    def test_pay_window_exclusion(self):
        agent = _make_agent('v2payout1')
        agent.status = 'inactive'
        agent.save()
        AgentSeparation.objects.create(
            agent=agent, status='finalized', separation_type='quit',
            last_day_worked=_WEEK_START - timedelta(days=10),
            remove_from_adherence_date=_WEEK_START,  # not > week_start -> already passed
        )
        rows = self._rows(self._export_ws())
        names = [(r[1], r[2]) for r in rows[1:]]
        first, last = _split_agent_display_name(agent.agent_name)
        self.assertNotIn((first, last), names)

    def test_non_super_admin_denied(self):
        self.client.logout()
        _make_agent('v2plainadmin', role='admin', role_type='supervisor')
        self.client.login(username='v2plainadmin', password='x')
        resp = self.client.get(reverse('billing_export_v2') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 302)

    def test_unauthenticated_denied(self):
        self.client.logout()
        resp = self.client.get(reverse('billing_export_v2') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 302)

    def test_existing_billing_report_and_export_untouched(self):
        # Regression guard: the original billing report/export must still work exactly
        # as before — same URL names, same content type, same header columns.
        resp_report = self.client.get(reverse('billing_report') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp_report.status_code, 200)

        resp_export = self.client.get(reverse('billing_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp_export.status_code, 200)
        self.assertEqual(resp_export['Content-Type'], XLSX_MIME)
        self.assertIn(f'billing_{_WEEK_START.isoformat()}.xlsx', resp_export['Content-Disposition'])
        ws = openpyxl.load_workbook(io.BytesIO(resp_export.content)).active
        self.assertEqual(ws.title, "Billing")
        header_row = [c.value for c in ws[3]]
        self.assertEqual(header_row, [
            'Agent Name', 'Legal Name', 'Employee ID', 'Five9 Username (Billable)',
            'Supervisor', 'Agent Type', 'Employer',
            'Worked Hrs (Final)', 'Billing Rate (USD)', 'Total Billing (USD)',
        ])

    def test_lcc_employer_agent_excluded(self):
        lcc_agent = _make_agent('lccworker2')
        lcc_agent.employer = 'LCC'
        lcc_agent.save()
        infinity_agent = _make_agent('infworker2')

        ws = self._export_ws()
        with self.assertRaises(AssertionError):
            self._row_index_for(ws, lcc_agent)
        self._row_index_for(ws, infinity_agent)  # still present


class AdminTabsAccessTests(TestCase):
    """
    Part 2: can_access_admin_tabs gates Admin Codings and Admin Adherence
    (now top-level tabs, previously Finance-only). Super admins keep implicit
    access; a plain holder of the flag gets access; anyone else is denied
    server-side, not just hidden from nav.
    """

    def test_holder_can_open_both_tabs(self):
        holder = _make_agent('tabsholder', role='admin', role_type='supervisor')
        holder.can_access_admin_tabs = True
        holder.save()
        self.client.login(username='tabsholder', password='x')
        resp1 = self.client.get(reverse('admin_codings') + f'?week={_WEEK_START.isoformat()}')
        resp2 = self.client.get(reverse('admin_adherence') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp1.status_code, 200)
        self.assertEqual(resp2.status_code, 200)

    def test_non_holder_denied_on_both_tabs(self):
        _make_agent('tabsnoholder', role='admin', role_type='supervisor')
        self.client.login(username='tabsnoholder', password='x')
        resp1 = self.client.get(reverse('admin_codings') + f'?week={_WEEK_START.isoformat()}')
        resp2 = self.client.get(reverse('admin_adherence') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp1.status_code, 302)
        self.assertEqual(resp2.status_code, 302)

    def test_non_holder_denied_on_ajax_endpoints_directly(self):
        # Server-side enforcement, not just hidden nav: hitting the save
        # endpoints directly without the permission must still be rejected.
        agent = _make_agent('tabsnoholder2', role='admin', role_type='supervisor')
        official = _make_agent('officialx', role='admin', role_type='supervisor')
        official.is_official_admin = True
        official.save()
        self.client.login(username='tabsnoholder2', password='x')
        resp = self.client.post(
            reverse('add_admin_coding_ajax'), data='{}', content_type='application/json'
        )
        self.assertEqual(resp.status_code, 302)

    def test_super_admin_has_access_without_the_flag(self):
        boss = _make_agent('tabssuper', role='admin', role_type='supervisor')
        boss.is_super_admin = True
        boss.can_access_admin_tabs = False
        boss.save()
        self.client.login(username='tabssuper', password='x')
        resp1 = self.client.get(reverse('admin_codings') + f'?week={_WEEK_START.isoformat()}')
        resp2 = self.client.get(reverse('admin_adherence') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp1.status_code, 200)
        self.assertEqual(resp2.status_code, 200)

    def test_unauthenticated_denied_on_both_tabs(self):
        resp1 = self.client.get(reverse('admin_codings'))
        resp2 = self.client.get(reverse('admin_adherence'))
        self.assertEqual(resp1.status_code, 302)
        self.assertEqual(resp2.status_code, 302)


class AdminTabsVisibilityTests(TestCase):
    """
    Part 4: team-scoped visibility for the Admin Codings / Admin Adherence
    rosters (_admin_tabs_access). Super admins/owners see every Official
    Admin regardless of who they supervise; a non-super-admin holder sees
    only themselves (if they're an Official Admin) plus the Official Admins
    who have them set as supervisor. Visibility only — no edit checks here.
    """

    def setUp(self):
        # Super admin — no restriction, sees everyone.
        self.boss = _make_agent('vteamboss', role='admin', role_type='supervisor')
        self.boss.is_super_admin = True
        self.boss.save()

        # A supervisor holding the new permission, and an Official Admin
        # herself — mirrors the "Vrenely" scenario in the task.
        self.vrenely = _make_agent('vrenely', role='admin', role_type='supervisor')
        self.vrenely.can_access_admin_tabs = True
        self.vrenely.is_official_admin = True
        self.vrenely.save()

        # An Official Admin supervised by Vrenely.
        self.supervised = _make_agent('supervisedadmin', role='admin', role_type='supervisor')
        self.supervised.is_official_admin = True
        self.supervised.supervisor = self.vrenely
        self.supervised.save()

        # An Official Admin supervised by someone else entirely.
        self.other_supervisor = _make_agent('othersupervisor', role='admin', role_type='supervisor')
        self.other_admin = _make_agent('otheradmin', role='admin', role_type='supervisor')
        self.other_admin.is_official_admin = True
        self.other_admin.supervisor = self.other_supervisor
        self.other_admin.save()

        # A non-Official-Admin on Vrenely's own team — must never appear.
        self.regular_teammate = _make_agent('regularteammate', role='agent', role_type='agent')
        self.regular_teammate.supervisor = self.vrenely
        self.regular_teammate.save()

        # Billable Five9 profiles so Official Admins show on BOTH tabs
        # (Admin Codings additionally requires a billable profile).
        for a in (self.vrenely, self.supervised, self.other_admin):
            Five9Profile.objects.create(agent=a, five9_username=a.agent_name, billable=True)

    def _pks(self, url_name, username):
        self.client.login(username=username, password='x')
        resp = self.client.get(reverse(url_name) + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 200)
        pks = [row['agent'].pk for row in resp.context['rows']]
        self.client.logout()
        return pks

    def test_super_admin_sees_all_official_admins_on_both_tabs(self):
        for url_name in ('admin_codings', 'admin_adherence'):
            pks = self._pks(url_name, 'vteamboss')
            self.assertIn(self.vrenely.pk, pks)
            self.assertIn(self.supervised.pk, pks)
            self.assertIn(self.other_admin.pk, pks)

    def test_supervisor_sees_only_self_and_supervised_admins(self):
        for url_name in ('admin_codings', 'admin_adherence'):
            pks = self._pks(url_name, 'vrenely')
            self.assertIn(self.vrenely.pk, pks)
            self.assertIn(self.supervised.pk, pks)

    def test_official_admin_of_different_supervisor_is_invisible(self):
        for url_name in ('admin_codings', 'admin_adherence'):
            pks = self._pks(url_name, 'vrenely')
            self.assertNotIn(self.other_admin.pk, pks)
            self.assertNotIn(self.other_supervisor.pk, pks)

    def test_non_official_admins_never_appear(self):
        # Regular teammate is on Vrenely's own team but isn't an Official
        # Admin — never shows, for her or for the super admin.
        for username in ('vteamboss', 'vrenely'):
            for url_name in ('admin_codings', 'admin_adherence'):
                pks = self._pks(url_name, username)
                self.assertNotIn(self.regular_teammate.pk, pks)


class AdminCodingsEditScopeTests(TestCase):
    """
    Part 5: server-side team-scoped edit enforcement on the Admin Codings
    save endpoints. A supervisor can add/edit/delete a coding for a
    supervised Official Admin; the same supervisor is rejected server-side
    (403) for an out-of-team Official Admin via a direct POST, even bypassing
    the UI entirely; a super admin can edit any Official Admin.
    """

    def setUp(self):
        self.boss = _make_agent('codeeditboss', role='admin', role_type='supervisor')
        self.boss.is_super_admin = True
        self.boss.save()

        self.vrenely = _make_agent('codeeditvrenely', role='admin', role_type='supervisor')
        self.vrenely.can_access_admin_tabs = True
        self.vrenely.is_official_admin = True
        self.vrenely.save()

        self.supervised = _make_agent('codeeditsupervised', role='admin', role_type='supervisor')
        self.supervised.is_official_admin = True
        self.supervised.supervisor = self.vrenely
        self.supervised.save()

        other_supervisor = _make_agent('codeeditothersup', role='admin', role_type='supervisor')
        self.other_admin = _make_agent('codeeditotheradmin', role='admin', role_type='supervisor')
        self.other_admin.is_official_admin = True
        self.other_admin.supervisor = other_supervisor
        self.other_admin.save()

    def _add(self, agent_id):
        return self.client.post(reverse('add_admin_coding_ajax'), data=json.dumps({
            'agent_id': agent_id, 'date': _WEEK_START.isoformat(),
            'start_time': '09:00:00', 'end_time': '10:00:00', 'notes': '',
        }), content_type='application/json')

    def test_supervisor_can_add_coding_for_supervised_admin(self):
        self.client.login(username='codeeditvrenely', password='x')
        resp = self._add(self.supervised.pk)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(Coding.objects.filter(agent=self.supervised, is_admin_coding=True).exists())

    def test_supervisor_denied_adding_coding_for_out_of_team_admin(self):
        self.client.login(username='codeeditvrenely', password='x')
        resp = self._add(self.other_admin.pk)
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(Coding.objects.filter(agent=self.other_admin, is_admin_coding=True).exists())

    def test_super_admin_can_add_coding_for_any_official_admin(self):
        self.client.login(username='codeeditboss', password='x')
        resp = self._add(self.other_admin.pk)
        self.assertEqual(resp.status_code, 200)

    def test_supervisor_can_edit_coding_for_supervised_admin(self):
        coding = Coding.objects.create(
            agent=self.supervised, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
            is_admin_coding=True,
        )
        self.client.login(username='codeeditvrenely', password='x')
        resp = self.client.post(reverse('edit_admin_coding_ajax'), data=json.dumps({
            'coding_id': coding.pk, 'start_time': '09:00:00', 'end_time': '11:00:00', 'notes': '',
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        coding.refresh_from_db()
        self.assertEqual(coding.end_time, time(11, 0))

    def test_supervisor_denied_editing_coding_for_out_of_team_admin(self):
        coding = Coding.objects.create(
            agent=self.other_admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
            is_admin_coding=True,
        )
        self.client.login(username='codeeditvrenely', password='x')
        resp = self.client.post(reverse('edit_admin_coding_ajax'), data=json.dumps({
            'coding_id': coding.pk, 'start_time': '09:00:00', 'end_time': '11:00:00', 'notes': '',
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 403)
        coding.refresh_from_db()
        self.assertEqual(coding.end_time, time(10, 0))

    def test_super_admin_can_edit_coding_for_any_official_admin(self):
        coding = Coding.objects.create(
            agent=self.other_admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
            is_admin_coding=True,
        )
        self.client.login(username='codeeditboss', password='x')
        resp = self.client.post(reverse('edit_admin_coding_ajax'), data=json.dumps({
            'coding_id': coding.pk, 'start_time': '09:00:00', 'end_time': '11:00:00', 'notes': '',
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)

    def test_supervisor_can_delete_coding_for_supervised_admin(self):
        coding = Coding.objects.create(
            agent=self.supervised, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
            is_admin_coding=True,
        )
        self.client.login(username='codeeditvrenely', password='x')
        resp = self.client.post(reverse('delete_admin_coding_ajax'), data=json.dumps({
            'coding_id': coding.pk,
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(Coding.objects.filter(pk=coding.pk).exists())

    def test_supervisor_denied_deleting_coding_for_out_of_team_admin(self):
        coding = Coding.objects.create(
            agent=self.other_admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
            is_admin_coding=True,
        )
        self.client.login(username='codeeditvrenely', password='x')
        resp = self.client.post(reverse('delete_admin_coding_ajax'), data=json.dumps({
            'coding_id': coding.pk,
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(Coding.objects.filter(pk=coding.pk).exists())

    def test_super_admin_can_delete_coding_for_any_official_admin(self):
        coding = Coding.objects.create(
            agent=self.other_admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(10, 0),
            is_admin_coding=True,
        )
        self.client.login(username='codeeditboss', password='x')
        resp = self.client.post(reverse('delete_admin_coding_ajax'), data=json.dumps({
            'coding_id': coding.pk,
        }), content_type='application/json')
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(Coding.objects.filter(pk=coding.pk).exists())


class AdherenceExportTests(TestCase):
    """Combined Adherence export — regular Adherence roster + Official Admins,
    one row per person, values sourced from adherence._build_rows exactly."""

    def setUp(self):
        # _get_adherence_agent_pks caches its roster per week_start for 5 minutes;
        # clear it so each test sees only the agents it creates for _WEEK_START.
        cache.clear()
        self.boss = _make_agent('adhexportboss', role='admin', role_type='supervisor')
        self.boss.is_super_admin = True
        self.boss.save()
        self.client.login(username='adhexportboss', password='x')

    def _export_wb(self, week_start=_WEEK_START):
        resp = self.client.get(reverse('adherence_export') + f'?week={week_start.isoformat()}')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        return openpyxl.load_workbook(io.BytesIO(resp.content))

    def _export_ws(self, week_start=_WEEK_START):
        return self._export_wb(week_start).active

    def _display_name(self, agent):
        return agent.user.get_full_name() or agent.agent_name or agent.user.username

    def _rows_for(self, ws, agent):
        """All data-row indices (1-indexed) whose Legal Name column matches this agent."""
        name = self._display_name(agent)
        return [r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=3).value == name]

    def _row_for(self, ws, agent):
        rows = self._rows_for(ws, agent)
        self.assertEqual(len(rows), 1, f"Expected exactly one row for {agent}, found {len(rows)}")
        return rows[0]

    def test_roster_union_regular_and_admin_no_duplicates(self):
        regular = _make_agent('adhexpreg1')
        Shift.objects.create(agent=regular, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        admin = _make_agent('adhexpadmin1')
        admin.is_official_admin = True
        admin.save()

        ws = self._export_ws()
        self._row_for(ws, regular)   # raises if not exactly one row
        self._row_for(ws, admin)     # raises if not exactly one row

    def test_met_day_shows_status_and_green_completed_hours(self):
        agent = _make_agent('adhexpmet1')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        cell = ws.cell(row=r, column=8)  # Monday
        self.assertEqual(cell.value, "P\n8:00:00")
        self.assertTrue(cell.font.color.rgb.upper().endswith('166534'))
        self.assertTrue(cell.fill.fgColor.rgb.upper().endswith('E8F5E9'))

    def test_short_day_shows_status_and_red_shortfall(self):
        agent = _make_agent('adhexpshort1')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='T', actual_hours=Decimal('7.75'))

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        cell = ws.cell(row=r, column=8)  # Monday: 8h scheduled, 7h45m worked -> 15 min short
        self.assertEqual(cell.value, "T\n-0:15:00")
        self.assertTrue(cell.font.color.rgb.upper().endswith('C0392B'))

    def test_official_admin_row_uses_admin_codings_not_regular_codings(self):
        admin = _make_agent('adhexpadmin2')
        admin.is_official_admin = True
        admin.save()
        Shift.objects.create(agent=admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=admin, date=_WEEK_START, status='P', actual_hours=None)
        # Regular coding (must be ignored) vs. admin coding (must be the one used)
        Coding.objects.create(agent=admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(9, 30),
                               is_admin_coding=False)
        Coding.objects.create(agent=admin, date=_WEEK_START, start_time=time(9, 0), end_time=time(11, 0),
                               is_admin_coding=True)

        ws = self._export_ws()
        r = self._row_for(ws, admin)
        cell = ws.cell(row=r, column=8)  # Monday: 8h scheduled, 2h admin-coded -> 6h short
        self.assertEqual(cell.value, "P\n-6:00:00")

    def test_identity_columns_populate(self):
        supervisor = _make_agent('adhexpsupv1')
        agent = _make_agent('adhexpid1')
        agent.employee_id = 'EMP-500'
        agent.agent_name = 'Johnny P'
        agent.supervisor = supervisor
        agent.save()
        Five9Profile.objects.create(agent=agent, five9_username='idtest.f9', billable=True, is_primary=True)
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        self.assertEqual(ws.cell(row=r, column=1).value, 'idtest.f9')
        self.assertEqual(ws.cell(row=r, column=2).value, 'EMP-500')
        self.assertEqual(ws.cell(row=r, column=4).value, 'Johnny P')
        self.assertEqual(ws.cell(row=r, column=5).value, str(supervisor))

    def test_identity_columns_blank_safe_when_missing(self):
        agent = _make_agent('adhexpid2')
        agent.agent_name = ''
        agent.employee_id = None
        agent.save()  # no Five9Profile, no supervisor; User has no first/last name
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        self.assertFalse(ws.cell(row=r, column=1).value)  # Username blank
        self.assertFalse(ws.cell(row=r, column=2).value)  # Employee ID blank
        self.assertEqual(ws.cell(row=r, column=3).value, 'adhexpid2')  # Legal Name falls back to username
        self.assertFalse(ws.cell(row=r, column=5).value)  # Supervisor blank

    def test_commission_deduction_populates_and_defaults_zero(self):
        agent = _make_agent('adhexpcomm1')
        PayrollAdjustment.objects.create(agent=agent, week_start=_WEEK_START, commission_deduction=Decimal('5.5'))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))
        agent_no_adj = _make_agent('adhexpcomm2')
        AdherenceRecord.objects.create(agent=agent_no_adj, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        ws = self._export_ws()
        self.assertEqual(ws.cell(row=self._row_for(ws, agent), column=6).value, 5.5)
        self.assertEqual(ws.cell(row=self._row_for(ws, agent_no_adj), column=6).value, 0.0)

    def test_scheduled_hours_reflects_vto_zeroing(self):
        agent = _make_agent('adhexpvto1')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        Shift.objects.create(agent=agent, date=_WEEK_START + timedelta(days=1), start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START + timedelta(days=1), status='VTO', actual_hours=None)

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        self.assertEqual(ws.cell(row=r, column=7).value, timedelta(hours=8))

    def test_scheduled_hours_is_real_duration_not_text(self):
        agent = _make_agent('adhexpdur1')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        cell = ws.cell(row=r, column=7)
        self.assertIsInstance(cell.value, timedelta)
        self.assertEqual(cell.number_format, '[h]:mm:ss')

    def test_off_day_cell_is_blank(self):
        agent = _make_agent('adhexpoff1')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        Shift.objects.create(agent=agent, date=_WEEK_START + timedelta(days=5), is_off=True,
                              start_time=time(0, 0), end_time=time(0, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        ws = self._export_ws()
        r = self._row_for(ws, agent)
        self.assertIsNone(ws.cell(row=r, column=13).value)  # Saturday, day 6 -> column 8+5

    def test_lcc_employer_excluded_regular_and_official_admin(self):
        lcc_regular = _make_agent('adhexplccreg1')
        lcc_regular.employer = 'LCC'
        lcc_regular.save()
        Shift.objects.create(agent=lcc_regular, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))

        lcc_admin = _make_agent('adhexplccadmin1')
        lcc_admin.employer = 'LCC'
        lcc_admin.is_official_admin = True
        lcc_admin.save()

        infinity_regular = _make_agent('adhexpinfreg1')
        Shift.objects.create(agent=infinity_regular, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))

        ws = self._export_ws()
        self.assertEqual(self._rows_for(ws, lcc_regular), [])
        self.assertEqual(self._rows_for(ws, lcc_admin), [])
        self._row_for(ws, infinity_regular)  # still present

    def test_vto_status_appears_on_vto_sheet(self):
        agent = _make_agent('adhexpvtosheet1')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='VTO', actual_hours=None)

        wb = self._export_wb()
        self._row_for(wb['VTO Agents'], agent)

    def test_p_vto_status_appears_on_vto_sheet(self):
        agent = _make_agent('adhexpvtosheet2')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P+VTO', actual_hours=Decimal('4'))

        wb = self._export_wb()
        self._row_for(wb['VTO Agents'], agent)

    def test_t_vto_status_appears_on_vto_sheet(self):
        agent = _make_agent('adhexpvtosheet3')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='T+VTO', actual_hours=Decimal('6'))

        wb = self._export_wb()
        self._row_for(wb['VTO Agents'], agent)

    def test_non_vto_agent_absent_from_vto_sheet(self):
        agent = _make_agent('adhexpvtosheet4')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        wb = self._export_wb()
        self.assertEqual(self._rows_for(wb['VTO Agents'], agent), [])

    def test_official_admin_with_vto_excluded_from_vto_sheet(self):
        admin = _make_agent('adhexpvtosheet5')
        admin.is_official_admin = True
        admin.save()
        AdherenceRecord.objects.create(agent=admin, date=_WEEK_START, status='VTO', actual_hours=None)

        wb = self._export_wb()
        self._row_for(wb.active, admin)  # still present on the main "Adherence" sheet
        self.assertEqual(self._rows_for(wb['VTO Agents'], admin), [])

    def test_main_sheet_unchanged_by_vto_sheet_addition(self):
        vto_agent = _make_agent('adhexpvtosheet6')
        Shift.objects.create(agent=vto_agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=vto_agent, date=_WEEK_START, status='VTO', actual_hours=None)
        plain_agent = _make_agent('adhexpvtosheet7')
        Shift.objects.create(agent=plain_agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=plain_agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        wb = self._export_wb()
        main_ws = wb.active
        self.assertEqual(main_ws.title, "Adherence")
        self._row_for(main_ws, vto_agent)
        self._row_for(main_ws, plain_agent)

    def test_empty_week_vto_sheet_shows_header_and_note(self):
        agent = _make_agent('adhexpvtosheet8')
        Shift.objects.create(agent=agent, date=_WEEK_START, start_time=time(9, 0), end_time=time(17, 0))
        AdherenceRecord.objects.create(agent=agent, date=_WEEK_START, status='P', actual_hours=Decimal('8'))

        wb = self._export_wb()
        vto_ws = wb['VTO Agents']
        self.assertEqual(
            [vto_ws.cell(row=3, column=c).value for c in range(1, 8)],
            ['Username', 'Employee ID', 'Legal Name', 'Agent Name', 'Supervisor',
             'Commission Deduction %', 'Scheduled Hours'],
        )
        self.assertIn('No VTO was recorded for this week.', vto_ws.cell(row=4, column=1).value)
        self.assertEqual(vto_ws.max_row, 4)

    def test_filename_and_content_type(self):
        resp = self.client.get(reverse('adherence_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        self.assertIn(f'adherence_{_WEEK_START.isoformat()}.xlsx', resp['Content-Disposition'])

    def test_non_super_admin_denied(self):
        self.client.logout()
        _make_agent('adhexpplain1', role='admin', role_type='supervisor')
        self.client.login(username='adhexpplain1', password='x')
        resp = self.client.get(reverse('adherence_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 302)

    def test_unauthenticated_denied(self):
        self.client.logout()
        resp = self.client.get(reverse('adherence_export') + f'?week={_WEEK_START.isoformat()}')
        self.assertEqual(resp.status_code, 302)


class AdminAdherenceLiveLoginTests(TestCase):
    """The admin adherence tab/export must reflect LIVE billable login hours
    (same source as Billing v2), NOT the stale, double-NR-deducted stored
    AdherenceRecord.actual_hours that made Official Admins show false
    'missing time'. Also fixes summing across multiple billable Five9 profiles."""

    def setUp(self):
        cache.clear()
        _settings()

    def _official_admin(self, username):
        a = _make_agent(username, role='admin', role_type='qa')
        a.is_official_admin = True
        a.save()
        return a

    def test_live_login_overwrites_stale_stored_value(self):
        from finance.views import _apply_live_login_hours
        admin = self._official_admin('liveadmin')
        Five9Profile.objects.create(agent=admin, five9_username=admin.agent_name, billable=True, is_primary=True)
        _add_hours(admin, login_seconds=8 * 3600, nr_seconds=0, week_day=0)  # 8h connected
        # Stale stored value (undercounted — simulates the double-deduction bug)
        rec = AdherenceRecord.objects.create(agent=admin, date=_WEEK[0], status='P', actual_hours=Decimal('3.00'))
        record_map = {(admin.pk, _WEEK[0]): rec}
        _apply_live_login_hours([admin], _WEEK, record_map)
        # In-memory record now reflects the true 8h billable login…
        self.assertEqual(record_map[(admin.pk, _WEEK[0])].actual_hours, Decimal('8'))
        # …but the DB row is untouched (request-scoped, never saved).
        rec.refresh_from_db()
        self.assertEqual(rec.actual_hours, Decimal('3.00'))

    def test_sums_multiple_billable_profiles(self):
        from finance.views import _apply_live_login_hours
        admin = self._official_admin('multiadmin')
        Five9Profile.objects.create(agent=admin, five9_username='acct_a', billable=True, is_primary=True)
        Five9Profile.objects.create(agent=admin, five9_username='acct_b', billable=True)
        upload, _ = DailyUpload.objects.get_or_create(date=_WEEK[0])
        DailyAgentHours.objects.create(upload=upload, agent=admin, five9_username='acct_a', login_seconds=5 * 3600, not_ready_seconds=0)
        DailyAgentHours.objects.create(upload=upload, agent=admin, five9_username='acct_b', login_seconds=3 * 3600, not_ready_seconds=0)
        record_map = {}
        _apply_live_login_hours([admin], _WEEK, record_map)
        # Both billable profiles summed = 8h (the old write path stored only the last).
        self.assertEqual(record_map[(admin.pk, _WEEK[0])].actual_hours, Decimal('8'))
        # A record was fabricated in memory only — nothing persisted.
        self.assertFalse(AdherenceRecord.objects.filter(agent=admin, date=_WEEK[0]).exists())

    def test_excludes_non_billable_username(self):
        from finance.views import _apply_live_login_hours
        admin = self._official_admin('nbadmin')
        Five9Profile.objects.create(agent=admin, five9_username='billable_u', billable=True, is_primary=True)
        upload, _ = DailyUpload.objects.get_or_create(date=_WEEK[0])
        DailyAgentHours.objects.create(upload=upload, agent=admin, five9_username='billable_u', login_seconds=6 * 3600, not_ready_seconds=0)
        DailyAgentHours.objects.create(upload=upload, agent=admin, five9_username='other_nb', login_seconds=2 * 3600, not_ready_seconds=0)
        record_map = {}
        _apply_live_login_hours([admin], _WEEK, record_map)
        # Only the billable username's 6h counts — matches Billing v2's billable filter.
        self.assertEqual(record_map[(admin.pk, _WEEK[0])].actual_hours, Decimal('6'))
