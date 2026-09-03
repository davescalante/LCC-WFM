import json
from datetime import date, timedelta, time as time_cls
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q, Sum

from scheduling.models import Agent, Five9Profile, OvertimeShift, log_action
from adherence.models import AdherenceRecord, DailyAgentHours, DailyUpload, PayrollAdjustment, Coding
from .models import BillingSettings, BillingSettingsHistory
from wfm.constants import BONUS_QUALIFYING, BONUS_DISQUALIFYING, VTO_TYPE_STATUSES
from wfm.utils import get_week_start, parse_week_param, get_billable_username_map

# ─── Access control ───────────────────────────────────────────────────────────
# Finance is visible only to users with is_super_admin=True, plus Django superusers.

def _has_finance_access(user):
    if user.is_superuser:
        return True
    try:
        return user.agent.is_super_admin
    except Exception:
        return False


def finance_access_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not _has_finance_access(request.user):
            messages.error(request, "Access denied.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped


# Admin Codings / Admin Adherence are visible to super admins (like Finance)
# plus anyone individually granted can_access_admin_tabs. Team-scoping a
# non-super-admin holder down to their own supervised Official Admins is a
# separate, later concern from this gate.

def _has_admin_tabs_access(user):
    if user.is_superuser:
        return True
    try:
        agent = user.agent
    except Exception:
        return False
    return bool(agent.is_super_admin or agent.can_access_admin_tabs)


def admin_tabs_access_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not _has_admin_tabs_access(request.user):
            messages.error(request, "Access denied.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped


def _admin_tabs_access(user):
    """
    Determine Admin Codings / Admin Adherence visibility scope for a user.

    Returns (has_access, team_pks):
      - has_access: whether the user may see these tabs at all.
      - team_pks:   None  => may see ALL Official Admins (super admin /
                             superuser — no restriction, even if they also
                             happen to supervise some);
                    a set => may see only these agent PKs (their own team:
                             direct reports + self). Since the roster query
                             is always additionally filtered to
                             is_official_admin=True, including a non-admin's
                             own pk here is harmless — it just won't match.
    """
    if user.is_superuser:
        return True, None
    try:
        agent = user.agent
    except Exception:
        return False, set()
    if agent.is_super_admin:
        return True, None
    if agent.can_access_admin_tabs:
        team = set(Agent.objects.filter(supervisor=agent).values_list('pk', flat=True))
        team.add(agent.pk)
        return True, team
    return False, set()


# ─── Week helpers ─────────────────────────────────────────────────────────────

def _get_week_start(request):
    """Parse ?week= param (ISO Monday) or default to current Monday."""
    return parse_week_param(request.GET.get('week', '')) or get_week_start()


def _week_dates(week_start):
    return [week_start + timedelta(days=i) for i in range(7)]


def _fmt_hrs(h):
    """Format Decimal hours as '12.75' (2 decimal places)."""
    if h is None:
        return '—'
    return f"{h:.2f}"


def _fmt_mxn(v):
    if v is None:
        return '—'
    return f"${v:,.2f}"


def _fmt_usd(v):
    if v is None:
        return '—'
    return f"${v:,.2f}"


# ─── Core weekly calculation ───────────────────────────────────────────────────

def _get_billable_weekly_data(agents, week_dates, settings):
    """
    Compute weekly pay and billing figures for a list of billable agents.

    Sums each agent's login seconds and NR seconds from Five9 DailyAgentHours rows
    (restricted to that agent's billable Five9 usernames), adds coded hours from
    exactly one path per person — Official Admins from admin codings, everyone else
    from regular codings, never both — then applies ONE weekly not-ready allowance:
      nr_allowed    = min(cap, connected × settings.nr_ratio)   # connected = login + coded
      nr_deduction  = max(0, weekly_NR − nr_allowed)
      final_hrs     = max(0, connected − nr_deduction)
    Cap = nr_cap_regular_hours (6 h) / nr_cap_kill_team_hours (7 h for kill_team). Any
    VTO-type day that week substitutes the flat cap for the ratio. There is no 48-hour
    threshold and no larger-of-two comparison. Then runs the pay calculations (base pay,
    OT top-ups, adherence or admin bonus, billing).

    Returns a dict keyed by agent.pk, each value containing:
      agent, five9_username
      total_nr_hrs, nr_cap_hrs, nr_allowed_hrs, nr_deduction
      actual_hrs (raw login), coded_hrs, pre_cap_total, final_hrs
      ot_regular_hrs, ot_1_5_hrs, ot_power_hrs
      hourly_mxn, billing_rate_usd
      base_pay_mxn, ph_topup_mxn, ot_1_5_topup_mxn
      bonus_qualifies (bool), is_official_admin (bool), bonus_mxn, admin_bonus_mxn
      commission_pct, total_pay_mxn, total_pay_usd, billing_usd
    """
    agent_ids = [a.pk for a in agents]
    week_start = week_dates[0]

    # ── Billable username lookup ───────────────────────────────────────────
    billable_map, primary_billable_map = get_billable_username_map(agent_ids)

    # ── Sum login + NR seconds from billable DailyAgentHours ──────────────
    nr_secs_map = {}    # agent_id -> total NR seconds
    login_secs_map = {} # agent_id -> total login seconds
    for row in DailyAgentHours.objects.filter(
        upload__date__in=week_dates, agent__in=agent_ids
    ).values('agent_id', 'five9_username', 'login_seconds', 'not_ready_seconds'):
        aid = row['agent_id']
        if aid is None:
            continue
        uname = row['five9_username'].strip().lower()
        bnames = billable_map.get(aid)
        if bnames is None or uname in bnames:
            nr_secs_map[aid] = nr_secs_map.get(aid, 0) + row['not_ready_seconds']
            login_secs_map[aid] = login_secs_map.get(aid, 0) + row['login_seconds']

    # ── Coding hours ──────────────────────────────────────────────────────
    # Hard partition (is_admin_coding / is_official_admin): each person's coded time comes
    # from exactly ONE place — official admins from their ADMIN codings, everyone else from
    # REGULAR codings. Summing both (the old behaviour) double-counted anyone with entries in
    # both tabs — e.g. an admin whose time was also coded on the regular Codings tab — which
    # inflated their hours here and everywhere this engine feeds (billing report + Nómina).
    admin_ids = {a.pk for a in agents if getattr(a, 'is_official_admin', False)}
    coded_hrs_map = {}
    for coding in Coding.objects.filter(agent__in=agent_ids, date__in=week_dates):
        if coding.is_admin_coding != (coding.agent_id in admin_ids):
            continue   # this coding belongs to the other path for this person — skip it
        coded_hrs_map[coding.agent_id] = coded_hrs_map.get(coding.agent_id, Decimal('0')) + Decimal(str(coding.total_hours()))

    # ── Adherence bonus (already tracked in adherence tab) ────────────────
    bonus_map = {}    # agent_id -> True/False/None
    has_status = set()
    vto_week_agents = set()  # agent_id -> had a VTO/P+VTO/T+VTO day this week
    for rec in AdherenceRecord.objects.filter(agent__in=agent_ids, date__in=week_dates).values('agent_id', 'status'):
        aid = rec['agent_id']
        if rec['status']:
            has_status.add(aid)
            if rec['status'] in VTO_TYPE_STATUSES:
                vto_week_agents.add(aid)
            if rec['status'] in BONUS_DISQUALIFYING:
                bonus_map[aid] = False
            elif aid not in bonus_map and rec['status'] in BONUS_QUALIFYING:
                bonus_map[aid] = True
            elif aid not in bonus_map:
                bonus_map[aid] = False
    # OT no-show disqualifies
    for ot in OvertimeShift.objects.filter(agent__in=agent_ids, date__in=week_dates, status='no_show'):
        bonus_map[ot.agent_id] = False

    # ── OT hours by type ──────────────────────────────────────────────────
    ot_regular_map = {}
    ot_1_5_map = {}
    ot_power_map = {}
    # Collapse exact-duplicate OT rows per agent/day, keyed on (start_time, end_time,
    # status) — the identical key adherence.views._build_maps uses. Identical rows are the
    # same slot recorded twice, and summing them paid the incentive premium twice. Split OT
    # is untouched by construction: two rows on one day at different hours have different
    # keys and both still count. incentive_type is deliberately NOT in the key — including
    # it would keep both rows of a slot recorded once as power_hour and once as
    # time_and_a_half and pay both premiums, the exact overpay this closes. order_by('pk')
    # makes "the first row wins" deterministic when a duplicated slot's rows disagree on
    # incentive_type, matching how schedule_data_inventory attributes a duplicated slot.
    # Proved a no-op on production over 52 weeks by the verify_ot_topups parity command.
    _ot_seen = {}
    for ot in OvertimeShift.objects.filter(
        agent__in=agent_ids, date__in=week_dates, status='completed'
    ).order_by('pk'):
        seen = _ot_seen.setdefault((ot.agent_id, ot.date), set())
        dedup_key = (ot.start_time, ot.end_time, ot.status)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        hrs = ot.total_shift_hours()
        if ot.incentive_type == 'none':
            ot_regular_map[ot.agent_id] = ot_regular_map.get(ot.agent_id, Decimal('0')) + hrs
        elif ot.incentive_type == 'time_and_a_half':
            ot_1_5_map[ot.agent_id] = ot_1_5_map.get(ot.agent_id, Decimal('0')) + hrs
        elif ot.incentive_type == 'power_hour':
            ot_power_map[ot.agent_id] = ot_power_map.get(ot.agent_id, Decimal('0')) + hrs

    # ── Commission deductions ─────────────────────────────────────────────
    commission_map = {}
    for pa in PayrollAdjustment.objects.filter(agent__in=agent_ids, week_start=week_start):
        commission_map[pa.agent_id] = pa.commission_deduction

    results = {}
    for agent in agents:
        aid = agent.pk
        hourly_mxn = agent.hourly_rate or Decimal('0')
        billing_rate = agent.billing_rate_usd or settings.billing_rate_usd
        usd_to_mxn = settings.usd_to_mxn

        # Raw login hours — no daily NR deductions; weekly checks only
        raw_login_hrs = Decimal(str(login_secs_map.get(aid, 0))) / Decimal('3600')
        coded_hrs = coded_hrs_map.get(aid, Decimal('0'))
        pre_total = raw_login_hrs + coded_hrs

        # Weekly Not-Ready allowance = 12.5% of CONNECTED time (login + coded), capped at the
        # weekly cap (6h regular / 7h Kill Team). A VTO/P+VTO/T+VTO day that week grants the
        # flat cap instead of the ratio. Anything above the allowance is deducted — one rule,
        # no competing checks.
        total_nr_hrs = Decimal(str(nr_secs_map.get(aid, 0))) / Decimal('3600')
        nr_cap = settings.nr_cap_kill_team_hours if agent.role_type == 'kill_team' else settings.nr_cap_regular_hours
        has_vto = aid in vto_week_agents
        nr_allowed_hrs = nr_cap if has_vto else min(nr_cap, pre_total * settings.nr_ratio)
        nr_deduction = max(Decimal('0'), total_nr_hrs - nr_allowed_hrs)
        final_hrs = max(Decimal('0'), pre_total - nr_deduction)

        # OT
        ot_reg = ot_regular_map.get(aid, Decimal('0'))
        ot_1_5 = ot_1_5_map.get(aid, Decimal('0'))
        ot_pow = ot_power_map.get(aid, Decimal('0'))

        # Pay calculations
        # OT base pay is already captured in final_hrs (agents log into Five9 during OT)
        # These columns are top-up only: the LCC incentive premium above the regular rate
        base_pay = (final_hrs * hourly_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP)
        ph_topup_mxn = (ot_pow * hourly_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP)
        ot_1_5_topup_mxn = (ot_1_5 * hourly_mxn * Decimal('0.5')).quantize(Decimal('0.01'), ROUND_HALF_UP)

        is_official_admin = getattr(agent, 'is_official_admin', False)
        bonus_qualifies = (not is_official_admin) and bonus_map.get(aid) is True and aid in has_status
        # Per-agent adherence-bonus cap overrides the global default when set (e.g. a
        # higher individual cap). Proration below the full-hours threshold is unchanged.
        adherence_cap = (agent.adherence_bonus_max_mxn if agent.adherence_bonus_max_mxn is not None
                         else settings.adherence_bonus_max_mxn)
        if bonus_qualifies and settings.adherence_bonus_full_hours > 0:
            bonus_mxn = min(
                adherence_cap,
                (final_hrs / settings.adherence_bonus_full_hours * adherence_cap)
            ).quantize(Decimal('0.01'), ROUND_HALF_UP)
        else:
            bonus_mxn = Decimal('0')

        if is_official_admin:
            admin_bonus = (agent.admin_bonus_mxn if agent.admin_bonus_mxn is not None
                           else settings.default_admin_bonus_mxn)
        else:
            admin_bonus = Decimal('0')

        comm_pct = commission_map.get(aid, Decimal('0'))

        total_pay_mxn = base_pay + ph_topup_mxn + ot_1_5_topup_mxn + bonus_mxn + admin_bonus
        total_pay_usd = (total_pay_mxn / usd_to_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP) if usd_to_mxn else Decimal('0')

        # Billing (what Infinity charges LCC)
        billing_usd = (final_hrs * billing_rate).quantize(Decimal('0.01'), ROUND_HALF_UP)

        results[aid] = {
            'agent': agent,
            'five9_username': primary_billable_map.get(aid, ''),
            'total_nr_hrs': total_nr_hrs,
            'nr_cap_hrs': nr_cap,
            'nr_allowed_hrs': nr_allowed_hrs,
            'nr_deduction': nr_deduction,
            'actual_hrs': raw_login_hrs,
            'coded_hrs': coded_hrs,
            'pre_cap_total': pre_total,
            'final_hrs': final_hrs,
            'ot_regular_hrs': ot_reg,
            'ot_1_5_hrs': ot_1_5,
            'ot_power_hrs': ot_pow,
            'hourly_mxn': hourly_mxn,
            'billing_rate_usd': billing_rate,
            'base_pay_mxn': base_pay,
            'ph_topup_mxn': ph_topup_mxn,
            'ot_1_5_topup_mxn': ot_1_5_topup_mxn,
            'bonus_qualifies': bonus_qualifies,
            'is_official_admin': is_official_admin,
            'bonus_mxn': bonus_mxn,
            'admin_bonus_mxn': admin_bonus,
            'commission_pct': comm_pct,
            'total_pay_mxn': total_pay_mxn,
            'total_pay_usd': total_pay_usd,
            'billing_usd': billing_usd,
        }
    return results


# ─── Views ────────────────────────────────────────────────────────────────────

@login_required
@finance_access_required
def finance_dashboard(request):
    today = date.today()
    week_start = _get_week_start(request)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]
    settings = BillingSettings.get_for_week(week_start)

    agents = Agent.objects.filter(
        status='active',
        five9_profiles__billable=True,
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles', 'separations').distinct()

    data = _get_billable_weekly_data(list(agents), week_dates, settings)

    total_hrs = sum(d['final_hrs'] for d in data.values())
    total_billing_usd = sum(d['billing_usd'] for d in data.values())
    total_payroll_mxn = sum(d['total_pay_mxn'] for d in data.values())
    total_payroll_usd = sum(d['total_pay_usd'] for d in data.values())
    bonus_count = sum(1 for d in data.values() if d['bonus_qualifies'])
    bonus_total_mxn = sum(d['bonus_mxn'] for d in data.values())
    admin_bonus_total_mxn = sum(d['admin_bonus_mxn'] for d in data.values())
    ph_topup_total_mxn = sum(d['ph_topup_mxn'] for d in data.values())
    ot_1_5_topup_total_mxn = sum(d['ot_1_5_topup_mxn'] for d in data.values())

    prev_week = (week_start - timedelta(days=7)).isoformat()
    next_week = (week_start + timedelta(days=7)).isoformat()
    current_week = (today - timedelta(days=today.weekday())).isoformat()

    return render(request, 'finance/dashboard.html', {
        'settings': settings,
        'week_start': week_start,
        'week_end': week_end,
        'today': today,
        'prev_week': prev_week,
        'next_week': next_week,
        'current_week': current_week,
        'total_hrs': total_hrs,
        'total_billing_usd': total_billing_usd,
        'total_payroll_mxn': total_payroll_mxn,
        'total_payroll_usd': total_payroll_usd,
        'bonus_count': bonus_count,
        'bonus_total_mxn': bonus_total_mxn,
        'admin_bonus_total_mxn': admin_bonus_total_mxn,
        'ph_topup_total_mxn': ph_topup_total_mxn,
        'ot_1_5_topup_total_mxn': ot_1_5_topup_total_mxn,
        'agent_count': len(data),
    })


@login_required
@finance_access_required
def billing_report(request):
    week_start = _get_week_start(request)
    settings = BillingSettings.get_for_week(week_start)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    # All billed agents (Infinity employer or LCC) with separation filter
    agents = Agent.objects.filter(
        five9_profiles__billable=True,
    ).exclude(
        Q(status='inactive') &
        Q(separations__status='finalized') &
        Q(separations__remove_from_adherence_date__lte=week_start)
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles', 'separations').distinct()

    data = _get_billable_weekly_data(list(agents), week_dates, settings)

    # Group by employer, then by role_type
    _ROLE_GROUPS = [
        ('regular_agent', 'Regular Agents'),
        ('kill_team', 'Kill Team'),
        ('incubation', 'Incubation'),
        ('night_shift', 'Night Shift'),
        ('training', 'Training'),
        ('qa', 'QA'),
        ('cs', 'CS'),
        ('tester', 'Testers'),
        ('sms_email', 'SMS / Email'),
        ('supervisor', 'Supervisors'),
        ('coordinator', 'Coordinators'),
        ('trainer', 'Trainers'),
    ]

    infinity_rows = []
    lcc_rows = []
    for agent in agents:
        d = data.get(agent.pk, {})
        row = {**d, 'agent': agent}
        if agent.employer == 'LCC':
            lcc_rows.append(row)
        else:
            infinity_rows.append(row)

    def _group_rows(rows):
        groups = []
        for role_key, role_label in _ROLE_GROUPS:
            group_rows = [r for r in rows if r['agent'].role_type == role_key]
            if group_rows:
                subtotal_hrs = sum(r.get('final_hrs', Decimal('0')) for r in group_rows)
                subtotal_usd = sum(r.get('billing_usd', Decimal('0')) for r in group_rows)
                groups.append({
                    'label': role_label,
                    'rows': group_rows,
                    'subtotal_hrs': subtotal_hrs,
                    'subtotal_usd': subtotal_usd,
                })
        # Catch-all for any role_types not in the list
        listed_keys = {k for k, _ in _ROLE_GROUPS}
        other_rows = [r for r in rows if r['agent'].role_type not in listed_keys]
        if other_rows:
            groups.append({
                'label': 'Other',
                'rows': other_rows,
                'subtotal_hrs': sum(r.get('final_hrs', Decimal('0')) for r in other_rows),
                'subtotal_usd': sum(r.get('billing_usd', Decimal('0')) for r in other_rows),
            })
        return groups

    infinity_groups = _group_rows(infinity_rows)
    lcc_groups = _group_rows(lcc_rows)

    infinity_total_hrs = sum(r.get('final_hrs', Decimal('0')) for r in infinity_rows)
    infinity_total_usd = sum(r.get('billing_usd', Decimal('0')) for r in infinity_rows)
    ph_topup_total_mxn = sum(d.get('ph_topup_mxn', Decimal('0')) for d in data.values())
    ot_1_5_topup_total_mxn = sum(d.get('ot_1_5_topup_mxn', Decimal('0')) for d in data.values())
    ph_topup_total_usd = (ph_topup_total_mxn / settings.usd_to_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP) if settings.usd_to_mxn else Decimal('0')
    ot_1_5_topup_total_usd = (ot_1_5_topup_total_mxn / settings.usd_to_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP) if settings.usd_to_mxn else Decimal('0')
    bonus_total_mxn = sum(d.get('bonus_mxn', Decimal('0')) for d in data.values())
    bonus_total_usd = (bonus_total_mxn / settings.usd_to_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP) if settings.usd_to_mxn else Decimal('0')

    # Week navigation
    prev_week = (week_start - timedelta(days=7)).isoformat()
    next_week = (week_start + timedelta(days=7)).isoformat()

    return render(request, 'finance/billing.html', {
        'settings': settings,
        'week_start': week_start,
        'week_end': week_end,
        'prev_week': prev_week,
        'next_week': next_week,
        'infinity_groups': infinity_groups,
        'lcc_groups': lcc_groups,
        'infinity_total_hrs': infinity_total_hrs,
        'infinity_total_usd': infinity_total_usd,
        'ph_topup_total_usd': ph_topup_total_usd,
        'ot_1_5_topup_total_usd': ot_1_5_topup_total_usd,
        'bonus_total_usd': bonus_total_usd,
    })


@login_required
@finance_access_required
def billing_export(request):
    """Export billing report as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    week_start = _get_week_start(request)
    settings = BillingSettings.get_for_week(week_start)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    agents = Agent.objects.filter(
        five9_profiles__billable=True,
    ).exclude(
        Q(status='inactive') &
        Q(separations__status='finalized') &
        Q(separations__remove_from_adherence_date__lte=week_start)
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles', 'separations').distinct()

    data = _get_billable_weekly_data(list(agents), week_dates, settings)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    subheader_fill = PatternFill("solid", fgColor="E5E7EB")
    subheader_font = Font(bold=True, color="374151")
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')

    headers = [
        'Agent Name', 'Legal Name', 'Employee ID', 'Five9 Username (Billable)',
        'Supervisor', 'Agent Type', 'Employer',
        'Worked Hrs (Final)', 'Billing Rate (USD)', 'Total Billing (USD)',
    ]

    # Title
    ws.append([f"Billing Report — Week of {week_start.strftime('%B %d, %Y')} to {week_end.strftime('%B %d, %Y')}"])
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])

    # Headers
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    def _primary_billable_username(agent):
        # Primary first, then any billable
        profiles = list(agent.five9_profiles.all())
        primary = next((p for p in profiles if p.is_primary and p.billable), None)
        if primary:
            return primary.five9_username
        billable = next((p for p in profiles if p.billable), None)
        return billable.five9_username if billable else ''

    row_num = 4
    for employer_label, employer_agents in [('Infinity', [a for a in agents if a.employer == 'Infinity']),
                                             ('LCC Direct', [a for a in agents if a.employer == 'LCC'])]:
        if not employer_agents:
            continue
        ws.append([employer_label])
        ws.cell(row=row_num, column=1).font = subheader_font
        ws.cell(row=row_num, column=1).fill = subheader_fill
        ws.merge_cells(f'A{row_num}:{get_column_letter(len(headers))}{row_num}')
        row_num += 1

        for agent in sorted(employer_agents, key=lambda a: (a.role_type or '', str(a))):
            d = data.get(agent.pk, {})
            ws.append([
                str(agent),
                agent.user.get_full_name(),
                agent.employee_id or '',
                _primary_billable_username(agent),
                str(agent.supervisor) if agent.supervisor else '',
                agent.get_role_type_display() or '',
                agent.employer,
                float(d.get('final_hrs', 0)),
                float(d.get('billing_rate_usd', settings.billing_rate_usd)),
                float(d.get('billing_usd', 0)),
            ])
            row_num += 1

    # Totals
    ws.append([])
    row_num += 1
    total_hrs = sum(d.get('final_hrs', Decimal('0')) for d in data.values())
    total_usd = sum(d.get('billing_usd', Decimal('0')) for d in data.values())
    ph_topup_mxn = sum(d.get('ph_topup_mxn', Decimal('0')) for d in data.values())
    ot_1_5_topup_mxn = sum(d.get('ot_1_5_topup_mxn', Decimal('0')) for d in data.values())
    ph_topup_usd = (ph_topup_mxn / settings.usd_to_mxn).quantize(Decimal('0.01')) if settings.usd_to_mxn else Decimal('0')
    ot_1_5_topup_usd = (ot_1_5_topup_mxn / settings.usd_to_mxn).quantize(Decimal('0.01')) if settings.usd_to_mxn else Decimal('0')
    bonus_mxn = sum(d.get('bonus_mxn', Decimal('0')) for d in data.values())
    bonus_usd = (bonus_mxn / settings.usd_to_mxn).quantize(Decimal('0.01')) if settings.usd_to_mxn else Decimal('0')

    ws.append(['TOTAL', '', '', '', '', '', '', float(total_hrs), '', float(total_usd)])
    ws.append(['Power Hour Top-up (MXN→USD equiv.)', '', '', '', '', '', '', '', '', float(ph_topup_usd)])
    ws.append(['1.5x OT Top-up (MXN→USD equiv.)', '', '', '', '', '', '', '', '', float(ot_1_5_topup_usd)])
    ws.append(['Adherence Bonus Total (USD equiv.)', '', '', '', '', '', '', '', '', float(bonus_usd)])

    # Column widths
    col_widths = [22, 22, 14, 24, 20, 18, 12, 16, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="billing_{week_start.isoformat()}.xlsx"'
    )
    wb.save(response)
    log_action(request.user, 'Billing report exported', f'Week {week_start.isoformat()}')
    return response


@login_required
@finance_access_required
def payroll_report(request):
    week_start = _get_week_start(request)
    settings = BillingSettings.get_for_week(week_start)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    agents = Agent.objects.filter(
        Q(track_attendance=True) | Q(five9_profiles__billable=True) |
        Q(status='inactive', separations__status='finalized',
          separations__remove_from_adherence_date__gt=week_start),
    ).exclude(
        Q(status='inactive') &
        Q(separations__status='finalized') &
        Q(separations__remove_from_adherence_date__lte=week_start)
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles', 'separations').distinct()

    data = _get_billable_weekly_data(list(agents), week_dates, settings)

    infinity_rows = []
    lcc_rows = []
    for agent in agents:
        d = data.get(agent.pk, {})
        row = {**d, 'agent': agent}
        if agent.employer == 'LCC':
            lcc_rows.append(row)
        else:
            infinity_rows.append(row)

    infinity_rows.sort(key=lambda r: str(r['agent']))
    lcc_rows.sort(key=lambda r: str(r['agent']))

    infinity_totals = {
        'base_pay_mxn': sum(r.get('base_pay_mxn', Decimal('0')) for r in infinity_rows),
        'ph_topup_mxn': sum(r.get('ph_topup_mxn', Decimal('0')) for r in infinity_rows),
        'ot_1_5_topup_mxn': sum(r.get('ot_1_5_topup_mxn', Decimal('0')) for r in infinity_rows),
        'bonus_mxn': sum(r.get('bonus_mxn', Decimal('0')) for r in infinity_rows),
        'admin_bonus_mxn': sum(r.get('admin_bonus_mxn', Decimal('0')) for r in infinity_rows),
        'total_pay_mxn': sum(r.get('total_pay_mxn', Decimal('0')) for r in infinity_rows),
        'total_pay_usd': sum(r.get('total_pay_usd', Decimal('0')) for r in infinity_rows),
    }

    prev_week = (week_start - timedelta(days=7)).isoformat()
    next_week = (week_start + timedelta(days=7)).isoformat()

    return render(request, 'finance/payroll.html', {
        'settings': settings,
        'week_start': week_start,
        'week_end': week_end,
        'prev_week': prev_week,
        'next_week': next_week,
        'infinity_rows': infinity_rows,
        'lcc_rows': lcc_rows,
        'infinity_totals': infinity_totals,
    })


@login_required
@finance_access_required
def payroll_export(request):
    """Export payroll report as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    week_start = _get_week_start(request)
    settings = BillingSettings.get_for_week(week_start)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    agents = Agent.objects.filter(
        Q(track_attendance=True) | Q(five9_profiles__billable=True) |
        Q(status='inactive', separations__status='finalized',
          separations__remove_from_adherence_date__gt=week_start),
    ).exclude(
        Q(status='inactive') &
        Q(separations__status='finalized') &
        Q(separations__remove_from_adherence_date__lte=week_start)
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles', 'separations').distinct()

    data = _get_billable_weekly_data(list(agents), week_dates, settings)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    section_fill = PatternFill("solid", fgColor="E5E7EB")
    section_font = Font(bold=True, color="374151")
    center = Alignment(horizontal='center')

    headers = [
        'Agent Name', 'Legal Name', 'Employee ID', 'Supervisor', 'Agent Type',
        'Worked Hrs', 'Hourly Rate (MXN)',
        'Base Pay (MXN)', 'Adh. Bonus (MXN)', 'Admin Bonus (MXN)',
        '1.5x Top-up (MXN)', 'PH Top-up (MXN)',
        'Comm. Ded. %', 'Comm. Earned (MXN)',
        'Total Pay (MXN)', 'Total Pay (USD equiv.)',
    ]

    ws.append([f"Payroll Report — Week of {week_start.strftime('%B %d, %Y')} to {week_end.strftime('%B %d, %Y')}"])
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    row_num = 4
    for section_label, section_agents in [
        ('Infinity Employees', [a for a in agents if a.employer == 'Infinity']),
        ('LCC Direct Employees', [a for a in agents if a.employer == 'LCC']),
    ]:
        if not section_agents:
            continue
        ws.append([section_label])
        ws.cell(row=row_num, column=1).font = section_font
        ws.cell(row=row_num, column=1).fill = section_fill
        ws.merge_cells(f'A{row_num}:{get_column_letter(len(headers))}{row_num}')
        row_num += 1

        for agent in sorted(section_agents, key=lambda a: str(a)):
            d = data.get(agent.pk, {})
            ws.append([
                str(agent),
                agent.user.get_full_name(),
                agent.employee_id or '',
                str(agent.supervisor) if agent.supervisor else '',
                agent.get_role_type_display() or '',
                float(d.get('final_hrs', 0)),
                float(agent.hourly_rate or 0),
                float(d.get('base_pay_mxn', 0)),
                float(d.get('bonus_mxn', 0)),
                float(d.get('admin_bonus_mxn', 0)),
                float(d.get('ot_1_5_topup_mxn', 0)),
                float(d.get('ph_topup_mxn', 0)),
                float(d.get('commission_pct', 0)),
                '—',
                float(d.get('total_pay_mxn', 0)),
                float(d.get('total_pay_usd', 0)),
            ])
            row_num += 1

    col_widths = [22, 22, 14, 20, 18, 12, 18, 16, 16, 16, 16, 16, 14, 18, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="payroll_{week_start.isoformat()}.xlsx"'
    )
    wb.save(response)
    log_action(request.user, 'Payroll report exported', f'Week {week_start.isoformat()}')
    return response


@login_required
@finance_access_required
def finance_settings(request):
    today = date.today()
    current_week = today - timedelta(days=today.weekday())
    singleton = BillingSettings.get()

    if request.method == 'POST':
        try:
            week_str = request.POST.get('effective_week', '').strip()
            try:
                effective_week = date.fromisoformat(week_str)
                effective_week = effective_week - timedelta(days=effective_week.weekday())
            except (ValueError, TypeError):
                effective_week = current_week

            new_vals = {
                'billing_rate_usd':       Decimal(request.POST.get('billing_rate_usd', str(singleton.billing_rate_usd))),
                'usd_to_mxn':             Decimal(request.POST.get('usd_to_mxn', str(singleton.usd_to_mxn))),
                'nr_cap_regular_hours':   Decimal(request.POST.get('nr_cap_regular_hours', str(singleton.nr_cap_regular_hours))),
                'nr_cap_kill_team_hours': Decimal(request.POST.get('nr_cap_kill_team_hours', str(singleton.nr_cap_kill_team_hours))),
                'default_admin_bonus_mxn':Decimal(request.POST.get('default_admin_bonus_mxn', str(singleton.default_admin_bonus_mxn))),
                'adherence_bonus_max_mxn':Decimal(request.POST.get('adherence_bonus_max_mxn', str(singleton.adherence_bonus_max_mxn))),
                'adherence_bonus_full_hours': Decimal(request.POST.get('adherence_bonus_full_hours', str(singleton.adherence_bonus_full_hours))),
                'nr_ratio':               Decimal(request.POST.get('nr_ratio', str(singleton.nr_ratio))),
                'default_tardy_hours':    Decimal(request.POST.get('default_tardy_hours', str(singleton.default_tardy_hours))),
            }

            with transaction.atomic():
                BillingSettingsHistory.objects.create(
                    week_start=effective_week,
                    changed_by=request.user,
                    **new_vals,
                )
                for field, value in new_vals.items():
                    setattr(singleton, field, value)
                singleton.save()

            log_action(request.user, 'Finance settings changed',
                       f'Effective week {effective_week.isoformat()}: ' +
                       ', '.join(f'{k}={v}' for k, v in new_vals.items()))
            messages.success(request, f"Settings saved — effective from week of {effective_week.strftime('%b %d, %Y')}.")
        except Exception as e:
            messages.error(request, f"Error saving settings: {e}")
        return redirect('finance_settings')

    # Current effective settings (latest history record)
    current = BillingSettings.get_for_week(current_week)
    history = BillingSettingsHistory.objects.order_by('-week_start', '-changed_at')[:50]

    return render(request, 'finance/settings.html', {
        'settings': current,
        'singleton': singleton,
        'history': history,
        'current_week': current_week,
    })


# ─── Admin Codings ────────────────────────────────────────────────────────────

@login_required
@admin_tabs_access_required
def admin_codings(request):
    """Codings for Official Admins and coordinators against their billable Five9 user."""
    week_start = _get_week_start(request)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    # Official Admins who have a billable Five9 profile
    agents = Agent.objects.filter(
        status='active',
        five9_profiles__billable=True,
        is_official_admin=True,
    ).distinct().select_related('user', 'supervisor__user').order_by(
        'user__last_name', 'user__first_name'
    )
    _, team_pks = _admin_tabs_access(request.user)
    if team_pks is not None:
        agents = agents.filter(pk__in=team_pks)

    # Billable username display map
    agent_ids = [a.pk for a in agents]
    billable_display_map = {}
    for p in Five9Profile.objects.filter(
        agent__in=agent_ids, billable=True
    ).values('agent_id', 'five9_username', 'is_primary').order_by('agent_id', '-is_primary', 'id'):
        if p['agent_id'] not in billable_display_map:
            billable_display_map[p['agent_id']] = p['five9_username']

    # Build coding map (admin codings only)
    codings_qs = Coding.objects.filter(
        date__in=week_dates, agent__in=agents, is_admin_coding=True
    ).select_related('agent__user').order_by('start_time')
    coding_map = {}
    for c in codings_qs:
        coding_map.setdefault((c.agent_id, c.date), []).append(c)

    rows = []
    for agent in agents:
        cells = []
        agent_total_seconds = 0
        for day_date in week_dates:
            entries = coding_map.get((agent.pk, day_date), [])
            day_seconds = sum(e.total_seconds_count() for e in entries)
            agent_total_seconds += day_seconds
            cells.append({'date': day_date, 'entries': entries, 'total_seconds': day_seconds})
        rows.append({
            'agent': agent,
            'billable_five9_username': billable_display_map.get(agent.pk, ''),
            'cells': cells,
            'total_seconds': agent_total_seconds,
        })

    return render(request, 'finance/admin_codings.html', {
        'rows': rows,
        'week_dates': week_dates,
        'week_start': week_start,
        'week_end': week_end,
        'today': timezone.localdate(),
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
    })


@login_required
@finance_access_required
def codings_export(request):
    """Export a one-row-per-agent weekly coding matrix (regular + admin combined),
    with Mon-Sun day totals and a weekly total. Read-only — no coding/adherence/payroll
    data is touched."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    week_start = _get_week_start(request)
    week_dates = _week_dates(week_start)

    # Same active-OR-still-in-pay-window rule as _get_adherence_agent_pks /
    # the Users export's pay_window_q / billing_report — a finalized separation
    # still counts while remove_from_adherence_date is after this week's Monday.
    # Applied once, with no role/admin/billable narrowing: every agent in the
    # pay window gets a row, zero-filled if they have no codings that week.
    pay_window_q = Q(status='active') | Q(
        status='inactive', separations__status='finalized',
        separations__remove_from_adherence_date__gt=week_start,
    )
    agents = Agent.objects.filter(pay_window_q).exclude(employer='LCC').distinct().select_related(
        'user', 'supervisor__user'
    )

    # Both regular and admin codings summed together, per agent per day.
    codings = Coding.objects.filter(date__in=week_dates, agent__in=agents)
    day_seconds = {}
    for c in codings:
        key = (c.agent_id, c.date)
        day_seconds[key] = day_seconds.get(key, 0) + c.total_seconds_count()

    primary_map = {}
    for p in Five9Profile.objects.filter(
        agent__in=agents, is_primary=True
    ).values('agent_id', 'five9_username'):
        primary_map.setdefault(p['agent_id'], p['five9_username'])

    def _display_name(agent):
        return agent.user.get_full_name() or agent.agent_name or ''

    def _supervisor_name(agent):
        return str(agent.supervisor) if agent.supervisor_id else ''

    agents_sorted = sorted(agents, key=lambda a: (_supervisor_name(a), _display_name(a)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Codings"

    header_font = Font(bold=True, color="FFFFFF")
    identity_header_fill = PatternFill("solid", fgColor="5B9BD5")
    day_header_fill = PatternFill("solid", fgColor="A6A6A6")
    identity_data_fill = PatternFill("solid", fgColor="BDD7EE")
    day_data_fill = PatternFill("solid", fgColor="D9D9D9")
    day_data_font = Font(bold=True)

    headers = [
        'Username', 'ID', 'LEGAL NAMES', 'Supervisor',
        'Mon Total', 'Tues Total', 'Wed Total', 'Thur Total', 'Fri Total', 'Sat Total', 'Sun Total',
        None,
        'Total', 'Total Decimal',
    ]
    IDENTITY_COLS = (1, 2, 3, 4, 13, 14)
    DAY_COLS = range(5, 12)  # E-K

    ws.append(headers)
    for col in IDENTITY_COLS:
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = identity_header_fill
    for col in DAY_COLS:
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = day_header_fill

    for agent in agents_sorted:
        day_values = []
        week_seconds = 0
        for day_date in week_dates:
            secs = day_seconds.get((agent.pk, day_date), 0)
            week_seconds += secs
            day_values.append(timedelta(seconds=secs))

        row = (
            [primary_map.get(agent.pk, '') or '', agent.employee_id or '',
             _display_name(agent), _supervisor_name(agent)]
            + day_values
            + [None, timedelta(seconds=week_seconds), round(week_seconds / 3600, 2)]
        )
        ws.append(row)
        r = ws.max_row
        for col in IDENTITY_COLS:
            ws.cell(row=r, column=col).fill = identity_data_fill
        for col in DAY_COLS:
            cell = ws.cell(row=r, column=col)
            cell.fill = day_data_fill
            cell.font = day_data_font
            cell.number_format = 'h:mm:ss'
        ws.cell(row=r, column=13).number_format = '[h]:mm:ss'
        ws.cell(row=r, column=14).number_format = '0.00'

    col_widths = {1: 17, 2: 6, 3: 51, 4: 14, 12: 4, 13: 9, 14: 12}
    for col in DAY_COLS:
        col_widths[col] = 9
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="codings_{week_start.isoformat()}.xlsx"'
    )
    wb.save(response)
    log_action(request.user, 'Codings exported', f'Week {week_start.isoformat()}')
    return response


def _split_agent_display_name(name):
    """Split a call-center display name into (first, last) on the LAST space."""
    name = (name or '').strip()
    if not name:
        return '', ''
    if ' ' not in name:
        return name, ''
    first, last = name.rsplit(' ', 1)
    return first, last


@login_required
@finance_access_required
def billing_export_v2(request):
    """Billing Report v2 — one row per agent, weekly payroll matrix (login, not-ready,
    coded, allowed-NR, deduction, final hours). Reuses _get_billable_weekly_data's
    NR/final-hours math exactly; adds no new not-ready calculations. Read-only —
    does not touch billing_report/billing_export or any hours/payroll data."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from adherence.views import _build_maps, _compute_shift_hours, _compute_effective_scheduled_hours

    week_start = _get_week_start(request)
    settings = BillingSettings.get_for_week(week_start)
    week_dates = _week_dates(week_start)

    # Same active-OR-still-in-pay-window rule as codings_export / the Users
    # export's pay_window_q / billing_report — a finalized separation still
    # counts while remove_from_adherence_date is after this week's Monday.
    # Applied once, with no role/admin/billable narrowing: every agent in the
    # pay window gets a row, zero-filled if they have no hours that week.
    pay_window_q = Q(status='active') | Q(
        status='inactive', separations__status='finalized',
        separations__remove_from_adherence_date__gt=week_start,
    )
    agents = Agent.objects.filter(pay_window_q).exclude(employer='LCC').distinct().select_related(
        'user', 'supervisor__user'
    ).prefetch_related('five9_profiles', 'separations')

    data = _get_billable_weekly_data(list(agents), week_dates, settings)

    # Shift Hours: schedule-only data via _build_maps (pure data gathering, no bonus/
    # NR-cap/variance math) + the shared _compute_shift_hours calculation — the same
    # single source of truth the combined Adherence export's Shift Hours column uses.
    shift_map, record_map, _, ot_map, extra_hrs_map, _, tmpl_by_agent_dow = _build_maps(list(agents), week_dates)

    primary_map = {}
    for p in Five9Profile.objects.filter(
        agent__in=agents, is_primary=True
    ).values('agent_id', 'five9_username'):
        primary_map.setdefault(p['agent_id'], p['five9_username'])

    agents_sorted = sorted(agents, key=lambda a: (a.agent_name or ''))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing v2"

    header_font = Font(bold=True)

    headers = [
        'AGENT/ADMIN user name', 'AGENT FIRST NAME', 'AGENT LAST NAME', 'Shift Hours',
        'LOGIN TIME', 'NOT READY TIME', 'Coded time', 'Total connected time',
        'Allowed Not Ready', 'Time that should be deducted for going over NR Allowed',
        'Total work time after deduction of Not Ready in Decimal',
        'Total work time after deduction of Not Ready',
        'Scheduled Hours',
    ]
    ws.append(headers)
    for col in range(2, 14):  # B-M bold, A left default
        ws.cell(row=1, column=col).font = header_font

    for agent in agents_sorted:
        d = data.get(agent.pk, {})
        first, last = _split_agent_display_name(agent.agent_name)

        shift_hrs = _compute_shift_hours(agent.pk, week_dates, shift_map, ot_map, extra_hrs_map, tmpl_by_agent_dow)
        # "Scheduled Hours" — the adherence-page value: regular schedule + OT, cut by VTO.
        sched_hrs_eff = _compute_effective_scheduled_hours(
            agent.pk, week_dates, shift_map, ot_map, extra_hrs_map, tmpl_by_agent_dow, record_map)
        login_hrs = d.get('actual_hrs', Decimal('0'))
        nr_hrs = d.get('total_nr_hrs', Decimal('0'))
        coded_hrs = d.get('coded_hrs', Decimal('0'))
        connected_hrs = d.get('pre_cap_total', login_hrs + coded_hrs)
        deduction_hrs = d.get('nr_deduction', Decimal('0'))
        final_hrs = d.get('final_hrs', connected_hrs - deduction_hrs)
        allowed_hrs = d.get('nr_allowed_hrs', Decimal('0'))

        def _secs(hrs):
            return timedelta(seconds=round(float(hrs) * 3600))

        row = [
            primary_map.get(agent.pk, '') or '',
            first,
            last,
            _secs(shift_hrs),
            _secs(login_hrs),
            _secs(nr_hrs),
            _secs(coded_hrs),
            _secs(connected_hrs),
            _secs(allowed_hrs),
            _secs(deduction_hrs),
            round(float(final_hrs), 2),
            _secs(final_hrs),
            _secs(sched_hrs_eff),
        ]
        ws.append(row)
        r = ws.max_row
        ws.cell(row=r, column=4).number_format = '[h]:mm:ss'
        ws.cell(row=r, column=5).number_format = '[h]:mm:ss'
        ws.cell(row=r, column=6).number_format = 'h:mm:ss'
        ws.cell(row=r, column=7).number_format = '[h]:mm:ss;@'
        ws.cell(row=r, column=8).number_format = '[h]:mm:ss;@'
        ws.cell(row=r, column=9).number_format = '[h]:mm:ss'
        ws.cell(row=r, column=10).number_format = '[h]:mm:ss'
        ws.cell(row=r, column=11).number_format = '0.00'
        ws.cell(row=r, column=12).number_format = '[h]:mm:ss'
        ws.cell(row=r, column=13).number_format = '[h]:mm:ss'

    col_widths = {1: 22, 2: 20, 3: 19, 4: 14, 5: 12, 6: 17, 7: 12, 8: 21, 9: 19, 10: 55, 11: 56, 12: 44, 13: 16}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="billing_v2_{week_start.isoformat()}.xlsx"'
    )
    wb.save(response)
    log_action(request.user, 'Billing report v2 exported', f'Week {week_start.isoformat()}')
    return response


@login_required
@admin_tabs_access_required
@require_POST
def add_admin_coding_ajax(request):
    from adherence.views import _attendance_edit_denied

    data = json.loads(request.body)
    agent_id = data.get('agent_id')
    date_str = data.get('date')
    start_time = data.get('start_time', '').strip()
    end_time = data.get('end_time', '').strip()
    notes = data.get('notes', '')

    if not all([agent_id, date_str, start_time, end_time]):
        return JsonResponse({'ok': False, 'error': 'missing fields'}, status=400)

    if _attendance_edit_denied(request.user, agent_id):
        return JsonResponse({'ok': False, 'error': 'Not permitted for this agent.'}, status=403)

    def _pad(s):
        parts = s.split(':')
        if parts:
            parts[0] = parts[0].zfill(2)
        return ':'.join(parts)

    start_time = _pad(start_time)
    end_time = _pad(end_time)

    try:
        start_t = time_cls.fromisoformat(start_time)
        end_t = time_cls.fromisoformat(end_time)
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Invalid time format. Use H:MM:SS'}, status=400)

    if end_t <= start_t:
        return JsonResponse({'ok': False, 'error': 'End time must be after start time.'}, status=400)

    try:
        coding = Coding.objects.create(
            agent_id=agent_id, date=date_str,
            start_time=start_time, end_time=end_time,
            notes=notes, is_admin_coding=True,
        )
        coding.refresh_from_db()
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)

    log_action(request.user, 'Admin coding added',
               f'Agent {agent_id} on {date_str}: {start_time}–{end_time}')
    return JsonResponse({
        'ok': True, 'id': coding.pk,
        'hhmmss': coding.total_hhmmss(),
        'start': coding.start_time.strftime('%H:%M'),
        'end': coding.end_time.strftime('%H:%M'),
        'start_full': coding.start_time.strftime('%H:%M:%S'),
        'end_full': coding.end_time.strftime('%H:%M:%S'),
        'notes': coding.notes,
    })


@login_required
@admin_tabs_access_required
@require_POST
def edit_admin_coding_ajax(request):
    from adherence.views import _attendance_edit_denied

    data = json.loads(request.body)
    coding_id = data.get('coding_id')
    start_time = data.get('start_time', '').strip()
    end_time = data.get('end_time', '').strip()
    notes = data.get('notes', '')

    coding = Coding.objects.filter(pk=coding_id, is_admin_coding=True).first()
    if not coding:
        return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)

    if _attendance_edit_denied(request.user, coding.agent_id):
        return JsonResponse({'ok': False, 'error': 'Not permitted for this agent.'}, status=403)

    def _pad(s):
        parts = s.split(':')
        if parts:
            parts[0] = parts[0].zfill(2)
        return ':'.join(parts)

    try:
        start_t = time_cls.fromisoformat(_pad(start_time))
        end_t = time_cls.fromisoformat(_pad(end_time))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Invalid time format.'}, status=400)

    if end_t <= start_t:
        return JsonResponse({'ok': False, 'error': 'End time must be after start time.'}, status=400)

    coding.start_time = start_t
    coding.end_time = end_t
    coding.notes = notes
    coding.save()

    log_action(request.user, 'Admin coding edited',
               f'Coding #{coding_id}: {start_t}–{end_t}')
    return JsonResponse({
        'ok': True, 'id': coding.pk,
        'hhmmss': coding.total_hhmmss(),
        'start': coding.start_time.strftime('%H:%M'),
        'end': coding.end_time.strftime('%H:%M'),
        'start_full': coding.start_time.strftime('%H:%M:%S'),
        'end_full': coding.end_time.strftime('%H:%M:%S'),
        'notes': coding.notes,
    })


@login_required
@admin_tabs_access_required
@require_POST
def delete_admin_coding_ajax(request):
    from adherence.views import _attendance_edit_denied

    data = json.loads(request.body)
    coding_id = data.get('coding_id')

    coding = Coding.objects.filter(pk=coding_id, is_admin_coding=True).first()
    if not coding:
        return JsonResponse({'ok': True})

    if _attendance_edit_denied(request.user, coding.agent_id):
        return JsonResponse({'ok': False, 'error': 'Not permitted for this agent.'}, status=403)

    coding.delete()
    log_action(request.user, 'Admin coding deleted', f'Coding #{coding_id}')
    return JsonResponse({'ok': True})


# ─── Admin Adherence ──────────────────────────────────────────────────────────

def _apply_live_login_hours(agents, week_dates, record_map):
    """
    Overwrite in-memory AdherenceRecord.actual_hours with LIVE raw billable
    login hours per (agent, date), summed across ALL billable Five9 profiles —
    the exact source Billing Report v2 uses.

    Why: the admin adherence tab otherwise reads the STORED actual_hours, which
    already had a per-day not-ready deduction baked in at upload time;
    _build_rows then re-applies the weekly NR cap on top, so Official Admins
    show false "missing time" even though the connected time is present (and
    Billing v2 shows it correctly). This also fixes admins with 2+ billable
    Five9 profiles, whose stored value held only the last profile's login
    instead of the sum.

    The record_map objects are request-scoped and are NEVER saved — this
    changes only what this page/export displays, not the database, Billing v2,
    the pay engine, or the regular adherence dashboard.
    """
    if not agents:
        return
    billable_map, _ = get_billable_username_map([a.pk for a in agents])
    login_hrs = {}
    for r in DailyAgentHours.objects.filter(
        upload__date__in=week_dates, agent__in=agents
    ).values('agent_id', 'upload__date', 'five9_username', 'login_seconds'):
        aid = r['agent_id']
        if aid is None:
            continue
        bnames = billable_map.get(aid)
        if bnames is None or r['five9_username'].strip().lower() in bnames:
            key = (aid, r['upload__date'])
            login_hrs[key] = login_hrs.get(key, Decimal('0')) + Decimal(str(r['login_seconds'])) / Decimal('3600')
    for (aid, d), hrs in login_hrs.items():
        rec = record_map.get((aid, d))
        if rec is not None:
            rec.actual_hours = hrs
        else:
            record_map[(aid, d)] = AdherenceRecord(agent_id=aid, date=d, actual_hours=hrs, status='')


@login_required
@admin_tabs_access_required
def admin_adherence(request):
    """Adherence tab for Official Admins only. Super admins/owners see all
    Official Admins; other permission holders see only their own team
    (themselves + the Official Admins who have them as supervisor)."""
    from adherence.views import _build_maps, _build_rows
    from adherence.models import AdherenceNote
    from django.db.models import Count as _Count

    week_start = _get_week_start(request)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    agents = Agent.objects.filter(
        status='active',
        is_official_admin=True,
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles').order_by(
        'user__last_name', 'user__first_name'
    )
    _, team_pks = _admin_tabs_access(request.user)
    if team_pks is not None:
        agents = agents.filter(pk__in=team_pks)
    agents = list(agents)

    # Shift/record/OT maps from adherence logic
    shift_map, record_map, _, ot_map, extra_hrs_map, split_labels_map, tmpl_by_agent_dow = _build_maps(agents, week_dates)

    # Coded map uses admin codings, not regular codings
    coded_map = {}
    for c in Coding.objects.filter(date__in=week_dates, agent__in=agents, is_admin_coding=True):
        coded_map[(c.agent_id, c.date)] = coded_map.get((c.agent_id, c.date), Decimal('0')) + Decimal(str(c.total_hours()))

    # Reflect LIVE billable login hours (same source as Billing v2) instead of
    # the stored, double-NR-deducted actual_hours that made Official Admins show
    # false "missing time". Request-scoped; never saved.
    _apply_live_login_hours(agents, week_dates, record_map)

    rows = _build_rows(agents, week_dates, shift_map, record_map, coded_map, ot_map=ot_map,
                       extra_hrs_map=extra_hrs_map, split_labels_map=split_labels_map,
                       tmpl_by_agent_dow=tmpl_by_agent_dow)

    # Replace adherence bonus with fixed admin bonus for each row
    billing_settings = BillingSettings.get_for_week(week_start)
    billable_five9_map = {}
    for p in Five9Profile.objects.filter(
        agent__in=[a.pk for a in agents], billable=True
    ).values('agent_id', 'five9_username', 'is_primary').order_by('agent_id', '-is_primary', 'id'):
        if p['agent_id'] not in billable_five9_map:
            billable_five9_map[p['agent_id']] = p['five9_username']

    for row in rows:
        agent = row['agent']
        admin_bonus = (
            agent.admin_bonus_mxn if agent.admin_bonus_mxn is not None
            else billing_settings.default_admin_bonus_mxn
        )
        row['bonus'] = 'Admin'
        row['bonus_mxn'] = admin_bonus
        row['admin_bonus_mxn'] = admin_bonus
        row['billable_five9_username'] = billable_five9_map.get(agent.pk, '')

    # Note counts
    note_count_map = {
        (n['agent_id'], n['date']): n['count']
        for n in AdherenceNote.objects.filter(
            agent__in=agents, date__in=week_dates
        ).values('agent_id', 'date').annotate(count=_Count('pk'))
    }
    for row in rows:
        for cell in row['cells']:
            cell['note_count'] = note_count_map.get((row['agent'].pk, cell['date']), 0)

    from nomina.models import Holiday, AdminBonusDeduction
    holiday_dates = set(Holiday.objects.filter(date__in=week_dates).values_list('date', flat=True))
    # Coder-entered weekly bonus deduction % per admin (guide-recommended, manual).
    ded_map = {x.agent_id: x.deduction_pct for x in AdminBonusDeduction.objects.filter(
        agent__in=agents, week_start=week_start)}
    for row in rows:
        row['bonus_ded_pct'] = ded_map.get(row['agent'].pk, Decimal('0'))
    return render(request, 'finance/admin_adherence.html', {
        'rows': rows,
        'week_dates': week_dates,
        'week_start': week_start,
        'week_end': week_end,
        'today': timezone.localdate(),
        'holiday_dates': holiday_dates,
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
        'status_choices': AdherenceRecord.STATUS_CHOICES,
        # 0,5,10,…,100 for the deduction dropdown.
        'ded_options': list(range(0, 105, 5)),
    })


@login_required
@admin_tabs_access_required
def admin_penalty_reco(request):
    """GET agent + week → JSON recommended admin-bonus deduction (guide for the alert)."""
    from nomina.views import admin_bonus_penalty
    try:
        agent = Agent.objects.get(pk=request.GET.get('agent'))
    except (Agent.DoesNotExist, ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'agent not found'}, status=404)
    ws = parse_week_param(request.GET.get('week')) or get_week_start()
    override = None
    raw_date = request.GET.get('date')
    if raw_date:
        try:                       # reflect the just-set cell even if its save hasn't landed yet
            override = (date.fromisoformat(raw_date), request.GET.get('status') or '')
        except ValueError:
            override = None
    reco = admin_bonus_penalty(agent, ws, override=override)
    return JsonResponse({'ok': True, 'pct': str(reco['pct']),
                         'reasons': reco['reasons'], 'hours_note': reco['hours_note']})


@login_required
@admin_tabs_access_required
@require_POST
def save_admin_deduction(request):
    """Save the coder-entered weekly admin-bonus deduction % for one admin/week."""
    from nomina.models import AdminBonusDeduction
    data = json.loads(request.body)
    try:
        agent = Agent.objects.get(pk=data.get('agent_id'))
    except (Agent.DoesNotExist, ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'agent not found'}, status=404)
    ws = parse_week_param(data.get('week')) or get_week_start()
    try:
        pct = Decimal(str(data.get('pct') or '0'))
    except (InvalidOperation, ValueError):
        return JsonResponse({'ok': False, 'error': 'bad pct'}, status=400)
    pct = max(Decimal('0'), min(Decimal('100'), pct))
    from wfm.utils import retry_on_locked
    retry_on_locked(lambda: AdminBonusDeduction.objects.update_or_create(
        agent=agent, week_start=ws,
        defaults={'deduction_pct': pct, 'note': (data.get('note') or '').strip()[:255],
                  'updated_by': getattr(request.user, 'agent', None)}))
    return JsonResponse({'ok': True, 'pct': str(pct)})


@login_required
@admin_tabs_access_required
def admin_adherence_export(request):
    """Export admin adherence payroll as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from adherence.views import _build_maps, _build_rows

    week_start = _get_week_start(request)
    billing_settings = BillingSettings.get_for_week(week_start)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    agents = Agent.objects.filter(
        status='active',
        is_official_admin=True,
    ).select_related('user', 'supervisor__user').prefetch_related('five9_profiles').order_by(
        'user__last_name', 'user__first_name'
    )
    _, team_pks = _admin_tabs_access(request.user)
    if team_pks is not None:
        agents = agents.filter(pk__in=team_pks)
    agents = list(agents)

    shift_map, record_map, _, ot_map, extra_hrs_map, split_labels_map, tmpl_by_agent_dow = _build_maps(agents, week_dates)

    coded_map = {}
    for c in Coding.objects.filter(date__in=week_dates, agent__in=agents, is_admin_coding=True):
        coded_map[(c.agent_id, c.date)] = coded_map.get((c.agent_id, c.date), Decimal('0')) + Decimal(str(c.total_hours()))

    # Reflect LIVE billable login hours (same source as Billing v2) instead of
    # the stored, double-NR-deducted actual_hours that made Official Admins show
    # false "missing time". Request-scoped; never saved.
    _apply_live_login_hours(agents, week_dates, record_map)

    rows = _build_rows(agents, week_dates, shift_map, record_map, coded_map, ot_map=ot_map,
                       extra_hrs_map=extra_hrs_map, split_labels_map=split_labels_map,
                       tmpl_by_agent_dow=tmpl_by_agent_dow)

    billable_five9_map = {}
    for p in Five9Profile.objects.filter(
        agent__in=[a.pk for a in agents], billable=True
    ).values('agent_id', 'five9_username', 'is_primary').order_by('agent_id', '-is_primary', 'id'):
        if p['agent_id'] not in billable_five9_map:
            billable_five9_map[p['agent_id']] = p['five9_username']

    usd_to_mxn = billing_settings.usd_to_mxn

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admin Payroll"

    headers = [
        'Agent Name', 'Legal Name', 'Employee ID', 'Five9 User',
        'Sch Hrs', 'Login Hrs', 'Coded Hrs', 'NR Cap Adj', 'Total Hrs',
        'Hourly Rate (MXN)', 'Base Pay (MXN)', 'Admin Bonus (MXN)',
        'Total Pay (MXN)', 'Total Pay (USD)',
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1A3A5C')
    center = Alignment(horizontal='center')

    ws.append([f"Admin Payroll — Week of {week_start.strftime('%B %d, %Y')} to {week_end.strftime('%B %d, %Y')}"])
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for row in rows:
        agent = row['agent']
        admin_bonus = (
            agent.admin_bonus_mxn if agent.admin_bonus_mxn is not None
            else billing_settings.default_admin_bonus_mxn
        )
        hourly_mxn = agent.hourly_rate or Decimal('0')
        final_hrs = row['final_adjusted']
        base_pay = (final_hrs * hourly_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP)
        total_pay_mxn = base_pay + admin_bonus
        total_pay_usd = (total_pay_mxn / usd_to_mxn).quantize(Decimal('0.01'), ROUND_HALF_UP) if usd_to_mxn else Decimal('0')

        ws.append([
            str(agent),
            agent.user.get_full_name(),
            agent.employee_id or '',
            billable_five9_map.get(agent.pk, ''),
            float(row['sched_hours'] or 0),
            float(row['actual_hours'] or 0),
            float(row['coded_hours'] or 0),
            float(row.get('nr_cap_adj') or 0),
            float(final_hrs or 0),
            float(hourly_mxn),
            float(base_pay),
            float(admin_bonus),
            float(total_pay_mxn),
            float(total_pay_usd),
        ])

    col_widths = [22, 22, 14, 20, 10, 10, 10, 12, 10, 18, 16, 18, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="admin_payroll_{week_start.isoformat()}.xlsx"'
    )
    wb.save(response)
    return response


# ─── Combined Adherence export (regular + Official Admins) ───────────────────

def _adherence_export_hhmmss(decimal_hours):
    """Decimal hours -> 'H:MM:SS' string. Mirrors adherence.templatetags.adherence_filters.to_hhmmss."""
    if not decimal_hours:
        return ''
    total_seconds = round(float(decimal_hours) * 3600)
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f'{h}:{m:02d}:{s:02d}'


def _adherence_export_hex(color):
    """Normalize a CSS hex color (e.g. '#fff' or '#e8f5e9') to a 6-digit hex for openpyxl."""
    hex_str = color.lstrip('#')
    if len(hex_str) == 3:
        hex_str = ''.join(ch * 2 for ch in hex_str)
    return hex_str


def _adherence_export_display_name(agent):
    return agent.user.get_full_name() or agent.agent_name or agent.user.username


def _adherence_export_supervisor_name(agent):
    return str(agent.supervisor) if agent.supervisor_id else ''


def _write_adherence_sheet(ws, rows, week_dates, week_start, week_end, commission_map,
                            empty_note=None, title_suffix=None):
    """Write one Adherence-style worksheet (title, headers, per-agent daily-status rows,
    column widths) into `ws`. Shared by the combined export's main sheet and its VTO-only
    sheet so both stay pixel-identical except for which rows they contain."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    day_headers = [d.strftime('%a') for d in week_dates]
    headers = [
        'Username', 'Employee ID', 'Legal Name', 'Agent Name', 'Supervisor',
        'Commission Deduction %', 'Scheduled Hours', 'Shift Hours',
    ] + day_headers

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1A3A5C')
    center = Alignment(horizontal='center')
    wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    met_font = Font(color='166534')
    short_font = Font(color='C0392B')

    ws.append([
        f"Adherence — Week of {week_start.strftime('%B %d, %Y')} to "
        f"{week_end.strftime('%B %d, %Y')}{title_suffix or ''}"
    ])
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    if not rows and empty_note:
        ws.append([empty_note])
        ws.merge_cells(f'A4:{get_column_letter(len(headers))}4')
        ws['A4'].font = Font(italic=True)
        ws['A4'].alignment = center
    else:
        for row in rows:
            agent = row['agent']
            commission_pct = commission_map.get(agent.pk, Decimal('0'))

            ws.append([
                row.get('billable_five9_username', ''),
                agent.employee_id or '',
                _adherence_export_display_name(agent),
                agent.agent_name or '',
                _adherence_export_supervisor_name(agent),
                float(commission_pct),
                timedelta(hours=float(row['sched_hours'] or 0)),
                timedelta(hours=float(row['shift_hours'] or 0)),
            ] + [None] * 7)
            r = ws.max_row
            ws.cell(row=r, column=6).number_format = '0.0"%"'
            ws.cell(row=r, column=7).number_format = '[h]:mm:ss'
            ws.cell(row=r, column=8).number_format = '[h]:mm:ss'

            for i, cell in enumerate(row['cells']):
                col = 9 + i
                status = cell['status']
                if cell['missing_hrs'] is not None:
                    hrs_line = f"-{_adherence_export_hhmmss(cell['missing_hrs'])}"
                    font = short_font
                elif cell['display_hrs'] is not None:
                    hrs_line = _adherence_export_hhmmss(cell['display_hrs'])
                    font = met_font
                else:
                    hrs_line = None
                    font = Font()

                if status and hrs_line:
                    text = f"{status}\n{hrs_line}"
                elif status:
                    text = status
                    font = Font()
                else:
                    text = None

                xl_cell = ws.cell(row=r, column=col, value=text)
                xl_cell.fill = PatternFill('solid', fgColor=_adherence_export_hex(cell['color']))
                xl_cell.font = font
                xl_cell.alignment = wrap_center

    col_widths = [17, 14, 30, 22, 20, 12, 14, 14] + [11] * 7
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30
    for r in range(4, ws.max_row + 1):
        ws.row_dimensions[r].height = 30


@login_required
@finance_access_required
def adherence_export(request):
    """Export a single combined Adherence workbook: the regular Adherence tab's
    roster plus Official Admins from Admin Adherence, one row per person.
    Sources every value from adherence._build_maps/_build_rows exactly as both
    on-screen tabs already do — no new adherence/bonus/scheduling math."""
    import openpyxl
    from adherence.views import _build_maps, _build_rows, _get_adherence_agent_pks

    week_start = _get_week_start(request)
    week_dates = _week_dates(week_start)
    week_end = week_dates[-1]

    # Regular roster — same PKs the regular Adherence tab shows (excludes
    # Official Admins, applies the pay-window rule for recent separations).
    regular_pks = _get_adherence_agent_pks(week_dates, week_start)
    regular_agents = list(
        Agent.objects.filter(pk__in=regular_pks).exclude(employer='LCC')
        .select_related('user', 'supervisor__user')
        .prefetch_related('five9_profiles')
    )
    shift_map, record_map, coded_map, ot_map, extra_hrs_map, split_labels_map, tmpl_by_agent_dow = (
        _build_maps(regular_agents, week_dates)
    )
    rows = _build_rows(
        regular_agents, week_dates, shift_map, record_map, coded_map, ot_map=ot_map,
        extra_hrs_map=extra_hrs_map, split_labels_map=split_labels_map,
        tmpl_by_agent_dow=tmpl_by_agent_dow,
    )

    # Official Admin roster — same query Admin Adherence uses. No team-scoping:
    # this view is gated by finance_access_required (super admin only), which
    # _admin_tabs_access always resolves to full access for anyway.
    admin_agents = list(
        Agent.objects.filter(status='active', is_official_admin=True).exclude(employer='LCC')
        .select_related('user', 'supervisor__user')
        .prefetch_related('five9_profiles')
    )
    if admin_agents:
        a_shift_map, a_record_map, _, a_ot_map, a_extra_hrs_map, a_split_labels_map, a_tmpl_by_agent_dow = (
            _build_maps(admin_agents, week_dates)
        )
        admin_coded_map = {}
        for c in Coding.objects.filter(date__in=week_dates, agent__in=admin_agents, is_admin_coding=True):
            key = (c.agent_id, c.date)
            admin_coded_map[key] = admin_coded_map.get(key, Decimal('0')) + Decimal(str(c.total_hours()))
        rows += _build_rows(
            admin_agents, week_dates, a_shift_map, a_record_map, admin_coded_map, ot_map=a_ot_map,
            extra_hrs_map=a_extra_hrs_map, split_labels_map=a_split_labels_map,
            tmpl_by_agent_dow=a_tmpl_by_agent_dow,
        )

    rows.sort(key=lambda r: (
        _adherence_export_supervisor_name(r['agent']), _adherence_export_display_name(r['agent'])
    ))

    commission_map = {
        pa.agent_id: pa.commission_deduction
        for pa in PayrollAdjustment.objects.filter(
            agent__in=[r['agent'].pk for r in rows], week_start=week_start
        )
    }

    # VTO-only subset: regular roster (Official Admins excluded), any day this week
    # in VTO_TYPE_STATUSES. Filtered from the already-computed `rows` — no second
    # _build_rows call, no new query.
    vto_rows = [
        row for row in rows
        if not row['agent'].is_official_admin
        and any(cell['status'] in VTO_TYPE_STATUSES for cell in row['cells'])
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adherence"
    _write_adherence_sheet(ws, rows, week_dates, week_start, week_end, commission_map)

    ws_vto = wb.create_sheet(title="VTO Agents")
    _write_adherence_sheet(
        ws_vto, vto_rows, week_dates, week_start, week_end, commission_map,
        empty_note=None if vto_rows else "No VTO was recorded for this week.",
        title_suffix=" — VTO Only",
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="adherence_{week_start.isoformat()}.xlsx"'
    )
    wb.save(response)
    log_action(request.user, 'Adherence report exported', f'Week {week_start.isoformat()}')
    return response


# ─── User Setup Audit export ──────────────────────────────────────────────────

@login_required
@finance_access_required
def user_audit_export(request):
    """One row per Five9 account (or one row per zero-account person), every active and inactive user."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    today = get_week_start()
    settings = BillingSettings.get_for_week(today)

    agents = Agent.objects.all().select_related(
        'user', 'supervisor__user'
    ).prefetch_related('five9_profiles').order_by('user__last_name', 'user__first_name')

    headers = [
        'Legal Name', 'Agent Name', 'Employee ID', 'Login Username', 'Email', 'Phone',
        'Role', 'Role Type', 'Status', 'Supervisor', 'Employer', 'Billing Status',
        'Tracked in Attendance', 'Official Admin', 'Super Admin', 'Admin-Tabs Access',
        'Has Working Login', 'Effective Billing Rate (USD)', 'Rate Source (USD)',
        'Effective Hourly Rate (MXN)', 'Admin Bonus Override (MXN)', 'Team Password',
        '# Billable Five9 Accounts',
        'Five9 Username', 'Five9 Role Type', 'Billable', 'Is Primary', 'Five9 Password',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Audit"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    center = Alignment(horizontal='center')

    ws.append([f"User Setup Audit — as of {today.strftime('%B %d, %Y')}"])
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    def yn(value):
        return 'Yes' if value else 'No'

    for agent in agents:
        supervisor_name = str(agent.supervisor) if agent.supervisor else ''
        billing_override = agent.billing_rate_usd not in (None, Decimal('0'))
        billing_rate = agent.billing_rate_usd if billing_override else settings.billing_rate_usd
        hourly_mxn = agent.hourly_rate or Decimal('0')
        phone = f"{agent.phone_country_code} {agent.phone_number}".strip() if agent.phone_number else ''

        person_cols = [
            agent.user.get_full_name() or '',
            agent.agent_name or '',
            agent.employee_id or '',
            agent.user.username,
            agent.user.email or '',
            phone,
            agent.get_role_display(),
            agent.get_role_type_display() or '',
            agent.get_status_display(),
            supervisor_name,
            agent.employer,
            agent.billing_status,
            yn(agent.track_attendance),
            yn(agent.is_official_admin),
            yn(agent.is_super_admin),
            yn(agent.can_access_admin_tabs),
            yn(agent.user.has_usable_password()),
            float(billing_rate),
            'Override' if billing_override else 'Default',
            float(hourly_mxn),
            float(agent.admin_bonus_mxn) if agent.admin_bonus_mxn is not None else '',
            agent.teams_password or '',
        ]

        five9_profiles = sorted(
            agent.five9_profiles.all(),
            key=lambda p: (not p.is_primary, p.label or p.five9_username or '')
        )
        billable_count = sum(1 for p in five9_profiles if p.billable)

        if five9_profiles:
            for p in five9_profiles:
                ws.append(person_cols + [billable_count, p.five9_username,
                                          p.get_role_type_display() or '',
                                          yn(p.billable), yn(p.is_primary),
                                          p.five9_password or ''])
        else:
            ws.append(person_cols + [0, '— none —', '— none —', '— none —', '— none —', '— none —'])

    col_widths = [22, 20, 12, 16, 26, 16, 10, 16, 10, 20, 10, 12, 12, 10, 10, 12,
                  12, 14, 12, 14, 14, 16, 12, 22, 16, 10, 10, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(headers))}{ws.max_row}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="user_audit_{today.isoformat()}.xlsx"'
    )
    wb.save(response)
    log_action(request.user, 'User setup audit exported', f'As of {today.isoformat()}')
    return response
