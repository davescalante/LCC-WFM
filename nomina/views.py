from datetime import date as date_cls, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from wfm.utils import get_week_start, parse_week_param
from .access import loan_access_required, nomina_access_required
from .models import (
    AdminBonusDeduction, BreakAbuseIncident, Holiday, Loan, NominaOverride, NominaWeek,
    PayrollRun, UnmatchedInputRow, WeeklyPayInput, WelcomeBonusEnrollment,
)


def _snap_rows(rows):
    """Serialize computed nómina rows to JSON-safe dicts (drop the Agent object, keep
    its pk; Decimals → floats). Rows are already dicts, so the templates/export render
    a snapshot identically to a live row."""
    out = []
    for r in rows:
        d = {}
        for k, v in r.items():
            if k == 'agent':
                d['agent_pk'] = v.pk
            else:
                d[k] = float(v) if hasattr(v, 'quantize') else v
        out.append(d)
    return out


def _snap_totals(totals):
    return {k: (float(v) if hasattr(v, 'quantize') else v) for k, v in totals.items()}


def _finalized_run(week_start):
    """The PayrollRun for a finalized week, or None if the week is still open."""
    return PayrollRun.objects.filter(week_start=week_start).first()

# The manual paste-in fields, in display order. (label, field, is_deduction)
# Per-type input modules. Each maps to one WeeklyPayInput field. `aggregate`
# sums all matched rows per person (e.g. a Comedor POS export); otherwise one
# value per person. `unit` is display only (USD spiffs convert at the week rate).
INPUT_TYPES = [
    {'key': 'lpo', 'field': 'lpo', 'label': 'LPO', 'unit': 'MXN', 'deduction': False,
     'aggregate': False, 'match': 'auto',
     'desc': 'Sales commission. Upload with columns Username + Amount.'},
    {'key': 'spiff', 'field': 'spiff_usd', 'label': 'Spiffs', 'unit': 'USD', 'deduction': False,
     'aggregate': True, 'match': 'auto', 'add_more': True,
     'desc': 'Spiffs in USD — a person can appear on many rows and all are summed, '
             'then converted at the week rate. Columns: Agent Username / Agent ID + the $ amount.'},
    {'key': 'hours', 'field': 'extra_hours', 'label': 'Extra Hours', 'unit': 'hrs', 'deduction': False,
     'aggregate': False, 'match': 'auto', 'manual_add': True, 'add_more': True, 'totals': True,
     'show_emp': False, 'name_only_agent': True, 'is_hours': True,
     'desc': 'Manual hours correction (e.g. a Tuesday touch-up) — add hours for someone and they '
             'fold into Hours Worked and Pay (48) on the nómina. Per week; starts empty.'},
    {'key': 'referral', 'field': 'referral', 'label': 'Referral', 'unit': 'MXN', 'deduction': False,
     'aggregate': False, 'match': 'auto', 'manual_add': True, 'totals': True,
     'show_emp': False, 'name_only_agent': True,
     'desc': 'Manual, per-week — add an agent and enter the referral amount. The list starts empty each week.'},
    {'key': 'killqa', 'field': 'kill_team_qa', 'label': 'Kill Team QA', 'unit': 'MXN', 'deduction': False,
     'aggregate': False, 'match': 'auto', 'roles': ['kill_team'],
     'default_amount': 400, 'show_emp': False, 'name_only_agent': True,
     'desc': 'Kill Team QA bonus — Kill Team agents only. Defaults to $400 each; change the exceptions and Save.'},
    {'key': 'comedor', 'field': 'comedor', 'label': 'Comedor', 'unit': 'MXN', 'deduction': True,
     'aggregate': True, 'match': 'auto', 'totals': True, 'sort_by_value': True,
     'desc': 'Cafeteria POS export — matched by Employee # (EMP), all charges per person summed. '
             'Columns: EMP # + Precio.'},
    {'key': 'transportation', 'field': 'transportation', 'label': 'Transportation', 'unit': 'MXN', 'deduction': True,
     'aggregate': False, 'match': 'auto', 'manual_add': True, 'totals': True,
     'show_emp': False, 'name_only_agent': True,
     'desc': 'Manual, per-week — add an agent and enter the amount. The list starts empty each week.'},
]
INPUT_TYPE_BY_KEY = {t['key']: t for t in INPUT_TYPES}


def _week(request):
    raw = request.GET.get('week_start') or request.GET.get('week') or request.POST.get('week_start')
    week_start = parse_week_param(raw) or get_week_start()
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    return week_start, week_dates


def _dec(val):
    """Parse a money-ish cell → Decimal. Tolerates '$', thousands commas, blank
    cells and accounting-style negatives like '(50.00)'. Bad input → 0."""
    s = str(val if val is not None else '').strip()
    if not s:
        return Decimal('0')
    neg = s.startswith('(') and s.endswith(')')
    s = s.replace('$', '').replace(',', '').replace('(', '').replace(')', '').strip()
    try:
        d = Decimal(s or '0')
    except (InvalidOperation, ValueError):
        return Decimal('0')
    return -d if neg else d


def _parse_rate(raw):
    """Parse a scalar rate (e.g. the weekly USD→MXN spiff rate). A rate is always
    a small number, so a comma can ONLY be a decimal separator — never a thousands
    grouping. This is the opposite of _dec (which strips commas as thousands),
    so a Mexican-style '18,50' correctly reads as 18.50, not 1850. Returns None
    for blank/unparseable input so the caller can reject it."""
    s = str(raw or '').strip().replace('$', '').replace(' ', '')
    if not s:
        return None
    if ',' in s and '.' not in s:      # '18,50' → decimal comma
        s = s.replace(',', '.')
    else:                              # '1,850.00' → comma is (unexpected) grouping
        s = s.replace(',', '')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _pay_window(week_start):
    return Q(status='active') | Q(
        status='inactive', separations__status='finalized',
        separations__remove_from_adherence_date__gt=week_start,
    )


def _infinity_agents(week_start):
    """Everyone on the Agent Nómina: ALL Infinity, non-official-admin people in the pay
    window — regardless of attendance tracking or a billable Five9 profile. Sales-type
    agents earn LPO but take no calls (no tracking/Five9), so they must still appear.
    Official admins go on the Admin Nómina instead."""
    from scheduling.models import Agent
    return list(
        Agent.objects.filter(_pay_window(week_start))
        .filter(employer='Infinity')       # INFINITY only — LCC excluded
        .exclude(is_official_admin=True)   # admins go on the Admin Nómina
        .distinct()
        .select_related('user', 'supervisor__user')
        .prefetch_related('five9_profiles')
        .order_by('user__last_name', 'user__first_name')
    )


def _admin_agents(week_start):
    from scheduling.models import Agent
    return list(
        Agent.objects.filter(_pay_window(week_start))
        .filter(is_official_admin=True)
        .filter(employer='Infinity')
        .distinct()
        .select_related('user', 'supervisor__user')
        .prefetch_related('five9_profiles')
        .order_by('user__last_name', 'user__first_name')
    )


def _unrostered_infinity(week_start):
    """Infinity people in the pay window who land on NEITHER nómina — not an official
    admin (so not on the Admin sheet) yet with no attendance tracking and no billable
    profile (so filtered off the Agent sheet). Surfaced as a warning so nobody is
    silently left unpaid; the operator fixes their profile or confirms the exclusion."""
    from scheduling.models import Agent
    rostered = {a.pk for a in _infinity_agents(week_start)} | {a.pk for a in _admin_agents(week_start)}
    return [a for a in Agent.objects.filter(_pay_window(week_start))
            .filter(employer='Infinity').distinct().select_related('user')
            .order_by('user__last_name', 'user__first_name')
            if a.pk not in rostered]


def _holiday_worked_hours(agents, holiday_dates, nr_ratio=Decimal('0.125')):
    """{agent_id: NR-adjusted billable hours worked on the holiday dates}.

    Per holiday day, the not-ready time *in excess* of the allowance
    (nr_ratio × that day's login) is discounted, so a holiday worked with heavy
    not-ready earns its 2× premium on fewer than the raw logged hours. With no
    (or within-allowance) not-ready, this equals the logged hours."""
    if not agents or not holiday_dates:
        return {}
    from adherence.models import DailyAgentHours, AdherenceRecord
    from wfm.utils import get_billable_username_map
    bmap, _ = get_billable_username_map([a.pk for a in agents])
    # Days marked 'Holiday' (scheduled, NOT worked) are paid the 1× not-worked way —
    # never also as worked hours, even if stray Five9 login exists on that date.
    notworked = {(r['agent_id'], r['date']) for r in AdherenceRecord.objects.filter(
        agent__in=agents, date__in=holiday_dates, status='Holiday').values('agent_id', 'date')}
    per_day = {}   # (agent_id, date) -> [login_seconds, nr_seconds]
    for r in DailyAgentHours.objects.filter(
        upload__date__in=holiday_dates, agent__in=agents
    ).values('agent_id', 'upload__date', 'five9_username', 'login_seconds', 'not_ready_seconds'):
        aid = r['agent_id']
        if aid is None or (aid, r['upload__date']) in notworked:
            continue
        bn = bmap.get(aid)
        if bn is None or r['five9_username'].strip().lower() in bn:
            k = (aid, r['upload__date'])
            acc = per_day.setdefault(k, [0, 0])
            acc[0] += r['login_seconds']
            acc[1] += r['not_ready_seconds']
    out = {}
    for (aid, _d), (lsec, nsec) in per_day.items():
        login_h = Decimal(str(lsec)) / Decimal('3600')
        nr_h = Decimal(str(nsec)) / Decimal('3600')
        excess_nr = max(Decimal('0'), nr_h - login_h * nr_ratio)   # only NR over the allowance
        worked = max(Decimal('0'), login_h - excess_nr)
        out[aid] = out.get(aid, Decimal('0')) + worked
    return out


def _holiday_not_worked_hours(agents, holiday_dates, week_dates):
    """{agent_id: scheduled hours on holidays the agent was scheduled for but did NOT
    work (status 'Holiday')}. These earn Holiday Pay at 1× (rate × hours) and add 0
    to Hours Worked — the paid-but-not-worked company-holiday case (decision 5b)."""
    if not agents or not holiday_dates:
        return {}
    from adherence.models import AdherenceRecord
    from adherence.views import _build_maps, _scheduled_hours
    hol_in_week = [d for d in holiday_dates if d in set(week_dates)]
    if not hol_in_week:
        return {}
    marked = {}
    for r in AdherenceRecord.objects.filter(
            agent__in=agents, date__in=hol_in_week, status='Holiday').values('agent_id', 'date'):
        marked.setdefault(r['agent_id'], []).append(r['date'])
    if not marked:
        return {}
    maps = _build_maps(agents, week_dates)
    shift_map, extra_hrs_map = maps[0], maps[4]
    out = {}
    for aid, days in marked.items():
        total = Decimal('0')
        for d in days:
            shift = shift_map.get((aid, d))
            sched = _scheduled_hours(shift) if shift else Decimal('0')
            sched += extra_hrs_map.get((aid, d), Decimal('0'))
            total += sched
        out[aid] = total
    return out


def _vacation_hours(agents, week_dates):
    """{agent_id: paid vacation hours this week}. Each 'V' day pays min(scheduled
    hours, 8); a 'V' on an unscheduled day (day off) pays a flat 8."""
    from adherence.models import AdherenceRecord
    v_days = {}
    for r in AdherenceRecord.objects.filter(
            agent__in=agents, date__in=week_dates, status='V').values('agent_id', 'date'):
        v_days.setdefault(r['agent_id'], []).append(r['date'])
    if not v_days:
        return {}
    from adherence.views import _build_maps, _scheduled_hours
    maps = _build_maps(agents, week_dates)
    shift_map, extra_hrs_map = maps[0], maps[4]   # [4] = split-shift extra blocks
    out = {}
    for aid, days in v_days.items():
        total = Decimal('0')
        for d in days:
            shift = shift_map.get((aid, d))
            sched = _scheduled_hours(shift) if shift else Decimal('0')
            sched += extra_hrs_map.get((aid, d), Decimal('0'))   # add split-shift blocks so a <8h main shift isn't undercounted
            total += min(sched, Decimal('8')) if sched > 0 else Decimal('8')
        out[aid] = total
    return out


def _nav(week_start, week_dates):
    return {
        'week_start': week_start,
        'week_end': week_dates[-1],
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
    }


@login_required
@nomina_access_required
def dashboard(request):
    """Nómina landing page — super admin only."""
    week_start, week_dates = _week(request)
    ctx = _nav(week_start, week_dates)
    ctx['today'] = timezone.localdate()
    return render(request, 'nomina/dashboard.html', ctx)


@login_required
@nomina_access_required
def inputs(request):
    """Inputs hub — one module per input type + the week's spiff FX rate."""
    week_start, week_dates = _week(request)
    nweek, _ = NominaWeek.objects.get_or_create(week_start=week_start)

    if request.method == 'POST' and _finalized_run(week_start):
        messages.error(request, "This week is finalized (locked) — the spiff rate can't be changed.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")
    if request.method == 'POST' and 'spiff_fx_rate' in request.POST:
        raw = (request.POST.get('spiff_fx_rate') or '').strip()
        if not raw:
            nweek.spiff_fx_rate = None       # blank stays empty
            nweek.save()
            messages.success(request, "Spiff rate cleared.")
        else:
            rate = _parse_rate(raw)          # comma = decimal point, never thousands
            if rate is None or not (Decimal('0') < rate < Decimal('1000')):
                messages.error(request, "Enter a valid exchange rate (e.g. 18.50). "
                                        "It looks off — check for a stray comma or typo.")
            else:
                nweek.spiff_fx_rate = rate
                nweek.save()
                messages.success(request, f"Spiff rate saved: {rate:.4f}")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    # Per-type filled count (agents with a non-zero value this week). Covers the whole
    # roster — agents AND admins — so the hub counts match what each module now accepts.
    agents = _infinity_agents(week_start) + _admin_agents(week_start)
    existing = {wi.agent_id: wi for wi in WeeklyPayInput.objects.filter(
        agent__in=agents, week_start=week_start)}
    cards = []
    for t in INPUT_TYPES:
        filled = sum(1 for a in agents if existing.get(a.pk) and getattr(existing[a.pk], t['field']))
        cards.append({**t, 'filled': filled})

    ctx = _nav(week_start, week_dates)
    ctx.update({'cards': cards, 'spiff_fx_rate': nweek.spiff_fx_rate})
    return render(request, 'nomina/inputs.html', ctx)


def _read_rows(uploaded):
    """Read an uploaded .csv or .xlsx into (headers, list-of-row-lists)."""
    name = (uploaded.name or '').lower()
    if name.endswith('.csv'):
        import csv, io
        text = uploaded.read().decode('utf-8-sig', errors='replace')
        # Sniff the delimiter — Excel in some locales exports ';' or tab, which a
        # comma-only reader would read as ONE column and silently import nothing.
        first = next((ln for ln in text.splitlines() if ln.strip()), '')
        delim = ','
        for cand in (';', '\t'):
            if first.count(cand) > first.count(delim):
                delim = cand
        r = list(csv.reader(io.StringIO(text), delimiter=delim))
        return (r[0] if r else []), r[1:]
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True, read_only=True)
    ws = wb.active
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    return (rows[0] if rows else []), rows[1:]


def _find_col(headers, keywords):
    for i, h in enumerate(headers):
        hl = str(h or '').strip().lower()
        if any(k in hl for k in keywords):
            return i
    return None


def _detect_amount_col(headers, data, skip):
    """Fallback when the amount header is blank/unlabelled (e.g. the Spiffs file,
    where the $ column has no header). Score each candidate column by how many of
    its cells look like money — a leading '$' counts strongest, a plain number
    counts too — and pick the best (rightmost wins ties)."""
    ncols = max([len(headers)] + [len(r) for r in data]) if data else len(headers)
    best_i, best_score = None, 0
    for i in range(ncols):
        if i in skip:
            continue
        strong = numeric = 0
        for row in data:
            if i >= len(row):
                continue
            raw = str(row[i] if row[i] is not None else '').strip()
            if not raw:
                continue
            if '$' in raw:
                strong += 1
            if _dec(raw) != 0:
                numeric += 1
        score = strong * 1000 + numeric
        if score and score >= best_score:   # >= so ties favour the rightmost column
            best_i, best_score = i, score
    return best_i


@login_required
@nomina_access_required
def input_type(request, key):
    """One input module: upload a file (mapped by headers) or edit manually."""
    t = INPUT_TYPE_BY_KEY.get(key)
    if not t:
        messages.error(request, "Unknown input type.")
        return redirect('nomina:inputs')
    field = t['field']
    week_start, week_dates = _week(request)
    if request.method == 'POST' and _finalized_run(week_start):
        messages.error(request, "This week is finalized (locked) — inputs can't be changed.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")
    # Inputs cover the WHOLE roster — agents AND official admins — so one uploaded file
    # matches everyone; an admin's value flows to the Admin Nómina (which reads the same
    # WeeklyPayInput). Role-scoped modules (Kill Team QA) still filter by role below.
    agents = _infinity_agents(week_start) + _admin_agents(week_start)
    if t.get('roles'):   # some modules are role-scoped (e.g. Kill Team QA → Kill Team only)
        agents = [a for a in agents if a.role_type in t['roles']]
    by_username, dupe_usernames = {}, set()
    for a in agents:
        k = a.user.username.strip().lower()
        if k in by_username:
            dupe_usernames.add(k)   # case-only collision — ambiguous
        by_username[k] = a
    for k in dupe_usernames:        # drop ambiguous keys → fall back to ID / unmatched
        by_username.pop(k, None)
    by_empid = {(a.employee_id or '').strip(): a for a in agents if a.employee_id}
    unmatched = []

    if request.method == 'POST' and request.FILES.get('file'):
        headers, data = _read_rows(request.FILES['file'])
        user_col = _find_col(headers, ['username', 'user', 'agent', 'login'])
        id_col = _find_col(headers, ['employee id', 'emp id', 'empid', 'employee', 'emp', 'id'])
        amt_col = _find_col(headers, ['amount', 'total', 'dollar', 'pesos', 'price', 'precio', 'value', 'monto'])
        if amt_col is None:   # header unlabelled (e.g. Spiffs) — find the $ column by content
            amt_col = _detect_amount_col(headers, data, skip={user_col, id_col})
        if amt_col is None or (user_col is None and id_col is None):
            messages.error(request, "Couldn't find the columns. Need a Username or Employee ID column and an Amount column.")
            return redirect(f"{request.path}?week_start={week_start.isoformat()}")
        totals = {}
        for row in data:
            if amt_col >= len(row):
                continue
            amt = _dec(row[amt_col])
            agent = None
            if user_col is not None and user_col < len(row) and row[user_col]:
                agent = by_username.get(str(row[user_col]).strip().lower())
            if agent is None and id_col is not None and id_col < len(row) and row[id_col] not in (None, ''):
                agent = by_empid.get(str(row[id_col]).strip())
            if agent is None:
                key_shown = row[user_col] if user_col is not None and user_col < len(row) else (row[id_col] if id_col is not None and id_col < len(row) else '?')
                if amt:
                    unmatched.append({'who': key_shown, 'amount': amt})
                continue
            totals[agent.pk] = (totals.get(agent.pk, Decimal('0')) + amt) if t['aggregate'] else amt
        # Wipe-and-replace: a fresh upload is authoritative, so clear this module's
        # values for the whole week first, then load only what's in the file. (Manual
        # "Save" is a separate, additive path — this replacement is upload-only.)
        WeeklyPayInput.objects.filter(week_start=week_start, agent__in=agents).update(**{field: 0})
        for aid, val in totals.items():
            WeeklyPayInput.objects.update_or_create(
                agent_id=aid, week_start=week_start, defaults={field: val})
        # Persist unmatched rows so they stay visible until acknowledged — a fresh
        # upload rebuilds this week/type's set from scratch.
        UnmatchedInputRow.objects.filter(week_start=week_start, input_key=key).delete()
        UnmatchedInputRow.objects.bulk_create([
            UnmatchedInputRow(week_start=week_start, input_key=key,
                              who=str(u['who'])[:200], amount=u['amount'])
            for u in unmatched])
        messages.success(request, f"Imported {len(totals)} {t['label']} value(s); "
                                  f"{len(unmatched)} row(s) need acknowledgement." if unmatched
                                  else f"Imported {len(totals)} {t['label']} value(s); all rows matched.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    if request.method == 'POST' and request.POST.get('ack'):  # acknowledge unmatched row(s)
        val = request.POST['ack']
        qs = UnmatchedInputRow.objects.filter(week_start=week_start, input_key=key, acknowledged=False)
        n = qs.update(acknowledged=True) if val == 'all' else qs.filter(pk=val).update(acknowledged=True)
        messages.success(request, f"Acknowledged {n} row(s)." if val == 'all' else "Acknowledged.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    if request.method == 'POST' and request.POST.get('add_agent'):  # manual-add modules
        aid, amt = request.POST.get('add_agent'), _dec(request.POST.get('add_amount'))
        if aid:
            WeeklyPayInput.objects.update_or_create(
                agent_id=aid, week_start=week_start, defaults={field: amt})
            messages.success(request, f"Added to {t['label']}.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    if request.method == 'POST' and request.POST.get('add_more_agent'):  # additive add (Spiffs / Extra Hours)
        from django.db.models import F
        try:
            aid_int = int(request.POST.get('add_more_agent'))
        except (TypeError, ValueError):
            aid_int = None
        amt = _dec(request.POST.get('add_more_amount'))
        if aid_int in {a.pk for a in agents} and amt:
            obj, _ = WeeklyPayInput.objects.get_or_create(agent_id=aid_int, week_start=week_start)
            WeeklyPayInput.objects.filter(pk=obj.pk).update(**{field: F(field) + amt})   # accumulate, don't replace
            obj.refresh_from_db(fields=[field])
            unit = t.get('unit', '')
            messages.success(request, f"Added {amt:g} {unit} — new total {getattr(obj, field):g} {unit}.")
        else:
            messages.error(request, "Pick an agent and enter an amount.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    if request.method == 'POST' and request.POST.get('remove'):  # drop from a manual-add list
        WeeklyPayInput.objects.filter(
            agent_id=request.POST['remove'], week_start=week_start).update(**{field: 0})
        messages.success(request, "Removed.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    if request.method == 'POST':  # manual save — only the agents rendered in the grid
        agent_pks = {a.pk for a in agents}
        for k, val in request.POST.items():
            if not k.startswith('v_'):
                continue
            try:
                aid = int(k[2:])
            except ValueError:
                continue
            if aid in agent_pks:
                WeeklyPayInput.objects.update_or_create(
                    agent_id=aid, week_start=week_start, defaults={field: _dec(val)})
        messages.success(request, f"{t['label']} saved.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    existing = {wi.agent_id: wi for wi in WeeklyPayInput.objects.filter(
        agent__in=agents, week_start=week_start)}
    # Per-module display flags.
    show_pesos = (key == 'spiff')                    # only Spiffs converts USD→MXN
    show_emp = t.get('show_emp', True)
    show_total = t.get('totals', False) or show_pesos
    sort_by_value = t.get('sort_by_value', False) or show_pesos
    default_amt = t.get('default_amount')            # e.g. Kill Team QA prefills $400
    fx_rate = None
    if show_pesos:
        nweek, _ = NominaWeek.objects.get_or_create(week_start=week_start)
        fx_rate = nweek.spiff_fx_rate
    rows = []
    amt_total = pesos_total = Decimal('0')
    for a in agents:
        wi = existing.get(a.pk)
        raw = getattr(wi, field) if wi else None     # None = never entered (nullable field or no row)
        v = raw if raw is not None else Decimal('0')  # numeric value for totals
        if default_amt is not None and raw is None:  # never entered → show the standard default
            display = f'{Decimal(str(default_amt)):.2f}'
        else:                                         # entered value wins — including an explicit $0 (blank)
            display = '' if v == 0 else f'{v:.2f}'
        name = (a.agent_name or a.user.username) if t.get('name_only_agent') \
            else (a.agent_name or a.user.get_full_name() or a.user.username)
        row = {'agent': a, 'name': name, 'username': a.user.username,
               'emp': a.employee_id or '', 'val': v, 'display': display}
        amt_total += v
        if show_pesos:
            pesos = (v * fx_rate).quantize(Decimal('0.01')) if fx_rate else None
            row['pesos'] = '' if (pesos is None or v == 0) else f'{pesos:,.2f}'
            if pesos:
                pesos_total += pesos
        rows.append(row)
    if sort_by_value:   # people with a value float to the top
        rows.sort(key=lambda r: (r['val'] == 0, -r['val'], r['name'].lower()))
    add_more = t.get('add_more', False)
    # The additive "add to a person" dropdown offers EVERYONE in scope (built before
    # the manual-add filter trims rows to only people who already have a value).
    all_options = sorted(
        ({'pk': r['agent'].pk, 'name': r['name'], 'username': r['username']} for r in rows),
        key=lambda o: o['name'].lower()) if add_more else None
    manual_add = t.get('manual_add', False)
    addable = None
    if manual_add:   # list = only the people added this week; dropdown = everyone else
        addable = [{'pk': r['agent'].pk, 'name': r['name'], 'username': r['username']}
                   for r in rows if r['val'] == 0]
        addable.sort(key=lambda a: a['name'].lower())
        rows = [r for r in rows if r['val'] != 0]
        rows.sort(key=lambda r: r['name'].lower())
        amt_total = sum((r['val'] for r in rows), Decimal('0'))
    unmatched = list(UnmatchedInputRow.objects.filter(
        week_start=week_start, input_key=key, acknowledged=False))
    ctx = _nav(week_start, week_dates)
    ctx.update({'t': t, 'rows': rows, 'unmatched': unmatched,
                'show_pesos': show_pesos, 'show_emp': show_emp, 'show_total': show_total,
                'default_amt': default_amt, 'fx_rate': fx_rate,
                'manual_add': manual_add, 'addable': addable,
                'add_more': add_more, 'all_options': all_options,
                'finalized': _finalized_run(week_start),
                'amt_total': amt_total if show_total else None,
                'pesos_total': pesos_total if (show_pesos and fx_rate) else None})
    return render(request, 'nomina/input_type.html', ctx)


def _agent_nomina_data(week_start, week_dates, corrected=True):
    """Compute the Agent Nómina rows + totals (shared by the view and the export).

    `corrected` selects the export variant:
      True  ("Mine")  — after edits: overrides applied, NET LPO (commission deducted),
                        vacation hours + manual extra hours folded into pay/hours.
      False ("Yours") — before edits: no overrides, GROSS LPO, no vacation, no manual
                        hours. The on-screen view always uses the corrected numbers."""
    from finance.views import _get_billable_weekly_data
    from finance.models import BillingSettings

    settings = BillingSettings.get_for_week(week_start)
    nweek, _ = NominaWeek.objects.get_or_create(week_start=week_start)
    fx = nweek.spiff_fx_rate or Decimal('0')  # empty rate → spiffs stay 0 until set

    agents = _infinity_agents(week_start)
    data = _get_billable_weekly_data(agents, week_dates, settings)
    # The roster now includes sales-type agents who take no calls — no attendance
    # tracking AND no billable Five9 profile. If such a person happens to have a
    # NON-billable Five9 profile that logged time, the billing engine's fallback
    # (bnames is None → count everything) would sum those hours as base pay, and the
    # adherence bonus is derived from the same hours — an overpay on hours deliberately
    # marked non-billable. Zero the call-derived figures (base pay, hours, adherence
    # bonus) for exactly those agents; their pay is LPO / manual inputs, not calls.
    # Their hourly rate is left intact, so manual Extra Hours / vacation / holiday pay
    # still compute. Tracked agents and anyone with a billable profile are untouched, so
    # no previously-paid agent's numbers change.
    for a in agents:
        if not a.track_attendance and not any(p.billable for p in a.five9_profiles.all()):
            d = data.get(a.pk)
            if d:
                d['base_pay_mxn'] = Decimal('0')
                d['bonus_mxn'] = Decimal('0')
                d['final_hrs'] = Decimal('0')
    inputs_map = {wi.agent_id: wi for wi in WeeklyPayInput.objects.filter(
        agent__in=agents, week_start=week_start)}
    ba_agents = set(BreakAbuseIncident.objects.filter(
        agent__in=agents, date__in=week_dates).values_list('agent_id', flat=True))
    holiday_dates = list(Holiday.objects.filter(date__in=week_dates).values_list('date', flat=True))
    hol_hours = _holiday_worked_hours(agents, holiday_dates, settings.nr_ratio)
    hol_nw_hours = _holiday_not_worked_hours(agents, holiday_dates, week_dates)  # scheduled, not worked (1×)
    vac_hours = _vacation_hours(agents, week_dates)   # paid vacation hours ('V' days)
    from adherence.models import AdherenceRecord
    vac_days_map = {}                                 # count of 'V' days (for the Yours note)
    for r in AdherenceRecord.objects.filter(
            agent__in=agents, date__in=week_dates, status='V').values('agent_id'):
        vac_days_map[r['agent_id']] = vac_days_map.get(r['agent_id'], 0) + 1
    enrolls = {e.agent_id: e for e in WelcomeBonusEnrollment.objects.filter(agent__in=agents)}
    loan_ded = {}
    for ln in Loan.objects.filter(agent__in=agents):
        inst = ln.installment_for_week(week_start)
        if inst:
            loan_ded[ln.agent_id] = loan_ded.get(ln.agent_id, Decimal('0')) + inst
    overrides = {(o.agent_id, o.field): o.value
                 for o in NominaOverride.objects.filter(agent__in=agents, week_start=week_start)}

    def ov(aid, field, computed):
        # Overrides are a "Mine" correction — the raw "Yours" sheet ignores them.
        if not corrected:
            return computed
        return overrides.get((aid, field), computed)

    rows = []
    tot_base = tot_bonus = tot_lpo = tot_spiff = tot_hol = tot_sub = tot_ded = tot_total = Decimal('0')
    tot_worked_hrs = tot_hol_hrs = tot_total_hrs = tot_vac_hrs = Decimal('0')
    tot_referral = tot_welcome = tot_kill = tot_comedor = tot_transport = tot_loan = Decimal('0')
    on_vacation = 0
    for a in agents:
        d = data.get(a.pk, {})
        wi = inputs_map.get(a.pk)
        broke = a.pk in ba_agents
        comm_pct = d.get('commission_pct', Decimal('0'))
        rate = d.get('hourly_mxn', Decimal('0'))

        base = ov(a.pk, 'base_pay', d.get('base_pay_mxn', Decimal('0')))
        bonus = ov(a.pk, 'adherence', Decimal('0') if broke else d.get('bonus_mxn', Decimal('0')))
        # LPO: "Yours" shows the GROSS amount; "Mine" applies the commission deduction (+ any override).
        gross_lpo = wi.lpo if wi else Decimal('0')
        net_lpo = ov(a.pk, 'net_lpo', (gross_lpo * (Decimal('1') - comm_pct / Decimal('100'))).quantize(Decimal('0.01'))) \
            if corrected else gross_lpo
        spiff_mxn = ov(a.pk, 'spiff', ((wi.spiff_usd if wi else Decimal('0')) * fx).quantize(Decimal('0.01')))
        # Welcome: enrollment-driven (paid only if they earned an adherence bonus), else the manual input.
        enroll = enrolls.get(a.pk)
        welcome_default = enroll.amount if (enroll and enroll.covers_week(week_start) and bonus > 0) else (wi.welcome if wi else Decimal('0'))
        welcome = ov(a.pk, 'welcome', welcome_default)
        referral = ov(a.pk, 'referral', wi.referral if wi else Decimal('0'))
        # Kill Team QA is a standard $400 for Kill Team agents ONLY when nothing was ever
        # entered (kill_team_qa is NULL). An entered value wins — including an explicit $0,
        # so a Kill Team agent can be zeroed out and it sticks (mirrors the module's prefill).
        kq_stored = wi.kill_team_qa if wi else None
        if kq_stored is None and a.role_type == 'kill_team':
            kq_raw = Decimal(str(INPUT_TYPE_BY_KEY['killqa'].get('default_amount') or 0))
        else:
            kq_raw = kq_stored if kq_stored is not None else Decimal('0')
        kill_qa = ov(a.pk, 'kill_qa', kq_raw)
        hol_hrs = hol_hours.get(a.pk, Decimal('0'))            # worked holiday hours → 2×
        hol_nw_hrs = hol_nw_hours.get(a.pk, Decimal('0'))      # scheduled, not worked → 1×
        holiday_pay = ov(a.pk, 'holiday', (hol_hrs * rate * 2 + hol_nw_hrs * rate).quantize(Decimal('0.01')))
        comedor = ov(a.pk, 'comedor', wi.comedor if wi else Decimal('0'))
        transport = ov(a.pk, 'transport', wi.transportation if wi else Decimal('0'))
        loan = ov(a.pk, 'loan', loan_ded.get(a.pk, Decimal('0')))

        # Manual hours + vacation are "Mine" corrections: folded into pay/hours only
        # in the corrected variant. "Yours" shows the raw pre-correction numbers.
        extra_hrs = wi.extra_hours if wi else Decimal('0')
        extra_pay = (extra_hrs * rate).quantize(Decimal('0.01'))
        vac_hrs = vac_hours.get(a.pk, Decimal('0'))
        vac_pay = (vac_hrs * rate).quantize(Decimal('0.01'))
        if vac_hrs > 0:
            on_vacation += 1
        applied_extra = extra_hrs if corrected else Decimal('0')
        applied_vac = vac_hrs if corrected else Decimal('0')
        pay48 = base + (extra_pay if corrected else Decimal('0')) + (vac_pay if corrected else Decimal('0'))

        subtotal = pay48 + bonus + net_lpo + spiff_mxn + welcome + referral + kill_qa + holiday_pay
        total = subtotal - comedor - transport - loan  # may go negative (G6)

        final_hrs = d.get('final_hrs', Decimal('0'))
        # Hours Worked INCLUDES holiday hours (paid 1× here via base; the Holiday column
        # shows that subset, Holiday Pay adds the +2× = triple) plus any applied extra hours.
        worked_hrs = final_hrs + applied_extra
        total_hrs = worked_hrs + applied_vac                  # + paid vacation hours (Mine only)
        rows.append({
            'agent': a, 'emp': a.employee_id or '',
            'legal_name': a.user.get_full_name() or a.user.username,
            'username': a.user.username, 'break_abuse': broke,
            'hours': final_hrs, 'rate': rate,
            'worked_hrs': worked_hrs, 'holiday_hrs': hol_hrs, 'total_hrs': total_hrs,
            'vac_hrs': vac_hrs, 'vac_pay': vac_pay, 'vac_days': vac_days_map.get(a.pk, 0),
            'extra_hrs': extra_hrs, 'extra_pay': extra_pay,
            'base_pay': pay48, 'adherence_bonus': bonus,
            'net_lpo': net_lpo, 'comm_pct': comm_pct, 'spiff_mxn': spiff_mxn,
            'welcome': welcome, 'referral': referral, 'kill_qa': kill_qa, 'holiday_pay': holiday_pay,
            'subtotal': subtotal, 'comedor': comedor, 'transport': transport, 'loan': loan, 'total': total,
        })
        tot_base += pay48; tot_bonus += bonus; tot_lpo += net_lpo; tot_spiff += spiff_mxn
        tot_hol += holiday_pay; tot_sub += subtotal; tot_ded += (comedor + transport + loan); tot_total += total
        tot_worked_hrs += worked_hrs; tot_hol_hrs += hol_hrs; tot_total_hrs += total_hrs; tot_vac_hrs += vac_hrs
        tot_referral += referral; tot_welcome += welcome; tot_kill += kill_qa
        tot_comedor += comedor; tot_transport += transport; tot_loan += loan

    # Spiff FX guard: an unset weekly rate silently converts every spiff to $0.
    # Flag it (with a count) so the nómina warns instead of quietly underpaying.
    spiff_unpaid = sum(1 for wi in inputs_map.values() if wi.spiff_usd) if not nweek.spiff_fx_rate else 0
    totals = {'base': tot_base, 'bonus': tot_bonus, 'net_lpo': tot_lpo, 'spiff': tot_spiff,
              'holiday': tot_hol, 'subtotal': tot_sub, 'total': tot_total,
              'worked_hrs': tot_worked_hrs, 'holiday_hrs': tot_hol_hrs, 'total_hrs': tot_total_hrs,
              'vac_hrs': tot_vac_hrs, 'on_vacation': on_vacation,
              'referral': tot_referral, 'welcome': tot_welcome, 'kill_qa': tot_kill,
              'comedor': tot_comedor, 'transport': tot_transport, 'loan': tot_loan,
              'spiff_needs_rate': bool(spiff_unpaid), 'spiff_unpaid_count': spiff_unpaid}
    return rows, totals


@login_required
@nomina_access_required
def agent_nomina(request):
    """Agent Nómina — auto columns (existing engine) + manual inputs + modules
    (break abuse, holiday, welcome, loans) + overrides → subtotal/total."""
    week_start, week_dates = _week(request)
    run = _finalized_run(week_start)
    if run:
        rows, totals = run.agent_rows, run.agent_totals
    else:
        rows, totals = _agent_nomina_data(week_start, week_dates)
    ctx = _nav(week_start, week_dates)
    ctx.update({'rows': rows, 'totals': totals, 'finalized': run,
                'unrostered': [] if run else _unrostered_infinity(week_start)})
    return render(request, 'nomina/agent_nomina.html', ctx)


# Agent export column order — mirrors the LCC AGENT NOMINA file.
AGENT_EXPORT_COLS = [
    ('EMP', 'emp'), ('Legal name', 'legal_name'), ('User', 'username'),
    ('Hours Worked', 'worked_hrs'), ('Holiday', 'holiday_hrs'), ('Holiday Pay', 'holiday_pay'),
    ('Total Hours', 'total_hrs'), ('Pay (48)', 'base_pay'),
    ('LPO', 'net_lpo'), ('Referral', 'referral'), ('Welcome Bonus', 'welcome'),
    ('Kill Team QA Bonus', 'kill_qa'), ('Spiff', 'spiff_mxn'), ('Adherence', 'adherence_bonus'),
    ('Sub Total', 'subtotal'), ('Cafeteria', 'comedor'), ('Prestamo', 'loan'),
    ('Transportation', 'transport'), ('Total', 'total'),
]

# Excel cell number formats for the export.
MONEY_FMT = '$#,##0.00'          # e.g. $3,000.00
HOURS_FMT = '#,##0.00'           # e.g. 25.11 — always two decimals, no currency
INT_FMT = '0'                    # whole number, no decimals (employee IDs only)
# A — whole-number ID; D, E, G — hours (two decimals); F and H–S — currency.
AGENT_INT_FIELDS = {'emp'}
AGENT_HOURS_FIELDS = {'worked_hrs', 'holiday_hrs', 'total_hrs'}
AGENT_MONEY_FIELDS = {'holiday_pay', 'base_pay', 'net_lpo', 'referral', 'welcome', 'kill_qa',
                      'spiff_mxn', 'adherence_bonus', 'subtotal', 'comedor', 'loan', 'transport', 'total'}


def _write_nomina_sheet(ws, cols, rows, int_fields, money_fields, hours_fields=frozenset(), note_fn=None):
    """Append the header + data rows to `ws`, writing numeric cells as real numbers
    and applying whole-number / two-decimal / currency formats per column."""
    from openpyxl.styles import Font
    headers = [h for h, _ in cols] + (['Notes'] if note_fn else [])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        values = []
        for _h, f in cols:
            v = r.get(f)
            if f in ('emp',):
                v = int(v) if str(v).isdigit() else (v or '')
            elif hasattr(v, 'quantize'):        # Decimal (live rows)
                # Round to two decimals so the stored value matches the displayed value —
                # money is already 2dp; hours (raw login fractions) become clean 2dp too.
                v = float(v.quantize(Decimal('0.01')))
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                v = round(float(v), 2)          # float (finalized snapshot rows) → same 2dp
            elif v is None:
                v = ''
            values.append(v)
        if note_fn:
            values.append(note_fn(r))
        ws.append(values)
        ri = ws.max_row
        for ci, (_h, f) in enumerate(cols, start=1):
            if f in money_fields:
                ws.cell(row=ri, column=ci).number_format = MONEY_FMT
            elif f in hours_fields:
                ws.cell(row=ri, column=ci).number_format = HOURS_FMT
            elif f in int_fields:
                ws.cell(row=ri, column=ci).number_format = INT_FMT


def _agent_note(yours, mine):
    """Notes for the 'Yours' sheet — what the corrected 'Mine' sheet changes. Fires
    only on real corrections, in two categories: LPO (commission/override) and hours
    (vacation and/or manual hours)."""
    if mine is None:
        return ''
    notes = []
    if yours['net_lpo'] != mine['net_lpo']:            # Yours holds gross LPO here
        notes.append(f"LPO should be ${mine['net_lpo']:.2f}")
    if mine['total_hrs'] != yours['total_hrs']:        # vacation and/or manual hours
        vd = mine.get('vac_days', 0)
        prefix = f"Agent had {vd} day{'s' if vd != 1 else ''} of vacation, " if vd else ""
        notes.append(f"{prefix}total hours worked should be {mine['total_hrs']:.2f}")
    return " · ".join(notes)


@login_required
@nomina_access_required
def agent_export(request):
    """Export the Agent Nómina: 'Yours' (raw, before edits, with a Notes column) +
    'Mine' (corrected) sheets — two genuinely different sheets."""
    import openpyxl
    week_start, week_dates = _week(request)
    run = _finalized_run(week_start)
    if run:
        rows_mine, rows_yours = run.agent_rows, run.agent_yours_rows
        note_of = lambda r: r.get('note', '')
    else:
        rows_mine, _ = _agent_nomina_data(week_start, week_dates, corrected=True)
        rows_yours, _ = _agent_nomina_data(week_start, week_dates, corrected=False)
        mine_by_id = {r['agent'].pk: r for r in rows_mine}
        note_of = lambda r: _agent_note(r, mine_by_id.get(r['agent'].pk))

    wb = openpyxl.Workbook()
    ws_yours = wb.active
    ws_yours.title = 'Yours'
    _write_nomina_sheet(ws_yours, AGENT_EXPORT_COLS, rows_yours, AGENT_INT_FIELDS, AGENT_MONEY_FIELDS,
                        hours_fields=AGENT_HOURS_FIELDS, note_fn=note_of)
    ws_mine = wb.create_sheet('Mine')
    _write_nomina_sheet(ws_mine, AGENT_EXPORT_COLS, rows_mine, AGENT_INT_FIELDS, AGENT_MONEY_FIELDS,
                        hours_fields=AGENT_HOURS_FIELDS)

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="LCC AGENT NOMINA {week_start:%m%d%Y}.xlsx"'
    wb.save(resp)
    return resp


@login_required
@nomina_access_required
def break_abuse(request):
    """Log break-abuse incidents. Any incident in a pay week zeroes that agent's
    adherence bonus for the week (applied in the Agent Nómina)."""
    from scheduling.models import Agent
    week_start, week_dates = _week(request)

    if request.method == 'POST':
        if request.POST.get('delete'):
            BreakAbuseIncident.objects.filter(pk=request.POST['delete']).delete()
            messages.success(request, "Incident removed.")
        else:
            agent_id = request.POST.get('agent')
            try:
                d = date_cls.fromisoformat(request.POST.get('date'))
            except (ValueError, TypeError):
                d = None
            if agent_id and d:
                BreakAbuseIncident.objects.create(
                    agent_id=agent_id, date=d, note=request.POST.get('note', '').strip())
                messages.success(request, "Break-abuse incident logged.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    incidents = list(BreakAbuseIncident.objects.filter(
        date__in=week_dates).select_related('agent__user').order_by('date'))
    agents = Agent.objects.filter(status='active').select_related('user').order_by(
        'user__last_name', 'user__first_name')

    ctx = _nav(week_start, week_dates)
    ctx.update({'incidents': incidents, 'agents': agents, 'week_dates': week_dates})
    return render(request, 'nomina/break_abuse.html', ctx)


@login_required
@nomina_access_required
def holidays(request):
    """Manage the company holiday calendar. A holiday worked earns a 2× premium."""
    if request.method == 'POST':
        if request.POST.get('delete'):
            Holiday.objects.filter(pk=request.POST['delete']).delete()
            messages.success(request, "Holiday removed.")
        else:
            try:
                d = date_cls.fromisoformat(request.POST.get('date'))
                Holiday.objects.update_or_create(date=d, defaults={'name': request.POST.get('name', '').strip()})
                messages.success(request, "Holiday saved.")
            except (ValueError, TypeError):
                messages.error(request, "Enter a valid date.")
        return redirect(request.path)

    week_start, week_dates = _week(request)
    ctx = _nav(week_start, week_dates)
    ctx['holidays'] = Holiday.objects.all()
    return render(request, 'nomina/holidays.html', ctx)


@login_required
@nomina_access_required
def exports(request):
    """Download the Agent (Yours/Mine) and Admin (Yours) nómina Excel files."""
    week_start, week_dates = _week(request)
    ctx = _nav(week_start, week_dates)
    return render(request, 'nomina/exports.html', ctx)


@login_required
@loan_access_required
def loans(request):
    """Add loans and view active ones. 1wk ×1.25 / 2wk ×1.35; weekly drawdown.
    Accessible to super admins and users granted `can_manage_loans`."""
    from scheduling.models import Agent
    week_start, week_dates = _week(request)

    if request.method == 'POST':
        if request.POST.get('delete'):
            Loan.objects.filter(pk=request.POST['delete']).delete()
            messages.success(request, "Loan removed.")
        else:
            agent_id = request.POST.get('agent')
            principal = _dec(request.POST.get('principal'))
            term = 2 if request.POST.get('term') == '2' else 1
            if agent_id and principal > 0:
                Loan.objects.create(
                    agent_id=agent_id, principal=principal, term_weeks=term,
                    rate=Decimal('1.35') if term == 2 else Decimal('1.25'),
                    start_week=week_start,
                    granted_by=getattr(request.user, 'agent', None),   # loan manager
                )
                messages.success(request, "Loan added.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    # A loan appears ONLY in the week(s) it is actually being repaid — a 1-week loan in its
    # one week, a 2-week loan in both — so navigating to a week with no active loans shows an
    # empty list. installment_for_week() is 0 outside a loan's repayment window.
    loan_list = []
    week_total = total_balance = Decimal('0')
    for ln in Loan.objects.select_related('agent__user').all():
        inst = ln.installment_for_week(week_start)
        if inst <= 0:
            continue                                         # not repaid this week → not shown here
        ln.bal = ln.balance(week_start)
        ln.this_week = inst                                  # amount deducted THIS pay week
        week_total += inst
        total_balance += ln.bal
        loan_list.append(ln)
    # A–Z by the agent_name shown in the picker (not by last name — otherwise the list
    # reads scrambled since the picker shows the agent name, not "Last, First").
    agents = sorted(
        Agent.objects.filter(status='active').select_related('user'),
        key=lambda a: (a.agent_name or '').lower())
    ctx = _nav(week_start, week_dates)
    ctx.update({'loans': loan_list, 'agents': agents,
                'week_total': week_total, 'total_balance': total_balance})
    return render(request, 'nomina/loans.html', ctx)


@login_required
@nomina_access_required
def welcome(request):
    """The welcome-bonus eligibility table. Paid per covered week the agent earns
    an adherence bonus; every calendar week counts toward the term."""
    from scheduling.models import Agent
    week_start, week_dates = _week(request)

    if request.method == 'POST':
        if request.POST.get('delete'):
            WelcomeBonusEnrollment.objects.filter(pk=request.POST['delete']).delete()
            messages.success(request, "Removed.")
        else:
            agent_id = request.POST.get('agent')
            amount = _dec(request.POST.get('amount'))
            num_weeks = int(request.POST.get('num_weeks') or 4)
            if agent_id:
                WelcomeBonusEnrollment.objects.create(
                    agent_id=agent_id, amount=amount, num_weeks=num_weeks, start_week=week_start)
                messages.success(request, "Enrolled.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    enrolls = list(WelcomeBonusEnrollment.objects.select_related('agent__user').all())
    for e in enrolls:
        e.weeks_in = e.covers_week(week_start)
    agents = Agent.objects.filter(status='active').select_related('user').order_by(
        'user__last_name', 'user__first_name')
    ctx = _nav(week_start, week_dates)
    ctx.update({'enrolls': enrolls, 'agents': agents})
    return render(request, 'nomina/welcome.html', ctx)


# Mexican LFT "Vacaciones Dignas" vacation-day schedule by completed years of service.
def _lft_vacation_days(years):
    if years < 1:
        return 0
    table = {1: 12, 2: 14, 3: 16, 4: 18, 5: 20}
    if years in table:
        return table[years]
    if years <= 10:
        return 22
    return 22 + 2 * ((years - 6) // 5)  # +2 every 5 yrs beyond year 6-10 band


def _service_start(agent):
    starts = [ep.start_date for ep in agent.employment_periods.all() if ep.start_date]
    return min(starts) if starts else agent.start_date


def _vacation_year(agent, as_of=None):
    """The `year` key for VacationAdjustment: the calendar year in which the agent's
    CURRENT anniversary vacation period BEGAN. Keyed this way (not by plain calendar
    year) so a mid-year anniversary keeps its adjustment across the Dec→Jan boundary
    instead of silently losing it. No hire date → this year."""
    today = as_of or timezone.localdate()
    start = _service_start(agent)
    if not start:
        return today.year
    before_anniv = (today.month, today.day) < (start.month, start.day)
    return today.year - (1 if before_anniv else 0)


def vacation_balance(agent, as_of=None):
    """(accrued, used, remaining) LFT vacation for `agent`'s CURRENT work-anniversary
    year. Accrued = LFT days by completed years of service; Used = 'V' adherence days
    since their most recent hire anniversary; plus any super-admin manual adjustment.
    Everyone resets to their full days on their anniversary. Shared by the Vacations
    page, the request flow, and the adherence 'V' safety net."""
    from adherence.models import AdherenceRecord
    from .models import VacationAdjustment
    today = as_of or timezone.localdate()
    start = _service_start(agent)
    anniv_year = _vacation_year(agent, today)   # adjustment key = start-of-vac-year
    if start:
        before_anniv = (today.month, today.day) < (start.month, start.day)
        years = max(0, today.year - start.year - (1 if before_anniv else 0))
        accrued = _lft_vacation_days(years)
        try:
            anniversary = date_cls(anniv_year, start.month, start.day)
        except ValueError:                   # Feb 29 hire → treat as Feb 28
            anniversary = date_cls(anniv_year, start.month, 28)
        used = AdherenceRecord.objects.filter(
            agent=agent, status='V', date__gte=anniversary, date__lte=today).count()
    else:
        accrued, used = 0, 0                  # no hire date → nothing accrued/used yet
    adj = VacationAdjustment.objects.filter(agent=agent, year=anniv_year).first()
    adjustment = adj.days if adj else Decimal('0')
    remaining = Decimal(accrued) - used + adjustment
    return accrued, used, remaining


def vacation_request_check(agent, start_date, end_date):
    """For a vacation request over [start, end]: (accrued, used, remaining,
    new_days, overdraw). `new_days` = days in the current year not already marked
    'V'; `overdraw` = approving would push the balance negative."""
    from adherence.models import AdherenceRecord
    today = timezone.localdate()
    accrued, used, remaining = vacation_balance(agent, today)
    existing = set(AdherenceRecord.objects.filter(
        agent=agent, status='V', date__range=(start_date, end_date)
    ).values_list('date', flat=True))
    new_days, d = 0, start_date
    while d <= end_date:
        if d.year == today.year and d not in existing:
            new_days += 1
        d += timedelta(days=1)
    return accrued, used, remaining, new_days, new_days > remaining


@login_required
def vacations(request):
    """Top-level Vacations page. Admins (role='admin') see everyone; agents see
    only their own row. Read-only for everyone except super admins, who can adjust
    a person's available days (carryover / corrections)."""
    from scheduling.models import Agent
    from .models import VacationAdjustment
    viewer = getattr(request.user, 'agent', None)
    is_super = request.user.is_superuser or getattr(viewer, 'is_super_admin', False)
    is_admin = is_super or (viewer is not None and viewer.role == 'admin')
    today = timezone.localdate()
    year = today.year

    # Super-admin: save an available-days adjustment for one agent.
    if request.method == 'POST':
        if not is_super:
            messages.error(request, "Only super admins can edit vacation days.")
            return redirect('vacations')
        try:
            target = Agent.objects.get(pk=request.POST.get('agent'))
        except (Agent.DoesNotExist, ValueError, TypeError):
            messages.error(request, "Agent not found.")
            return redirect('vacations')
        accrued, used, _rem = vacation_balance(target)
        new_available = _dec(request.POST.get('available'))
        adjustment = (new_available - (Decimal(accrued) - used)).quantize(Decimal('0.1'))
        VacationAdjustment.objects.update_or_create(
            agent=target, year=_vacation_year(target),   # key by their current vac-year, not calendar year
            defaults={'days': adjustment, 'note': (request.POST.get('note') or '').strip(),
                      'updated_by': viewer})
        messages.success(request, f"{target}: available days set to {new_available:g}.")
        return redirect(f"{request.path}?{request.META.get('QUERY_STRING', '')}")

    # Which agents does the viewer see?
    if is_admin:
        agents = Agent.objects.filter(status='active')
        q = (request.GET.get('q') or '').strip()
        if q:
            agents = agents.filter(
                Q(agent_name__icontains=q) | Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) | Q(user__username__icontains=q))
        sup = request.GET.get('supervisor')
        if sup:
            try:
                agents = agents.filter(supervisor_id=int(sup))
            except (ValueError, TypeError):
                pass
    elif viewer is not None:
        agents = Agent.objects.filter(pk=viewer.pk)   # agents see only themselves
    else:
        agents = Agent.objects.none()
    agents = list(agents.select_related('user').prefetch_related('employment_periods')
                  .order_by('user__last_name', 'user__first_name'))

    rows = []
    for a in agents:
        accrued, used, remaining = vacation_balance(a)
        rows.append({
            'agent': a,
            'agent_name': a.agent_name or a.user.username,
            'legal_name': a.user.get_full_name() or '—',
            'start': _service_start(a),
            'accrued': accrued, 'used': used, 'available': remaining,
        })

    supervisors = Agent.objects.filter(status='active', role='admin').select_related('user').order_by(
        'user__last_name') if is_admin else []
    return render(request, 'nomina/vacations.html', {
        'rows': rows, 'is_admin': is_admin, 'is_super': is_super, 'year': year,
        'q': request.GET.get('q', ''), 'supervisors': supervisors,
        'selected_supervisor': request.GET.get('supervisor', ''),
    })


# Auto columns that can be overridden on the Agent Nómina.
OVERRIDE_FIELDS = [('base_pay', 'Base Pay'), ('adherence', 'Adherence'), ('holiday', 'Holiday')]
# Admins have no adherence bonus — they get the admin bonus instead.
ADMIN_OVERRIDE_FIELDS = [('base_pay', 'Base Pay'), ('admin_bonus', 'Admin Bonus'), ('holiday', 'Holiday')]


@login_required
@nomina_access_required
def overrides(request):
    """Override the auto-computed columns per person. Agents: base pay / adherence /
    holiday. Official admins: base pay / admin bonus / holiday. Blank = use computed."""
    from finance.views import _get_billable_weekly_data
    from finance.models import BillingSettings

    week_start, week_dates = _week(request)
    if request.method == 'POST' and _finalized_run(week_start):
        messages.error(request, "This week is finalized (locked) — overrides can't be changed.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")
    agents = _infinity_agents(week_start)
    admins = _admin_agents(week_start)

    if request.method == 'POST':
        for a, fields in ([(x, OVERRIDE_FIELDS) for x in agents]
                          + [(x, ADMIN_OVERRIDE_FIELDS) for x in admins]):
            for field, _label in fields:
                raw = (request.POST.get(f'{field}_{a.pk}') or '').strip()
                if raw == '':
                    NominaOverride.objects.filter(agent=a, week_start=week_start, field=field).delete()
                else:
                    NominaOverride.objects.update_or_create(
                        agent=a, week_start=week_start, field=field, defaults={'value': _dec(raw)})
        messages.success(request, "Overrides saved.")
        return redirect(f"{request.path}?week_start={week_start.isoformat()}")

    settings = BillingSettings.get_for_week(week_start)
    holiday_dates = list(Holiday.objects.filter(date__in=week_dates).values_list('date', flat=True))
    existing = {(o.agent_id, o.field): o.value for o in NominaOverride.objects.filter(
        agent__in=list(agents) + list(admins), week_start=week_start)}

    def _cells(a, computed, fields):
        out = []
        for field, _label in fields:
            o = existing.get((a.pk, field))
            out.append({'field': field, 'computed': computed[field],
                        'override': '' if o is None else f'{o:.2f}'})
        return out

    # Agents
    data = _get_billable_weekly_data(agents, week_dates, settings)
    ba_agents = set(BreakAbuseIncident.objects.filter(
        agent__in=agents, date__in=week_dates).values_list('agent_id', flat=True))
    hol_hours = _holiday_worked_hours(agents, holiday_dates, settings.nr_ratio)
    hol_nw_hours = _holiday_not_worked_hours(agents, holiday_dates, week_dates)
    rows = []
    for a in agents:
        d = data.get(a.pk, {})
        rate = d.get('hourly_mxn', Decimal('0'))
        computed = {
            'base_pay': d.get('base_pay_mxn', Decimal('0')),
            'adherence': Decimal('0') if a.pk in ba_agents else d.get('bonus_mxn', Decimal('0')),
            'holiday': (hol_hours.get(a.pk, Decimal('0')) * rate * 2
                        + hol_nw_hours.get(a.pk, Decimal('0')) * rate).quantize(Decimal('0.01')),
        }
        rows.append({'agent': a, 'name': a.agent_name or a.user.get_full_name() or a.user.username,
                     'cells': _cells(a, computed, OVERRIDE_FIELDS)})

    # Official admins
    adata = _get_billable_weekly_data(admins, week_dates, settings)
    ahol = _holiday_worked_hours(admins, holiday_dates, settings.nr_ratio)
    ahol_nw = _holiday_not_worked_hours(admins, holiday_dates, week_dates)
    admin_rows = []
    for a in admins:
        d = adata.get(a.pk, {})
        rate = d.get('hourly_mxn', Decimal('0'))
        computed = {
            'base_pay': d.get('base_pay_mxn', Decimal('0')),
            'admin_bonus': d.get('admin_bonus_mxn', Decimal('0')),
            'holiday': (ahol.get(a.pk, Decimal('0')) * rate * 2
                        + ahol_nw.get(a.pk, Decimal('0')) * rate).quantize(Decimal('0.01')),
        }
        admin_rows.append({'agent': a, 'name': a.agent_name or a.user.get_full_name() or a.user.username,
                           'cells': _cells(a, computed, ADMIN_OVERRIDE_FIELDS)})

    ctx = _nav(week_start, week_dates)
    ctx.update({'rows': rows, 'fields': OVERRIDE_FIELDS,
                'admin_rows': admin_rows, 'admin_fields': ADMIN_OVERRIDE_FIELDS})
    return render(request, 'nomina/overrides.html', ctx)


# ── Admin-bonus penalty matrix (from the user's "admin deduction" rules) ──────────
# Tardies and Incomplete shifts are SEPARATE tracks: each escalates on its OWN
# count-in-the-week AND its OWN consecutive-week recurrence, and the two STACK
# (one tardy 10% + one incomplete 10% = 20%). Clock-in "Issues" is a third track.
# Absent / NCNS / S each remove the whole bonus. Everything stacks, capped at 100%.
#
# Matrix: {consecutive-week run (incl. this week): {count in the week: deduction %}}.
_TRACK_ESCALATION = {
    1: {1: 10, 2: 30, 3: 60, 4: 100},   # this week only
    2: {1: 30, 2: 60, 3: 100},          # 2 weeks running
    3: {1: 60, 2: 100},                 # 3 weeks running
    4: {1: 100},                        # 4 weeks running
}
_ISSUES_ESCALATION = {
    1: {1: 25, 2: 50, 3: 100},
    2: {1: 50, 2: 100},
    3: {1: 100},
}
# Per-day hours guidance for clock-in Issues, keyed (run, count).
_ISSUES_HOURS_NOTE = {
    (1, 1): "Pay 8 hrs for the issue day.",
    (1, 2): "Pay 8 hrs the first day, 4 hrs the second.",
    (1, 3): "Pay 8 hrs day 1, 4 hrs day 2, 0 hrs day 3.",
    (2, 1): "Pay only 4 hrs for the issue day (2nd week running).",
    (2, 2): "Pay 4 hrs the first day, none the second (2nd week running).",
    (3, 1): "Pay no hours for the issue day (3rd week running).",
}
TARDY_STATUSES = frozenset({'T', 'T+VTO', 'T+I'})            # anything with a tardy component
INCOMPLETE_STATUSES = frozenset({'I', 'T+I'})               # anything with an incomplete component
FULL_PENALTY_STATUSES = frozenset({'Absent', 'NCNS', 'S'})   # each removes the whole bonus
ISSUES_STATUSES = frozenset({'Issues'})


def _matrix_pct(matrix, run, count):
    """Deduction % from an escalation matrix for a consecutive-week run and count in
    the week. Extra recurrence/count only ever means MORE penalty, so both clamp to
    the largest defined row/column (already 100%)."""
    if count <= 0:
        return 0
    row = matrix[min(max(run, 1), max(matrix))]
    return row[min(count, max(row))]


def admin_bonus_penalty(agent, week_start, override=None):
    """Recommended admin-bonus deduction for one admin/week, per the user's admin
    deduction rules. Tardies and Incompletes are SEPARATE escalating tracks that
    STACK; clock-in Issues is a third track; Absent/NCNS/S each remove the whole
    bonus. Every track escalates by its own count-in-week AND consecutive-week
    recurrence (reads history). Stacks, capped at 100%.

    Returns {'pct': Decimal 0–100, 'reasons': [str], 'hours_note': str}.
    `override` = optional (date, status) for the cell the coder just changed, applied
    on top of saved data so the recommendation is race-free. A GUIDE — the coder
    enters the final %."""
    from adherence.models import AdherenceRecord
    lookback = 5                                    # prior weeks scanned for recurrence
    window_start = week_start - timedelta(days=7 * lookback)
    week_end = week_start + timedelta(days=6)
    by_date = {d: s for d, s in AdherenceRecord.objects.filter(
        agent=agent, date__gte=window_start, date__lte=week_end
    ).values_list('date', 'status')}                # one query for the whole window
    wd = [week_start + timedelta(days=i) for i in range(7)]
    if override:                                    # reflect the just-set cell (race-free)
        od, ostatus = override
        if od in wd:
            if ostatus:
                by_date[od] = ostatus
            else:
                by_date.pop(od, None)

    def week_count(wstart, statuses):
        return sum(1 for i in range(7) if by_date.get(wstart + timedelta(days=i)) in statuses)

    def total_run(statuses):
        """This week + consecutive prior weeks that each had ≥1 of `statuses`."""
        run = 1
        for k in range(1, lookback + 1):
            if week_count(week_start - timedelta(days=7 * k), statuses) > 0:
                run += 1
            else:
                break
        return run

    pct, reasons, hours_note = 0, [], ''

    t_count = week_count(week_start, TARDY_STATUSES)
    if t_count:
        run = total_run(TARDY_STATUSES)
        p = _matrix_pct(_TRACK_ESCALATION, run, t_count)
        pct += p
        reasons.append(f"{t_count} tard{'y' if t_count == 1 else 'ies'}"
                       + (f" · {run} wks running" if run > 1 else "") + f" → −{p}%")

    i_count = week_count(week_start, INCOMPLETE_STATUSES)
    if i_count:
        run = total_run(INCOMPLETE_STATUSES)
        p = _matrix_pct(_TRACK_ESCALATION, run, i_count)
        pct += p
        reasons.append(f"{i_count} incomplete{'s' if i_count > 1 else ''}"
                       + (f" · {run} wks running" if run > 1 else "") + f" → −{p}%")

    full_hits = sorted({by_date.get(d) for d in wd if by_date.get(d) in FULL_PENALTY_STATUSES})
    if full_hits:
        pct += 100
        reasons.append(f"{', '.join(full_hits)} → −100%")

    iss_count = week_count(week_start, ISSUES_STATUSES)
    if iss_count:
        run = total_run(ISSUES_STATUSES)
        p = _matrix_pct(_ISSUES_ESCALATION, run, iss_count)
        pct += p
        reasons.append(f"{iss_count} log-in issue{'s' if iss_count > 1 else ''}"
                       + (f" · {run} wks running" if run > 1 else "") + f" → −{p}%")
        hours_note = _ISSUES_HOURS_NOTE.get((min(run, 3), min(iss_count, 3)),
                                            "Dock hours for the issue day(s) per the matrix.")

    return {'pct': Decimal(str(min(pct, 100))), 'reasons': reasons, 'hours_note': hours_note}


def _admin_bonus_factors(agents, week_dates):
    """{agent_id: (scheduled_days, worked_days)} for admin-bonus proration. Scheduled =
    days with a (non-off) shift; worked = scheduled minus vacation ('V') days that fell
    on a scheduled day. The admin bonus is prorated worked ÷ scheduled."""
    if not agents:
        return {}
    from adherence.models import AdherenceRecord
    from adherence.views import _build_maps
    shift_map = _build_maps(agents, week_dates)[0]
    v_by = {}
    for r in AdherenceRecord.objects.filter(
            agent__in=agents, date__in=week_dates, status='V').values('agent_id', 'date'):
        v_by.setdefault(r['agent_id'], set()).add(r['date'])
    out = {}
    for a in agents:
        sched = worked = 0
        for dte in week_dates:
            shift = shift_map.get((a.pk, dte))
            if shift and not getattr(shift, 'is_off', False):
                sched += 1
                if dte not in v_by.get(a.pk, set()):
                    worked += 1
        out[a.pk] = (sched, worked)
    return out


def _admin_note(r):
    """Notes for the single Admin sheet — the sheet shows UNMODIFIED numbers; each
    correction is stated here. Two categories: bonus (penalty % × vacation proration)
    and vacation hours."""
    notes = []
    if r['admin_bonus_corrected'] != r['admin_bonus']:
        reasons = []
        if r['bonus_ded_pct']:
            reasons.append(f"−{float(r['bonus_ded_pct']):g}% penalty")
        if r['vac_days']:
            reasons.append(f"{r['vac_days']} vacation day{'s' if r['vac_days'] != 1 else ''}")
        rtxt = f" ({', '.join(reasons)})" if reasons else ""
        notes.append(f"Bonus should be ${r['admin_bonus_corrected']:.2f}{rtxt}")
    if r['vac_hrs']:
        notes.append(f"Admin had {r['vac_days']} day{'s' if r['vac_days'] != 1 else ''} of vacation, "
                     f"total hours worked should be {r['corrected_hours']:.2f}")
    return " · ".join(notes)


def _admin_nomina_data(week_start, week_dates):
    """Compute the Admin Nómina rows + totals — a SINGLE sheet showing UNMODIFIED
    numbers (full admin bonus, hours before vacation) with each correction stated in
    the Notes. Super-admin overrides (base pay / admin bonus / holiday) ARE applied."""
    from finance.views import _get_billable_weekly_data
    from finance.models import BillingSettings

    settings = BillingSettings.get_for_week(week_start)
    nweek, _ = NominaWeek.objects.get_or_create(week_start=week_start)
    fx = nweek.spiff_fx_rate or Decimal('0')  # empty rate → spiffs stay 0 until set

    agents = _admin_agents(week_start)
    data = _get_billable_weekly_data(agents, week_dates, settings)
    inputs_map = {wi.agent_id: wi for wi in WeeklyPayInput.objects.filter(
        agent__in=agents, week_start=week_start)}

    # Holiday hours worked (NR-adjusted, per-day 12.5% allowance) + premium.
    holiday_dates = list(Holiday.objects.filter(date__in=week_dates).values_list('date', flat=True))
    hol_hours = _holiday_worked_hours(agents, holiday_dates, settings.nr_ratio)
    hol_nw_hours = _holiday_not_worked_hours(agents, holiday_dates, week_dates)  # scheduled, not worked (1×)
    # Prestamo GIVEN: the loan manager who fronted the cash (granted_by) is credited this
    # week's repayment for loans they granted — the ONE place a loan adds money to pay
    # (the same amount the borrower is deducted). Giving the cash itself is done in person.
    granted = {}
    for ln in Loan.objects.filter(granted_by__in=agents):
        inst = ln.installment_for_week(week_start)
        if inst:
            granted[ln.granted_by_id] = granted.get(ln.granted_by_id, Decimal('0')) + inst
    # A borrower is deducted their repayment regardless of who granted the loan, but the
    # offsetting credit only lands if the manager is on THIS week's admin nómina. Surface
    # any repayment whose manager is unset / no longer an official admin so it's never
    # silently unreconciled (the borrower is still being deducted).
    uncredited_repay, uncredited_loans = Decimal('0'), 0
    for ln in Loan.objects.exclude(granted_by__in=agents):
        inst = ln.installment_for_week(week_start)
        if inst:
            uncredited_repay += inst
            uncredited_loans += 1
    # Prestamo OWED: the admin's own loan repayment this week (deducted from the borrower).
    owed = {}
    for ln in Loan.objects.filter(agent__in=agents):
        inst = ln.installment_for_week(week_start)
        if inst:
            owed[ln.agent_id] = owed.get(ln.agent_id, Decimal('0')) + inst
    # Coder-entered admin-bonus deduction % this week (guide-recommended, manually applied).
    deductions = {x.agent_id: x.deduction_pct for x in AdminBonusDeduction.objects.filter(
        agent__in=agents, week_start=week_start)}
    vac_hours = _vacation_hours(agents, week_dates)
    from adherence.models import AdherenceRecord
    vac_days_map = {}
    for r in AdherenceRecord.objects.filter(
            agent__in=agents, date__in=week_dates, status='V').values('agent_id'):
        vac_days_map[r['agent_id']] = vac_days_map.get(r['agent_id'], 0) + 1
    bonus_factors = _admin_bonus_factors(agents, week_dates)     # {aid: (sched_days, worked_days)}
    overrides = {(o.agent_id, o.field): o.value
                 for o in NominaOverride.objects.filter(agent__in=agents, week_start=week_start)}

    def ov(aid, field, computed):
        return overrides.get((aid, field), computed)

    rows = []
    tot = {k: Decimal('0') for k in (
        'hours', 'holiday_hrs', 'holiday_pay', 'spiffs', 'lpo', 'referral', 'given',
        'subtotal', 'bonus', 'comedor', 'repay', 'transport', 'total')}
    for a in agents:
        d = data.get(a.pk, {})
        wi = inputs_map.get(a.pk)
        rate = d.get('hourly_mxn', Decimal('0'))
        extra_hrs = wi.extra_hours if wi else Decimal('0')          # manual hours correction
        hours = d.get('final_hrs', Decimal('0')) + extra_hrs
        base = ov(a.pk, 'base_pay', d.get('base_pay_mxn', Decimal('0')) + (extra_hrs * rate).quantize(Decimal('0.01')))
        # Bonus is shown FULL/unmodified; the penalty % + vacation proration land in the Note.
        gross_bonus = ov(a.pk, 'admin_bonus', d.get('admin_bonus_mxn', Decimal('0')))
        ded_pct = deductions.get(a.pk, Decimal('0'))
        hol_hrs = hol_hours.get(a.pk, Decimal('0'))
        hol_nw_hrs = hol_nw_hours.get(a.pk, Decimal('0'))
        holiday_pay = ov(a.pk, 'holiday', (hol_hrs * rate * 2 + hol_nw_hrs * rate).quantize(Decimal('0.01')))
        spiffs = ((wi.spiff_usd if wi else Decimal('0')) * fx).quantize(Decimal('0.01'))
        lpo = wi.lpo if wi else Decimal('0')
        referral = wi.referral if wi else Decimal('0')
        comedor = wi.comedor if wi else Decimal('0')
        transport = wi.transportation if wi else Decimal('0')
        prestamo_given = granted.get(a.pk, Decimal('0'))     # repayments credited to the manager (added)
        prestamo_repay = owed.get(a.pk, Decimal('0'))        # their own loan repayment (deducted)

        # Vacation prorates the bonus by worked ÷ scheduled days AND multiplies by
        # (1 − penalty%); vacation hours fold in only via the Note (sheet stays raw).
        sched_days, worked_days = bonus_factors.get(a.pk, (0, 0))
        prorate = (Decimal(worked_days) / Decimal(sched_days)) if sched_days else Decimal('1')
        corrected_bonus = max(Decimal('0'),
            (gross_bonus * prorate * (Decimal('1') - ded_pct / Decimal('100'))).quantize(Decimal('0.01')))
        vac_hrs = vac_hours.get(a.pk, Decimal('0'))
        vac_days = vac_days_map.get(a.pk, 0)
        corrected_hours = hours + vac_hrs

        subtotal = base + holiday_pay + spiffs + lpo + referral + prestamo_given
        total = subtotal + gross_bonus - comedor - prestamo_repay - transport   # raw sheet (full bonus)

        row = {
            'agent': a, 'emp': a.employee_id or '',
            'name': a.user.get_full_name() or a.user.username, 'username': a.user.username,
            'wage': rate, 'hours': hours, 'holiday_hrs': hol_hrs, 'holiday_pay': holiday_pay,
            'base_pay': base, 'spiffs': spiffs, 'lpo': lpo, 'referral': referral,
            'prestamo_given': prestamo_given, 'subtotal': subtotal, 'admin_bonus': gross_bonus,
            'bonus_ded_pct': ded_pct, 'admin_bonus_corrected': corrected_bonus,
            'vac_hrs': vac_hrs, 'vac_days': vac_days, 'corrected_hours': corrected_hours,
            'sched_days': sched_days, 'worked_days': worked_days,
            'comedor': comedor, 'prestamo_repay': prestamo_repay, 'transport': transport, 'total': total,
        }
        row['note'] = _admin_note(row)
        rows.append(row)
        tot['hours'] += hours; tot['holiday_hrs'] += hol_hrs; tot['holiday_pay'] += holiday_pay
        tot['spiffs'] += spiffs; tot['lpo'] += lpo; tot['referral'] += referral; tot['given'] += prestamo_given
        tot['subtotal'] += subtotal; tot['bonus'] += gross_bonus; tot['comedor'] += comedor
        tot['repay'] += prestamo_repay; tot['transport'] += transport; tot['total'] += total

    spiff_unpaid = sum(1 for wi in inputs_map.values() if wi.spiff_usd) if not nweek.spiff_fx_rate else 0
    tot['spiff_needs_rate'] = bool(spiff_unpaid)
    tot['spiff_unpaid_count'] = spiff_unpaid
    tot['uncredited_loans'] = uncredited_loans
    tot['uncredited_repay'] = uncredited_repay
    return rows, tot


@login_required
@nomina_access_required
def admin_nomina(request):
    """Admin Nómina — Official Admins (Infinity). Base = hours × admin wage,
    plus the fixed admin bonus, plus admin spiffs/commissions/referral, minus
    comedor/transport."""
    week_start, week_dates = _week(request)
    run = _finalized_run(week_start)
    if run:
        rows, totals = run.admin_rows, run.admin_totals
    else:
        rows, totals = _admin_nomina_data(week_start, week_dates)
    ctx = _nav(week_start, week_dates)
    ctx.update({'rows': rows, 'totals': totals, 'finalized': run})
    return render(request, 'nomina/admin_nomina.html', ctx)


ADMIN_EXPORT_COLS = [
    ('ID', 'emp'), ('Username', 'username'), ('Nombre', 'name'),
    ('Admin Wage', 'wage'), ('Hours Worked', 'hours'), ('Holiday', 'holiday_hrs'),
    ('Holiday Pay', 'holiday_pay'), ('Spiffs', 'spiffs'), ('LPO', 'lpo'), ('Refferal', 'referral'),
    ('Prestamo', 'prestamo_given'), ('Subtotal', 'subtotal'), ('Bonus', 'admin_bonus'),
    ('Cafeteria', 'comedor'), ('Prestamo', 'prestamo_repay'),
    ('Transportation', 'transport'), ('Total', 'total'),
]
ADMIN_INT_FIELDS = {'emp'}
ADMIN_HOURS_FIELDS = {'hours', 'holiday_hrs'}
ADMIN_MONEY_FIELDS = {'wage', 'holiday_pay', 'spiffs', 'lpo', 'referral', 'prestamo_given',
                      'subtotal', 'admin_bonus', 'comedor', 'prestamo_repay', 'transport', 'total'}


@login_required
@nomina_access_required
def admin_export(request):
    """Export the Admin Nómina — one sheet (with a Notes column)."""
    import openpyxl
    week_start, week_dates = _week(request)
    run = _finalized_run(week_start)
    rows = run.admin_rows if run else _admin_nomina_data(week_start, week_dates)[0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Yours'
    _write_nomina_sheet(ws, ADMIN_EXPORT_COLS, rows, ADMIN_INT_FIELDS, ADMIN_MONEY_FIELDS,
                        hours_fields=ADMIN_HOURS_FIELDS, note_fn=lambda r: r.get('note', ''))
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="LCC ADMIN NOMINA {week_start:%m%d%Y}.xlsx"'
    wb.save(resp)
    return resp


@login_required
@nomina_access_required
def finalize(request):
    """Permanently FINALIZE (freeze) a nómina week — snapshots exactly what was paid so
    the numbers never recompute afterward. Irreversible (user decision); the button is
    gated behind a confirmation on the front end."""
    from django.urls import reverse
    week_start, week_dates = _week(request)
    back = f"{reverse('nomina:agent_nomina')}?week_start={week_start.isoformat()}"
    if request.method != 'POST':
        return redirect(back)
    if _finalized_run(week_start):
        messages.info(request, "That week is already finalized.")
        return redirect(back)
    mine, agent_totals = _agent_nomina_data(week_start, week_dates, corrected=True)
    yours, _ = _agent_nomina_data(week_start, week_dates, corrected=False)
    mine_by_id = {r['agent'].pk: r for r in mine}
    yours_snap = _snap_rows(yours)
    for snap, raw in zip(yours_snap, yours):
        snap['note'] = _agent_note(raw, mine_by_id.get(raw['agent'].pk))
    admin_rows, admin_totals = _admin_nomina_data(week_start, week_dates)
    PayrollRun.objects.create(
        week_start=week_start,
        finalized_by=getattr(request.user, 'agent', None),
        agent_rows=_snap_rows(mine),
        agent_yours_rows=yours_snap,
        agent_totals=_snap_totals(agent_totals),
        admin_rows=_snap_rows(admin_rows),
        admin_totals=_snap_totals(admin_totals),
    )
    messages.success(request, f"Week of {week_start:%b %d, %Y} is FINALIZED — a permanent record of what was paid.")
    return redirect(back)
