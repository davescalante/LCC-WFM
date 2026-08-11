from datetime import date, time, timedelta
from decimal import Decimal

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from adherence.models import AdherenceRecord, Coding
from .models import (
    Agent, AgentRequest, OvertimeShift,
    OpenOTShift, OTShiftClaimRequest, OTCancellationRequest,
    Shift, ShiftTemplate,
)


def _make_agent(username, role='admin', role_type='supervisor', supervisor=None,
                official=False):
    user = User.objects.create_user(username=username, password='pw')
    return Agent.objects.create(
        user=user, role=role, role_type=role_type, agent_name=username.title(),
        supervisor=supervisor, is_official_admin=official,
    )


class StaffRequestTests(TestCase):
    def setUp(self):
        self.boss = _make_agent('boss', role_type='supervisor')
        self.other_sup = _make_agent('othersup', role_type='supervisor')
        self.coordinator = _make_agent('coord', role_type='coordinator', supervisor=self.boss)

    def _login(self, agent):
        self.client.login(username=agent.user.username, password='pw')

    def _submit_vacation(self, agent):
        self._login(agent)
        start = date.today() + timedelta(days=7)
        self.client.post(reverse('staff_my_requests'), {
            'request_type': 'vacation',
            'vacation_start': start.isoformat(),
            'vacation_end': (start + timedelta(days=1)).isoformat(),
            'notes': 'trip',
        })
        self.client.logout()
        return AgentRequest.objects.get(agent=agent)

    def test_staff_without_supervisor_cannot_submit(self):
        self._login(self.other_sup)
        resp = self.client.get(reverse('staff_my_requests'))
        self.assertContains(resp, "You need a supervisor assigned to your profile")

        resp = self.client.post(reverse('staff_my_requests'), {
            'request_type': 'vto', 'vto_date': date.today().isoformat(),
        }, follow=True)
        self.assertContains(resp, "You need a supervisor assigned to your profile")
        self.assertEqual(AgentRequest.objects.count(), 0)

    def test_staff_submission_snapshots_assigned_supervisor(self):
        ar = self._submit_vacation(self.coordinator)
        self.assertTrue(ar.is_staff_request)
        self.assertEqual(ar.assigned_supervisor, self.boss)
        self.assertEqual(ar.status, 'pending')
        self.assertFalse(ar.supervisor_read)

    def test_only_assigned_supervisor_sees_actions(self):
        ar = self._submit_vacation(self.coordinator)

        self._login(self.other_sup)
        resp = self.client.get(reverse('request_detail', kwargs={'pk': ar.pk}))
        self.assertIsNotNone(resp.context['action_block'])
        self.assertContains(resp, "can action this request")
        self.client.logout()

        self._login(self.boss)
        resp = self.client.get(reverse('request_detail', kwargs={'pk': ar.pk}))
        self.assertIsNone(resp.context['action_block'])

    def test_non_assigned_supervisor_cannot_approve(self):
        ar = self._submit_vacation(self.coordinator)
        self._login(self.other_sup)
        self.client.post(reverse('request_approve', kwargs={'pk': ar.pk}))
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'pending')

    def test_assigned_supervisor_approval_auto_applies(self):
        from .models import EmploymentPeriod
        # Give the coordinator real tenure so their 2-day request is within their
        # accrued vacation balance (else it's an over-balance request that only a
        # super admin could approve).
        EmploymentPeriod.objects.create(
            agent=self.coordinator, start_date=date.today() - timedelta(days=800))
        ar = self._submit_vacation(self.coordinator)
        self._login(self.boss)
        self.client.post(reverse('request_approve', kwargs={'pk': ar.pk}))
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'approved')
        self.assertFalse(ar.agent_read)  # requester gets a response badge
        self.assertEqual(
            AdherenceRecord.objects.filter(agent=self.coordinator, status='V').count(), 2
        )

    def test_self_approval_blocked(self):
        ar = self._submit_vacation(self.coordinator)
        self._login(self.coordinator)
        self.client.post(reverse('request_approve', kwargs={'pk': ar.pk}))
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'pending')

    def test_self_supervisor_blocked_for_everyone(self):
        loner = _make_agent('loner', role_type='coordinator')
        loner.supervisor = loner
        loner.save()
        ar = self._submit_vacation(loner)

        for viewer in (loner, self.boss):
            self._login(viewer)
            self.client.post(reverse('request_approve', kwargs={'pk': ar.pk}))
            ar.refresh_from_db()
            self.assertEqual(ar.status, 'pending')
            self.client.logout()

    def test_coding_request_approval_creates_no_coding(self):
        """Coding requests are status-only: supervisors enter the exact coded
        time manually, so neither approval nor mark-as-done creates a Coding."""
        admin = _make_agent('offadmin', role_type='qa', supervisor=self.boss, official=True)
        self._login(admin)
        self.client.post(reverse('staff_my_requests'), {
            'request_type': 'coding',
            'coding_date': date.today().isoformat(),
            'coding_start_time': '09:00',
            'coding_end_time': '11:00',
        })
        self.client.logout()
        ar = AgentRequest.objects.get(agent=admin)

        self._login(self.boss)
        self.client.post(reverse('request_approve', kwargs={'pk': ar.pk}))
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'approved')
        self.assertEqual(ar.auto_action_log, '')
        self.assertEqual(Coding.objects.count(), 0)

        self.client.post(reverse('request_mark_done', kwargs={'pk': ar.pk}))
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'done')
        self.assertEqual(Coding.objects.count(), 0)

    def test_coding_request_preserves_seconds(self):
        self._login(self.coordinator)
        self.client.post(reverse('staff_my_requests'), {
            'request_type': 'coding',
            'coding_date': date.today().isoformat(),
            'coding_start_time': '09:00:00',
            'coding_end_time': '11:30:45',
        })
        ar = AgentRequest.objects.get(agent=self.coordinator, request_type='coding')
        self.assertEqual(ar.coding_start_time, time(9, 0, 0))
        self.assertEqual(ar.coding_end_time, time(11, 30, 45))

    def test_coding_request_seconds_optional_defaults_to_zero(self):
        self._login(self.coordinator)
        self.client.post(reverse('staff_my_requests'), {
            'request_type': 'coding',
            'coding_date': date.today().isoformat(),
            'coding_start_time': '09:00',
            'coding_end_time': '11:30',
        })
        ar = AgentRequest.objects.get(agent=self.coordinator, request_type='coding')
        self.assertEqual(ar.coding_start_time, time(9, 0, 0))
        self.assertEqual(ar.coding_end_time, time(11, 30, 0))

    def test_coding_request_invalid_time_rejected_server_side(self):
        """Out-of-range hour/minute always fails, regardless of Python version
        (unlike a padding-only issue, whose acceptance can vary by version)."""
        self._login(self.coordinator)
        self.client.post(reverse('staff_my_requests'), {
            'request_type': 'coding',
            'coding_date': date.today().isoformat(),
            'coding_start_time': '25:99:00',
            'coding_end_time': '11:30',
        })
        self.assertFalse(
            AgentRequest.objects.filter(agent=self.coordinator, request_type='coding').exists()
        )

    def test_list_view_only_clears_badge_for_assigned_supervisor(self):
        ar = self._submit_vacation(self.coordinator)

        self._login(self.other_sup)
        self.client.get(reverse('requests_list'))
        ar.refresh_from_db()
        self.assertFalse(ar.supervisor_read)
        self.client.logout()

        self._login(self.boss)
        resp = self.client.get(reverse('requests_list'))
        self.assertEqual(resp.wsgi_request.supervisor_request_badge, 1)
        ar.refresh_from_db()
        self.assertTrue(ar.supervisor_read)

    def test_agent_request_flow_unchanged(self):
        agent = _make_agent('agent1', role='agent', role_type='regular_agent',
                            supervisor=self.boss)
        self._login(agent)
        self.client.post(reverse('agent_my_requests'), {
            'request_type': 'vto', 'vto_date': date.today().isoformat(),
        })
        self.client.logout()
        ar = AgentRequest.objects.get(agent=agent)
        self.assertFalse(ar.is_staff_request)
        self.assertIsNone(ar.assigned_supervisor)

        # Any staff user (not just the profile supervisor) can approve
        self._login(self.other_sup)
        self.client.post(reverse('request_approve', kwargs={'pk': ar.pk}))
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'approved')
        self.assertTrue(
            AdherenceRecord.objects.filter(agent=agent, status='VTO').exists()
        )


class OpenOTShiftTests(TestCase):
    def setUp(self):
        self.sup = _make_agent('sup', role_type='supervisor')
        self.sup2 = _make_agent('sup2', role_type='supervisor')
        self.coord = _make_agent('coord2', role_type='coordinator')
        self.qa = _make_agent('qa1', role_type='qa')
        self.agent = _make_agent('agent2', role='agent', role_type='regular_agent',
                                 supervisor=self.sup)
        self.tomorrow = date.today() + timedelta(days=1)

    def _login(self, agent):
        self.client.login(username=agent.user.username, password='pw')

    def _post_open(self, poster=None, count=1, incentive='power_hour'):
        self._login(poster or self.sup)
        self.client.post(reverse('open_ot_create'), {
            'date': self.tomorrow.isoformat(),
            'start_time': '13:00',
            'end_time': '14:00',
            'incentive_type': incentive,
            'notes': 'Need coverage for the 5pm spike',
            'count': str(count),
        })
        self.client.logout()
        return list(OpenOTShift.objects.order_by('pk'))

    def _claim(self, posting, requester):
        self._login(requester)
        self.client.post(reverse('open_ot_claim', kwargs={'pk': posting.pk}))
        self.client.logout()
        return OTShiftClaimRequest.objects.filter(open_shift=posting, requester=requester).latest('pk')

    # ── Part 1: posting ──────────────────────────────────────────────

    def test_supervisor_posts_multiple_identical_open_shifts(self):
        postings = self._post_open(count=3)
        self.assertEqual(len(postings), 3)
        for p in postings:
            self.assertEqual(p.status, 'open')
            self.assertEqual(p.incentive_type, 'power_hour')
            self.assertEqual((p.date, p.start_time, p.end_time),
                             (self.tomorrow, time(13, 0), time(14, 0)))
        # Open postings are NOT OvertimeShift rows → never counted as coverage
        self.assertEqual(OvertimeShift.objects.count(), 0)

    def test_super_admin_and_superuser_can_post_and_approve(self):
        boss = _make_agent('boss2', role_type='qa')
        boss.is_super_admin = True
        boss.save()
        postings = self._post_open(poster=boss)
        self.assertEqual(len(postings), 1)
        claim = self._claim(postings[0], self.agent)
        self._login(boss)
        self.client.post(reverse('ot_claim_approve', kwargs={'pk': claim.pk}))
        claim.refresh_from_db()
        self.assertEqual(claim.status, 'approved')
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.wsgi_request.ot_request_badge, 0)  # cleared after actioning
        self.client.logout()

        # Django superuser with no Agent profile can approve too
        User.objects.create_superuser('root', 'root@example.com', 'pw')
        posting2 = self._post_open()[0]
        claim2 = self._claim(posting2, self.agent)
        self.client.login(username='root', password='pw')
        self.client.post(reverse('ot_claim_approve', kwargs={'pk': claim2.pk}))
        claim2.refresh_from_db()
        self.assertEqual(claim2.status, 'approved')

    def test_non_approver_staff_cannot_post(self):
        self._login(self.qa)
        self.client.post(reverse('open_ot_create'), {
            'date': self.tomorrow.isoformat(), 'start_time': '13:00', 'end_time': '14:00',
        })
        self.assertEqual(OpenOTShift.objects.count(), 0)

    def test_open_shift_delete_rejects_pending_claims(self):
        posting = self._post_open()[0]
        claim = self._claim(posting, self.agent)
        self._login(self.sup)
        self.client.post(reverse('open_ot_delete', kwargs={'pk': posting.pk}))
        claim.refresh_from_db()
        posting.refresh_from_db()
        self.assertEqual(claim.status, 'rejected')
        self.assertEqual(claim.rejection_reason, 'The open shift posting was removed.')
        self.assertFalse(claim.requester_read)
        self.assertEqual(posting.status, 'removed')  # soft-deleted, history kept
        # Removed postings vanish from the available list
        self._login(self.agent)
        resp = self.client.get(reverse('agent_available_ot'))
        self.assertNotContains(resp, 'Request This Shift')

    # ── Part 2: claiming and approvals ───────────────────────────────

    def test_agent_sees_and_requests_open_shift(self):
        posting = self._post_open()[0]
        self._login(self.agent)
        resp = self.client.get(reverse('agent_available_ot'))
        self.assertContains(resp, 'Request This Shift')
        self.client.post(reverse('open_ot_claim', kwargs={'pk': posting.pk}))
        claim = OTShiftClaimRequest.objects.get()
        self.assertEqual(claim.status, 'pending')
        self.assertFalse(claim.supervisor_read)
        # Duplicate request blocked
        self.client.post(reverse('open_ot_claim', kwargs={'pk': posting.pk}))
        self.assertEqual(OTShiftClaimRequest.objects.count(), 1)
        # Others' pending requests are shown
        resp = self.client.get(reverse('agent_available_ot'))
        self.assertContains(resp, 'pending approval')

    def test_approval_assigns_real_shift_and_rejects_backups(self):
        posting = self._post_open()[0]
        claim1 = self._claim(posting, self.agent)
        claim2 = self._claim(posting, self.qa)  # backup request from staff

        self._login(self.sup)
        self.client.post(reverse('ot_claim_approve', kwargs={'pk': claim1.pk}))

        shift = OvertimeShift.objects.get()
        self.assertEqual(shift.agent, self.agent)
        self.assertEqual(shift.status, 'pending')
        self.assertEqual(shift.incentive_type, 'power_hour')
        self.assertEqual(shift.base_hourly_rate, self.agent.hourly_rate)
        self.assertEqual(shift.incentivized_hours, shift.total_shift_hours())

        posting.refresh_from_db()
        self.assertEqual(posting.status, 'filled')
        self.assertEqual(posting.filled_by, self.agent)
        self.assertEqual(posting.assigned_shift, shift)

        claim2.refresh_from_db()
        self.assertEqual(claim2.status, 'rejected')
        self.assertFalse(claim2.requester_read)

        # Filled shifts disappear from the available list
        self.client.logout()
        self._login(self.agent)
        resp = self.client.get(reverse('agent_available_ot'))
        self.assertNotContains(resp, 'Request This Shift')

    def test_self_approval_of_claim_blocked(self):
        posting = self._post_open()[0]
        claim = self._claim(posting, self.coord)

        self._login(self.coord)
        self.client.post(reverse('ot_claim_approve', kwargs={'pk': claim.pk}))
        claim.refresh_from_db()
        self.assertEqual(claim.status, 'pending')
        self.client.logout()

        self._login(self.sup2)
        self.client.post(reverse('ot_claim_approve', kwargs={'pk': claim.pk}))
        claim.refresh_from_db()
        self.assertEqual(claim.status, 'approved')
        self.assertEqual(OvertimeShift.objects.get().agent, self.coord)

    def test_non_approver_cannot_action_claims(self):
        posting = self._post_open()[0]
        claim = self._claim(posting, self.agent)
        self._login(self.qa)
        self.client.post(reverse('ot_claim_approve', kwargs={'pk': claim.pk}))
        claim.refresh_from_db()
        self.assertEqual(claim.status, 'pending')
        self.assertEqual(OvertimeShift.objects.count(), 0)

    def test_rejected_claim_keeps_shift_open(self):
        posting = self._post_open()[0]
        claim = self._claim(posting, self.agent)
        self._login(self.sup)
        self.client.post(reverse('ot_claim_reject', kwargs={'pk': claim.pk}),
                         {'rejection_reason': 'Covered already'})
        claim.refresh_from_db()
        posting.refresh_from_db()
        self.assertEqual(claim.status, 'rejected')
        self.assertEqual(claim.rejection_reason, 'Covered already')
        self.assertFalse(claim.requester_read)
        self.assertEqual(posting.status, 'open')

    # ── Part 3: cancellation requests ────────────────────────────────

    def _make_shift(self, owner):
        return OvertimeShift.objects.create(
            agent=owner, date=self.tomorrow, start_time=time(13, 0), end_time=time(14, 0),
        )

    def test_cancellation_request_does_not_change_shift(self):
        shift = self._make_shift(self.agent)
        self._login(self.agent)
        self.client.post(reverse('ot_cancel_request', kwargs={'pk': shift.pk}),
                         {'reason': 'Family emergency'})
        shift.refresh_from_db()
        self.assertEqual(shift.status, 'pending')  # still assigned and counted
        cr = OTCancellationRequest.objects.get()
        self.assertEqual(cr.status, 'pending')
        self.assertFalse(cr.supervisor_read)
        # Requester sees the pending state on their My OT Shifts page
        week_start = self.tomorrow - timedelta(days=self.tomorrow.weekday())
        resp = self.client.get(reverse('agent_my_ot_shifts') + f'?week_start={week_start.isoformat()}')
        self.assertContains(resp, 'Cancellation requested')

    def test_cancellation_requires_reason_and_ownership(self):
        shift = self._make_shift(self.agent)
        self._login(self.agent)
        self.client.post(reverse('ot_cancel_request', kwargs={'pk': shift.pk}), {'reason': '  '})
        self.assertEqual(OTCancellationRequest.objects.count(), 0)
        self.client.logout()
        self._login(self.qa)
        self.client.post(reverse('ot_cancel_request', kwargs={'pk': shift.pk}),
                         {'reason': 'not mine'})
        self.assertEqual(OTCancellationRequest.objects.count(), 0)

    def test_approved_cancellation_cancels_shift_with_reason(self):
        shift = self._make_shift(self.agent)
        self._login(self.agent)
        self.client.post(reverse('ot_cancel_request', kwargs={'pk': shift.pk}),
                         {'reason': 'Family emergency'})
        self.client.logout()
        cr = OTCancellationRequest.objects.get()

        self._login(self.sup)
        self.client.post(reverse('ot_cancel_approve', kwargs={'pk': cr.pk}))
        shift.refresh_from_db()
        cr.refresh_from_db()
        self.assertEqual(shift.status, 'cancelled')
        self.assertEqual(shift.cancellation_reason, 'Family emergency')
        self.assertEqual(cr.status, 'approved')
        self.assertFalse(cr.requester_read)

    def test_rejected_cancellation_keeps_shift_assigned(self):
        shift = self._make_shift(self.agent)
        self._login(self.agent)
        self.client.post(reverse('ot_cancel_request', kwargs={'pk': shift.pk}),
                         {'reason': 'Sick'})
        self.client.logout()
        cr = OTCancellationRequest.objects.get()

        self._login(self.sup)
        self.client.post(reverse('ot_cancel_reject', kwargs={'pk': cr.pk}),
                         {'review_note': 'No coverage available'})
        shift.refresh_from_db()
        cr.refresh_from_db()
        self.assertEqual(shift.status, 'pending')
        self.assertEqual(cr.status, 'rejected')
        self.assertEqual(cr.review_note, 'No coverage available')

    def test_supervisor_cannot_approve_own_cancellation(self):
        shift = self._make_shift(self.sup)
        self._login(self.sup)
        self.client.post(reverse('ot_cancel_request', kwargs={'pk': shift.pk}),
                         {'reason': 'Conflict'})
        cr = OTCancellationRequest.objects.get()
        self.client.post(reverse('ot_cancel_approve', kwargs={'pk': cr.pk}))
        shift.refresh_from_db()
        self.assertEqual(shift.status, 'pending')
        self.client.logout()

        self._login(self.sup2)
        self.client.post(reverse('ot_cancel_approve', kwargs={'pk': cr.pk}))
        shift.refresh_from_db()
        self.assertEqual(shift.status, 'cancelled')

    # ── Part 2/3: badges ─────────────────────────────────────────────

    def test_ot_badges(self):
        posting = self._post_open()[0]
        self._claim(posting, self.agent)

        # Approver sees badge; QA staff (non-approver) does not
        self._login(self.sup)
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.wsgi_request.ot_request_badge, 1)
        self.client.logout()
        self._login(self.qa)
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.wsgi_request.ot_request_badge, 0)
        self.client.logout()

        # Viewing the OT page clears the approver badge
        self._login(self.sup)
        self.client.get(reverse('overtime_list'))
        resp = self.client.get(reverse('dashboard'))
        self.assertEqual(resp.wsgi_request.ot_request_badge, 0)

        # Rejection badges the requester until they view Available OT
        claim = OTShiftClaimRequest.objects.get()
        self.client.post(reverse('ot_claim_reject', kwargs={'pk': claim.pk}))
        self.client.logout()
        self._login(self.agent)
        resp = self.client.get(reverse('agent_my_shifts'))
        self.assertEqual(resp.wsgi_request.agent_ot_claim_badge, 1)
        self.client.get(reverse('agent_available_ot'))
        resp = self.client.get(reverse('agent_my_shifts'))
        self.assertEqual(resp.wsgi_request.agent_ot_claim_badge, 0)


class AgentListRoleTypeFilterTests(TestCase):
    """Part 1: role_type filter on the Users page, combining with the
    existing supervisor and status filters."""

    def setUp(self):
        self.sup = _make_agent('lsup', role_type='supervisor')
        self.viewer = _make_agent('lviewer', role_type='coordinator')
        self.ra1 = _make_agent('ra1', role='agent', role_type='regular_agent',
                               supervisor=self.sup)
        self.ra2 = _make_agent('ra2', role='agent', role_type='regular_agent')
        self.kt = _make_agent('kt1', role='agent', role_type='kill_team',
                              supervisor=self.sup)
        self.inactive_ra = _make_agent('ra3', role='agent', role_type='regular_agent',
                                       supervisor=self.sup)
        self.inactive_ra.status = 'inactive'
        self.inactive_ra.save()
        self.client.login(username='lviewer', password='pw')

    def _names(self, query=''):
        resp = self.client.get(reverse('agent_list') + query)
        return {a.user.username for a in resp.context['agents']}

    def test_role_type_filter_narrows(self):
        names = self._names('?role_type=kill_team')
        self.assertEqual(names, {'kt1'})

    def test_filters_combine(self):
        # role_type + supervisor + status all narrow together
        names = self._names(f'?role_type=regular_agent&supervisor={self.sup.pk}&status_filter=active')
        self.assertEqual(names, {'ra1'})
        names = self._names(f'?role_type=regular_agent&supervisor={self.sup.pk}&status_filter=inactive')
        self.assertEqual(names, {'ra3'})

    def test_invalid_role_type_ignored(self):
        names = self._names('?role_type=bogus')
        self.assertIn('ra1', names)  # unfiltered active list
        self.assertIn('kt1', names)

    def test_existing_filters_and_pagination_unchanged(self):
        resp = self.client.get(reverse('agent_list'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn('page_obj', resp.context)
        names = self._names('?status_filter=inactive')
        self.assertEqual(names, {'ra3'})


class AgentListOfficialAdminFilterTests(TestCase):
    """The Official Admin filter narrows purely on is_official_admin, is not
    coupled to role, and combines with the existing filters."""

    def setUp(self):
        self.sup = _make_agent('oasup', role_type='supervisor')
        self.viewer = _make_agent('oaviewer', role_type='coordinator')
        self.official1 = _make_agent('oaoff1', role='admin', role_type='supervisor',
                                     supervisor=self.sup, official=True)
        self.official_agent_role = _make_agent('oaoff2', role='agent', role_type='regular_agent',
                                               official=True)
        self.regular1 = _make_agent('oareg1', role='admin', role_type='coordinator',
                                    supervisor=self.sup, official=False)
        self.inactive_official = _make_agent('oaoff3', role='admin', role_type='supervisor',
                                             supervisor=self.sup, official=True)
        self.inactive_official.status = 'inactive'
        self.inactive_official.save()
        self.client.login(username='oaviewer', password='pw')

    def _names(self, query=''):
        resp = self.client.get(reverse('agent_list') + query)
        return {a.user.username for a in resp.context['agents']}

    def test_yes_filter_returns_only_flagged(self):
        names = self._names('?official_admin=yes')
        self.assertEqual(names, {'oaoff1', 'oaoff2'})

    def test_no_filter_returns_only_unflagged(self):
        names = self._names('?official_admin=no')
        self.assertIn('oareg1', names)
        self.assertIn('oaviewer', names)
        self.assertNotIn('oaoff1', names)
        self.assertNotIn('oaoff2', names)

    def test_all_default_unchanged(self):
        baseline = self._names()
        self.assertEqual(self._names('?official_admin='), baseline)

    def test_not_coupled_to_role(self):
        # oaoff2 has role='agent' with is_official_admin=True — must still
        # surface under the filter, proving it isn't gated on the role field.
        names = self._names('?official_admin=yes')
        self.assertIn('oaoff2', names)

    def test_invalid_value_ignored(self):
        baseline = self._names()
        self.assertEqual(self._names('?official_admin=bogus'), baseline)

    def test_combines_with_existing_filters(self):
        names = self._names(f'?official_admin=yes&supervisor={self.sup.pk}&status_filter=active')
        self.assertEqual(names, {'oaoff1'})


import io

import openpyxl

from .models import EmploymentPeriod, Five9Profile

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class AgentListExportTests(TestCase):
    """Part 2: filter-respecting Excel export with export-only columns."""

    HEADER = ['Legal name', 'Role type', 'Supervisor', 'Status',
              'Primary Five9 username', 'Start date',
              'Complete years with us', 'Full phone number']

    def setUp(self):
        self.sup = _make_agent('xsup', role_type='supervisor')
        self.viewer = _make_agent('xviewer', role_type='coordinator')
        self.client.login(username='xviewer', password='pw')
        self.today = date.today()

    def _make_full_agent(self, username, **kwargs):
        a = _make_agent(username, role='agent', role_type='regular_agent', **kwargs)
        a.user.first_name = 'Test'
        a.user.last_name = username.title()
        a.user.save()
        return a

    def _export_ws(self, query=''):
        resp = self.client.get(reverse('agent_list') + '?export=1' + query)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        return openpyxl.load_workbook(io.BytesIO(resp.content)).active

    def _export_rows(self, query=''):
        ws = self._export_ws(query)
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def _row_for(self, rows, full_name):
        return next(r for r in rows[1:] if r[0] == full_name)

    def test_header_and_column_order(self):
        rows = self._export_rows()
        self.assertEqual(rows[0], self.HEADER)

    def test_export_respects_all_filters(self):
        a1 = self._make_full_agent('exp1', supervisor=self.sup)
        self._make_full_agent('exp2')  # different supervisor
        kt = _make_agent('exp3', role='agent', role_type='kill_team', supervisor=self.sup)
        rows = self._export_rows(f'&supervisor={self.sup.pk}&role_type=regular_agent&status_filter=active')
        names = [r[0] for r in rows[1:]]
        self.assertEqual(names, ['Test Exp1'])

    def test_export_respects_official_admin_filter(self):
        official = self._make_full_agent('oaexp1', official=True)
        self._make_full_agent('oaexp2', official=False)
        rows = self._export_rows('&official_admin=yes')
        self.assertEqual([r[0] for r in rows[1:]], ['Test Oaexp1'])
        rows = self._export_rows('&official_admin=no')
        names = [r[0] for r in rows[1:]]
        self.assertIn('Test Oaexp2', names)
        self.assertNotIn('Test Oaexp1', names)

    def test_official_admin_pay_window_agent_included(self):
        # Mirrors AgentListExportPayWindowTests._separated_agent: inactive with
        # a finalized separation whose pay window is still open must still be
        # exported when official_admin=yes — proving the filter composes with
        # pay_window_q instead of replacing/short-circuiting it.
        a = self._make_full_agent('oaowed1', official=True)
        a.status = 'inactive'
        a.save()
        today = date.today()
        this_monday = today - timedelta(days=today.weekday())
        next_monday = this_monday + timedelta(days=7)
        AgentSeparation.objects.create(
            agent=a, status='finalized', separation_type='quit',
            last_day_worked=this_monday - timedelta(days=1),
            remove_from_adherence_date=next_monday,
        )
        rows = self._export_rows('&official_admin=yes')
        names = [r[0] for r in rows[1:]]
        self.assertIn('Test Oaowed1', names)

    def test_full_row_content(self):
        a = self._make_full_agent('exp4', supervisor=self.sup)
        a.phone_number = '662-369-2710'
        a.phone_country_code = '+52'
        a.save()
        Five9Profile.objects.create(agent=a, five9_username='other.acct', is_primary=False)
        Five9Profile.objects.create(agent=a, five9_username='exp4.primary', is_primary=True)
        # Two periods — the LATEST start (rehire) must win
        EmploymentPeriod.objects.create(agent=a, start_date=self.today - timedelta(days=2000),
                                        end_date=self.today - timedelta(days=1500))
        rehire = self.today - timedelta(days=912)  # ~2.5 years ago
        EmploymentPeriod.objects.create(agent=a, start_date=rehire)

        ws = self._export_ws()
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        row = self._row_for(rows, 'Test Exp4')
        self.assertEqual(row[1], 'Regular Agent')
        self.assertEqual(row[2], str(self.sup))
        self.assertEqual(row[3], 'Active')
        self.assertEqual(row[4], 'exp4.primary')
        self.assertEqual(row[5], rehire.isoformat())
        self.assertEqual(row[6], 2)  # 2.5 years → floored to 2
        self.assertEqual(row[7], '+5216623692710')  # clean text — no tab, no sci-notation
        # The phone cell is explicitly Text-formatted so Excel never coerces it
        row_idx = rows.index(row) + 1
        self.assertEqual(ws.cell(row=row_idx, column=8).number_format, '@')

    def test_years_floor_edge_cases(self):
        # 11 months → 0 complete years
        a = self._make_full_agent('exp5')
        EmploymentPeriod.objects.create(agent=a, start_date=self.today - timedelta(days=335))
        # Exactly N years today → N
        b = self._make_full_agent('exp6')
        try:
            exact = self.today.replace(year=self.today.year - 3)
        except ValueError:  # Feb 29 edge
            exact = self.today.replace(year=self.today.year - 3, day=28)
        EmploymentPeriod.objects.create(agent=b, start_date=exact)
        # One day short of a full year → 0
        c = self._make_full_agent('exp7')
        EmploymentPeriod.objects.create(agent=c, start_date=self.today - timedelta(days=364))

        rows = self._export_rows()
        self.assertEqual(self._row_for(rows, 'Test Exp5')[6], 0)
        self.assertEqual(self._row_for(rows, 'Test Exp6')[6], 3)
        self.assertEqual(self._row_for(rows, 'Test Exp7')[6], 0)

    def test_phone_formats_all_normalize(self):
        cases = {
            'exp8': '662-369-2710',
            'exp9': '(662) 369 2710',
            'exp10': '6623692710',
            'exp11': '+52 662 369 2710',
            'exp12': '+521 662 369 2710',
        }
        for username, raw in cases.items():
            a = self._make_full_agent(username)
            a.phone_number = raw
            a.phone_country_code = '+52'
            a.save()
        rows = self._export_rows()
        for username in cases:
            row = self._row_for(rows, f'Test {username.title()}')
            self.assertEqual(row[7], '+5216623692710', f'failed for {username}')

    def test_phone_us_country_code(self):
        a = self._make_full_agent('exp14')
        a.phone_country_code = '+1'
        a.phone_number = '520-369-2710'
        a.save()
        ws = self._export_ws()
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        row = self._row_for(rows, 'Test Exp14')
        self.assertEqual(row[7], '+15203692710')
        row_idx = rows.index(row) + 1
        self.assertEqual(ws.cell(row=row_idx, column=8).number_format, '@')

    def test_phone_us_country_normalizes_variants(self):
        cases = {
            'exp15': '520-369-2710',
            'exp16': '(520) 369 2710',
            'exp17': '+1 520 369 2710',
        }
        for username, raw in cases.items():
            a = self._make_full_agent(username)
            a.phone_country_code = '+1'
            a.phone_number = raw
            a.save()
        rows = self._export_rows()
        for username in cases:
            row = self._row_for(rows, f'Test {username.title()}')
            self.assertEqual(row[7], '+15203692710', f'failed for {username}')

    def test_phone_blank_country_code_defaults_to_mexico(self):
        a = self._make_full_agent('exp18')
        a.phone_country_code = ''
        a.phone_number = '662-369-2710'
        a.save()
        row = self._row_for(self._export_rows(), 'Test Exp18')
        self.assertEqual(row[7], '+5216623692710')

    def test_bare_agent_exports_blanks_without_error(self):
        self._make_full_agent('exp13')  # no phone, no periods, no five9 profiles
        row = self._row_for(self._export_rows(), 'Test Exp13')
        self.assertIsNone(row[4])   # primary five9
        self.assertIsNone(row[5])   # start date
        self.assertIsNone(row[6])   # years
        self.assertIsNone(row[7])   # phone

    def test_export_button_opens_column_picker(self):
        # The export button opens a column-picker popup listing the available
        # fields. xviewer is a coordinator (not a super admin), so financial
        # fields like Hourly rate are NOT offered.
        resp = self.client.get(reverse('agent_list'))
        self.assertContains(resp, 'Export Excel')
        self.assertContains(resp, 'choose columns')             # the picker modal
        self.assertContains(resp, 'Agent name')                 # agent name is offered
        self.assertContains(resp, 'Legal name')                 # legal name (was "Full name")
        self.assertContains(resp, 'Employee ID')                # a non-financial newly-available field
        self.assertContains(resp, 'value="full_name" checked')  # default field pre-checked
        self.assertNotContains(resp, 'Hourly rate')             # financial — gated to super admins
        self.assertNotContains(resp, 'First name')              # removed — no such field in the app
        self.assertNotContains(resp, 'Phone country code')      # removed — full number instead

    def test_hourly_rate_hidden_and_blocked_for_non_super(self):
        # A non-super admin can't see the financial column and can't force it
        # into the export by crafting the query — the server drops it.
        ws = self._export_ws('&fields=hourly_rate&fields=full_name')
        self.assertNotIn('Hourly rate (MXN)', [c.value for c in ws[1]])

    def test_hourly_rate_available_to_super_admin(self):
        self.viewer.is_super_admin = True
        self.viewer.save()
        resp = self.client.get(reverse('agent_list'))
        self.assertContains(resp, 'Hourly rate (MXN)')
        ws = self._export_ws('&fields=username&fields=hourly_rate')
        header = [c.value for c in ws[1]]
        self.assertIn('Hourly rate (MXN)', header)
        self.assertIn('Username', header)


from .models import AgentSeparation


class AgentListExportPayWindowTests(TestCase):
    """Separated agents stay in the export while their pay window is open —
    the same remove_from_adherence_date rule Finance/Adherence use."""

    def setUp(self):
        self.viewer = _make_agent('pwviewer', role_type='coordinator')
        self.client.login(username='pwviewer', password='pw')
        today = date.today()
        self.this_monday = today - timedelta(days=today.weekday())
        self.next_monday = self.this_monday + timedelta(days=7)
        self.last_monday = self.this_monday - timedelta(days=7)

    def _separated_agent(self, username, remove_date, sep_status='finalized',
                         role_type='regular_agent'):
        a = _make_agent(username, role='agent', role_type=role_type)
        a.user.first_name = 'Pay'
        a.user.last_name = username.title()
        a.user.save()
        a.status = 'inactive'
        a.save()
        AgentSeparation.objects.create(
            agent=a, status=sep_status, separation_type='quit',
            last_day_worked=self.this_monday - timedelta(days=1),
            remove_from_adherence_date=remove_date,
        )
        return a

    def _export_names(self, query=''):
        resp = self.client.get(reverse('agent_list') + '?export=1' + query)
        self.assertEqual(resp['Content-Type'], XLSX_MIME)
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        rows = list(ws.iter_rows(values_only=True))
        return [r[0] or '' for r in rows[1:]]   # Legal name may be blank (no legal name set)

    def test_still_owed_agent_included(self):
        self._separated_agent('owed1', self.next_monday)
        names = self._export_names()
        self.assertIn('Pay Owed1', names)
        # Also caught when a role-type filter is active
        names = self._export_names('&role_type=regular_agent')
        self.assertIn('Pay Owed1', names)
        # But NOT with a non-matching role filter
        names = self._export_names('&role_type=kill_team')
        self.assertNotIn('Pay Owed1', names)
        # On-screen table unchanged: default active view hides them
        resp = self.client.get(reverse('agent_list'))
        screen = {a.user.username for a in resp.context['agents']}
        self.assertNotIn('owed1', screen)

    def test_window_closed_agent_excluded(self):
        self._separated_agent('gone1', self.last_monday)
        self._separated_agent('gone2', self.this_monday)  # boundary: == Monday → out
        names = self._export_names()
        self.assertNotIn('Pay Gone1', names)
        self.assertNotIn('Pay Gone2', names)
        # Explicitly choosing Inactive still shows them (user's explicit pick)
        names = self._export_names('&status_filter=inactive')
        self.assertIn('Pay Gone1', names)

    def test_missing_data_never_errors(self):
        # Inactive with no separation record at all
        bare = _make_agent('bare1', role='agent', role_type='regular_agent')
        bare.status = 'inactive'
        bare.save()
        # Separation without a remove date
        self._separated_agent('nodate1', None)
        names = self._export_names()
        self.assertNotIn('bare1', ''.join(names).lower())
        self.assertNotIn('Pay Nodate1', names)


from unittest.mock import patch

from django.db import IntegrityError


class AgentCreateEditAtomicityTests(TestCase):
    """agent_create/agent_edit wrap their multi-step save (User, Agent,
    RoleHistory, Five9 profiles, EmploymentPeriod) in transaction.atomic() so
    a downstream failure can't leave an orphaned User (create) or a
    half-applied edit (edit)."""

    def setUp(self):
        self.admin = _make_agent('atomicadmin', role_type='supervisor')
        self.client.login(username='atomicadmin', password='pw')

    def _create_payload(self, username):
        return {
            'username': username,
            'email': f'{username}@example.com',
            'legal_name': 'New Agent',
            'password': '',
            'agent_name': 'New Agent',
            'employee_id': '',
            'role': 'agent',
            'role_type': 'regular_agent',
            'status': 'active',
            'employer': 'Infinity',
            'billing_status': 'Not Billed',
            'phone_country_code': '+1',
            'phone_number': '',
            'teams_password': '',
            'hourly_rate': '62.50',
            'billing_rate_usd': '',
            'admin_bonus_mxn': '',
            'notes': '',
        }

    def test_agent_create_rolls_back_user_on_downstream_failure(self):
        payload = self._create_payload('orphancandidate')
        with patch('scheduling.views.RoleHistory.objects.create', side_effect=IntegrityError('boom')):
            with self.assertRaises(IntegrityError):
                self.client.post(reverse('agent_create'), payload)
        self.assertFalse(User.objects.filter(username='orphancandidate').exists())
        self.assertFalse(Agent.objects.filter(user__username='orphancandidate').exists())

    def test_agent_create_succeeds_normally(self):
        payload = self._create_payload('realagent1')
        resp = self.client.post(reverse('agent_create'), payload)
        self.assertRedirects(resp, reverse('agent_list'))
        self.assertTrue(User.objects.filter(username='realagent1').exists())
        agent = Agent.objects.get(user__username='realagent1')
        self.assertEqual(agent.role_history.count(), 1)

    def _edit_payload(self, username, email, agent_name):
        return {
            'username': username,
            'email': email,
            'legal_name': agent_name,
            'password': '',
            'agent_name': agent_name,
            'employee_id': '',
            'role': 'agent',
            'role_type': 'regular_agent',
            'status': 'active',
            'employer': 'Infinity',
            'billing_status': 'Not Billed',
            'phone_country_code': '+1',
            'phone_number': '',
            'teams_password': '',
            'hourly_rate': '62.50',
            'billing_rate_usd': '',
            'admin_bonus_mxn': '',
            'notes': '',
        }

    def test_agent_edit_rolls_back_on_downstream_failure(self):
        target = _make_agent('editvictim', role='agent', role_type='regular_agent')
        target.user.email = 'original@example.com'
        target.user.save()
        payload = self._edit_payload('editvictim', 'changed@example.com', target.agent_name)
        with patch('scheduling.views._save_five9_profiles', side_effect=IntegrityError('boom')):
            with self.assertRaises(IntegrityError):
                self.client.post(reverse('agent_edit', args=[target.pk]), payload)
        target.user.refresh_from_db()
        self.assertEqual(target.user.email, 'original@example.com')

    def test_agent_edit_succeeds_normally(self):
        target = _make_agent('editok', role='agent', role_type='regular_agent')
        payload = self._edit_payload('editok', 'editok-new@example.com', target.agent_name)
        resp = self.client.post(reverse('agent_edit', args=[target.pk]), payload)
        self.assertRedirects(resp, reverse('agent_detail', args=[target.pk]))
        target.user.refresh_from_db()
        self.assertEqual(target.user.email, 'editok-new@example.com')


class AgentAdherenceStartDateValidationTests(TestCase):
    """adherence_start_date must be a Monday, enforced server-side via
    AgentForm.clean_adherence_start_date. Both agent_create and agent_edit gate
    their POST handling on AgentForm.is_valid() identically (no separate
    manual-parsing path for this field, unlike the legacy start_date), so both
    write paths must reject a non-Monday submission — tested explicitly rather
    than assuming create inherits edit's behavior."""

    def setUp(self):
        self.admin = _make_agent('floorformadmin', role_type='supervisor')
        self.client.login(username='floorformadmin', password='pw')

    def _payload(self, username, adherence_start_date=''):
        return {
            'username': username,
            'email': f'{username}@example.com',
            'legal_name': 'Floor Test',
            'password': '',
            'agent_name': 'Floor Test',
            'employee_id': '',
            'role': 'agent',
            'role_type': 'regular_agent',
            'status': 'active',
            'employer': 'Infinity',
            'billing_status': 'Not Billed',
            'phone_country_code': '+1',
            'phone_number': '',
            'teams_password': '',
            'hourly_rate': '62.50',
            'billing_rate_usd': '',
            'admin_bonus_mxn': '',
            'adherence_start_date': adherence_start_date,
            'notes': '',
        }

    def test_agent_create_rejects_non_monday(self):
        payload = self._payload('floorcreatebad', adherence_start_date='2025-01-07')  # Tuesday
        resp = self.client.post(reverse('agent_create'), payload)
        self.assertEqual(resp.status_code, 200)  # form re-rendered with errors, no redirect
        self.assertFalse(User.objects.filter(username='floorcreatebad').exists())

    def test_agent_create_accepts_monday(self):
        payload = self._payload('floorcreateok', adherence_start_date='2025-01-06')  # Monday
        resp = self.client.post(reverse('agent_create'), payload)
        self.assertRedirects(resp, reverse('agent_list'))
        agent = Agent.objects.get(user__username='floorcreateok')
        self.assertEqual(agent.adherence_start_date, date(2025, 1, 6))

    def test_agent_edit_rejects_non_monday(self):
        target = _make_agent('floorexistbad', role='agent', role_type='regular_agent')
        payload = self._payload('floorexistbad', adherence_start_date='2025-01-07')  # Tuesday
        resp = self.client.post(reverse('agent_edit', args=[target.pk]), payload)
        self.assertEqual(resp.status_code, 200)
        target.refresh_from_db()
        self.assertIsNone(target.adherence_start_date)


class AgentSeparationMondayValidationTests(TestCase):
    """remove_from_adherence_date must be a Monday, enforced server-side in
    process_separation and update_separation (both hand-parse request.POST —
    there's no Form for AgentSeparation — so each needs its own check, added
    right after each view's existing date.fromisoformat parse)."""

    MONDAY = date(2025, 1, 6)
    TUESDAY = date(2025, 1, 7)
    LAST_DAY = date(2025, 1, 2)

    def setUp(self):
        self.admin = _make_agent('sepvalidator', role_type='supervisor')
        self.client.login(username='sepvalidator', password='pw')

    def test_process_separation_accepts_monday(self):
        target = _make_agent('sepokmon', role='agent', role_type='regular_agent')
        resp = self.client.post(reverse('process_separation', args=[target.pk]), {
            'separation_status': 'finalized',
            'separation_type': 'quit',
            'last_day_worked': self.LAST_DAY.isoformat(),
            'remove_from_adherence_date': self.MONDAY.isoformat(),
            'confirm': 'on',
        })
        self.assertRedirects(resp, reverse('agent_detail', args=[target.pk]))
        sep = AgentSeparation.objects.get(agent=target)
        self.assertEqual(sep.status, 'finalized')
        self.assertEqual(sep.remove_from_adherence_date, self.MONDAY)
        target.refresh_from_db()
        self.assertEqual(target.status, 'inactive')

    def test_process_separation_rejects_non_monday(self):
        target = _make_agent('sepbadtue', role='agent', role_type='regular_agent')
        resp = self.client.post(reverse('process_separation', args=[target.pk]), {
            'separation_status': 'finalized',
            'separation_type': 'quit',
            'last_day_worked': self.LAST_DAY.isoformat(),
            'remove_from_adherence_date': self.TUESDAY.isoformat(),
            'confirm': 'on',
        }, follow=True)
        self.assertContains(resp, 'must be a Monday')
        self.assertFalse(AgentSeparation.objects.filter(agent=target).exists())
        target.refresh_from_db()
        self.assertEqual(target.status, 'active')

    def test_update_separation_finalize_rejects_non_monday(self):
        target = _make_agent('sepupdbad', role='agent', role_type='regular_agent')
        sep = AgentSeparation.objects.create(
            agent=target, status='in_progress', separation_type='quit',
            last_day_worked=self.LAST_DAY, processed_by=self.admin.user,
        )
        resp = self.client.post(reverse('update_separation', args=[target.pk]), {
            'action': 'finalize',
            'remove_from_adherence_date': self.TUESDAY.isoformat(),
            'confirm': 'on',
        }, follow=True)
        self.assertContains(resp, 'must be a Monday')
        sep.refresh_from_db()
        self.assertEqual(sep.status, 'in_progress')
        self.assertIsNone(sep.remove_from_adherence_date)
        target.refresh_from_db()
        self.assertEqual(target.status, 'active')

    def test_update_separation_finalize_accepts_monday(self):
        target = _make_agent('sepupdok', role='agent', role_type='regular_agent')
        sep = AgentSeparation.objects.create(
            agent=target, status='in_progress', separation_type='quit',
            last_day_worked=self.LAST_DAY, processed_by=self.admin.user,
        )
        resp = self.client.post(reverse('update_separation', args=[target.pk]), {
            'action': 'finalize',
            'remove_from_adherence_date': self.MONDAY.isoformat(),
            'confirm': 'on',
        })
        self.assertRedirects(resp, reverse('agent_detail', args=[target.pk]))
        sep.refresh_from_db()
        self.assertEqual(sep.status, 'finalized')
        self.assertEqual(sep.remove_from_adherence_date, self.MONDAY)
        target.refresh_from_db()
        self.assertEqual(target.status, 'inactive')

    def test_legacy_non_monday_separation_untouched_by_cancel(self):
        """A pre-existing in_progress separation with a legacy non-Monday date
        (created directly via the ORM, as production rows already are) must still
        load and be actionable. There's no re-finalize/re-submit path that would
        replay that stored value through the new check, so the closest reachable
        proof is: the page still renders, and the 'cancel' action — which never
        touches remove_from_adherence_date — still works and leaves it untouched."""
        target = _make_agent('seplegacy', role='agent', role_type='regular_agent')
        sep = AgentSeparation.objects.create(
            agent=target, status='in_progress', separation_type='quit',
            last_day_worked=self.LAST_DAY, remove_from_adherence_date=self.TUESDAY,
            processed_by=self.admin.user,
        )
        resp = self.client.get(reverse('agent_detail', args=[target.pk]))
        self.assertEqual(resp.status_code, 200)

        resp = self.client.post(reverse('update_separation', args=[target.pk]), {
            'action': 'cancel',
        })
        self.assertRedirects(resp, reverse('agent_detail', args=[target.pk]))
        sep.refresh_from_db()
        self.assertEqual(sep.status, 'cancelled')
        self.assertEqual(sep.remove_from_adherence_date, self.TUESDAY)


class AgentFormFinanceGatingTests(TestCase):
    """The Finance fields (pay/billing rate, admin bonus) and the Super Admin flag are
    stripped from the user form server-side for non-super-admins — not merely hidden."""

    def test_non_super_editor_form_strips_finance_and_super_flag(self):
        from scheduling.forms import AgentForm
        f = AgentForm(can_grant_admin_tabs=False)          # non-super editor
        for field in ('hourly_rate', 'billing_rate_usd', 'admin_bonus_mxn', 'is_super_admin'):
            self.assertNotIn(field, f.fields)
        self.assertNotIn('can_access_admin_tabs', f.fields)   # existing behavior preserved

    def test_super_editor_form_keeps_finance_and_super_flag(self):
        from scheduling.forms import AgentForm
        f = AgentForm(can_grant_admin_tabs=True)           # super-admin editor
        for field in ('hourly_rate', 'billing_rate_usd', 'admin_bonus_mxn', 'is_super_admin'):
            self.assertIn(field, f.fields)

    def test_non_super_post_cannot_write_rate_or_super_admin(self):
        # A crafted POST with these fields is ignored — the fields don't exist on the form.
        from scheduling.forms import AgentForm
        f = AgentForm(data={'hourly_rate': '999', 'is_super_admin': 'on'}, can_grant_admin_tabs=False)
        self.assertNotIn('hourly_rate', f.fields)
        self.assertNotIn('is_super_admin', f.fields)


class ShiftSaveTests(TestCase):
    """Shifts tab save path (shift_week: Permanent / One-time / Specific Date). Covers the
    retroactive-recurring-save bug (a stale Shift override shadowed the new template so the
    week being edited never visibly changed), the past-week Permanent-mode default bug, and
    the Phase 2 approved-day-off warning data. All week math is relative to date.today() since
    the view itself resolves "today" at request time — there is no fixed historical week that
    would exercise the past/current/future branches correctly."""

    def setUp(self):
        self.boss = _make_agent('shiftboss', role_type='supervisor')
        self.agent = _make_agent('shiftagent', role='agent', role_type='regular_agent')
        self.client.login(username='shiftboss', password='pw')
        self.today = date.today()
        self.week = self.today - timedelta(days=self.today.weekday())
        self.prev_week = self.week - timedelta(days=7)
        self.next_week = self.week + timedelta(days=7)

    # ---------- helpers ----------

    def _seed_recurring(self, agent, eff_from, off_days, start='06:00', end='14:00'):
        for dow in range(7):
            ShiftTemplate.objects.create(
                agent=agent, day_of_week=dow,
                start_time=None if dow in off_days else start,
                end_time=None if dow in off_days else end,
                is_off=dow in off_days, effective_from=eff_from,
            )

    def _day_payload(self, off_days, start='06:00', end='14:00'):
        payload = {}
        for i in range(7):
            if i in off_days:
                payload[f'day_{i}_off'] = 'on'
                payload[f'day_{i}_start'] = ''
                payload[f'day_{i}_end'] = ''
            else:
                payload[f'day_{i}_start'] = start
                payload[f'day_{i}_end'] = end
            payload[f'day_{i}_notes'] = ''
        return payload

    def _post(self, agent, week_start, edit_type, **extra):
        body = {'agent': agent.pk, 'week_start': week_start.isoformat(), 'edit_type': edit_type}
        body.update(extra)
        return self.client.post(reverse('shift_week'), body)

    def _week_cells(self, agent, week_start):
        """[(kind, is_off, start_time, end_time), ...] for the 7 days of week_start, read
        straight from the Shifts grid's own context — the same resolution the grid renders."""
        resp = self.client.get(reverse('shift_list') + f'?week_start={week_start.isoformat()}')
        row = next(r for r in resp.context['rows'] if r['agent'].pk == agent.pk)
        out = []
        for cell in row['cells']:
            if cell['shift']:
                s = cell['shift']
                out.append(('override', s.is_off, s.start_time, s.end_time))
            elif cell['template']:
                t = cell['template']
                out.append(('template', t.is_off, t.start_time, t.end_time))
            else:
                out.append(('empty', None, None, None))
        return out

    def _template_snapshot(self, agent):
        return sorted(
            (t.day_of_week, t.is_off, t.start_time, t.end_time, t.effective_from, t.effective_until)
            for t in ShiftTemplate.objects.filter(agent=agent)
        )

    # ---------- the core scenario ----------

    def test_recurring_save_flips_current_week_all_seven_days(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})  # Sat+Sun off
        self._post(self.agent, self.week, 'permanent',
                   effective_date=self.week.isoformat(), **self._day_payload({1, 6}))  # Tue+Sun off

        current = self._week_cells(self.agent, self.week)
        expected_off = [False, True, False, False, False, False, True]  # Mon..Sun, Tue+Sun off
        self.assertEqual([c[1] for c in current], expected_off)
        for kind, is_off, start, end in current:
            self.assertEqual(kind, 'template')
            if not is_off:
                self.assertEqual((start, end), (time(6, 0), time(14, 0)))

        prior = self._week_cells(self.agent, self.prev_week)
        old_expected_off = [False, False, False, False, False, True, True]  # Sat+Sun off
        self.assertEqual([c[1] for c in prior], old_expected_off)
        for kind, is_off, start, end in prior:
            self.assertEqual(kind, 'template')
            if not is_off:
                self.assertEqual((start, end), (time(6, 0), time(14, 0)))

        following = self._week_cells(self.agent, self.next_week)
        self.assertEqual([c[1] for c in following], expected_off)

    # ---------- the production shape (Daniela) ----------

    def test_recurring_save_overrides_stale_shift_rows_in_current_week(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        # Stale one-time overrides across the whole current week, still carrying the OLD
        # schedule — exactly the shape left behind by Copy-week / quick-edit / a prior save.
        for i in range(7):
            Shift.objects.create(
                agent=self.agent, date=self.week + timedelta(days=i),
                start_time=time(6, 0), end_time=time(14, 0), is_off=i in (5, 6),
            )

        self._post(self.agent, self.week, 'permanent',
                   effective_date=self.week.isoformat(), **self._day_payload({1, 6}))

        current = self._week_cells(self.agent, self.week)
        self.assertEqual([c[1] for c in current], [False, True, False, False, False, False, True])
        for kind, is_off, start, end in current:
            self.assertEqual(kind, 'template', "a stale override is still shadowing the new schedule")

    # ---------- later-week override survives ----------

    def test_later_week_one_time_override_survives_recurring_save(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        survivor_date = self.next_week + timedelta(days=2)
        Shift.objects.create(agent=self.agent, date=survivor_date,
                             start_time=time(10, 0), end_time=time(18, 0), is_off=False)

        self._post(self.agent, self.week, 'permanent',
                   effective_date=self.week.isoformat(), **self._day_payload({1, 6}))

        survivor = Shift.objects.get(agent=self.agent, date=survivor_date)
        self.assertEqual((survivor.start_time, survivor.end_time), (time(10, 0), time(18, 0)))
        following = self._week_cells(self.agent, self.next_week)
        self.assertEqual(following[2], ('override', False, time(10, 0), time(18, 0)))

    # ---------- effective date = today pins existing partial-week behavior ----------

    def test_recurring_effective_today_leaves_earlier_days_on_old_schedule(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        self._post(self.agent, self.week, 'permanent',
                   effective_date=self.today.isoformat(), **self._day_payload({1, 6}))

        cells = self._week_cells(self.agent, self.week)
        week_dates = [self.week + timedelta(days=i) for i in range(7)]
        for day_date, (kind, is_off, start, end) in zip(week_dates, cells):
            self.assertEqual(kind, 'template')
            if day_date < self.today:
                self.assertEqual(is_off, day_date.weekday() in (5, 6))       # old: Sat+Sun off
            else:
                self.assertEqual(is_off, day_date.weekday() in (1, 6))       # new: Tue+Sun off

    # ---------- one-time leaves templates untouched ----------

    def test_one_time_save_leaves_templates_untouched(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        before = self._template_snapshot(self.agent)

        self._post(self.agent, self.week, 'one_time', **self._day_payload({1, 6}))

        self.assertEqual(self._template_snapshot(self.agent), before)
        prior = self._week_cells(self.agent, self.prev_week)
        self.assertEqual([c[1] for c in prior], [False, False, False, False, False, True, True])
        following = self._week_cells(self.agent, self.next_week)
        self.assertEqual([c[1] for c in following], [False, False, False, False, False, True, True])
        current = self._week_cells(self.agent, self.week)
        self.assertEqual([c[1] for c in current], [False, True, False, False, False, False, True])
        self.assertTrue(all(kind == 'override' for kind, *_ in current))

    def test_one_time_on_past_week_writes_only_that_week_no_templates(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        before_count = ShiftTemplate.objects.filter(agent=self.agent).count()

        self._post(self.agent, self.prev_week, 'one_time', **self._day_payload({1, 6}))

        self.assertEqual(ShiftTemplate.objects.filter(agent=self.agent).count(), before_count)
        self.assertEqual(
            Shift.objects.filter(agent=self.agent,
                                 date__range=(self.prev_week, self.prev_week + timedelta(days=6))).count(),
            7,
        )
        self.assertFalse(Shift.objects.filter(agent=self.agent, date__gte=self.week).exists())

    # ---------- specific date (range) ----------

    def test_date_range_applies_to_every_week_then_falls_back(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        before = self._template_snapshot(self.agent)

        self._post(self.agent, self.week, 'date_range',
                   range_start=self.week.isoformat(), range_end=self.next_week.isoformat(),
                   **self._day_payload({1, 6}))

        self.assertEqual(self._template_snapshot(self.agent), before)
        for wk in (self.week, self.next_week):
            cells = self._week_cells(self.agent, wk)
            self.assertEqual([c[1] for c in cells], [False, True, False, False, False, False, True])
            self.assertTrue(all(kind == 'override' for kind, *_ in cells))

        after_range = self._week_cells(self.agent, self.next_week + timedelta(days=7))
        self.assertEqual([c[1] for c in after_range], [False, False, False, False, False, True, True])
        self.assertTrue(all(kind == 'template' for kind, *_ in after_range))

    def test_date_range_leaves_earlier_weeks_untouched(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})

        self._post(self.agent, self.week, 'date_range',
                   range_start=self.week.isoformat(), range_end=self.next_week.isoformat(),
                   **self._day_payload({1, 6}))

        prior = self._week_cells(self.agent, self.prev_week)
        self.assertEqual([c[1] for c in prior], [False, False, False, False, False, True, True])
        self.assertTrue(all(kind == 'template' for kind, *_ in prior))
        self.assertFalse(Shift.objects.filter(
            agent=self.agent, date__range=(self.prev_week, self.prev_week + timedelta(days=6))
        ).exists())

    # ---------- Bug 2a regression: past-week radio default ----------

    def test_opening_past_week_defaults_to_one_time_not_permanent(self):
        self._seed_recurring(self.agent, self.week, off_days={5, 6})  # schedule starts THIS week
        resp = self.client.get(
            reverse('shift_week') + f'?agent={self.agent.pk}&week_start={self.prev_week.isoformat()}'
        )
        html = resp.content.decode()
        self.assertTrue(resp.context['agent_has_recurring'])
        self.assertIn('value="one_time"\n                    checked', html)
        self.assertNotIn('value="permanent"\n                    checked', html)

    def test_new_agent_with_no_schedule_defaults_to_permanent(self):
        blank_agent = _make_agent('blankagent', role='agent', role_type='regular_agent')
        resp = self.client.get(
            reverse('shift_week') + f'?agent={blank_agent.pk}&week_start={self.week.isoformat()}'
        )
        html = resp.content.decode()
        self.assertFalse(resp.context['agent_has_recurring'])
        self.assertIn('value="permanent"\n                    checked', html)
        self.assertNotIn('value="one_time"\n                    checked', html)

    # ---------- Bug 2b regression: server-side edit_type validation ----------

    def test_missing_edit_type_writes_nothing(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        before_count = ShiftTemplate.objects.filter(agent=self.agent).count()
        body = {'agent': self.agent.pk, 'week_start': self.week.isoformat(),
               'effective_date': self.week.isoformat(), **self._day_payload({1, 6})}
        resp = self.client.post(reverse('shift_week'), body, follow=True)
        self.assertEqual(ShiftTemplate.objects.filter(agent=self.agent).count(), before_count)
        self.assertTrue(any('choose how to apply' in str(m).lower() for m in resp.context['messages']))

    def test_garbage_edit_type_writes_nothing(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        before_count = ShiftTemplate.objects.filter(agent=self.agent).count()
        resp = self.client.post(reverse('shift_week'), {
            'agent': self.agent.pk, 'week_start': self.week.isoformat(), 'edit_type': 'sometotallybadvalue',
            'effective_date': self.week.isoformat(), **self._day_payload({1, 6}),
        }, follow=True)
        self.assertEqual(ShiftTemplate.objects.filter(agent=self.agent).count(), before_count)
        self.assertTrue(any('choose how to apply' in str(m).lower() for m in resp.context['messages']))

    # ---------- ripple guard: the Adherence engine sees the same schedule ----------

    def test_build_maps_resolves_new_schedule_after_recurring_save(self):
        from adherence.views import _build_maps
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        self._post(self.agent, self.week, 'permanent',
                  effective_date=self.week.isoformat(), **self._day_payload({1, 6}))

        week_dates = [self.week + timedelta(days=i) for i in range(7)]
        shift_map = _build_maps([self.agent], week_dates)[0]
        tuesday = self.week + timedelta(days=1)
        saturday = self.week + timedelta(days=5)
        tue_shift = shift_map.get((self.agent.pk, tuesday))
        sat_shift = shift_map.get((self.agent.pk, saturday))
        self.assertIsNotNone(tue_shift)
        self.assertTrue(tue_shift.is_off, "Adherence engine still sees the OLD schedule on Tuesday")
        self.assertIsNotNone(sat_shift)
        self.assertFalse(sat_shift.is_off, "Adherence engine still sees the OLD schedule on Saturday")

    # ---------- adherence bonus invariance ----------

    def test_adherence_bonus_unaffected_by_schedule_change(self):
        from adherence.views import _build_maps, _build_rows
        from finance.models import BillingSettings

        for i in range(5):
            AdherenceRecord.objects.create(
                agent=self.agent, date=self.week + timedelta(days=i),
                status='P', actual_hours=Decimal('7.5'),
            )
        settings = BillingSettings.get()
        week_dates = [self.week + timedelta(days=i) for i in range(7)]

        def snapshot():
            maps = _build_maps([self.agent], week_dates)
            row = _build_rows([self.agent], week_dates, *maps, settings)[0]
            return row['bonus_mxn'], row['final_adjusted']

        self._seed_recurring(self.agent, self.week - timedelta(days=28), off_days={5, 6})
        before = snapshot()

        ShiftTemplate.objects.filter(agent=self.agent).delete()
        self._seed_recurring(self.agent, self.week - timedelta(days=28), off_days={1, 6})
        after = snapshot()

        self.assertEqual(before, after)

    # ---------- nomina admin-bonus proration ----------

    def test_admin_bonus_prorate_is_one_without_vacation_regardless_of_schedule(self):
        from nomina.views import _admin_bonus_factors
        admin = _make_agent('nominaadmin', role='admin', role_type='supervisor', official=True)
        week_dates = [self.week + timedelta(days=i) for i in range(7)]

        self._seed_recurring(admin, self.week - timedelta(days=28), off_days={5, 6})
        sched, worked = _admin_bonus_factors([admin], week_dates)[admin.pk]
        self.assertEqual(sched, worked)
        self.assertEqual(Decimal(worked) / Decimal(sched), Decimal('1'))

        ShiftTemplate.objects.filter(agent=admin).delete()
        self._seed_recurring(admin, self.week - timedelta(days=28), off_days={1, 3, 5, 6})
        sched2, worked2 = _admin_bonus_factors([admin], week_dates)[admin.pk]
        self.assertEqual(sched2, worked2)
        self.assertEqual(Decimal(worked2) / Decimal(sched2), Decimal('1'))

    # ---------- Phase 2: override delete is scoped to the week on screen ----------

    def test_override_delete_scoped_to_week_on_screen_only(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        for i in range(7):
            Shift.objects.create(agent=self.agent, date=self.week + timedelta(days=i),
                                 start_time=time(6, 0), end_time=time(14, 0), is_off=i in (5, 6))
        outside = Shift.objects.create(agent=self.agent, date=self.next_week + timedelta(days=2),
                                       start_time=time(10, 0), end_time=time(18, 0))

        self._post(self.agent, self.week, 'permanent',
                  effective_date=self.week.isoformat(), **self._day_payload({1, 6}))

        self.assertEqual(
            Shift.objects.filter(agent=self.agent,
                                 date__range=(self.week, self.week + timedelta(days=6))).count(),
            0,
        )
        self.assertTrue(Shift.objects.filter(pk=outside.pk).exists())

    # ---------- Phase 2: approved day-off warning data ----------

    def test_day_off_warnings_includes_approved_one_time_in_week(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        off_date = self.week + timedelta(days=1)
        AgentRequest.objects.create(
            agent=self.agent, request_type='day_off_change', status='approved',
            day_off_type='one_time', effective_date=off_date,
        )
        resp = self.client.get(
            reverse('shift_week') + f'?agent={self.agent.pk}&week_start={self.week.isoformat()}'
        )
        warnings = resp.context['day_off_warnings']
        self.assertEqual(len(warnings), 1)
        self.assertEqual(warnings[0]['trigger_date'], off_date.isoformat())

    def test_day_off_warnings_includes_future_recurring_request(self):
        self._seed_recurring(self.agent, self.week - timedelta(days=56), off_days={5, 6})
        future_eff = self.week + timedelta(days=21)
        AgentRequest.objects.create(
            agent=self.agent, request_type='day_off_change', status='approved',
            day_off_type='permanent', effective_date=future_eff, requested_day_off=3,
        )
        resp = self.client.get(
            reverse('shift_week') + f'?agent={self.agent.pk}&week_start={self.week.isoformat()}'
        )
        warnings = resp.context['day_off_warnings']
        self.assertEqual(len(warnings), 1)
        eff_mon = future_eff - timedelta(days=future_eff.weekday())
        self.assertEqual(warnings[0]['trigger_date'], eff_mon.isoformat())

    def test_day_off_warnings_excludes_pending_and_rejected(self):
        off_date = self.week + timedelta(days=1)
        AgentRequest.objects.create(
            agent=self.agent, request_type='day_off_change', status='pending',
            day_off_type='one_time', effective_date=off_date,
        )
        AgentRequest.objects.create(
            agent=self.agent, request_type='day_off_change', status='rejected',
            day_off_type='one_time', effective_date=off_date + timedelta(days=1),
        )
        resp = self.client.get(
            reverse('shift_week') + f'?agent={self.agent.pk}&week_start={self.week.isoformat()}'
        )
        self.assertEqual(resp.context['day_off_warnings'], [])

    def test_both_warning_triggers_can_be_present_together_independently(self):
        """The past-dated trigger and the approved-day-off trigger fire independently in the
        page's own JS, gated on separate data. This exercises the server side of that: a past
        week (whose week_start_iso sorts before current_week_start_iso, so a default Permanent
        save on it is past-dated) that ALSO has an approved day-off in range — both pieces of
        data must be present at once, neither suppressing the other. The JS trigger logic
        itself isn't exercised here (no JS test runner in this project); this confirms the data
        both triggers read is assembled together rather than one crowding out the other."""
        past_week = self.week - timedelta(days=56)
        self._seed_recurring(self.agent, past_week - timedelta(days=28), off_days={5, 6})
        off_date = past_week + timedelta(days=1)
        AgentRequest.objects.create(
            agent=self.agent, request_type='day_off_change', status='approved',
            day_off_type='one_time', effective_date=off_date,
        )
        resp = self.client.get(
            reverse('shift_week') + f'?agent={self.agent.pk}&week_start={past_week.isoformat()}'
        )
        self.assertLess(resp.context['week_start_iso'], resp.context['current_week_start_iso'])
        self.assertEqual(len(resp.context['day_off_warnings']), 1)
