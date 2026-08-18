from datetime import date, timedelta
from django.core.paginator import Paginator
from django.db.models import Q, F
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from .models import Agent, Shift, ShiftBlock, EmploymentPeriod, Five9Profile, ShiftTemplate, ShiftTemplateBlock, OvertimeShift, RoleHistory, ScheduledRoleChange, LoginLogoutUpload, AgentLoginSession, OTShiftVerification, AgentRequest, AgentSeparation, OpenOTShift, OTShiftClaimRequest, OTCancellationRequest, log_action
from .forms import AgentUserForm, AgentForm, ShiftForm


_ADMIN_ROLE_TYPES = {'supervisor', 'qa', 'cs', 'tester', 'sms_email', 'coordinator', 'trainer'}


def _sync_pending_schedule(src):
    """
    Immediately create ShiftTemplates for a pending role change's new schedule
    so that the Shifts view and Adherence tab reflect the upcoming schedule
    before the effective date arrives. Idempotent — safe to call repeatedly.
    Also expires old open templates so they don't bleed past the effective date.
    """
    if not (src.new_shift_days and src.new_shift_start_time and src.new_shift_end_time):
        return
    ag = src.agent
    eff = src.effective_date
    # Expire old open templates so they stop at the effective date
    ShiftTemplate.objects.filter(
        agent=ag, effective_until__isnull=True
    ).exclude(effective_from=eff).update(effective_until=eff)
    # Create new templates for all 7 days — working days with times, rest as off
    existing_days = set(
        ShiftTemplate.objects.filter(agent=ag, effective_from=eff)
        .values_list('day_of_week', flat=True)
    )
    working_days = set(src.new_shift_days)
    for day_num in range(7):
        if day_num in existing_days:
            continue
        if day_num in working_days:
            ShiftTemplate.objects.create(
                agent=ag, day_of_week=day_num,
                start_time=src.new_shift_start_time,
                end_time=src.new_shift_end_time,
                is_off=False, effective_from=eff,
            )
        else:
            ShiftTemplate.objects.create(
                agent=ag, day_of_week=day_num,
                start_time=None, end_time=None,
                is_off=True, effective_from=eff,
            )


def apply_due_role_changes(agent=None):
    """
    Apply all pending ScheduledRoleChanges whose effective_date has arrived.
    Pass agent= to limit to a single agent (used lazily on profile load).
    Returns the number of changes applied.
    """
    today = timezone.localdate()
    qs = ScheduledRoleChange.objects.filter(
        effective_date__lte=today,
        applied_at__isnull=True,
        cancelled_at__isnull=True,
    ).select_related('agent__supervisor')
    if agent is not None:
        qs = qs.filter(agent=agent)

    count = 0
    for src in qs:
        ag = src.agent
        old_role_type = ag.role_type
        new_role_type = src.new_role_type
        new_role = 'admin' if new_role_type in _ADMIN_ROLE_TYPES else 'agent'

        # Update Agent
        ag.role_type = new_role_type
        ag.role = new_role
        update_fields = ['role_type', 'role']
        if src.new_supervisor is not None:
            ag.supervisor = src.new_supervisor
            update_fields.append('supervisor')
        ag.save(update_fields=update_fields)

        # Update matching Five9Profiles
        ag.five9_profiles.filter(role_type=old_role_type).update(role_type=new_role_type)

        # Close open RoleHistory entry and open a new one
        open_entry = ag.role_history.filter(effective_to__isnull=True).first()
        if open_entry:
            open_entry.effective_to = src.effective_date
            open_entry.save(update_fields=['effective_to'])
        RoleHistory.objects.create(
            agent=ag,
            role=new_role,
            role_type=new_role_type,
            supervisor=src.new_supervisor if src.new_supervisor else ag.supervisor,
            employer=ag.employer,
            billing_status=ag.billing_status,
            effective_from=src.effective_date,
            changed_by=src.scheduled_by,
        )

        # Apply new schedule if provided — idempotent, _sync_pending_schedule may have run already
        if src.new_shift_days and src.new_shift_start_time and src.new_shift_end_time:
            ShiftTemplate.objects.filter(
                agent=ag, effective_until__isnull=True
            ).exclude(effective_from=src.effective_date).update(effective_until=src.effective_date)
            existing_days = set(
                ShiftTemplate.objects.filter(agent=ag, effective_from=src.effective_date)
                .values_list('day_of_week', flat=True)
            )
            for day_num in src.new_shift_days:
                if day_num not in existing_days:
                    ShiftTemplate.objects.create(
                        agent=ag,
                        day_of_week=day_num,
                        start_time=src.new_shift_start_time,
                        end_time=src.new_shift_end_time,
                        is_off=False,
                        effective_from=src.effective_date,
                    )

        log_action(
            src.scheduled_by,
            'Scheduled role change applied',
            f'Automatically applied: role changed to {src.get_new_role_type_display()} effective {src.effective_date}',
            agent=ag,
        )

        src.applied_at = timezone.now()
        src.save(update_fields=['applied_at'])
        count += 1

    return count


@login_required
def dashboard(request):
    from adherence.models import AdherenceRecord
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    # Pending requests — fetch top 5 for actionable list
    pending_qs = AgentRequest.objects.filter(
        status='pending'
    ).select_related('agent__user').order_by('-submitted_at')
    pending_count = pending_qs.count()
    pending_unread = AgentRequest.objects.filter(status='pending', supervisor_read=False).count()
    pending_top5 = list(pending_qs[:5])

    # Today's attendance from AdherenceRecord
    today_records = list(AdherenceRecord.objects.filter(
        date=today, agent__track_attendance=True
    ).values_list('status', flat=True))
    attendance_on_time = sum(1 for s in today_records if s == 'on_time')
    attendance_tardy   = sum(1 for s in today_records if s in ('tardy', 'ti'))
    attendance_absent  = sum(1 for s in today_records if s == 'absent')
    attendance_vto     = sum(1 for s in today_records if s == 'vto')
    attendance_total   = len(today_records)

    # Agents with no adherence record today (up to 5 for the actionable list)
    tracked_agents = Agent.objects.filter(
        track_attendance=True, status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    recorded_ids = set(AdherenceRecord.objects.filter(
        date=today
    ).values_list('agent_id', flat=True))
    missing_agents = [a for a in tracked_agents if a.pk not in recorded_ids]
    missing_count = len(missing_agents)
    missing_top5 = missing_agents[:5]

    # Active agents (for context)
    active_agents_count = Agent.objects.filter(status='active').count()

    return render(request, 'scheduling/dashboard.html', {
        'today': today,
        'week_start': week_start,
        'week_end': week_end,
        'pending_count': pending_count,
        'pending_unread': pending_unread,
        'pending_top5': pending_top5,
        'attendance_on_time': attendance_on_time,
        'attendance_tardy': attendance_tardy,
        'attendance_absent': attendance_absent,
        'attendance_vto': attendance_vto,
        'attendance_total': attendance_total,
        'missing_count': missing_count,
        'missing_top5': missing_top5,
        'active_agents_count': active_agents_count,
    })


# Columns available in the Users export. The user picks which to include each time
# (defaults to the classic set). (key, label, default column width).
USER_EXPORT_FIELDS = [
    ('agent_name', 'Agent name', 22),
    ('full_name', 'Legal name', 24),
    ('username', 'Username', 16),
    ('email', 'Email', 26),
    ('employee_id', 'Employee ID', 12),
    ('employer', 'Employer', 14),
    ('role', 'Role', 12),
    ('role_type', 'Role type', 16),
    ('status', 'Status', 12),
    ('supervisor', 'Supervisor', 20),
    ('primary_five9', 'Primary Five9 username', 22),
    ('all_five9', 'All Five9 usernames', 28),
    ('start_date', 'Start date', 12),
    ('years', 'Complete years with us', 10),
    ('phone', 'Full phone number', 18),
    ('hourly_rate', 'Hourly rate (MXN)', 12),
]
# The original 8-column export — the default when nothing is picked.
USER_EXPORT_DEFAULTS = ['full_name', 'role_type', 'supervisor', 'status',
                        'primary_five9', 'start_date', 'years', 'phone']
# Financial columns — offered/exported only to super admins (financial info is
# hidden from everyone else). Enforced server-side, not just in the picker.
USER_EXPORT_FINANCIAL = {'hourly_rate'}


@login_required
def agent_list(request):
    # Financial columns (e.g. hourly rate) are for super admins only.
    is_super = request.user.is_superuser or getattr(
        getattr(request.user, 'agent', None), 'is_super_admin', False)
    allowed_export_fields = [f for f in USER_EXPORT_FIELDS
                             if f[0] not in USER_EXPORT_FINANCIAL or is_super]
    allowed_keys = {f[0] for f in allowed_export_fields}
    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    if 'supervisor' in request.GET:
        supervisor_id = request.GET.get('supervisor', '')
        request.session['supervisor_filter'] = supervisor_id
    else:
        supervisor_id = request.session.get('supervisor_filter', '')

    # Status filter — default to active so separated agents don't clutter daily view
    status_filter = request.GET.get('status_filter', 'active')
    if status_filter not in ('active', 'inactive', 'in_progress', 'all'):
        status_filter = 'active'

    role_type_filter = request.GET.get('role_type', '')
    if role_type_filter not in dict(Agent.ROLE_TYPE_CHOICES):
        role_type_filter = ''

    official_admin_filter = request.GET.get('official_admin', '')
    if official_admin_filter not in ('yes', 'no'):
        official_admin_filter = ''

    agents = Agent.objects.select_related('user', 'supervisor__user').prefetch_related('separations').order_by(
        'user__last_name', 'user__first_name'
    )
    if supervisor_id:
        try:
            agents = agents.filter(supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass

    if role_type_filter:
        agents = agents.filter(role_type=role_type_filter)

    if official_admin_filter == 'yes':
        agents = agents.filter(is_official_admin=True)
    elif official_admin_filter == 'no':
        agents = agents.filter(is_official_admin=False)

    if request.GET.get('export') == '1':
        # Excel export of the full filtered list (not just the current page).
        # These columns are export-only — the on-screen table stays as is.
        # Modeled on finance billing_export; .xlsx so the phone column can be
        # true text (no CSV scientific-notation workarounds needed).
        import re as _re
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        today = timezone.localdate()
        # Pay runs a week in arrears, so the export must keep separated agents
        # until their pay window closes — the same rule Finance/Adherence use
        # (billing_report, _get_adherence_agent_pks): a finalized-separation
        # agent still counts while remove_from_adherence_date > week Monday.
        week_start = today - timedelta(days=today.weekday())
        pay_window_q = Q(status='inactive', separations__status='finalized',
                         separations__remove_from_adherence_date__gt=week_start)
        status_q = {
            'active': Q(status='active'),
            'inactive': Q(status='inactive'),
            'in_progress': Q(separations__status='in_progress'),
        }.get(status_filter)
        if status_q is not None:  # 'all' needs no narrowing
            agents = agents.filter(status_q | pay_window_q).distinct()

        # Which columns the user picked in the export popup (preserve registry
        # order; drop anything they're not allowed to see; fall back to the
        # classic set — minus any restricted field — if nothing was selected).
        picked = set(request.GET.getlist('fields'))
        selected = [k for k, _l, _w in allowed_export_fields if k in picked] \
            or [k for k in USER_EXPORT_DEFAULTS if k in allowed_keys]
        label_by_key = {k: l for k, l, _w in USER_EXPORT_FIELDS}
        width_by_key = {k: w for k, _l, w in USER_EXPORT_FIELDS}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = [label_by_key[k] for k in selected]
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A3A5C")
        center = Alignment(horizontal='center')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for a in agents.prefetch_related('five9_profiles', 'employment_periods'):
            primary_five9 = next(
                (p.five9_username for p in a.five9_profiles.all() if p.is_primary), None
            )
            # Most recent employment period start = latest rehire date
            start = max((ep.start_date for ep in a.employment_periods.all()), default=None)
            years = None
            if start:
                years = max(0, today.year - start.year
                            - ((today.month, today.day) < (start.month, start.day)))
            # Digits only, last 10 (drops any typed +52/521/1 prefix), then the
            # agent's country code (+1 US, else +521 Mexico incl. blank default)
            phone = None
            digits = _re.sub(r'\D', '', a.phone_number or '')
            if digits:
                prefix = '+1' if a.phone_country_code == '+1' else '+521'
                phone = prefix + digits[-10:]
            vals = {
                'agent_name': a.agent_name or None,
                'full_name': a.user.get_full_name() or None,   # legal name
                'username': a.user.username,
                'email': a.user.email or None,
                'employee_id': a.employee_id or None,
                'employer': a.employer or None,
                'role': a.get_role_display() if a.role else None,
                'role_type': a.get_role_type_display() or None,
                'status': a.get_status_display(),
                'supervisor': str(a.supervisor) if a.supervisor else None,
                'primary_five9': primary_five9,
                'all_five9': ', '.join(p.five9_username for p in a.five9_profiles.all()) or None,
                'start_date': start.isoformat() if start else None,
                'years': years,
                'phone': phone,
                'hourly_rate': float(a.hourly_rate) if a.hourly_rate is not None else None,
            }
            ws.append([vals[k] for k in selected])
            if 'phone' in selected:   # text format ('@') keeps Excel from coercing the phone
                ws.cell(row=ws.max_row, column=selected.index('phone') + 1).number_format = '@'

        for i, k in enumerate(selected, 1):
            ws.column_dimensions[get_column_letter(i)].width = width_by_key.get(k, 16)

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="users_{today}.xlsx"'
        wb.save(resp)
        return resp

    if status_filter == 'active':
        agents = agents.filter(status='active')
    elif status_filter == 'inactive':
        agents = agents.filter(status='inactive')
    elif status_filter == 'in_progress':
        agents = agents.filter(separations__status='in_progress').distinct()
    # 'all' — no status filter

    paginator = Paginator(agents, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'scheduling/agent_list.html', {
        'agents': page_obj,
        'page_obj': page_obj,
        'supervisors': supervisors,
        'selected_supervisor': str(supervisor_id) if supervisor_id else '',
        'status_filter': status_filter,
        'role_type_choices': Agent.ROLE_TYPE_CHOICES,
        'selected_role_type': role_type_filter,
        'selected_official_admin': official_admin_filter,
        'export_fields': [{'key': k, 'label': l, 'checked': k in USER_EXPORT_DEFAULTS}
                          for k, l, _w in allowed_export_fields],
    })


@login_required
def agent_detail(request, pk):
    agent = get_object_or_404(Agent.objects.prefetch_related('separations'), pk=pk)
    apply_due_role_changes(agent=agent)
    shifts = Shift.objects.filter(agent=agent).order_by('-date')[:30]
    pending_role_change = agent.scheduled_role_changes.filter(
        applied_at__isnull=True, cancelled_at__isnull=True
    ).select_related('new_supervisor__user').first()
    if pending_role_change:
        _sync_pending_schedule(pending_role_change)
    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    agent_separation = agent.separation  # current separation (via property)

    # Termination documentation tracker (for in_progress separations)
    tracker_data = None
    if agent_separation and agent_separation.status == 'in_progress':
        from adherence.models import AdherenceRecord
        today = date.today()
        days_since = (today - agent_separation.last_day_worked).days
        thirty_ago = today - timedelta(days=30)
        ncns_count = AdherenceRecord.objects.filter(
            agent=agent, status='NCNS', date__gte=thirty_ago
        ).count()
        absent_count = AdherenceRecord.objects.filter(
            agent=agent, status='Absent', date__gte=thirty_ago
        ).count()
        tracker_data = {
            'last_day': agent_separation.last_day_worked,
            'days_since': days_since,
            'ncns_count': ncns_count,
            'absent_count': absent_count,
        }

    return render(request, 'scheduling/agent_detail.html', {
        'agent': agent,
        'shifts': shifts,
        'pending_role_change': pending_role_change,
        'role_type_choices': Agent.ROLE_TYPE_CHOICES,
        'day_choices': ShiftTemplate.DAY_CHOICES,
        'supervisors': supervisors,
        'separation_type_choices': AgentSeparation.SEPARATION_TYPE_CHOICES,
        'tracker_data': tracker_data,
    })


def _save_five9_profiles(request, agent):
    """Process Five9Profile rows from POST: update existing, delete flagged, create new."""
    primary_pk = request.POST.get('five9_primary', '')

    for profile in list(agent.five9_profiles.all()):
        if request.POST.get(f'five9_{profile.pk}_delete'):
            profile.delete()
            continue
        username = request.POST.get(f'five9_{profile.pk}_username', '').strip()
        if username:
            profile.label = request.POST.get(f'five9_{profile.pk}_label', '').strip()
            profile.five9_username = username
            profile.five9_password = request.POST.get(f'five9_{profile.pk}_password', '').strip()
            profile.role_type = request.POST.get(f'five9_{profile.pk}_role_type', '')
            profile.billable = bool(request.POST.get(f'five9_{profile.pk}_billable'))
            profile.is_primary = (str(profile.pk) == primary_pk)
            profile.save()

    i = 0
    while f'new_five9_{i}_username' in request.POST:
        username = request.POST.get(f'new_five9_{i}_username', '').strip()
        if username:
            Five9Profile.objects.create(
                agent=agent,
                label=request.POST.get(f'new_five9_{i}_label', '').strip(),
                five9_username=username,
                five9_password=request.POST.get(f'new_five9_{i}_password', '').strip(),
                role_type=request.POST.get(f'new_five9_{i}_role_type', ''),
                billable=bool(request.POST.get(f'new_five9_{i}_billable')),
                is_primary=(primary_pk == f'new_{i}'),
            )
        i += 1


@login_required
def agent_create(request):
    user_form = AgentUserForm(request.POST or None)
    agent_form = AgentForm(request.POST or None, can_grant_admin_tabs=request.has_finance_access)
    if request.method == 'POST' and user_form.is_valid() and agent_form.is_valid():
        from django.db import transaction
        with transaction.atomic():
            user = user_form.save(commit=False)
            password = user_form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            else:
                user.set_unusable_password()
            user.save()
            agent = agent_form.save(commit=False)
            agent.user = user
            agent.save()
            from django.utils import timezone as _tz
            RoleHistory.objects.create(
                agent=agent,
                role=agent.role,
                role_type=agent.role_type or '',
                supervisor=agent.supervisor,
                employer=agent.employer,
                billing_status=agent.billing_status,
                effective_from=agent.start_date or _tz.localdate(),
                changed_by=request.user,
            )
            _save_five9_profiles(request, agent)
            if agent.role == 'admin' and agent.role_type not in ('supervisor', 'coordinator', 'cs', 'tester', 'sms_email'):
                user.set_unusable_password()
                user.save()
            start_date = request.POST.get('start_date', '').strip()
            if start_date:
                from .models import EmploymentPeriod
                EmploymentPeriod.objects.create(agent=agent, start_date=start_date)
        log_action(request.user, 'Created agent profile', f'Created {user.get_full_name()}', agent=agent)
        messages.success(request, f"User {user.get_full_name()} created successfully.")
        return redirect('agent_list')
    from finance.models import BillingSettings as _BS
    return render(request, 'scheduling/agent_form.html', {
        'user_form': user_form,
        'agent_form': agent_form,
        'title': 'Add User',
        'five9_profiles': [],
        'role_type_choices': Agent.ROLE_TYPE_CHOICES,
        'is_own_profile': False,
        'default_admin_bonus_mxn': str(_BS.get().default_admin_bonus_mxn),
        'default_adherence_bonus_max_mxn': str(_BS.get().adherence_bonus_max_mxn),
    })


@login_required
def agent_edit(request, pk):
    agent = get_object_or_404(Agent, pk=pk)

    # Auto-seed period from legacy start_date if no periods exist yet
    if agent.start_date and not agent.employment_periods.exists():
        EmploymentPeriod.objects.create(
            agent=agent,
            start_date=agent.start_date,
            end_date=agent.termination_date,
            reason_ended='terminated' if agent.termination_date else '',
        )

    if request.method == 'POST':
        user_form = AgentUserForm(request.POST, instance=agent.user)
        agent_form = AgentForm(request.POST, instance=agent, can_grant_admin_tabs=request.has_finance_access)
        if user_form.is_valid() and agent_form.is_valid():
            from django.db import transaction
            with transaction.atomic():
                # Capture user-level values before save
                _old_user = {
                    'username': agent.user.username,
                    'legal_name': f"{agent.user.first_name} {agent.user.last_name}".strip(),
                    'email': agent.user.email,
                }
                user = user_form.save(commit=False)
                password = user_form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                user.save()
                # Capture before save
                _old = {
                    'role': agent.role, 'role_type': agent.role_type,
                    'supervisor_id': agent.supervisor_id,
                    'employer': agent.employer, 'billing_status': agent.billing_status,
                    'agent_name': agent.agent_name,
                }
                agent = agent_form.save()
                # Record role history if tracked fields changed
                _new = {
                    'role': agent.role, 'role_type': agent.role_type,
                    'supervisor_id': agent.supervisor_id,
                    'employer': agent.employer, 'billing_status': agent.billing_status,
                    'agent_name': agent.agent_name,
                }
                if _old['role_type'] != _new['role_type'] and _old['role_type']:
                    agent.five9_profiles.filter(role_type=_old['role_type']).update(role_type=_new['role_type'])
                if _old != _new:
                    from django.utils import timezone as _tz
                    today = _tz.localdate()
                    if not agent.role_history.exists():
                        # Seed initial entry from old values
                        RoleHistory.objects.create(
                            agent=agent, role=_old['role'], role_type=_old['role_type'] or '',
                            supervisor_id=_old['supervisor_id'], employer=_old['employer'],
                            billing_status=_old['billing_status'],
                            effective_from=agent.start_date or today,
                            effective_to=today, changed_by=request.user,
                        )
                    else:
                        open_entry = agent.role_history.filter(effective_to__isnull=True).first()
                        if open_entry:
                            open_entry.effective_to = today
                            open_entry.save(update_fields=['effective_to'])
                    RoleHistory.objects.create(
                        agent=agent, role=agent.role, role_type=agent.role_type or '',
                        supervisor=agent.supervisor, employer=agent.employer,
                        billing_status=agent.billing_status,
                        effective_from=today, changed_by=request.user,
                    )
                if agent.role == 'admin' and agent.role_type not in ('supervisor', 'coordinator', 'cs', 'tester', 'sms_email'):
                    user.set_unusable_password()
                    user.save()

                # Update or delete existing periods
                for period in list(agent.employment_periods.all()):
                    if request.POST.get(f'period_{period.pk}_delete'):
                        period.delete()
                        continue
                    start = request.POST.get(f'period_{period.pk}_start', '').strip()
                    if start:
                        period.start_date = start
                        period.end_date = request.POST.get(f'period_{period.pk}_end', '').strip() or None
                        period.reason_ended = request.POST.get(f'period_{period.pk}_reason', '')
                        period.notes = request.POST.get(f'period_{period.pk}_notes', '')
                        period.save()

                # Create new periods (indexed rows added via JS)
                i = 0
                while f'new_{i}_start' in request.POST:
                    start = request.POST.get(f'new_{i}_start', '').strip()
                    if start:
                        EmploymentPeriod.objects.create(
                            agent=agent,
                            start_date=start,
                            end_date=request.POST.get(f'new_{i}_end', '').strip() or None,
                            reason_ended=request.POST.get(f'new_{i}_reason', ''),
                            notes=request.POST.get(f'new_{i}_notes', ''),
                        )
                    i += 1

                _save_five9_profiles(request, agent)
            # Build detailed change description
            _change_parts = []
            _new_user = {
                'username': user.username,
                'legal_name': f"{user.first_name} {user.last_name}".strip(),
                'email': user.email,
            }
            for _field, _old_val, _new_val in [
                ('username', _old_user['username'], _new_user['username']),
                ('legal name', _old_user['legal_name'], _new_user['legal_name']),
                ('email', _old_user['email'], _new_user['email']),
                ('display name', _old['agent_name'], _new['agent_name']),
                ('role', _old['role'], _new['role']),
                ('role type', _old['role_type'], _new['role_type']),
                ('employer', _old['employer'], _new['employer']),
                ('billing status', _old['billing_status'], _new['billing_status']),
            ]:
                if str(_old_val or '') != str(_new_val or ''):
                    _change_parts.append(f'{_field}: "{_old_val}" → "{_new_val}"')
            if password:
                _change_parts.append('password changed')
            _detail = '; '.join(_change_parts) if _change_parts else 'no tracked fields changed'
            log_action(request.user, 'Edited agent profile',
                       f'Edited {user.get_full_name()} — {_detail}', agent=agent)
            messages.success(request, f"User {user.get_full_name()} updated successfully.")
            return redirect('agent_detail', pk=agent.pk)
    else:
        user_form = AgentUserForm(instance=agent.user)
        agent_form = AgentForm(instance=agent, can_grant_admin_tabs=request.has_finance_access)

    from finance.models import BillingSettings as _BS
    return render(request, 'scheduling/agent_form.html', {
        'user_form': user_form,
        'agent_form': agent_form,
        'title': 'Edit User',
        'agent': agent,
        'periods': agent.employment_periods.all(),
        'reason_choices': EmploymentPeriod.REASON_CHOICES,
        'five9_profiles': agent.five9_profiles.all(),
        'role_type_choices': Agent.ROLE_TYPE_CHOICES,
        'is_own_profile': (agent.user == request.user),
        'default_admin_bonus_mxn': str(_BS.get().default_admin_bonus_mxn),
        'default_adherence_bonus_max_mxn': str(_BS.get().adherence_bonus_max_mxn),
    })


@login_required
def agent_delete(request, pk):
    agent = get_object_or_404(Agent, pk=pk)
    if request.method == 'POST':
        name = agent.user.get_full_name()
        log_action(request.user, 'Deleted agent profile', f'Deleted {name}')
        agent.user.delete()
        messages.success(request, f"User {name} deleted.")
        return redirect('agent_list')
    return render(request, 'scheduling/confirm_delete.html', {
        'object': agent,
        'cancel_url': reverse('agent_list'),
    })


def _save_shift_template(agent, day_of_week, effective_date, is_off, start, end, notes):
    """
    Cap any existing active template before effective_date and create a new one from effective_date.
    If an active template already starts on this exact date, just update it in place.
    Any templates that start AFTER effective_date are deleted — they are fully superseded.
    effective_date is the specific date the change takes effect (not necessarily week Monday).
    Returns the (template, created) tuple.
    """
    active = (
        ShiftTemplate.objects
        .filter(agent=agent, day_of_week=day_of_week)
        .filter(Q(effective_from__isnull=True) | Q(effective_from__lte=effective_date))
        .filter(Q(effective_until__isnull=True) | Q(effective_until__gte=effective_date))
        .order_by(F('effective_from').desc(nulls_last=True))
        .first()
    )
    if active and active.effective_from == effective_date:
        # Same date — update in place; no history to preserve
        active.start_time = start or '09:00'
        active.end_time = end or '17:00'
        active.is_off = is_off
        active.notes = notes
        active.effective_until = None
        active.save()
        new_tmpl = active
        created = False
    else:
        if active:
            # Earlier date — cap it so prior days keep their correct schedule
            active.effective_until = effective_date - timedelta(days=1)
            active.save(update_fields=['effective_until'])
        new_tmpl = ShiftTemplate.objects.create(
            agent=agent, day_of_week=day_of_week,
            start_time=start or '09:00',
            end_time=end or '17:00',
            is_off=is_off, notes=notes,
            effective_from=effective_date, effective_until=None,
        )
        created = True
    # Delete any templates that start after effective_date — the new template supersedes them.
    ShiftTemplate.objects.filter(
        agent=agent, day_of_week=day_of_week, effective_from__gt=effective_date,
    ).exclude(pk=new_tmpl.pk).delete()
    return new_tmpl, created


@login_required
def shift_list(request):
    today = timezone.localdate()
    default_week_start = today - timedelta(days=today.weekday())

    week_start_str = request.GET.get('week_start')
    if week_start_str:
        try:
            week_start = date.fromisoformat(week_start_str)
            week_start = week_start - timedelta(days=week_start.weekday())
            request.session['shift_list_week_start'] = week_start.isoformat()
        except ValueError:
            week_start = default_week_start
    else:
        saved = request.session.get('shift_list_week_start')
        if saved:
            try:
                week_start = date.fromisoformat(saved)
            except ValueError:
                week_start = default_week_start
        else:
            week_start = default_week_start

    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    week_end = week_dates[-1]

    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    if 'supervisor' in request.GET:
        supervisor_id = request.GET.get('supervisor', '')
        request.session['supervisor_filter'] = supervisor_id
    else:
        supervisor_id = request.session.get('supervisor_filter', '')

    if 'section' in request.GET:
        section_filter = request.GET.get('section', '')
        request.session['shift_section_filter'] = section_filter
    else:
        section_filter = request.session.get('shift_section_filter', '')

    agents = Agent.objects.filter(status='active').select_related('user', 'supervisor__user').order_by('user__last_name', 'user__first_name')
    if supervisor_id:
        try:
            agents = agents.filter(supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass

    prev_week_start = week_start - timedelta(days=7)
    prev_week_dates = [prev_week_start + timedelta(days=i) for i in range(7)]

    # Override Shift records (specific dates)
    shifts_qs = Shift.objects.filter(date__in=week_dates, agent__in=agents)
    shift_map = {(s.agent_id, s.date): s for s in shifts_qs}

    # Recurring templates — per-day effective lookup so mid-week changes show correctly
    # and the Shifts tab stays consistent with the Attendance tab.
    all_templates_ls = list(ShiftTemplate.objects.filter(agent__in=agents))
    # Pre-group by (agent_id, day_of_week) for fast per-day iteration
    from collections import defaultdict as _dd
    _tmpl_by_key = _dd(list)
    for _t in all_templates_ls:
        _tmpl_by_key[(_t.agent_id, _t.day_of_week)].append(_t)

    templates_qs = ShiftTemplate.objects.filter(agent__in=agents)  # kept for .exists() checks below
    template_map = {}  # (agent_id, day_date) -> ShiftTemplate
    for day_date in week_dates:
        dow = day_date.weekday()
        for _ag in agents:
            candidates = _tmpl_by_key.get((_ag.pk, dow), [])
            best = None
            for _t in candidates:
                if _t.effective_from is not None and _t.effective_from > day_date:
                    continue
                if _t.effective_until is not None and _t.effective_until < day_date:
                    continue
                if best is None or (_t.effective_from or date.min) > (best.effective_from or date.min):
                    best = _t
            if best:
                template_map[(_ag.pk, day_date)] = best

    has_prev_week = Shift.objects.filter(date__in=prev_week_dates).exists()
    has_this_week = shifts_qs.exists() or templates_qs.exists()

    prev_week_agent_ids = set(
        Shift.objects.filter(date__in=prev_week_dates).values_list('agent_id', flat=True)
    )
    this_week_override_ids = {s.agent_id for s in shifts_qs}

    rows = []
    for agent in agents:
        cells = []
        for day_date in week_dates:
            override = shift_map.get((agent.pk, day_date))
            t = template_map.get((agent.pk, day_date))
            cells.append({
                'date': day_date,
                'shift': override,
                'template': t if not override else None,
            })
        rows.append({
            'agent': agent,
            'cells': cells,
            'has_prev_week_shifts': agent.pk in prev_week_agent_ids,
            'has_this_week_shifts': agent.pk in this_week_override_ids,
            'has_template': any(
                template_map.get((agent.pk, d)) is not None
                for d in week_dates
            ),
        })

    # Classify each agent into one of four sections
    _GROUP_ORDER = {'morning': 0, 'afternoon': 1, 'kill_team': 2, 'admin': 3}
    # Fixed sort priority for admin supervisors; others sort alphabetically after
    _ADMIN_SUP_PRIORITY = {'Jesus Urbina': 0, 'Andrea Jones': 1}

    def _shift_group(agent, tmpl_map, cells):
        if agent.role == 'admin':
            return 'admin'
        if agent.role_type == 'kill_team':
            return 'kill_team'
        for d in week_dates:
            t = tmpl_map.get((agent.pk, d))
            if t and not t.is_off and t.start_time:
                return 'morning' if t.start_time.hour < 10 else 'afternoon'
        for cell in cells:
            s = cell['shift']
            if s and not s.is_off and s.start_time:
                return 'morning' if s.start_time.hour < 10 else 'afternoon'
        return 'afternoon'

    for row in rows:
        row['group'] = _shift_group(row['agent'], template_map, row['cells'])

    def _sort_key(r):
        g = _GROUP_ORDER[r['group']]
        a = r['agent']
        if r['group'] == 'admin':
            if not a.supervisor:
                sup_sort = (2, '', '')
            else:
                name = a.supervisor.user.get_full_name()
                if name in _ADMIN_SUP_PRIORITY:
                    sup_sort = (_ADMIN_SUP_PRIORITY[name], '', '')
                else:
                    sup_sort = (3, a.supervisor.user.last_name, a.supervisor.user.first_name)
        else:
            sup_sort = (0, '', '')
        return (g,) + sup_sort + (a.user.last_name, a.user.first_name)

    rows.sort(key=_sort_key)

    # Apply section filter
    if section_filter:
        rows = [r for r in rows if r['group'] == section_filter]

    # Flag first admin row per supervisor so template can draw sub-dividers
    prev_admin_sup_id = None
    for row in rows:
        if row['group'] == 'admin':
            row['show_supervisor_header'] = (row['agent'].supervisor_id != prev_admin_sup_id)
            prev_admin_sup_id = row['agent'].supervisor_id
        else:
            row['show_supervisor_header'] = False

    current_week = today - timedelta(days=today.weekday())
    return render(request, 'scheduling/shift_list.html', {
        'rows': rows,
        'week_dates': week_dates,
        'week_start': week_start,
        'week_end': week_end,
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
        'current_week': current_week,
        'is_current_week': week_start == current_week,
        'has_prev_week': has_prev_week,
        'has_this_week': has_this_week,
        'supervisors': supervisors,
        'selected_supervisor': str(supervisor_id) if supervisor_id else '',
        'section_filter': section_filter,
        'today': today,
    })


@login_required
def shift_copy_from_prev(request):
    if request.method != 'POST':
        return redirect('shift_list')

    week_start_str = request.POST.get('week_start')
    try:
        week_start = date.fromisoformat(week_start_str)
        week_start = week_start - timedelta(days=week_start.weekday())
    except (ValueError, TypeError):
        messages.error(request, "Invalid week.")
        return redirect('shift_list')

    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    prev_week_start = week_start - timedelta(days=7)
    prev_week_dates = [prev_week_start + timedelta(days=i) for i in range(7)]

    prev_shifts = list(Shift.objects.filter(date__in=prev_week_dates).select_related('agent'))
    if not prev_shifts:
        messages.warning(request, "No shifts found in the previous week to copy.")
        return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")

    # Delete existing shifts for this week, then bulk-create from previous week
    Shift.objects.filter(date__in=week_dates).delete()
    day_offset_map = {src: tgt for src, tgt in zip(prev_week_dates, week_dates)}
    Shift.objects.bulk_create([
        Shift(
            agent=s.agent,
            date=day_offset_map[s.date],
            start_time=s.start_time,
            end_time=s.end_time,
            is_off=s.is_off,
            notes=s.notes,
        )
        for s in prev_shifts
    ])

    messages.success(request, f"Schedule copied from week of {prev_week_start.strftime('%B %d')} to {week_start.strftime('%B %d, %Y')}.")
    return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")


@login_required
def shift_copy_agent_from_prev(request):
    """Copy one specific agent's shifts from the previous week to the target week."""
    if request.method != 'POST':
        return redirect('shift_list')

    week_start_str = request.POST.get('week_start')
    agent_id = request.POST.get('agent_id')
    try:
        week_start = date.fromisoformat(week_start_str)
        week_start = week_start - timedelta(days=week_start.weekday())
        agent = get_object_or_404(Agent, pk=agent_id)
    except (ValueError, TypeError):
        messages.error(request, "Invalid parameters.")
        return redirect('shift_list')

    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    prev_week_start = week_start - timedelta(days=7)
    prev_week_dates = [prev_week_start + timedelta(days=i) for i in range(7)]

    prev_shifts = list(Shift.objects.filter(agent=agent, date__in=prev_week_dates))
    if not prev_shifts:
        messages.warning(request, f"No shifts found for {agent} in the previous week.")
        return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")

    Shift.objects.filter(agent=agent, date__in=week_dates).delete()
    day_offset_map = {src: tgt for src, tgt in zip(prev_week_dates, week_dates)}
    Shift.objects.bulk_create([
        Shift(
            agent=agent,
            date=day_offset_map[s.date],
            start_time=s.start_time,
            end_time=s.end_time,
            is_off=s.is_off,
            notes=s.notes,
        )
        for s in prev_shifts
    ])

    messages.success(request, f"Copied last week's schedule for {agent} to {week_start.strftime('%B %d, %Y')}.")
    return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")


@login_required
def shift_week(request):
    agents = Agent.objects.select_related('user').order_by('user__last_name')
    DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    today = timezone.localdate()
    default_week_start = today - timedelta(days=today.weekday())

    selected_agent_id = request.GET.get('agent') or request.POST.get('agent')
    week_start_str = request.GET.get('week_start') or request.POST.get('week_start')

    try:
        week_start = date.fromisoformat(week_start_str) if week_start_str else default_week_start
        week_start = week_start - timedelta(days=week_start.weekday())
    except ValueError:
        week_start = default_week_start

    week_dates = [week_start + timedelta(days=i) for i in range(7)]

    if request.method == 'POST' and selected_agent_id:
        agent = get_object_or_404(Agent, pk=selected_agent_id)
        edit_type = request.POST.get('edit_type', 'permanent')

        def _day_configs():
            for i, day_date in enumerate(week_dates):
                is_off = request.POST.get(f'day_{i}_off') == 'on'
                start = request.POST.get(f'day_{i}_start', '').strip()
                end = request.POST.get(f'day_{i}_end', '').strip()
                notes = request.POST.get(f'day_{i}_notes', '')
                yield i, day_date, is_off, start, end, notes

        if edit_type == 'permanent':
            eff_str = request.POST.get('effective_date', '').strip()
            if not eff_str or eff_str == 'today':
                eff_date = today
            else:
                try:
                    eff_date = date.fromisoformat(eff_str)
                except (ValueError, TypeError):
                    eff_date = today
            partial_days = []
            for i, day_date, is_off, start, end, notes in _day_configs():
                if is_off or (start and end):
                    tmpl, _ = _save_shift_template(agent, i, eff_date, is_off, start, end, notes)
                    ShiftTemplateBlock.objects.filter(shift_template=tmpl).delete()
                    for n in (2, 3):
                        b_start = request.POST.get(f'day_{i}_block{n}_start', '').strip()
                        b_end = request.POST.get(f'day_{i}_block{n}_end', '').strip()
                        if b_start and b_end:
                            ShiftTemplateBlock.objects.create(
                                shift_template=tmpl, block_number=n,
                                start_time=b_start, end_time=b_end
                            )
                elif (start and not end) or (end and not start):
                    partial_days.append(DAYS[i])
            if partial_days:
                messages.warning(request, f"⚠ {', '.join(partial_days)} not saved — both a start and end time are required. Please re-enter those days.")
            log_action(request.user, 'Saved recurring schedule', f'Permanent schedule for {agent} week of {week_start}', agent=agent)
            messages.success(request, f"Recurring schedule saved for {agent}. This schedule will appear every week.")

        elif edit_type == 'one_time':
            partial_days = []
            for i, day_date, is_off, start, end, notes in _day_configs():
                if is_off or (start and end):
                    shift_obj, _ = Shift.objects.update_or_create(
                        agent=agent, date=day_date,
                        defaults={
                            'start_time': start or '09:00',
                            'end_time': end or '17:00',
                            'is_off': is_off,
                            'notes': notes,
                        }
                    )
                    ShiftBlock.objects.filter(shift=shift_obj).delete()
                    for n in (2, 3):
                        b_start = request.POST.get(f'day_{i}_block{n}_start', '').strip()
                        b_end = request.POST.get(f'day_{i}_block{n}_end', '').strip()
                        if b_start and b_end:
                            ShiftBlock.objects.create(
                                shift=shift_obj, block_number=n,
                                start_time=b_start, end_time=b_end
                            )
                elif (start and not end) or (end and not start):
                    partial_days.append(DAYS[i])
            if partial_days:
                messages.warning(request, f"⚠ {', '.join(partial_days)} not saved — both a start and end time are required. Please re-enter those days.")
            log_action(request.user, 'Saved one-time schedule', f'One-time schedule for {agent} week of {week_start}', agent=agent)
            messages.success(request, f"One-time schedule saved for week of {week_start.strftime('%B %d, %Y')} for {agent}.")

        elif edit_type == 'date_range':
            range_start_str = request.POST.get('range_start', '').strip()
            range_end_str = request.POST.get('range_end', '').strip()
            try:
                range_start = date.fromisoformat(range_start_str)
                range_start -= timedelta(days=range_start.weekday())
                range_end = date.fromisoformat(range_end_str)
                range_end -= timedelta(days=range_end.weekday())
            except (ValueError, TypeError):
                messages.error(request, "Please select a valid start and end date for the range.")
                return redirect(f"{reverse('shift_week')}?agent={selected_agent_id}&week_start={week_start.isoformat()}")

            configs = list(_day_configs())
            current_week = range_start
            week_count = 0
            while current_week <= range_end:
                for i, day_date, is_off, start, end, notes in configs:
                    target = current_week + timedelta(days=i)
                    if is_off or (start and end):
                        shift_obj, _ = Shift.objects.update_or_create(
                            agent=agent, date=target,
                            defaults={
                                'start_time': start or '09:00',
                                'end_time': end or '17:00',
                                'is_off': is_off,
                                'notes': notes,
                            }
                        )
                        ShiftBlock.objects.filter(shift=shift_obj).delete()
                        for n in (2, 3):
                            b_start = request.POST.get(f'day_{i}_block{n}_start', '').strip()
                            b_end = request.POST.get(f'day_{i}_block{n}_end', '').strip()
                            if b_start and b_end:
                                ShiftBlock.objects.create(
                                    shift=shift_obj, block_number=n,
                                    start_time=b_start, end_time=b_end
                                )
                current_week += timedelta(days=7)
                week_count += 1
            log_action(request.user, 'Saved date-range schedule',
                       f'Schedule for {agent} from {range_start} to {range_end} ({week_count} weeks)', agent=agent)
            messages.success(
                request,
                f"Schedule applied to {week_count} week(s) "
                f"({range_start.strftime('%b %d')} – {range_end.strftime('%b %d, %Y')}) for {agent}."
            )

        return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")

    # ── GET: pre-fill from overrides first, then templates ───────────────────
    overrides = {}
    templates = {}
    if selected_agent_id:
        for s in Shift.objects.filter(agent_id=selected_agent_id, date__in=week_dates):
            overrides[s.date] = s
        for t in ShiftTemplate.objects.filter(agent_id=selected_agent_id):
            in_range = (
                (t.effective_from is None or t.effective_from <= week_start)
                and (t.effective_until is None or t.effective_until >= week_start)
            )
            if not in_range:
                continue
            existing = templates.get(t.day_of_week)
            if existing is None or (t.effective_from or date.min) > (existing.effective_from or date.min):
                templates[t.day_of_week] = t

    # Fetch extra blocks for pre-filling
    tmpl_extra_map = {}  # day_of_week -> list of ShiftTemplateBlock
    for tmpl in templates.values():
        blocks = list(tmpl.extra_blocks.all())
        if blocks:
            tmpl_extra_map[tmpl.day_of_week] = blocks

    shift_extra_map = {}  # date -> list of ShiftBlock
    for shift_obj in overrides.values():
        blocks = list(shift_obj.extra_blocks.all())
        if blocks:
            shift_extra_map[shift_obj.date] = blocks

    days = []
    for i, day_date in enumerate(week_dates):
        override = overrides.get(day_date)
        tmpl = templates.get(i)
        src = override or tmpl
        # Extra blocks for this day
        if override:
            raw_extra = shift_extra_map.get(day_date, [])
        elif tmpl:
            raw_extra = tmpl_extra_map.get(i, [])
        else:
            raw_extra = []
        extra_blocks = [
            {'n': b.block_number, 'start': b.start_time.strftime('%H:%M'), 'end': b.end_time.strftime('%H:%M')}
            for b in raw_extra
        ]
        days.append({
            'index': i,
            'name': DAYS[i],
            'date': day_date,
            'start': src.start_time.strftime('%H:%M') if src and src.start_time and not src.is_off else '',
            'end': src.end_time.strftime('%H:%M') if src and src.end_time and not src.is_off else '',
            'is_off': src.is_off if src else False,
            'notes': src.notes if src else '',
            'from_template': bool(tmpl and not override),
            'has_override': bool(override),
            'extra_blocks': extra_blocks,
        })

    has_any_template = bool(templates)

    return render(request, 'scheduling/shift_week.html', {
        'agents': agents,
        'selected_agent_id': int(selected_agent_id) if selected_agent_id else None,
        'week_start': week_start,
        'week_end': week_dates[-1],
        'days': days,
        'has_any_template': has_any_template,
        'week_start_iso': week_start.isoformat(),
        'today': today,
        'today_iso': today.isoformat(),
    })


@login_required
def shift_edit(request, pk):
    shift = get_object_or_404(Shift, pk=pk)
    form = ShiftForm(request.POST or None, instance=shift)
    if request.method == 'POST' and form.is_valid():
        old_start = shift.start_time.strftime('%H:%M') if shift.start_time else '—'
        old_end = shift.end_time.strftime('%H:%M') if shift.end_time else '—'
        saved = form.save()
        new_start = saved.start_time.strftime('%H:%M') if saved.start_time else '—'
        new_end = saved.end_time.strftime('%H:%M') if saved.end_time else '—'
        week_start = shift.date - timedelta(days=shift.date.weekday())
        log_action(request.user, 'Edited shift override',
                   f'{shift.agent} on {shift.date.isoformat()}: '
                   f'{old_start}–{old_end} → {new_start}–{new_end}',
                   agent=shift.agent)
        messages.success(request, "Shift updated successfully.")
        return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")
    return render(request, 'scheduling/shift_form.html', {
        'form': form,
        'title': 'Edit Shift',
        'shift': shift,
    })


@login_required
def shift_delete(request, pk):
    from django.http import JsonResponse
    shift = get_object_or_404(Shift, pk=pk)
    if request.method == 'POST':
        if shift.date < date.today():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': 'Cannot delete a past shift.'}, status=400)
            messages.error(request, "Cannot delete a past shift — historical records are preserved.")
            return redirect(f"{reverse('shift_list')}?week_start={(shift.date - timedelta(days=shift.date.weekday())).isoformat()}")
        week_start = shift.date - timedelta(days=shift.date.weekday())
        agent = shift.agent
        log_action(request.user, 'Deleted shift override',
                   f'{agent} on {shift.date.isoformat()}: '
                   f'{shift.start_time.strftime("%H:%M")}–{shift.end_time.strftime("%H:%M")}',
                   agent=agent)
        shift.delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'deleted': True, 'pk': pk})
        messages.success(request, "Shift deleted.")
        return redirect(f"{reverse('shift_list')}?week_start={week_start.isoformat()}")
    return render(request, 'scheduling/confirm_delete.html', {
        'object': shift,
        'cancel_url': reverse('shift_list'),
    })


@login_required
def shift_clear_recurring(request):
    """AJAX: clear recurring schedule for an agent from a given week forward."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    agent_pk = request.POST.get('agent_pk')
    week_start_str = request.POST.get('week_start', '')
    agent = get_object_or_404(Agent, pk=agent_pk)

    try:
        ws = date.fromisoformat(week_start_str)
        ws = ws - timedelta(days=ws.weekday())  # ensure Monday
    except (ValueError, TypeError):
        ws = date.today()
        ws = ws - timedelta(days=ws.weekday())

    end_date = ws - timedelta(days=1)  # last day before the clearing week

    count = 0
    for tmpl in ShiftTemplate.objects.filter(agent=agent):
        if tmpl.effective_from is None or tmpl.effective_from < ws:
            # Template was active before this week — cap it so past weeks keep their display
            tmpl.effective_until = end_date
            tmpl.save(update_fields=['effective_until'])
        else:
            # Template started this week or later — remove it entirely
            tmpl.delete()
        count += 1

    log_action(request.user, 'Cleared recurring schedule',
               f'{agent}: {count} template day(s) cleared from {ws}', agent=agent)
    return JsonResponse({'deleted': count})


def _get_week_start(request):
    """Return Monday of the selected week from GET param, session, or current week."""
    today = timezone.localdate()
    default = today - timedelta(days=today.weekday())
    raw = request.GET.get('week_start')
    if raw:
        try:
            ws = date.fromisoformat(raw)
            ws = ws - timedelta(days=ws.weekday())
            request.session['sched_week_start'] = ws.isoformat()
            return ws
        except (ValueError, TypeError):
            pass
    saved = request.session.get('sched_week_start')
    if saved:
        try:
            return date.fromisoformat(saved)
        except ValueError:
            pass
    return default


def _get_supervisor_filter(request):
    """Returns (supervisor_id_str, supervisors_qs). Reads GET param saving to session."""
    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    if 'supervisor' in request.GET:
        val = request.GET.get('supervisor', '')
        request.session['supervisor_filter'] = val
        return val, supervisors

    return request.session.get('supervisor_filter', ''), supervisors


def _apply_supervisor_filter(agents_qs, supervisor_id):
    if supervisor_id:
        try:
            return agents_qs.filter(supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass
    return agents_qs


@login_required
def overtime_list(request):
    week_start = _get_week_start(request)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    week_end = week_dates[-1]

    supervisor_id, supervisors = _get_supervisor_filter(request)
    agents = Agent.objects.filter(status='active').select_related('user').order_by(
        'user__last_name', 'user__first_name'
    )
    agents = _apply_supervisor_filter(agents, supervisor_id)

    status_filter = request.GET.get('status_filter', '')
    if 'status_filter' in request.GET:
        request.session['ot_status_filter'] = status_filter
    else:
        status_filter = request.session.get('ot_status_filter', '')

    ot_qs = OvertimeShift.objects.filter(date__in=week_dates, agent__in=agents).order_by('start_time')
    if status_filter:
        ot_qs = ot_qs.filter(status=status_filter)
    ot_map = {}
    for s in ot_qs:
        ot_map.setdefault((s.agent_id, s.date), []).append(s)

    # Fetch verifications for this week's OT shifts
    ot_pks = [s.pk for shifts in ot_map.values() for s in shifts]
    verif_map = {
        v.ot_shift_id: v
        for v in OTShiftVerification.objects.filter(ot_shift_id__in=ot_pks)
    }
    for shifts in ot_map.values():
        for s in shifts:
            s.ot_verif = verif_map.get(s.pk)

    # Last login-logout upload covering any day this week
    last_ll_upload = (
        LoginLogoutUpload.objects.filter(sessions__date__in=week_dates)
        .order_by('-uploaded_at').first()
    )

    viewer = _viewer_agent(request)
    is_ot_approver_flag = _is_ot_approver(viewer, request.user)
    today_val = timezone.localdate()

    # Open OT shift postings for this week (dashed "Open Shifts" row)
    open_shifts = list(
        OpenOTShift.objects.filter(status='open', date__in=week_dates)
        .order_by('start_time')
        .prefetch_related('claim_requests__requester')
    )
    open_map = {}
    for p in open_shifts:
        p.pending_claims = [c for c in p.claim_requests.all() if c.status == 'pending']
        p.my_pending = viewer is not None and any(c.requester_id == viewer.pk for c in p.pending_claims)
        open_map.setdefault(p.date, []).append(p)
    open_cells = [{'date': d, 'shifts': open_map.get(d, [])} for d in week_dates]

    # Daily posting totals for the day-column headers: "X open · Y pending · Z filled"
    filled_by_date = {}
    for d in OpenOTShift.objects.filter(status='filled', date__in=week_dates).values_list('date', flat=True):
        filled_by_date[d] = filled_by_date.get(d, 0) + 1
    day_headers = []
    for d in week_dates:
        day_postings = open_map.get(d, [])
        n_pending = sum(1 for p in day_postings if p.pending_claims)
        summary = {
            'open': len(day_postings) - n_pending,
            'pending': n_pending,
            'filled': filled_by_date.get(d, 0),
        }
        day_headers.append({'date': d, 'summary': summary if any(summary.values()) else None})

    # Approver inbox: all pending claim + cancellation requests (not week-scoped)
    pending_claims = []
    pending_cancels = []
    if is_ot_approver_flag:
        pending_claims = list(
            OTShiftClaimRequest.objects.filter(status='pending')
            .select_related('open_shift', 'requester__user')
            .order_by('open_shift__date', 'open_shift__start_time', 'submitted_at')
        )
        pending_cancels = list(
            OTCancellationRequest.objects.filter(status='pending')
            .select_related('shift__agent__user', 'requester__user')
            .order_by('shift__date', 'submitted_at')
        )
        OTShiftClaimRequest.objects.filter(status='pending', supervisor_read=False).update(supervisor_read=True)
        OTCancellationRequest.objects.filter(status='pending', supervisor_read=False).update(supervisor_read=True)

    if viewer is not None:
        # Staff viewing this page have seen the responses to their own requests
        OTShiftClaimRequest.objects.filter(requester=viewer, requester_read=False).update(requester_read=True)
        OTCancellationRequest.objects.filter(requester=viewer, requester_read=False).update(requester_read=True)

        # Cancellation-request state for the viewer's own shifts in the grid
        own_pks = [s.pk for shifts in ot_map.values() for s in shifts if s.agent_id == viewer.pk]
        cr_map = {}
        for cr in OTCancellationRequest.objects.filter(shift_id__in=own_pks).order_by('submitted_at'):
            cr_map[cr.shift_id] = cr  # latest per shift wins
        for shifts in ot_map.values():
            for s in shifts:
                if s.agent_id == viewer.pk:
                    s.my_cancel_req = cr_map.get(s.pk)
                    s.can_request_cancel = (
                        s.status == 'pending' and s.date >= today_val
                        and not (s.my_cancel_req and s.my_cancel_req.status == 'pending')
                    )

    rows = []
    for agent in agents:
        cells = []
        week_offered = None
        week_earned = None
        for day_date in week_dates:
            ot_shifts = ot_map.get((agent.pk, day_date), [])
            active_shifts = [s for s in ot_shifts if s.status != 'cancelled']
            offered_vals = [s.incentive_offered() for s in active_shifts if s.incentive_offered() is not None]
            earned_vals = [s.incentive_earned() for s in active_shifts if s.incentive_earned() is not None]
            cell_offered = sum(offered_vals) if offered_vals else None
            cell_earned = sum(earned_vals) if earned_vals else None
            if cell_offered is not None:
                week_offered = (week_offered or 0) + float(cell_offered)
            if cell_earned is not None:
                week_earned = (week_earned or 0) + float(cell_earned)
            cells.append({
                'date': day_date,
                'ot_shifts': ot_shifts,
                'has_active_ot': bool(active_shifts),
                'offered': cell_offered,
                'earned': cell_earned,
            })
        if any(c['ot_shifts'] for c in cells):
            rows.append({'agent': agent, 'cells': cells, 'week_offered': week_offered, 'week_earned': week_earned})

    return render(request, 'scheduling/overtime_list.html', {
        'rows': rows,
        'week_dates': week_dates,
        'week_start': week_start,
        'week_end': week_end,
        'today': today_val,
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
        'supervisors': supervisors,
        'selected_supervisor': str(supervisor_id) if supervisor_id else '',
        'status_filter': status_filter,
        'last_ll_upload': last_ll_upload,
        'viewer': viewer,
        'is_ot_approver': is_ot_approver_flag,
        'open_cells': open_cells,
        'day_headers': day_headers,
        'has_open_shifts': bool(open_shifts),
        'pending_claims': pending_claims,
        'pending_cancels': pending_cancels,
        'incentive_choices': OvertimeShift.INCENTIVE_CHOICES,
    })


@login_required
def verify_ot_upload(request):
    import csv, io, json as _json
    from datetime import datetime, timedelta as td
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    csv_file = request.FILES.get('file')
    if not csv_file:
        return JsonResponse({'ok': False, 'error': 'No file uploaded.'}, status=400)

    try:
        content = csv_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Could not read file: {e}'}, status=400)

    # Parse all sessions — index by (username, date)
    by_username_date = {}   # {(username, date): [(login_at, logout_at, secs)]}
    all_usernames = set()   # every username that appears anywhere in the file

    for row in rows:
        username = (row.get('AGENT') or '').strip().lower()
        date_str = (row.get('DATE') or '').strip()
        login_ts = (row.get('LOGIN TIMESTAMP') or '').strip()
        logout_ts = (row.get('LOGOUT TIMESTAMP') or '').strip()
        login_time = (row.get('LOGIN TIME') or '').strip()
        if not username or not date_str:
            continue
        try:
            row_date = datetime.strptime(date_str, '%Y/%m/%d').date()
        except ValueError:
            continue
        try:
            login_at = datetime.strptime(login_ts, '%a, %d %b %Y %H:%M:%S')
            logout_at = datetime.strptime(logout_ts, '%a, %d %b %Y %H:%M:%S')
        except ValueError:
            continue
        try:
            p = login_time.split(':')
            secs = int(p[0]) * 3600 + int(p[1]) * 60 + int(p[2])
        except (ValueError, IndexError):
            secs = max(0, int((logout_at - login_at).total_seconds()))
        all_usernames.add(username)
        by_username_date.setdefault((username, row_date), []).append((login_at, logout_at, secs))

    dates_in_file = {d for _, d in by_username_date.keys()}
    if not dates_in_file:
        return JsonResponse({'ok': False, 'error': 'No valid data found in file.'}, status=400)

    # Build agent → username(s) map from Five9Profile
    profiles = Five9Profile.objects.filter(five9_username__gt='').select_related('agent')
    agent_usernames = {}  # agent_id -> set of usernames
    for p in profiles:
        uname = p.five9_username.strip().lower()
        agent_usernames.setdefault(p.agent_id, set()).add(uname)

    # Replace sessions for these dates and create new upload record
    AgentLoginSession.objects.filter(date__in=dates_in_file).delete()
    upload = LoginLogoutUpload.objects.create(filename=csv_file.name, row_count=len(rows))

    # Build agent_map for matching
    agent_map = {
        p.five9_username.strip().lower(): p.agent
        for p in profiles if p.agent and p.agent.status == 'active'
    }

    session_objects = []
    for (username, row_date), sessions in by_username_date.items():
        agent = agent_map.get(username)
        for login_at, logout_at, secs in sessions:
            session_objects.append(AgentLoginSession(
                upload=upload, agent=agent, five9_username=username,
                date=row_date, login_at=login_at, logout_at=logout_at,
                session_seconds=secs,
            ))
    AgentLoginSession.objects.bulk_create(session_objects)

    # Verify OT shifts for dates covered
    ot_shifts = OvertimeShift.objects.filter(date__in=dates_in_file).select_related('agent')
    OTShiftVerification.objects.filter(ot_shift__in=ot_shifts).delete()

    # Fetch codings for all relevant agents on dates in the file
    from adherence.models import Coding
    coding_intervals = {}  # {(agent_id, date): [(start_dt, end_dt)]}
    for c in Coding.objects.filter(date__in=dates_in_file).values('agent_id', 'date', 'start_time', 'end_time'):
        dt_s = datetime.combine(c['date'], c['start_time'])
        dt_e = datetime.combine(c['date'], c['end_time'])
        if dt_e > dt_s:
            coding_intervals.setdefault((c['agent_id'], c['date']), []).append((dt_s, dt_e))

    def _merge_and_sum(intervals):
        if not intervals:
            return 0
        ivs = sorted(intervals)
        cs, ce = ivs[0]
        total = 0
        for s, e in ivs[1:]:
            if s <= ce:
                ce = max(ce, e)
            else:
                total += (ce - cs).total_seconds()
                cs, ce = s, e
        total += (ce - cs).total_seconds()
        return int(total)

    new_verifs = []
    for ot in ot_shifts:
        usernames = agent_usernames.get(ot.agent_id, set())
        username_found = bool(usernames & all_usernames)

        shift_start = datetime.combine(ot.date, ot.start_time)
        shift_end = datetime.combine(ot.date, ot.end_time)
        if shift_end <= shift_start:
            shift_end += td(days=1)
        shift_secs = int((shift_end - shift_start).total_seconds())

        # Clip each interval to the shift window
        def _clip(s, e):
            cs, ce = max(s, shift_start), min(e, shift_end)
            return (cs, ce) if ce > cs else None

        five9_ivs = []
        for uname in usernames:
            for login_at, logout_at, _ in by_username_date.get((uname, ot.date), []):
                iv = _clip(login_at, logout_at)
                if iv:
                    five9_ivs.append(iv)

        coding_ivs = []
        for dt_s, dt_e in coding_intervals.get((ot.agent_id, ot.date), []):
            iv = _clip(dt_s, dt_e)
            if iv:
                coding_ivs.append(iv)

        five9_secs = min(_merge_and_sum(five9_ivs), shift_secs)
        merged_secs = min(_merge_and_sum(five9_ivs + coding_ivs), shift_secs)
        coding_secs = merged_secs - five9_secs  # net additional seconds codings contributed

        new_verifs.append(OTShiftVerification(
            ot_shift=ot, upload=upload,
            verified_seconds=merged_secs,
            five9_seconds=five9_secs,
            coding_seconds=coding_secs,
            shift_seconds=shift_secs,
            username_found=username_found,
        ))
    OTShiftVerification.objects.bulk_create(new_verifs)

    log_action(request.user, 'Uploaded OT verification report',
               f'{csv_file.name} — {len(dates_in_file)} date(s), {len(new_verifs)} shifts verified')

    return JsonResponse({
        'ok': True,
        'dates': sorted(d.isoformat() for d in dates_in_file),
        'shifts_verified': len(new_verifs),
        'filename': csv_file.name,
    })


@login_required
def overtime_week(request):
    DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    agents = Agent.objects.filter(status='active').select_related('user').order_by(
        'user__last_name', 'user__first_name'
    )

    today = timezone.localdate()
    default_week_start = today - timedelta(days=today.weekday())

    selected_agent_id = request.GET.get('agent') or request.POST.get('agent')
    week_start_str = request.GET.get('week_start') or request.POST.get('week_start')

    try:
        week_start = date.fromisoformat(week_start_str) if week_start_str else default_week_start
        week_start = week_start - timedelta(days=week_start.weekday())
    except (ValueError, TypeError):
        week_start = default_week_start

    week_dates = [week_start + timedelta(days=i) for i in range(7)]

    if request.method == 'POST' and selected_agent_id:
        from decimal import Decimal, InvalidOperation
        agent = get_object_or_404(Agent, pk=selected_agent_id)

        for i, day_date in enumerate(week_dates):
            count = int(request.POST.get(f'day_{i}_ot_count', 0) or 0)
            submitted_pks = set()

            for j in range(count):
                pk_str = request.POST.get(f'day_{i}_ot_{j}_pk', '').strip()
                remove = request.POST.get(f'day_{i}_ot_{j}_remove', '').strip()
                start = request.POST.get(f'day_{i}_ot_{j}_start', '').strip()
                end = request.POST.get(f'day_{i}_ot_{j}_end', '').strip()
                notes = request.POST.get(f'day_{i}_ot_{j}_notes', '').strip()
                incentive_type = request.POST.get(f'day_{i}_ot_{j}_incentive_type', 'none').strip()
                if incentive_type not in ('none', 'time_and_a_half', 'power_hour'):
                    incentive_type = 'none'
                inc_hrs_str = request.POST.get(f'day_{i}_ot_{j}_incentivized_hours', '').strip()
                base_rate_str = request.POST.get(f'day_{i}_ot_{j}_base_hourly_rate', '').strip()
                try:
                    inc_hrs = Decimal(inc_hrs_str) if inc_hrs_str else None
                except InvalidOperation:
                    inc_hrs = None
                try:
                    base_rate = Decimal(base_rate_str) if base_rate_str else None
                except InvalidOperation:
                    base_rate = None

                if remove == 'true':
                    if pk_str:
                        OvertimeShift.objects.filter(pk=pk_str, agent=agent).delete()
                    continue

                if not (start and end):
                    if pk_str:
                        OvertimeShift.objects.filter(pk=pk_str, agent=agent).delete()
                    continue

                defaults = {
                    'start_time': start,
                    'end_time': end,
                    'notes': notes,
                    'incentive_type': incentive_type,
                    'incentivized_hours': inc_hrs,
                    'base_hourly_rate': base_rate,
                }

                if pk_str:
                    try:
                        ot_obj = OvertimeShift.objects.get(pk=pk_str, agent=agent)
                        for k, v in defaults.items():
                            setattr(ot_obj, k, v)
                        ot_obj.save()
                        submitted_pks.add(ot_obj.pk)
                        log_action(request.user, 'Updated OT shift',
                                   f'{agent} on {day_date.isoformat()}: {start}–{end}', agent=agent)
                    except OvertimeShift.DoesNotExist:
                        pass
                    except Exception as _e:
                        messages.error(request, f"Error saving OT shift for {day_date.strftime('%A, %b %d')} ({start}–{end}): {_e}")
                else:
                    try:
                        ot_obj = OvertimeShift.objects.create(agent=agent, date=day_date, **defaults)
                        submitted_pks.add(ot_obj.pk)
                        log_action(request.user, 'Added OT shift',
                                   f'{agent} on {day_date.isoformat()}: {start}–{end}', agent=agent)
                    except Exception as _e:
                        messages.error(request, f"Error saving OT shift for {day_date.strftime('%A, %b %d')} ({start}–{end}): {_e}")

        messages.success(request, f"OT shifts saved for {agent} — week of {week_start.strftime('%B %d, %Y')}.")
        return redirect(f"{reverse('overtime_list')}?week_start={week_start.isoformat()}")

    # GET: pre-fill from existing OT records for selected agent
    days = []
    agent_hourly_rate = ''
    agent_schedule = {}  # day_index -> list of {start, end} for overlap check

    if selected_agent_id:
        try:
            selected_agent = Agent.objects.get(pk=selected_agent_id)
            if selected_agent.hourly_rate:
                agent_hourly_rate = str(selected_agent.hourly_rate)
        except Agent.DoesNotExist:
            selected_agent = None

        # Build scheduled blocks for overlap check
        for tmpl in ShiftTemplate.objects.filter(agent_id=selected_agent_id):
            if not tmpl.is_off and tmpl.start_time and tmpl.end_time:
                blocks = [{'start': tmpl.start_time.strftime('%H:%M'), 'end': tmpl.end_time.strftime('%H:%M')}]
                for eb in tmpl.extra_blocks.all():
                    blocks.append({'start': eb.start_time.strftime('%H:%M'), 'end': eb.end_time.strftime('%H:%M')})
                agent_schedule[tmpl.day_of_week] = blocks

        # All OT shifts for this agent this week, grouped by date
        ot_by_date = {}
        for s in OvertimeShift.objects.filter(agent_id=selected_agent_id, date__in=week_dates).order_by('start_time'):
            ot_by_date.setdefault(s.date, []).append(s)

        for i, day_date in enumerate(week_dates):
            day_ot_shifts = ot_by_date.get(day_date, [])
            ot_entries = []
            for j, s in enumerate(day_ot_shifts):
                ot_entries.append({
                    'j': j,
                    'pk': s.pk,
                    'start': s.start_time.strftime('%H:%M'),
                    'end': s.end_time.strftime('%H:%M'),
                    'notes': s.notes,
                    'incentive_type': s.incentive_type,
                    'incentivized_hours': str(s.incentivized_hours) if s.incentivized_hours is not None else '',
                    'base_hourly_rate': str(s.base_hourly_rate) if s.base_hourly_rate is not None else agent_hourly_rate,
                })
            days.append({
                'index': i,
                'name': DAYS[i],
                'date': day_date,
                'ot_entries': ot_entries,
                'has_ot': bool(day_ot_shifts),
                'ot_count': len(day_ot_shifts),
            })

    import json
    return render(request, 'scheduling/overtime_week.html', {
        'agents': agents,
        'selected_agent_id': int(selected_agent_id) if selected_agent_id else None,
        'days': days,
        'week_start': week_start,
        'week_end': week_dates[-1],
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
        'week_start_iso': week_start.isoformat(),
        'agent_hourly_rate': agent_hourly_rate,
        'agent_schedule_json': json.dumps(agent_schedule),
    })


@login_required
def overtime_delete(request, pk):
    from django.http import JsonResponse
    ot_shift = get_object_or_404(OvertimeShift, pk=pk)
    if request.method == 'POST':
        week_start = ot_shift.date - timedelta(days=ot_shift.date.weekday())
        agent = ot_shift.agent
        log_action(request.user, 'Deleted OT shift',
                   f'{agent} on {ot_shift.date.isoformat()}: '
                   f'{ot_shift.start_time.strftime("%H:%M")}–{ot_shift.end_time.strftime("%H:%M")}',
                   agent=agent)
        ot_shift.delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'deleted': True, 'pk': pk})
        messages.success(request, "OT shift deleted.")
        return redirect(f"{reverse('overtime_list')}?week_start={week_start.isoformat()}")
    return render(request, 'scheduling/confirm_delete.html', {
        'object': ot_shift,
        'cancel_url': reverse('overtime_list'),
    })


@login_required
def shift_quick_edit(request):
    """AJAX: create/update a one-time Shift override (or permanent ShiftTemplate) for a specific agent+date."""
    from django.http import JsonResponse
    import json as _json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = _json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    agent_pk = data.get('agent_pk')
    date_str = data.get('date', '')
    start = (data.get('start') or '').strip()
    end = (data.get('end') or '').strip()
    is_off = bool(data.get('is_off'))
    permanent = bool(data.get('permanent'))
    extra_blocks = data.get('extra_blocks') or []

    agent = get_object_or_404(Agent, pk=agent_pk)
    try:
        day_date = date.fromisoformat(date_str)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date'}, status=400)

    if not is_off and not (start and end):
        return JsonResponse({'error': 'Start and end time required'}, status=400)

    if permanent:
        day_of_week = day_date.weekday()
        eff_date = timezone.localdate()  # Permanent changes take effect from today
        tmpl, _ = _save_shift_template(agent, day_of_week, eff_date, is_off, start, end, '')
        ShiftTemplateBlock.objects.filter(shift_template=tmpl).delete()
        for n, block in enumerate(extra_blocks[:2], start=2):
            bs = (block.get('start') or '').strip()
            be = (block.get('end') or '').strip()
            if bs and be:
                ShiftTemplateBlock.objects.create(
                    shift_template=tmpl, block_number=n, start_time=bs, end_time=be
                )
        log_action(request.user, 'Quick-edit shift (permanent)',
                   f'{agent} {day_date.strftime("%A")}: {"OFF" if is_off else f"{start}–{end}"}',
                   agent=agent)
    else:
        defaults = {
            'is_off': is_off,
            'start_time': start if start else '00:00',
            'end_time': end if end else '00:00',
        }
        shift_obj, _ = Shift.objects.update_or_create(agent=agent, date=day_date, defaults=defaults)
        ShiftBlock.objects.filter(shift=shift_obj).delete()
        for n, block in enumerate(extra_blocks[:2], start=2):
            bs = (block.get('start') or '').strip()
            be = (block.get('end') or '').strip()
            if bs and be:
                ShiftBlock.objects.create(
                    shift=shift_obj, block_number=n, start_time=bs, end_time=be
                )
        log_action(request.user, 'Quick-edit shift (one-time)',
                   f'{agent} on {day_date.isoformat()}: {"OFF" if is_off else f"{start}–{end}"}',
                   agent=agent)

    return JsonResponse({'ok': True, 'scheduled': '' if is_off else f'{start}–{end}'})


@login_required
def overtime_export(request):
    import csv
    from django.http import HttpResponse

    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    supervisor_id = request.GET.get('supervisor', '')

    try:
        date_from = date.fromisoformat(date_from_str)
    except (ValueError, TypeError):
        date_from = timezone.localdate() - timedelta(days=30)
    try:
        date_to = date.fromisoformat(date_to_str)
    except (ValueError, TypeError):
        date_to = timezone.localdate()

    ot_qs = OvertimeShift.objects.filter(
        date__range=[date_from, date_to]
    ).select_related('agent__user', 'agent__supervisor__user', 'verification').order_by('date', 'agent__user__last_name')

    if supervisor_id:
        try:
            ot_qs = ot_qs.filter(agent__supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ot_payroll_{date_from}_{date_to}.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Agent Name', 'Employee ID', 'Employer', 'Supervisor',
        'Week Start', 'Date', 'Day of Week',
        'Start Time', 'End Time', 'Total Hours',
        'Incentive Type', 'Incentivized Hours', 'Base Hourly Rate ($)',
        'Base Pay ($)', 'Incentive Bonus ($)', 'Total Pay Offered ($)',
        'Status', 'Cancellation Reason', 'Coverage %',
    ])
    for ot in ot_qs:
        from decimal import Decimal
        is_cancelled = ot.status == 'cancelled'
        total_hrs = Decimal('0') if is_cancelled else ot.total_shift_hours()
        base_rate = ot.base_hourly_rate or Decimal('0')
        inc_hrs = ot.incentivized_hours or Decimal('0')
        if is_cancelled:
            base_pay = incentive_bonus = total_offered = '0.00'
        else:
            if ot.incentive_type == 'time_and_a_half':
                premium = Decimal('0.5')
            elif ot.incentive_type == 'power_hour':
                premium = Decimal('1.0')
            else:
                premium = Decimal('0')
            base_pay = (total_hrs * base_rate).quantize(Decimal('0.01')) if base_rate else ''
            incentive_bonus = (inc_hrs * base_rate * premium).quantize(Decimal('0.01')) if base_rate else ''
            total_offered = ot.incentive_offered() or ''
        week_start = ot.date - timedelta(days=ot.date.weekday())
        writer.writerow([
            str(ot.agent),
            ot.agent.employee_id or '',
            ot.agent.employer,
            str(ot.agent.supervisor) if ot.agent.supervisor else '',
            week_start.strftime('%Y-%m-%d'),
            ot.date.strftime('%Y-%m-%d'),
            ot.date.strftime('%A'),
            ot.start_time.strftime('%H:%M'),
            ot.end_time.strftime('%H:%M'),
            str(total_hrs),
            ot.get_incentive_type_display(),
            str(inc_hrs) if ot.incentivized_hours is not None else '',
            str(ot.base_hourly_rate) if ot.base_hourly_rate is not None else '',
            str(base_pay),
            str(incentive_bonus),
            str(total_offered),
            ot.get_status_display(),
            ot.cancellation_reason,
            str(ot.verification.coverage_pct) + '%' if hasattr(ot, 'verification') and ot.verification and ot.verification.coverage_pct is not None else '',
        ])
    return response


@login_required
def overtime_set_status(request, pk):
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    import json as _json
    try:
        body = _json.loads(request.body)
        new_status = body.get('status', '')
        cancellation_reason = body.get('cancellation_reason', '').strip()
    except (ValueError, KeyError):
        new_status = ''
        cancellation_reason = ''
    valid = {'pending', 'completed', 'no_show', 'cancelled'}
    if new_status not in valid:
        return JsonResponse({'error': 'Invalid status'}, status=400)
    if new_status == 'cancelled' and not cancellation_reason:
        return JsonResponse({'error': 'Cancellation reason is required.'}, status=400)
    ot_shift = get_object_or_404(OvertimeShift, pk=pk)
    ot_shift.status = new_status
    update_fields = ['status']
    if new_status == 'cancelled':
        ot_shift.cancellation_reason = cancellation_reason
        update_fields.append('cancellation_reason')
    elif ot_shift.cancellation_reason:
        ot_shift.cancellation_reason = ''
        update_fields.append('cancellation_reason')
    ot_shift.save(update_fields=update_fields)
    log_action(request.user, 'Updated OT status',
               f'{ot_shift.agent} on {ot_shift.date}: status={new_status}'
               + (f' (reason: {cancellation_reason})' if cancellation_reason else ''),
               agent=ot_shift.agent)
    return JsonResponse({'status': new_status, 'pk': pk, 'cancellation_reason': ot_shift.cancellation_reason})


# ── Open OT shift postings, claim requests, cancellation requests ─────────────

def _viewer_agent(request):
    try:
        return request.user.agent
    except Exception:
        return None


def _is_ot_approver(agent, user=None):
    """Supervisors, coordinators, super admins, and Django superusers can post
    open shifts and action claim/cancellation requests."""
    if agent is not None:
        if agent.role == 'admin' and agent.role_type in ('supervisor', 'coordinator'):
            return True
        if agent.is_super_admin:
            return True
    return user is not None and user.is_superuser


def _redirect_after_ot_action(request, target_date=None):
    """Send portal users back to their portal, staff back to the OT week grid.
    A local `next` path in the POST (e.g. from the Staffing Calculator) wins."""
    nxt = request.POST.get('next', '')
    if nxt.startswith('/') and not nxt.startswith('//'):
        return redirect(nxt)
    viewer = _viewer_agent(request)
    if viewer is not None and _is_portal_user(viewer):
        return redirect('agent_available_ot')
    url = reverse('overtime_list')
    if target_date:
        week_start = target_date - timedelta(days=target_date.weekday())
        url += f'?week_start={week_start.isoformat()}'
    return redirect(url)


def _parse_open_shift_fields(post):
    """Returns (fields_dict, error_message)."""
    from django.utils.dateparse import parse_time
    try:
        shift_date = date.fromisoformat(post.get('date', ''))
    except (ValueError, TypeError):
        return None, "Please enter a valid date."
    start = parse_time(post.get('start_time') or '')
    end = parse_time(post.get('end_time') or '')
    if not start or not end:
        return None, "Please enter valid start and end times."
    incentive = post.get('incentive_type', 'none')
    if incentive not in dict(OvertimeShift.INCENTIVE_CHOICES):
        incentive = 'none'
    return {
        'date': shift_date,
        'start_time': start,
        'end_time': end,
        'incentive_type': incentive,
        'notes': post.get('notes', '').strip(),
    }, None


@login_required
def open_ot_create(request):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can post open shifts.")
        return redirect('overtime_list')

    fields, err = _parse_open_shift_fields(request.POST)
    if err:
        messages.error(request, err)
        return redirect('overtime_list')

    try:
        count = max(1, min(20, int(request.POST.get('count', '1'))))
    except (ValueError, TypeError):
        count = 1

    for _ in range(count):
        OpenOTShift.objects.create(posted_by=request.user, **fields)

    label = f"{fields['date']} {fields['start_time']:%H:%M}–{fields['end_time']:%H:%M}"
    log_action(request.user, 'Posted open OT shift',
               f"{label} ×{count}" + (f" — {fields['notes']}" if fields['notes'] else ''))
    messages.success(request, f"Posted {count} open shift{'s' if count > 1 else ''} for {label}.")
    return _redirect_after_ot_action(request, fields['date'])


@login_required
def open_ot_update(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can edit open shifts.")
        return redirect('overtime_list')

    posting = get_object_or_404(OpenOTShift, pk=pk)
    if posting.status != 'open':
        messages.error(request, "This shift has already been claimed and can no longer be edited.")
        return _redirect_after_ot_action(request, posting.date)

    fields, err = _parse_open_shift_fields(request.POST)
    if err:
        messages.error(request, err)
        return _redirect_after_ot_action(request, posting.date)

    for k, v in fields.items():
        setattr(posting, k, v)
    posting.save()
    log_action(request.user, 'Updated open OT shift',
               f"{posting.date} {posting.time_label()}")
    messages.success(request, "Open shift updated.")
    return _redirect_after_ot_action(request, posting.date)


@login_required
def open_ot_delete(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can delete open shifts.")
        return redirect('overtime_list')

    posting = get_object_or_404(OpenOTShift, pk=pk)
    if posting.status != 'open':
        messages.error(request, "This shift has already been claimed and can no longer be deleted.")
        return _redirect_after_ot_action(request, posting.date)

    # Notify anyone still waiting on this posting
    posting.claim_requests.filter(status='pending').update(
        status='rejected', rejection_reason='The open shift posting was removed.',
        reviewed_by=request.user, reviewed_at=timezone.now(), requester_read=False,
    )
    # Soft-delete so claim-request history and notifications survive
    posting.status = 'removed'
    posting.save(update_fields=['status'])
    log_action(request.user, 'Deleted open OT shift', f"{posting.date} {posting.time_label()}")
    messages.success(request, "Open shift deleted.")
    return _redirect_after_ot_action(request, posting.date)


@login_required
def open_ot_claim(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    if viewer is None or viewer.status != 'active':
        messages.error(request, "You need an active profile to request shifts.")
        return _redirect_after_ot_action(request)

    posting = get_object_or_404(OpenOTShift, pk=pk)
    if posting.status != 'open':
        messages.error(request, "This shift has already been filled.")
        return _redirect_after_ot_action(request, posting.date)
    if posting.claim_requests.filter(requester=viewer, status='pending').exists():
        messages.error(request, "You already have a pending request for this shift.")
        return _redirect_after_ot_action(request, posting.date)

    OTShiftClaimRequest.objects.create(
        open_shift=posting, requester=viewer,
        supervisor_read=False, requester_read=True,
    )
    log_action(request.user, 'Requested open OT shift',
               f"{posting.date} {posting.time_label()}", agent=viewer)
    messages.success(request, "Your request has been submitted and is pending approval.")
    return _redirect_after_ot_action(request, posting.date)


@login_required
def ot_claim_approve(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    claim = get_object_or_404(
        OTShiftClaimRequest.objects.select_related('open_shift', 'requester'), pk=pk
    )
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can approve shift requests.")
        return redirect('overtime_list')
    if viewer is not None and claim.requester_id == viewer.pk:
        messages.error(request, "You cannot approve your own shift request. "
                                "Another supervisor or coordinator must approve it.")
        return _redirect_after_ot_action(request, claim.open_shift.date)
    if claim.status != 'pending':
        messages.error(request, "This request has already been reviewed.")
        return _redirect_after_ot_action(request, claim.open_shift.date)
    posting = claim.open_shift
    if posting.status != 'open':
        messages.error(request, "This shift has already been filled.")
        return _redirect_after_ot_action(request, posting.date)

    from django.db import transaction
    now = timezone.now()
    with transaction.atomic():
        shift = OvertimeShift.objects.create(
            agent=claim.requester,
            date=posting.date,
            start_time=posting.start_time,
            end_time=posting.end_time,
            incentive_type=posting.incentive_type,
            notes=posting.notes,
            base_hourly_rate=claim.requester.hourly_rate,
        )
        if posting.incentive_type != 'none':
            shift.incentivized_hours = shift.total_shift_hours()
            shift.save(update_fields=['incentivized_hours'])

        posting.status = 'filled'
        posting.filled_by = claim.requester
        posting.filled_at = now
        posting.assigned_shift = shift
        posting.save(update_fields=['status', 'filled_by', 'filled_at', 'assigned_shift'])

        claim.status = 'approved'
        claim.reviewed_by = request.user
        claim.reviewed_at = now
        claim.requester_read = False
        claim.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'requester_read'])

        # Backup requests lose out once the shift is assigned
        posting.claim_requests.filter(status='pending').exclude(pk=claim.pk).update(
            status='rejected', rejection_reason='Shift was assigned to another requester.',
            reviewed_by=request.user, reviewed_at=now, requester_read=False,
        )

    log_action(request.user, 'Approved OT shift claim',
               f"{posting.date} {posting.time_label()} → {claim.requester}", agent=claim.requester)
    messages.success(request, f"Shift assigned to {claim.requester}.")
    return _redirect_after_ot_action(request, posting.date)


@login_required
def ot_claim_reject(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    claim = get_object_or_404(
        OTShiftClaimRequest.objects.select_related('open_shift', 'requester'), pk=pk
    )
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can reject shift requests.")
        return redirect('overtime_list')
    if claim.status != 'pending':
        messages.error(request, "This request has already been reviewed.")
        return _redirect_after_ot_action(request, claim.open_shift.date)

    claim.status = 'rejected'
    claim.reviewed_by = request.user
    claim.reviewed_at = timezone.now()
    claim.rejection_reason = request.POST.get('rejection_reason', '').strip()
    claim.requester_read = False
    claim.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'rejection_reason', 'requester_read'])

    log_action(request.user, 'Rejected OT shift claim',
               f"{claim.open_shift.date} {claim.open_shift.time_label()} — {claim.requester}",
               agent=claim.requester)
    messages.success(request, "Shift request rejected. The shift remains open.")
    return _redirect_after_ot_action(request, claim.open_shift.date)


@login_required
def ot_cancel_request(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    shift = get_object_or_404(OvertimeShift, pk=pk)
    if viewer is None or shift.agent_id != viewer.pk:
        messages.error(request, "You can only request cancellation of your own shifts.")
        return _redirect_after_ot_action(request, shift.date)
    if shift.status != 'pending':
        messages.error(request, "Only upcoming pending shifts can be cancelled.")
        return _redirect_after_ot_action(request, shift.date)
    if shift.date < timezone.localdate():
        messages.error(request, "This shift is in the past and can no longer be cancelled.")
        return _redirect_after_ot_action(request, shift.date)
    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, "Please provide a reason for the cancellation request.")
        return _redirect_after_ot_action(request, shift.date)
    if shift.cancellation_requests.filter(status='pending').exists():
        messages.error(request, "A cancellation request for this shift is already pending.")
        return _redirect_after_ot_action(request, shift.date)

    OTCancellationRequest.objects.create(
        shift=shift, requester=viewer, reason=reason,
        supervisor_read=False, requester_read=True,
    )
    log_action(request.user, 'Requested OT shift cancellation',
               f"{shift.date} {shift.start_time:%H:%M}–{shift.end_time:%H:%M} — {reason}",
               agent=viewer)
    messages.success(request, "Cancellation request submitted. The shift stays assigned to you "
                              "unless it is approved.")
    if viewer is not None and _is_portal_user(viewer):
        return redirect('agent_my_ot_shifts')
    return _redirect_after_ot_action(request, shift.date)


@login_required
def ot_cancel_approve(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    cr = get_object_or_404(
        OTCancellationRequest.objects.select_related('shift', 'requester'), pk=pk
    )
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can approve cancellation requests.")
        return redirect('overtime_list')
    if viewer is not None and cr.requester_id == viewer.pk:
        messages.error(request, "You cannot approve your own cancellation request. "
                                "Another supervisor or coordinator must approve it.")
        return _redirect_after_ot_action(request, cr.shift.date)
    if cr.status != 'pending':
        messages.error(request, "This request has already been reviewed.")
        return _redirect_after_ot_action(request, cr.shift.date)

    shift = cr.shift
    shift.status = 'cancelled'
    shift.cancellation_reason = cr.reason
    shift.save(update_fields=['status', 'cancellation_reason'])

    cr.status = 'approved'
    cr.reviewed_by = request.user
    cr.reviewed_at = timezone.now()
    cr.requester_read = False
    cr.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'requester_read'])

    log_action(request.user, 'Approved OT cancellation request',
               f"{shift.agent} on {shift.date}: cancelled (reason: {cr.reason})",
               agent=shift.agent)
    messages.success(request, "Cancellation approved — the shift is now cancelled.")
    return _redirect_after_ot_action(request, shift.date)


@login_required
def ot_cancel_reject(request, pk):
    if request.method != 'POST':
        return redirect('overtime_list')
    viewer = _viewer_agent(request)
    cr = get_object_or_404(
        OTCancellationRequest.objects.select_related('shift', 'requester'), pk=pk
    )
    if not _is_ot_approver(viewer, request.user):
        messages.error(request, "Only supervisors and coordinators can reject cancellation requests.")
        return redirect('overtime_list')
    if cr.status != 'pending':
        messages.error(request, "This request has already been reviewed.")
        return _redirect_after_ot_action(request, cr.shift.date)

    cr.status = 'rejected'
    cr.reviewed_by = request.user
    cr.reviewed_at = timezone.now()
    cr.review_note = request.POST.get('review_note', '').strip()
    cr.requester_read = False
    cr.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'review_note', 'requester_read'])

    log_action(request.user, 'Rejected OT cancellation request',
               f"{cr.shift.agent} on {cr.shift.date}"
               + (f" (note: {cr.review_note})" if cr.review_note else ''),
               agent=cr.shift.agent)
    messages.success(request, "Cancellation request rejected. The shift remains assigned.")
    return _redirect_after_ot_action(request, cr.shift.date)


@login_required
def agent_available_ot(request):
    """Agent portal: browse open OT shifts and request to claim them."""
    viewer = _viewer_agent(request)
    if viewer is None:
        return redirect('dashboard')
    if not _is_portal_user(viewer):
        return redirect('overtime_list')

    today = timezone.localdate()
    # Mark responses to my claim requests as seen
    OTShiftClaimRequest.objects.filter(requester=viewer, requester_read=False).update(requester_read=True)

    postings = list(
        OpenOTShift.objects.filter(status='open', date__gte=today)
        .order_by('date', 'start_time')
        .prefetch_related('claim_requests__requester')
    )
    for p in postings:
        pending = [c for c in p.claim_requests.all() if c.status == 'pending']
        p.pending_claims = pending
        p.my_pending = any(c.requester_id == viewer.pk for c in pending)

    my_requests = (
        OTShiftClaimRequest.objects.filter(requester=viewer)
        .select_related('open_shift', 'reviewed_by')
        .order_by('-submitted_at')[:30]
    )

    return render(request, 'agent/available_ot.html', {
        'agent': viewer,
        'postings': postings,
        'my_requests': my_requests,
        'today': today,
    })


@login_required
def live_poll(request):
    """Generic poll endpoint — returns the latest change timestamp for a given page type."""
    from django.db.models import Max
    from .models import AuditLog

    poll_type = request.GET.get('type', '')
    week_start_str = request.GET.get('week_start', '')

    try:
        ws = date.fromisoformat(week_start_str)
        ws -= timedelta(days=ws.weekday())
        week_end = ws + timedelta(days=6)
    except (ValueError, TypeError):
        ws = week_end = None

    latest = None

    if poll_type == 'codings' and ws:
        from adherence.models import Coding
        r = Coding.objects.filter(date__range=[ws, week_end]).aggregate(latest=Max('created_at'))
        latest = r['latest']
    elif poll_type == 'shifts':
        r = AuditLog.objects.filter(action__icontains='schedule').aggregate(latest=Max('timestamp'))
        latest = r['latest']
    elif poll_type == 'users':
        r = AuditLog.objects.filter(action__icontains='agent profile').aggregate(latest=Max('timestamp'))
        latest = r['latest']
    elif poll_type == 'daily' and ws:
        from adherence.models import DailyUpload
        dates = [ws + timedelta(days=i) for i in range(7)]
        r = DailyUpload.objects.filter(date__in=dates).aggregate(latest=Max('uploaded_at'))
        latest = r['latest']
    elif poll_type == 'overtime':
        r = AuditLog.objects.filter(action__icontains='OT shift').aggregate(latest=Max('timestamp'))
        latest = r['latest']

    from django.http import JsonResponse
    return JsonResponse({'latest': latest.isoformat() if latest else None})


@login_required
def activity_log(request):
    from .models import AuditLog
    from django.contrib.auth.models import User

    logs = AuditLog.objects.select_related('user', 'agent__user').order_by('-timestamp')

    # Filter by user
    user_filter = request.GET.get('user', '')
    if user_filter:
        try:
            logs = logs.filter(user_id=int(user_filter))
        except (ValueError, TypeError):
            pass

    # Filter by date
    date_filter = request.GET.get('date', '')
    if date_filter:
        try:
            from datetime import datetime
            filter_date = date.fromisoformat(date_filter)
            logs = logs.filter(timestamp__date=filter_date)
        except (ValueError, TypeError):
            pass

    logs = logs[:500]

    users = User.objects.filter(audit_logs__isnull=False).distinct().order_by('last_name', 'first_name')

    return render(request, 'scheduling/activity_log.html', {
        'logs': logs,
        'users': users,
        'selected_user': user_filter,
        'selected_date': date_filter,
    })


@login_required
def agent_history(request, pk):
    from datetime import date as date_cls, timedelta
    from decimal import Decimal
    from adherence.models import AdherenceRecord, Coding

    agent = get_object_or_404(Agent, pk=pk)
    today = date_cls.today()
    five_years_ago = today.replace(year=today.year - 5)

    date_from_str = request.GET.get('from', '')
    date_to_str = request.GET.get('to', '')
    try:
        date_from = date_cls.fromisoformat(date_from_str)
    except (ValueError, TypeError):
        date_from = today - timedelta(days=30)
    try:
        date_to = date_cls.fromisoformat(date_to_str)
    except (ValueError, TypeError):
        date_to = today
    date_from = max(date_from, five_years_ago)
    date_to = min(date_to, today)

    # Role history
    role_history = list(agent.role_history.select_related('supervisor__user', 'changed_by').all())

    # Seed initial entry if none exists (for legacy agents)
    if not role_history:
        RoleHistory.objects.create(
            agent=agent, role=agent.role, role_type=agent.role_type or '',
            supervisor=agent.supervisor, employer=agent.employer,
            billing_status=agent.billing_status,
            effective_from=agent.start_date or today,
            changed_by=None,
        )
        role_history = list(agent.role_history.select_related('supervisor__user', 'changed_by').all())

    # Schedule history: ShiftTemplates grouped by effective_from
    from collections import defaultdict, OrderedDict
    templates = list(ShiftTemplate.objects.filter(agent=agent).order_by('-effective_from', 'day_of_week'))
    sched_groups = OrderedDict()
    DAYS_ABBR = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for t in templates:
        key = t.effective_from or date_cls(2000, 1, 1)
        if key not in sched_groups:
            sched_groups[key] = []
        sched_groups[key].append(t)
    schedule_history = list(sched_groups.items())

    # Attendance & hours by week
    records = list(AdherenceRecord.objects.filter(
        agent=agent, date__gte=date_from, date__lte=date_to
    ).order_by('date'))
    codings_qs = list(Coding.objects.filter(
        agent=agent, date__gte=date_from, date__lte=date_to
    ).order_by('-date', 'start_time'))

    record_map = {r.date: r for r in records}
    coding_map = defaultdict(Decimal)
    for c in codings_qs:
        coding_map[c.date] += Decimal(str(c.total_hours()))

    # Build week rows (most recent first)
    ws = date_from - timedelta(days=date_from.weekday())
    attendance_weeks = []
    while ws <= date_to:
        week_dates = [ws + timedelta(days=i) for i in range(7)]
        days = []
        login_total = Decimal('0')
        coded_total = Decimal('0')
        has_data = False
        bonus = True
        bonus_det = False
        BONUS_Q = {'P', 'OT', 'MUT', 'VTO', 'P+VTO'}
        BONUS_DQ = {'Absent', 'NCNS', 'T', 'T+VTO', 'T+I', 'I', 'LOA', 'S'}
        for d in week_dates:
            r = record_map.get(d)
            c_hrs = coding_map.get(d, Decimal('0'))
            status = r.status if r else ''
            hrs = r.actual_hours if r else None
            if r or c_hrs:
                has_data = True
            if hrs:
                login_total += hrs
            coded_total += c_hrs
            if status in BONUS_Q:
                bonus_det = True
            elif status in BONUS_DQ:
                bonus = False
                bonus_det = True
            elif status:
                bonus = False
                bonus_det = True
            days.append({'date': d, 'status': status, 'hours': hrs, 'coded': c_hrs})
        if has_data:
            total = login_total + coded_total
            attendance_weeks.append({
                'week_start': ws,
                'days': days,
                'login_hrs': login_total,
                'coded_hrs': coded_total,
                'total_hrs': total,
                'bonus': 'Yes' if (bonus and bonus_det) else ('No' if not bonus else '—'),
            })
        ws += timedelta(days=7)
    attendance_weeks.reverse()

    # OT history
    ot_shifts = list(OvertimeShift.objects.filter(agent=agent).order_by('-date')[:500])

    # Export CSV
    if request.GET.get('export') == 'attendance':
        import csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="attendance_{agent.pk}_{date_from}_{date_to}.csv"'
        w = csv.writer(resp)
        w.writerow(['Week Start', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun', 'Login Hrs', 'Coded Hrs', 'Total Hrs', 'Bonus'])
        for wk in attendance_weeks:
            w.writerow([
                wk['week_start'],
                *[d['status'] for d in wk['days']],
                str(wk['login_hrs']), str(wk['coded_hrs']), str(wk['total_hrs']),
                wk['bonus'],
            ])
        return resp

    if request.GET.get('export') == 'codings':
        import csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="codings_{agent.pk}_{date_from}_{date_to}.csv"'
        w = csv.writer(resp)
        w.writerow(['Date', 'Start', 'End', 'Duration', 'Notes'])
        for c in codings_qs:
            w.writerow([c.date, c.start_time.strftime('%H:%M'), c.end_time.strftime('%H:%M'), c.total_hhmmss(), c.notes])
        return resp

    if request.GET.get('export') == 'ot':
        import csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="ot_{agent.pk}.csv"'
        w = csv.writer(resp)
        w.writerow(['Date', 'Start', 'End', 'Hours', 'Incentive', 'Completed'])
        for s in ot_shifts:
            w.writerow([s.date, s.start_time.strftime('%H:%M'), s.end_time.strftime('%H:%M'), str(s.total_shift_hours()), s.get_incentive_type_display(), s.get_status_display()])
        return resp

    return render(request, 'scheduling/agent_history.html', {
        'agent': agent,
        'role_history': role_history,
        'schedule_history': schedule_history,
        'attendance_weeks': attendance_weeks,
        'codings': codings_qs,
        'ot_shifts': ot_shifts,
        'date_from': date_from,
        'date_to': date_to,
        'five_years_ago': five_years_ago,
        'today': today,
        'days_abbr': DAYS_ABBR,
    })


# ── Floor-wide Records ────────────────────────────────────────────────────────

@login_required
def records_attendance(request):
    from datetime import date as date_cls, timedelta
    from adherence.models import AdherenceRecord

    today = date_cls.today()
    five_years_ago = today.replace(year=today.year - 5)

    date_from = _parse_date(request.GET.get('from'), today - timedelta(days=30))
    date_to   = _parse_date(request.GET.get('to'), today)
    date_from = max(date_from, five_years_ago)
    date_to   = min(date_to, today)

    if 'supervisor' in request.GET:
        supervisor_id = request.GET.get('supervisor', '')
        request.session['supervisor_filter'] = supervisor_id
    else:
        supervisor_id = request.session.get('supervisor_filter', '')
    role_type_f   = request.GET.get('role_type', '')
    employer_f    = request.GET.get('employer', '')
    status_f      = request.GET.get('status', '')

    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    qs = AdherenceRecord.objects.filter(
        date__gte=date_from, date__lte=date_to,
        agent__track_attendance=True,
    ).select_related('agent__user', 'agent__supervisor__user').order_by('-date', 'agent__user__last_name')

    if supervisor_id:
        try:
            qs = qs.filter(agent__supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass
    if role_type_f:
        qs = qs.filter(agent__role_type=role_type_f)
    if employer_f:
        qs = qs.filter(agent__employer=employer_f)
    if status_f:
        qs = qs.filter(status=status_f)

    records = list(qs[:2000])  # cap for performance

    if request.GET.get('export') == '1':
        import csv as _csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="attendance_records_{date_from}_{date_to}.csv"'
        w = _csv.writer(resp)
        w.writerow(['Date', 'Day', 'Agent', 'Supervisor', 'Role Type', 'Employer', 'Status', 'Hours'])
        for r in records:
            w.writerow([
                r.date, r.date.strftime('%A'),
                str(r.agent), str(r.agent.supervisor or ''),
                r.agent.get_role_type_display(), r.agent.employer,
                r.status, str(r.actual_hours or ''),
            ])
        return resp

    from adherence.models import AdherenceRecord as AR
    status_choices = AR.STATUS_CHOICES
    role_type_choices = Agent.ROLE_TYPE_CHOICES
    employer_choices = Agent.EMPLOYER_CHOICES

    return render(request, 'records/attendance.html', {
        'records': records,
        'date_from': date_from,
        'date_to': date_to,
        'five_years_ago': five_years_ago,
        'today': today,
        'supervisors': supervisors,
        'selected_supervisor': supervisor_id,
        'selected_role_type': role_type_f,
        'selected_employer': employer_f,
        'selected_status': status_f,
        'status_choices': status_choices,
        'role_type_choices': role_type_choices,
        'employer_choices': employer_choices,
        'count': len(records),
    })


@login_required
def records_hours(request):
    from datetime import date as date_cls, timedelta
    from decimal import Decimal
    from collections import defaultdict
    from adherence.models import AdherenceRecord, Coding, DailyAgentHours

    today = date_cls.today()
    five_years_ago = today.replace(year=today.year - 5)

    date_from = _parse_date(request.GET.get('from'), today - timedelta(days=30))
    date_to   = _parse_date(request.GET.get('to'), today)
    date_from = max(date_from, five_years_ago)
    date_to   = min(date_to, today)

    if 'supervisor' in request.GET:
        supervisor_id = request.GET.get('supervisor', '')
        request.session['supervisor_filter'] = supervisor_id
    else:
        supervisor_id = request.session.get('supervisor_filter', '')
    role_type_f   = request.GET.get('role_type', '')
    employer_f    = request.GET.get('employer', '')
    billing_f     = request.GET.get('billing', '')

    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    agents_qs = Agent.objects.filter(track_attendance=True).select_related('user', 'supervisor__user')
    if supervisor_id:
        try:
            agents_qs = agents_qs.filter(supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass
    if role_type_f:
        agents_qs = agents_qs.filter(role_type=role_type_f)
    if employer_f:
        agents_qs = agents_qs.filter(employer=employer_f)
    if billing_f:
        agents_qs = agents_qs.filter(billing_status=billing_f)

    agents = list(agents_qs.order_by('user__last_name', 'user__first_name'))
    agent_ids = [a.pk for a in agents]

    # Get all dates in range
    all_dates = []
    d = date_from
    while d <= date_to:
        all_dates.append(d)
        d += timedelta(days=1)

    adh_qs = AdherenceRecord.objects.filter(agent_id__in=agent_ids, date__gte=date_from, date__lte=date_to)
    cod_qs = Coding.objects.filter(agent_id__in=agent_ids, date__gte=date_from, date__lte=date_to)

    # Map (agent_id, date) -> hours
    adh_map = {}
    for r in adh_qs:
        adh_map[(r.agent_id, r.date)] = r.actual_hours or Decimal('0')

    cod_map = defaultdict(Decimal)
    for c in cod_qs:
        cod_map[(c.agent_id, c.date)] += Decimal(str(c.total_hours()))

    # Group by agent + week
    def week_start(d):
        return d - timedelta(days=d.weekday())

    rows = []
    for agent in agents:
        # Get all weeks
        weeks_seen = set()
        for d in all_dates:
            ws = week_start(d)
            if (agent.pk, ws) not in weeks_seen:
                weeks_seen.add((agent.pk, ws))
                week_dates = [ws + timedelta(days=i) for i in range(7)]
                login_hrs = sum((adh_map.get((agent.pk, wd), Decimal('0')) for wd in week_dates), Decimal('0'))
                coded_hrs = sum((cod_map.get((agent.pk, wd), Decimal('0')) for wd in week_dates), Decimal('0'))
                total_hrs = login_hrs + coded_hrs
                if login_hrs > 0 or coded_hrs > 0:
                    rows.append({
                        'agent': agent,
                        'week_start': ws,
                        'login_hrs': login_hrs,
                        'coded_hrs': coded_hrs,
                        'total_hrs': total_hrs,
                    })

    rows.sort(key=lambda r: (-r['week_start'].toordinal(), r['agent'].user.last_name))

    if request.GET.get('export') == '1':
        import csv as _csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="hours_records_{date_from}_{date_to}.csv"'
        w = _csv.writer(resp)
        w.writerow(['Week Start', 'Agent', 'Supervisor', 'Role Type', 'Employer', 'Login Hrs', 'Coded Hrs', 'Total Hrs'])
        for r in rows:
            w.writerow([
                r['week_start'], str(r['agent']),
                str(r['agent'].supervisor or ''),
                r['agent'].get_role_type_display(), r['agent'].employer,
                str(r['login_hrs']), str(r['coded_hrs']), str(r['total_hrs']),
            ])
        return resp

    return render(request, 'records/hours.html', {
        'rows': rows,
        'date_from': date_from,
        'date_to': date_to,
        'five_years_ago': five_years_ago,
        'today': today,
        'supervisors': supervisors,
        'selected_supervisor': supervisor_id,
        'selected_role_type': role_type_f,
        'selected_employer': employer_f,
        'selected_billing': billing_f,
        'role_type_choices': Agent.ROLE_TYPE_CHOICES,
        'employer_choices': Agent.EMPLOYER_CHOICES,
        'billing_choices': Agent.BILLING_STATUS_CHOICES,
        'count': len(rows),
    })


@login_required
def records_role_log(request):
    from datetime import date as date_cls, timedelta

    today = date_cls.today()
    five_years_ago = today.replace(year=today.year - 5)

    date_from = _parse_date(request.GET.get('from'), today - timedelta(days=30))
    date_to   = _parse_date(request.GET.get('to'), today)
    date_from = max(date_from, five_years_ago)
    date_to   = min(date_to, today)

    supervisor_id = request.GET.get('supervisor', '')

    supervisors = Agent.objects.filter(
        role_type__in=('supervisor', 'coordinator'), status='active'
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    qs = RoleHistory.objects.filter(
        effective_from__gte=date_from, effective_from__lte=date_to,
    ).select_related('agent__user', 'agent__supervisor__user', 'supervisor__user', 'changed_by').order_by('-changed_at')

    if supervisor_id:
        try:
            qs = qs.filter(agent__supervisor_id=int(supervisor_id))
        except (ValueError, TypeError):
            pass

    entries = list(qs[:1000])

    if request.GET.get('export') == '1':
        import csv as _csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="role_log_{date_from}_{date_to}.csv"'
        w = _csv.writer(resp)
        w.writerow(['Date', 'Agent', 'Supervisor', 'Role', 'Role Type', 'Employer', 'Effective From', 'Effective To', 'Changed By'])
        for e in entries:
            w.writerow([
                e.changed_at.strftime('%Y-%m-%d'),
                str(e.agent), str(e.agent.supervisor or ''),
                e.role, e.role_type, e.employer,
                str(e.effective_from), str(e.effective_to or ''),
                e.changed_by.get_full_name() if e.changed_by else '',
            ])
        return resp

    return render(request, 'records/role_log.html', {
        'entries': entries,
        'date_from': date_from,
        'date_to': date_to,
        'five_years_ago': five_years_ago,
        'today': today,
        'supervisors': supervisors,
        'selected_supervisor': supervisor_id,
        'count': len(entries),
    })


def _parse_date(val, default):
    from datetime import date as date_cls
    try:
        return date_cls.fromisoformat(val) if val else default
    except (ValueError, TypeError):
        return default


# ── Scheduled Role Changes ────────────────────────────────────────────────────

@login_required
def schedule_role_change(request, pk):
    from django.views.decorators.http import require_POST as _rp
    if request.method != 'POST':
        return redirect('agent_detail', pk=pk)

    agent = get_object_or_404(Agent, pk=pk)

    # Only one pending change at a time
    if agent.scheduled_role_changes.filter(applied_at__isnull=True, cancelled_at__isnull=True).exists():
        messages.error(request, "This agent already has a pending role change. Cancel it before scheduling a new one.")
        return redirect('agent_detail', pk=pk)

    new_role_type = request.POST.get('new_role_type', '').strip()
    effective_date_str = request.POST.get('effective_date', '').strip()

    valid_role_types = {k for k, _ in Agent.ROLE_TYPE_CHOICES}
    if new_role_type not in valid_role_types:
        messages.error(request, "Invalid role type selected.")
        return redirect('agent_detail', pk=pk)

    try:
        effective_date = date.fromisoformat(effective_date_str)
    except (ValueError, TypeError):
        messages.error(request, "Invalid effective date.")
        return redirect('agent_detail', pk=pk)

    today = timezone.localdate()
    if effective_date < today:
        messages.error(request, "Effective date cannot be in the past.")
        return redirect('agent_detail', pk=pk)

    # Optional new schedule
    days_raw = request.POST.getlist('new_shift_days')
    start_raw = request.POST.get('new_shift_start_time', '').strip()
    end_raw = request.POST.get('new_shift_end_time', '').strip()

    new_shift_days = None
    new_shift_start_time = None
    new_shift_end_time = None

    if days_raw and start_raw and end_raw:
        try:
            from datetime import time as time_cls
            new_shift_days = sorted(int(d) for d in days_raw)
            new_shift_start_time = time_cls.fromisoformat(start_raw)
            new_shift_end_time = time_cls.fromisoformat(end_raw)
        except (ValueError, TypeError):
            new_shift_days = new_shift_start_time = new_shift_end_time = None

    new_supervisor = None
    supervisor_id_raw = request.POST.get('new_supervisor_id', '').strip()
    if supervisor_id_raw:
        try:
            new_supervisor = Agent.objects.get(pk=int(supervisor_id_raw))
        except (Agent.DoesNotExist, ValueError, TypeError):
            pass

    src = ScheduledRoleChange.objects.create(
        agent=agent,
        new_role_type=new_role_type,
        effective_date=effective_date,
        new_shift_days=new_shift_days,
        new_shift_start_time=new_shift_start_time,
        new_shift_end_time=new_shift_end_time,
        new_supervisor=new_supervisor,
        scheduled_by=request.user,
    )
    _sync_pending_schedule(src)

    log_action(
        request.user,
        'Scheduled role change',
        f'Scheduled change to {src.get_new_role_type_display()} effective {effective_date}',
        agent=agent,
    )
    messages.success(request, f"Role change to {src.get_new_role_type_display()} scheduled for {effective_date.strftime('%b %d, %Y')}.")
    return redirect('agent_detail', pk=pk)


@login_required
def cancel_role_change(request, pk):
    if request.method != 'POST':
        return redirect('agent_list')

    src = get_object_or_404(ScheduledRoleChange, pk=pk)
    if not src.is_pending:
        messages.error(request, "This role change has already been applied or cancelled.")
        return redirect('agent_detail', pk=src.agent_id)

    # Undo pre-created templates from _sync_pending_schedule
    if src.new_shift_days:
        ShiftTemplate.objects.filter(agent=src.agent, effective_from=src.effective_date).delete()
        ShiftTemplate.objects.filter(
            agent=src.agent, effective_until=src.effective_date
        ).update(effective_until=None)

    src.cancelled_at = timezone.now()
    src.cancelled_by = request.user
    src.save(update_fields=['cancelled_at', 'cancelled_by'])

    log_action(
        request.user,
        'Cancelled scheduled role change',
        f'Cancelled planned change to {src.get_new_role_type_display()} (was scheduled for {src.effective_date})',
        agent=src.agent,
    )
    messages.success(request, "Scheduled role change cancelled.")
    return redirect('agent_detail', pk=src.agent_id)


# ── Agent self-service views ───────────────────────────────────────────────────

_DAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']


def _agent_week_start(request):
    today = timezone.localdate()
    default = today - timedelta(days=today.weekday())
    raw = request.GET.get('week_start')
    try:
        ws = date.fromisoformat(raw) if raw else default
        ws -= timedelta(days=ws.weekday())
        return ws
    except (ValueError, TypeError):
        return default


def _best_shift_template(all_templates, agent_id, d):
    dow = d.weekday()
    best = None
    for t in all_templates:
        if t.agent_id != agent_id or t.day_of_week != dow:
            continue
        if t.effective_from is not None and t.effective_from > d:
            continue
        if t.effective_until is not None and t.effective_until < d:
            continue
        if best is None or (t.effective_from or date.min) > (best.effective_from or date.min):
            best = t
    return best


from wfm.constants import PORTAL_ADMIN_TYPES as _PORTAL_ADMIN_TYPES


def _is_portal_user(agent):
    return agent.role == 'agent' or (agent.role == 'admin' and agent.role_type in _PORTAL_ADMIN_TYPES)


@login_required
def agent_my_shifts(request):
    try:
        agent = request.user.agent
    except Exception:
        return redirect('dashboard')
    if not _is_portal_user(agent):
        return redirect('dashboard')

    today = timezone.localdate()
    week_start = _agent_week_start(request)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]

    overrides = {s.date: s for s in Shift.objects.filter(agent=agent, date__in=week_dates)}
    all_templates = list(ShiftTemplate.objects.filter(agent=agent))

    days = []
    for i, d in enumerate(week_dates):
        override = overrides.get(d)
        tmpl = None if override else _best_shift_template(all_templates, agent.pk, d)
        src = override or tmpl
        days.append({
            'name': _DAY_NAMES[i],
            'date': d,
            'is_today': d == today,
            'is_off': src.is_off if src else None,
            'start': src.start_time.strftime('%H:%M') if src and src.start_time and not src.is_off else '',
            'end': src.end_time.strftime('%H:%M') if src and src.end_time and not src.is_off else '',
            'no_data': src is None,
        })

    return render(request, 'agent/my_shifts.html', {
        'agent': agent,
        'days': days,
        'week_start': week_start,
        'week_end': week_dates[-1],
        'today': today,
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
    })


@login_required
def agent_my_ot_shifts(request):
    try:
        agent = request.user.agent
    except Exception:
        return redirect('dashboard')
    if not _is_portal_user(agent):
        return redirect('dashboard')

    today = timezone.localdate()
    week_start = _agent_week_start(request)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    ot_shifts = list(
        OvertimeShift.objects.filter(agent=agent, date__in=week_dates).order_by('date', 'start_time')
    )

    # Cancellation requests: mark responses seen, attach state to each shift
    OTCancellationRequest.objects.filter(requester=agent, requester_read=False).update(requester_read=True)
    cr_map = {}
    for cr in OTCancellationRequest.objects.filter(shift__in=ot_shifts).order_by('submitted_at'):
        cr_map[cr.shift_id] = cr  # latest per shift wins
    for s in ot_shifts:
        s.my_cancel_req = cr_map.get(s.pk)
        s.can_request_cancel = (
            s.status == 'pending' and s.date >= today
            and not (s.my_cancel_req and s.my_cancel_req.status == 'pending')
        )

    # Group shifts by date
    ot_days = []
    from collections import defaultdict as _dd
    by_date = _dd(list)
    for s in ot_shifts:
        by_date[s.date].append(s)
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for d in week_dates:
        if d in by_date:
            ot_days.append({
                'date': d,
                'day_name': day_names[d.weekday()],
                'is_today': d == today,
                'shifts': by_date[d],
            })

    return render(request, 'agent/my_ot_shifts.html', {
        'agent': agent,
        'ot_days': ot_days,
        'week_start': week_start,
        'week_end': week_dates[-1],
        'today': today,
        'prev_week': (week_start - timedelta(days=7)).isoformat(),
        'next_week': (week_start + timedelta(days=7)).isoformat(),
    })


# ── Agent Request views ────────────────────────────────────────────────────────

def _fill_request_from_post(ar, request):
    """Populate the per-type fields of an AgentRequest from POST data.

    Returns an error message to show the user, or None on success.
    """
    from django.utils.dateparse import parse_time

    req_type = ar.request_type
    if req_type == 'coding':
        # Parse to time objects so summary()/strftime work on the fresh instance
        ar.coding_date = request.POST.get('coding_date') or None
        ar.coding_start_time = parse_time(request.POST.get('coding_start_time') or '')
        ar.coding_end_time = parse_time(request.POST.get('coding_end_time') or '')
    elif req_type == 'vacation':
        ar.vacation_start = request.POST.get('vacation_start') or None
        ar.vacation_end = request.POST.get('vacation_end') or request.POST.get('vacation_start') or None
    elif req_type == 'day_off_change':
        ar.day_off_type = request.POST.get('day_off_type', '')
        try:
            ar.current_day_off = int(request.POST.get('current_day_off', ''))
        except (ValueError, TypeError):
            pass
        try:
            ar.requested_day_off = int(request.POST.get('requested_day_off', ''))
        except (ValueError, TypeError):
            pass
        ar.effective_date = request.POST.get('effective_date') or None
    elif req_type == 'vto':
        ar.vto_date = request.POST.get('vto_date') or None
    elif req_type == 'loa':
        ar.loa_start = request.POST.get('loa_start') or None
        ar.loa_end = request.POST.get('loa_end') or None
    elif req_type == 'schedule_change':
        days_raw = request.POST.getlist('schedule_change_days')
        if not days_raw:
            return "Please select at least one working day for the schedule change."
        if not request.POST.get('schedule_new_start_time') or not request.POST.get('schedule_new_end_time'):
            return "Please enter both a start time and end time for the new schedule."
        if not request.POST.get('schedule_effective_date'):
            return "Please enter an effective date for the schedule change."
        ar.current_schedule_desc = request.POST.get('current_schedule_desc', '').strip()
        ar.requested_schedule_desc = request.POST.get('requested_schedule_desc', '').strip()
        ar.schedule_new_start_time = request.POST.get('schedule_new_start_time')
        ar.schedule_new_end_time = request.POST.get('schedule_new_end_time')
        ar.schedule_effective_date = request.POST.get('schedule_effective_date')
        try:
            ar.schedule_change_days = [int(d) for d in days_raw]
        except (ValueError, TypeError):
            return "Invalid day selection. Please try again."
    return None


def _request_action_block_reason(ar, viewer):
    """Why `viewer` (an Agent or None) may NOT action this request; None if allowed.

    Agent-submitted requests stay actionable by any staff user, except the
    requester themselves. Staff-submitted requests are only actionable by the
    requester's assigned supervisor (snapshotted at submission).
    """
    if viewer is not None and ar.agent_id == viewer.pk:
        return "You cannot action your own request. Another approver is required."
    if not ar.is_staff_request:
        return None
    sup = ar.assigned_supervisor
    if sup is None:
        return ("This staff request has no assigned supervisor. "
                "An admin must set a supervisor on the requester's profile.")
    if ar.agent_id == sup.pk:
        return f"{sup} is their own assigned supervisor — another approver is required."
    if viewer is None or viewer.pk != sup.pk:
        return f"Only {sup} can action this request."
    return None


@login_required
def agent_my_requests(request):
    try:
        agent = request.user.agent
    except Exception:
        return redirect('dashboard')
    if not _is_portal_user(agent):
        return redirect('dashboard')

    if request.method == 'POST':
        req_type = request.POST.get('request_type', '').strip()
        if not req_type:
            messages.error(request, "Please select a request type.")
            return redirect('agent_my_requests')

        ar = AgentRequest(agent=agent, request_type=req_type,
                          notes=request.POST.get('notes', '').strip(),
                          supervisor_read=False, agent_read=True)

        err = _fill_request_from_post(ar, request)
        if err:
            messages.error(request, err)
            return redirect('agent_my_requests')

        ar.save()
        log_action(request.user, f'Submitted agent request: {ar.get_request_type_display()}',
                   ar.summary(), agent=agent)
        messages.success(request, "Your request has been submitted and is pending review.")
        return redirect('agent_my_requests')

    # Mark agent's unread responses as seen
    AgentRequest.objects.filter(agent=agent, agent_read=False).update(agent_read=True)
    reqs = AgentRequest.objects.filter(agent=agent).order_by('-submitted_at')

    return render(request, 'agent/my_requests.html', {
        'agent': agent,
        'requests': reqs,
        'today': timezone.localdate(),
    })


@login_required
def staff_my_requests(request):
    """Self-service requests for staff users (supervisors, coordinators, admins).

    Mirrors the agent portal experience but routes approval to the staff
    member's assigned supervisor.
    """
    try:
        agent = request.user.agent
    except Exception:
        return redirect('dashboard')
    if _is_portal_user(agent):
        return redirect('agent_my_requests')

    supervisor = agent.supervisor

    if request.method == 'POST':
        if supervisor is None:
            messages.error(request, "You need a supervisor assigned to your profile before you can "
                                    "submit requests. Please contact an admin.")
            return redirect('staff_my_requests')

        req_type = request.POST.get('request_type', '').strip()
        if not req_type:
            messages.error(request, "Please select a request type.")
            return redirect('staff_my_requests')

        ar = AgentRequest(agent=agent, request_type=req_type,
                          notes=request.POST.get('notes', '').strip(),
                          supervisor_read=False, agent_read=True,
                          is_staff_request=True, assigned_supervisor=supervisor)

        err = _fill_request_from_post(ar, request)
        if err:
            messages.error(request, err)
            return redirect('staff_my_requests')

        ar.save()
        log_action(request.user, f'Submitted staff request: {ar.get_request_type_display()}',
                   ar.summary(), agent=agent)
        messages.success(request, "Your request has been submitted and is pending review "
                                  f"by {supervisor}.")
        return redirect('staff_my_requests')

    # Mark this staff member's unread responses as seen
    AgentRequest.objects.filter(agent=agent, agent_read=False).update(agent_read=True)
    reqs = AgentRequest.objects.filter(agent=agent).order_by('-submitted_at')

    return render(request, 'scheduling/staff_my_requests.html', {
        'agent': agent,
        'supervisor': supervisor,
        'requests': reqs,
        'today': timezone.localdate(),
    })


@login_required
def requests_list(request):
    # Agents are blocked by middleware; this is staff-only
    try:
        viewer = request.user.agent
    except Exception:
        viewer = None

    qs = AgentRequest.objects.select_related(
        'agent__user', 'agent__supervisor__user', 'reviewed_by', 'done_by'
    ).order_by('-submitted_at')

    # Build supervisor list for filter dropdown (agents with supervisor role_type)
    supervisors = Agent.objects.filter(
        role='admin', role_type='supervisor'
    ).select_related('user').order_by('agent_name')

    # Filters
    agent_search = request.GET.get('agent_search', '').strip()
    type_filter = request.GET.get('type', '').strip()
    status_filter = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if 'supervisor' in request.GET:
        supervisor_filter = request.GET.get('supervisor', '').strip()
        request.session['supervisor_filter'] = supervisor_filter
    else:
        supervisor_filter = request.session.get('supervisor_filter', '')

    if supervisor_filter:
        try:
            qs = qs.filter(agent__supervisor_id=int(supervisor_filter))
        except ValueError:
            pass

    if agent_search:
        from django.db.models import Q as _Q
        qs = qs.filter(
            _Q(agent__user__first_name__icontains=agent_search) |
            _Q(agent__user__last_name__icontains=agent_search) |
            _Q(agent__agent_name__icontains=agent_search)
        )
    if type_filter:
        qs = qs.filter(request_type=type_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        try:
            qs = qs.filter(submitted_at__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(submitted_at__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    # Mark visible pending requests as seen by supervisor. Staff requests only
    # count as seen when their assigned supervisor views the list — other staff
    # opening the page must not clear the approver's badge.
    mark_read = qs.filter(status='pending', supervisor_read=False)
    if viewer is not None:
        mark_read = mark_read.filter(Q(is_staff_request=False) | Q(assigned_supervisor=viewer))
    else:
        mark_read = mark_read.filter(is_staff_request=False)
    mark_read.update(supervisor_read=True)

    rows = list(qs)
    pending  = [r for r in rows if r.status == 'pending']
    approved = [r for r in rows if r.status == 'approved']
    rejected = [r for r in rows if r.status == 'rejected']
    done     = [r for r in rows if r.status == 'done']

    return render(request, 'scheduling/requests_list.html', {
        'pending': pending,
        'approved': approved,
        'rejected': rejected,
        'done': done,
        'type_choices': AgentRequest.REQUEST_TYPE_CHOICES,
        'status_choices': AgentRequest.STATUS_CHOICES,
        'supervisors': supervisors,
        'filters': {
            'agent_search': agent_search,
            'type': type_filter,
            'status': status_filter,
            'date_from': date_from,
            'date_to': date_to,
            'supervisor': supervisor_filter,
        },
    })


@login_required
def request_detail(request, pk):
    ar = get_object_or_404(AgentRequest, pk=pk)
    viewer = None
    try:
        viewer = request.user.agent
        if viewer.role == 'agent':
            return redirect('agent_my_requests')
    except Exception:
        pass
    vac = None
    if ar.request_type == 'vacation' and ar.vacation_start and ar.vacation_end:
        from nomina.views import vacation_request_check
        accrued, used, remaining, new_days, overdraw = vacation_request_check(
            ar.agent, ar.vacation_start, ar.vacation_end)
        is_super = request.user.is_superuser or getattr(viewer, 'is_super_admin', False)
        vac = {
            'accrued': accrued, 'used': used, 'remaining': remaining,
            'req_days': new_days, 'overdraw': overdraw,
            'block_for_viewer': overdraw and not is_super,   # supervisor can't approve
        }
    return render(request, 'scheduling/request_detail.html', {
        'ar': ar,
        'DAY_NAMES': _DAY_NAMES,
        'action_block': _request_action_block_reason(ar, viewer),
        'vac': vac,
    })


@login_required
def request_approve(request, pk):
    if request.method != 'POST':
        return redirect('requests_list')

    ar = get_object_or_404(AgentRequest, pk=pk)
    viewer = None
    try:
        viewer = request.user.agent
        if viewer.role == 'agent':
            return redirect('agent_my_requests')
    except Exception:
        pass

    block = _request_action_block_reason(ar, viewer)
    if block:
        messages.error(request, block)
        return redirect('request_detail', pk=pk)

    if ar.status != 'pending':
        messages.error(request, "This request has already been reviewed.")
        return redirect('request_detail', pk=pk)

    agent = ar.agent
    actions = []

    # Coding requests intentionally have NO auto-apply: agents only estimate
    # their missed time, so the supervisor pulls the exact amount and enters
    # the coding manually. Approval/Done only track the request's status.
    if ar.request_type == 'vacation' and ar.vacation_start and ar.vacation_end:
        from adherence.models import AdherenceRecord
        from nomina.views import vacation_request_check
        is_super = request.user.is_superuser or getattr(viewer, 'is_super_admin', False)
        _acc, _used, remaining, new_days, overdraw = vacation_request_check(
            agent, ar.vacation_start, ar.vacation_end)
        # Over-balance vacation can only be approved by a super admin (David/Jhon).
        if overdraw and not is_super:
            messages.error(
                request,
                f"{agent} is exceeding their available vacation days "
                f"({new_days} requested, {remaining} remaining). Only a super admin "
                f"(David or Jhon) can approve this — please contact them."
            )
            return redirect('request_detail', pk=pk)
        d = ar.vacation_start
        count = 0
        while d <= ar.vacation_end:
            AdherenceRecord.objects.update_or_create(
                agent=agent, date=d, defaults={'status': 'V'}
            )
            d += timedelta(days=1)
            count += 1
        actions.append(f"Set V (Vacation) for {count} day(s): {ar.vacation_start} – {ar.vacation_end}")

    elif ar.request_type == 'vto' and ar.vto_date:
        from adherence.models import AdherenceRecord
        AdherenceRecord.objects.update_or_create(
            agent=agent, date=ar.vto_date, defaults={'status': 'VTO'}
        )
        actions.append(f"Set VTO status for {ar.vto_date}")

    elif ar.request_type == 'day_off_change' and ar.effective_date:
        if ar.day_off_type == 'one_time':
            Shift.objects.update_or_create(
                agent=agent, date=ar.effective_date,
                defaults={'start_time': '00:00', 'end_time': '00:00', 'is_off': True}
            )
            actions.append(f"Shift override: {ar.effective_date} marked as day off")
            if ar.current_day_off is not None:
                week_mon = ar.effective_date - timedelta(days=ar.effective_date.weekday())
                working_date = week_mon + timedelta(days=ar.current_day_off)
                work_tmpl = ShiftTemplate.objects.filter(
                    agent=agent, is_off=False, start_time__isnull=False
                ).first()
                if work_tmpl:
                    Shift.objects.update_or_create(
                        agent=agent, date=working_date,
                        defaults={
                            'start_time': work_tmpl.start_time,
                            'end_time': work_tmpl.end_time,
                            'is_off': False,
                        }
                    )
                    actions.append(f"Shift override: {working_date} set as working ({work_tmpl.start_time}–{work_tmpl.end_time})")

        elif ar.day_off_type == 'permanent':
            eff_mon = ar.effective_date - timedelta(days=ar.effective_date.weekday())
            if ar.requested_day_off is not None:
                _save_shift_template(agent, ar.requested_day_off, eff_mon, True, None, None, '')
                actions.append(f"ShiftTemplate: {_DAY_NAMES[ar.requested_day_off]} set as day off from {eff_mon}")
            if ar.current_day_off is not None:
                work_tmpl = ShiftTemplate.objects.filter(
                    agent=agent, is_off=False, start_time__isnull=False
                ).first()
                if work_tmpl:
                    _save_shift_template(
                        agent, ar.current_day_off, eff_mon, False,
                        work_tmpl.start_time.strftime('%H:%M'),
                        work_tmpl.end_time.strftime('%H:%M'), ''
                    )
                    actions.append(f"ShiftTemplate: {_DAY_NAMES[ar.current_day_off]} set as working from {eff_mon}")

    elif ar.request_type == 'loa' and ar.loa_start and ar.loa_end:
        from adherence.models import AdherenceRecord
        d = ar.loa_start
        count = 0
        while d <= ar.loa_end:
            override = Shift.objects.filter(agent=agent, date=d).first()
            if override:
                day_is_off = override.is_off
            else:
                tmpl = ShiftTemplate.objects.filter(
                    agent=agent,
                    day_of_week=d.weekday(),
                ).filter(
                    Q(effective_from__isnull=True) | Q(effective_from__lte=d)
                ).filter(
                    Q(effective_until__isnull=True) | Q(effective_until__gte=d)
                ).order_by(F('effective_from').desc(nulls_last=True)).first()
                day_is_off = tmpl.is_off if tmpl else True
            if not day_is_off:
                AdherenceRecord.objects.update_or_create(
                    agent=agent, date=d, defaults={'status': 'LOA'}
                )
                count += 1
            d += timedelta(days=1)
        actions.append(f"Set LOA status for {count} scheduled working day(s): {ar.loa_start} – {ar.loa_end}")

    elif ar.request_type == 'schedule_change':
        if (ar.schedule_new_start_time and ar.schedule_new_end_time
                and ar.schedule_change_days and ar.schedule_effective_date):
            start_str = ar.schedule_new_start_time.strftime('%H:%M')
            end_str = ar.schedule_new_end_time.strftime('%H:%M')
            for dow in ar.schedule_change_days:
                _save_shift_template(
                    agent, dow, ar.schedule_effective_date, False, start_str, end_str, ''
                )
            day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            days_label = ', '.join(day_names[d] for d in sorted(ar.schedule_change_days))
            actions.append(
                f"Schedule updated: {days_label} {start_str}–{end_str} from {ar.schedule_effective_date}"
            )
        else:
            actions.append("Manual action required: please update the agent's schedule in the system.")

    ar.status = 'approved'
    ar.reviewed_by = request.user
    ar.reviewed_at = timezone.now()
    ar.auto_action_log = '\n'.join(actions)
    ar.agent_read = False
    ar.save()

    kind = 'staff' if ar.is_staff_request else 'agent'
    log_action(request.user, f'Approved {kind} request: {ar.get_request_type_display()}',
               ar.summary(), agent=agent)

    if ar.request_type == 'schedule_change' and 'Manual action required' in ar.auto_action_log:
        messages.warning(request, "Request approved. Manual update required — please update the agent's schedule.")
    else:
        messages.success(request, "Request approved.")
    return redirect('request_detail', pk=pk)


@login_required
def request_reject(request, pk):
    if request.method != 'POST':
        return redirect('requests_list')

    ar = get_object_or_404(AgentRequest, pk=pk)
    viewer = None
    try:
        viewer = request.user.agent
        if viewer.role == 'agent':
            return redirect('agent_my_requests')
    except Exception:
        pass

    block = _request_action_block_reason(ar, viewer)
    if block:
        messages.error(request, block)
        return redirect('request_detail', pk=pk)

    if ar.status != 'pending':
        messages.error(request, "This request has already been reviewed.")
        return redirect('request_detail', pk=pk)

    ar.status = 'rejected'
    ar.reviewed_by = request.user
    ar.reviewed_at = timezone.now()
    ar.rejection_reason = request.POST.get('rejection_reason', '').strip()
    ar.agent_read = False
    ar.save()

    kind = 'staff' if ar.is_staff_request else 'agent'
    log_action(request.user, f'Rejected {kind} request: {ar.get_request_type_display()}',
               ar.summary(), agent=ar.agent)
    messages.success(request, "Request rejected.")
    return redirect('request_detail', pk=pk)


@login_required
def request_mark_done(request, pk):
    if request.method != 'POST':
        return redirect('requests_list')

    ar = get_object_or_404(AgentRequest, pk=pk)
    viewer = None
    try:
        viewer = request.user.agent
        if viewer.role == 'agent':
            return redirect('agent_my_requests')
    except Exception:
        pass

    block = _request_action_block_reason(ar, viewer)
    if block:
        messages.error(request, block)
        return redirect('request_detail', pk=pk)

    if ar.status != 'approved':
        messages.error(request, "Only approved requests can be marked as done.")
        return redirect('request_detail', pk=pk)

    ar.status = 'done'
    ar.done_by = request.user
    ar.done_at = timezone.now()
    ar.save()

    kind = 'staff' if ar.is_staff_request else 'agent'
    log_action(request.user, f'Marked {kind} request as done: {ar.get_request_type_display()}',
               ar.summary(), agent=ar.agent)
    messages.success(request, "Request marked as done.")
    return redirect('request_detail', pk=pk)


# ── Agent Separation ──────────────────────────────────────────────────────────

def _finalize_separation(agent, sep, remove_date, last_day, separation_type, user):
    """Apply all side-effects of a finalized separation."""
    agent.status = 'inactive'
    agent.track_attendance = False
    agent.termination_date = last_day
    agent.save(update_fields=['status', 'track_attendance', 'termination_date'])

    # Update finalized_by/finalized_at on sep record
    sep.finalized_by = user
    sep.finalized_at = timezone.now()
    sep.save(update_fields=['finalized_by', 'finalized_at'])

    ep_reason = AgentSeparation._EP_REASON_MAP.get(separation_type, 'other')
    ep = agent.employment_periods.filter(end_date__isnull=True).order_by('-start_date').first()
    if ep:
        ep.end_date = remove_date
        ep.reason_ended = ep_reason
        if sep.notes and not ep.notes:
            ep.notes = sep.notes
        ep.save()

    OvertimeShift.objects.filter(
        agent=agent, date__gte=remove_date, status='pending'
    ).update(status='cancelled', cancellation_reason='Auto-cancelled due to agent separation')

    AgentRequest.objects.filter(agent=agent, status='pending').update(
        status='rejected',
        rejection_reason='Auto-rejected due to agent separation',
        reviewed_by=user,
        reviewed_at=timezone.now(),
    )

    ScheduledRoleChange.objects.filter(
        agent=agent, effective_date__gte=remove_date,
        applied_at__isnull=True, cancelled_at__isnull=True
    ).update(cancelled_at=timezone.now())

    _auto_code_separation_week(agent, last_day, separation_type, user)


@login_required
@require_POST
def process_separation(request, pk):
    agent = get_object_or_404(Agent.objects.prefetch_related('separations'), pk=pk)

    try:
        viewer = request.user.agent
        if viewer.role == 'agent':
            messages.error(request, "Access denied.")
            return redirect('agent_detail', pk=pk)
    except Exception:
        pass

    # Already has a non-cancelled separation?
    existing = agent.separation  # uses property
    if existing:
        messages.error(request, "This agent already has an active separation. Use Update to modify it.")
        return redirect('agent_detail', pk=pk)

    sep_status      = request.POST.get('separation_status', '').strip()
    separation_type = request.POST.get('separation_type', '').strip()
    last_day_str    = request.POST.get('last_day_worked', '').strip()
    remove_str      = request.POST.get('remove_from_adherence_date', '').strip()
    notes           = request.POST.get('notes', '').strip()
    confirmed       = request.POST.get('confirm', '')

    errors = []
    if sep_status not in ('in_progress', 'finalized'):
        errors.append("Separation Status is required.")
    if not separation_type:
        errors.append("Separation type is required.")
    if not last_day_str:
        errors.append("Last Day Worked is required.")
    if sep_status == 'finalized' and not remove_str:
        errors.append("Remove from Adherence date is required when finalizing.")
    if not confirmed:
        errors.append("Please check the confirmation checkbox.")

    valid_types = [c[0] for c in AgentSeparation.SEPARATION_TYPE_CHOICES]
    if separation_type and separation_type not in valid_types:
        errors.append("Invalid separation type.")

    last_day = None
    if last_day_str:
        try:
            last_day = date.fromisoformat(last_day_str)
        except ValueError:
            errors.append("Invalid Last Day Worked date.")

    remove_date = None
    if remove_str:
        try:
            remove_date = date.fromisoformat(remove_str)
        except ValueError:
            errors.append("Invalid Remove from Adherence date.")

    if errors:
        for e in errors:
            messages.error(request, e)
        return redirect('agent_detail', pk=pk)

    sep = AgentSeparation.objects.create(
        agent=agent,
        status=sep_status,
        separation_type=separation_type,
        last_day_worked=last_day,
        remove_from_adherence_date=remove_date,
        notes=notes,
        processed_by=request.user,
    )

    sep_display = dict(AgentSeparation.SEPARATION_TYPE_CHOICES).get(separation_type, separation_type)

    if sep_status == 'finalized':
        _finalize_separation(agent, sep, remove_date, last_day, separation_type, request.user)
        log_action(request.user, 'Processed agent separation (Finalized)',
                   f'{agent} — {sep_display} | Last day: {last_day} | Remove from adherence: {remove_date}',
                   agent=agent)
        messages.success(request, f"Separation finalized for {agent}.")
    else:
        log_action(request.user, 'Opened separation (In Progress)',
                   f'{agent} — {sep_display} | Last day: {last_day}',
                   agent=agent)
        messages.success(request, f"Separation marked In Progress for {agent}. Agent remains active for documentation.")

    return redirect('agent_detail', pk=pk)


@login_required
@require_POST
def update_separation(request, pk):
    agent = get_object_or_404(Agent.objects.prefetch_related('separations'), pk=pk)

    try:
        viewer = request.user.agent
        if viewer.role == 'agent':
            messages.error(request, "Access denied.")
            return redirect('agent_detail', pk=pk)
    except Exception:
        pass

    sep = agent.separation
    if not sep:
        messages.error(request, "No active separation found.")
        return redirect('agent_detail', pk=pk)

    action = request.POST.get('action', '').strip()  # 'finalize' or 'cancel'

    if action == 'cancel':
        sep.status = 'cancelled'
        sep.save(update_fields=['status'])
        sep_display = sep.get_separation_type_display()
        log_action(request.user, 'Cancelled separation (In Progress)',
                   f'{agent} — {sep_display} separation cancelled', agent=agent)
        messages.success(request, f"Separation for {agent} has been cancelled. Agent remains active.")
        return redirect('agent_detail', pk=pk)

    if action == 'finalize':
        remove_str = request.POST.get('remove_from_adherence_date', '').strip()
        confirmed  = request.POST.get('confirm', '')

        errors = []
        if not remove_str:
            errors.append("Remove from Adherence date is required to finalize.")
        if not confirmed:
            errors.append("Please check the confirmation checkbox.")

        remove_date = None
        if remove_str:
            try:
                remove_date = date.fromisoformat(remove_str)
            except ValueError:
                errors.append("Invalid date.")

        if errors:
            for e in errors:
                messages.error(request, e)
            return redirect('agent_detail', pk=pk)

        sep.status = 'finalized'
        sep.remove_from_adherence_date = remove_date
        sep.save(update_fields=['status', 'remove_from_adherence_date'])

        _finalize_separation(agent, sep, remove_date, sep.last_day_worked, sep.separation_type, request.user)

        sep_display = sep.get_separation_type_display()
        log_action(request.user, 'Finalized agent separation',
                   f'{agent} — {sep_display} | Remove from adherence: {remove_date}', agent=agent)
        messages.success(request, f"Separation finalized for {agent}.")
        return redirect('agent_detail', pk=pk)

    messages.error(request, "Invalid action.")
    return redirect('agent_detail', pk=pk)


def _auto_code_separation_week(agent, last_day, separation_type, user):
    """Create AdherenceRecord entries for remaining days in the separation week."""
    from adherence.models import AdherenceRecord

    # Determine status to apply (None = skip)
    status_map = {
        'quit':       'Quit',
        'terminated': 'Quit',   # spec: use Quit code, note Terminated
        'abandonment':'NCNS',
        # contract_end and resigned_notice: don't auto-code
    }
    status_code = status_map.get(separation_type)
    if not status_code:
        return

    notes = 'Terminated' if separation_type == 'terminated' else ''

    week_start = last_day - timedelta(days=last_day.weekday())
    week_end   = week_start + timedelta(days=6)

    # Only code days AFTER the last day worked, up to end of that week
    days_to_code = [
        week_start + timedelta(days=i)
        for i in range(7)
        if (week_start + timedelta(days=i)) > last_day
    ]

    for d in days_to_code:
        AdherenceRecord.objects.update_or_create(
            agent=agent,
            date=d,
            defaults={'status': status_code, 'notes': notes},
        )


@login_required
def records_separations(request):
    sep_type_f    = request.GET.get('sep_type', '').strip()
    sep_status_f  = request.GET.get('sep_status', '').strip()
    date_from_str = request.GET.get('from', '').strip()
    date_to_str   = request.GET.get('to', '').strip()

    qs = AgentSeparation.objects.select_related(
        'agent__user', 'agent__supervisor__user', 'processed_by', 'finalized_by'
    ).order_by('-processed_at')

    if sep_type_f:
        qs = qs.filter(separation_type=sep_type_f)
    if sep_status_f:
        qs = qs.filter(status=sep_status_f)
    if date_from_str:
        try:
            qs = qs.filter(last_day_worked__gte=date.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(last_day_worked__lte=date.fromisoformat(date_to_str))
        except ValueError:
            pass

    return render(request, 'records/separations.html', {
        'separations': list(qs),
        'separation_type_choices': AgentSeparation.SEPARATION_TYPE_CHOICES,
        'separation_status_choices': AgentSeparation.STATUS_CHOICES,
        'selected_type': sep_type_f,
        'selected_status': sep_status_f,
        'date_from': date_from_str,
        'date_to': date_to_str,
    })


def agent_inactive(request):
    """Shown to inactive agents who try to log in to the agent portal."""
    from django.contrib.auth import logout
    return render(request, 'agent/inactive.html', {})


@login_required
def agent_search(request):
    """Quick agent search for the nav bar — returns JSON list of up to 10 matches."""
    from django.db.models import Q as _Q
    from django.http import JsonResponse as _JsonResponse
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return _JsonResponse({'results': []})

    agents = Agent.objects.filter(
        _Q(agent_name__icontains=q) |
        _Q(employee_id__icontains=q) |
        _Q(user__first_name__icontains=q) |
        _Q(user__last_name__icontains=q) |
        _Q(five9_profiles__five9_username__icontains=q)
    ).filter(status='active').distinct().select_related('user')[:10]

    results = [
        {
            'pk': a.pk,
            'name': a.agent_name or a.user.get_full_name() or a.user.username,
            'employee_id': a.employee_id or '',
            'url': f'/agents/{a.pk}/',
        }
        for a in agents
    ]
    return _JsonResponse({'results': results})
